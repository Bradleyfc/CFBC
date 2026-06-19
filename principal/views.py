import logging
from typing import override
from django.contrib.auth.forms import UserCreationForm
from django.views import View
from django.core.mail import send_mail
from django.conf import settings
from django.views.generic import ListView, DetailView, TemplateView, CreateView, UpdateView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.contrib.auth import logout
from django.contrib import messages
from django.utils import timezone
from .forms import (
    CustomUserCreationForm, CourseForm, CalificacionesForm, NotaIndividualFormSet, NotaIndividualFormSetCustom,
    FormularioAplicacionForm, PreguntaFormularioForm, OpcionRespuestaForm,
    OpcionRespuestaFormSet, PreguntaFormularioFormSet, RespuestaEstudianteForm,
    ReglamentoCursoForm, ArticuloReglamentoFormSet,
    ReglamentoGeneralForm, ArticuloReglamentoGeneralFormSet,
)
from django.contrib.auth.models import Group, User
from django.db.models import Q, Max, Case, When, IntegerField
from datetime import date, datetime
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
import unicodedata
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from accounts.models import Registro
from blog.models import Noticia, Categoria
try:
    from course_documents.indicator_service import ContentIndicatorService
except ImportError:
    ContentIndicatorService = None
from .models import (
    CursoAcademico, Curso, Matriculas, Calificaciones, Asistencia,
    FormularioAplicacion, PreguntaFormulario, OpcionRespuesta, SolicitudInscripcion, RespuestaEstudiante, NotaIndividual,
    ReglamentoCurso, ArticuloReglamento,
    ReglamentoGeneral, ArticuloReglamentoGeneral,
)
from course_documents.mixins import DocumentsProfileMixin, DocumentsCourseMixin

logger = logging.getLogger(__name__)

# Create your views here.

class UsuariosRegistradosView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Registro
    template_name = 'usuarios_registrados.html'
    context_object_name = 'registros'
    paginate_by = 100  # Mostrar 100 usuarios por página como en el admin de Django
    
    def test_func(self):
        return self.request.user.groups.filter(name='Secretaría').exists()

    def get_queryset(self):
        queryset = Registro.objects.all().select_related('user').prefetch_related('user__groups')
        
        # Obtener parámetros de filtro
        grupo = self.request.GET.get('grupo')
        search = self.request.GET.get('search')
        anio = self.request.GET.get('anio')
        fecha_desde = self.request.GET.get('fecha_desde')
        fecha_hasta = self.request.GET.get('fecha_hasta')
        orden = self.request.GET.get('orden', 'desc')

        # Filtrar por grupo
        if grupo and grupo != 'todos':
            queryset = queryset.filter(user__groups__name=grupo)

        # Filtrar por búsqueda de texto
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__username__icontains=search) |
                Q(user__email__icontains=search) |
                Q(carnet__icontains=search)
            )

        # Filtrar por año de registro
        if anio:
            try:
                queryset = queryset.filter(user__date_joined__year=int(anio))
            except ValueError:
                pass

        # Filtrar por rango de fechas
        if fecha_desde:
            try:
                from datetime import datetime
                queryset = queryset.filter(user__date_joined__date__gte=fecha_desde)
            except Exception:
                pass
        if fecha_hasta:
            try:
                queryset = queryset.filter(user__date_joined__date__lte=fecha_hasta)
            except Exception:
                pass

        # Ordenar por fecha de registro.
        # En PostgreSQL, distinct() con order_by sobre campo relacionado causa conflictos.
        # Solución: obtener IDs únicos primero, luego ordenar.
        ids = queryset.values_list('id', flat=True).distinct()
        if orden == 'asc':
            return Registro.objects.filter(id__in=ids).select_related('user').prefetch_related('user__groups').order_by('user__date_joined', 'id')
        else:
            return Registro.objects.filter(id__in=ids).select_related('user').prefetch_related('user__groups').order_by('-user__date_joined', 'id')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        from django.contrib.auth.models import Group
        context['grupos_disponibles'] = Group.objects.all().order_by('name')
        context['grupo_seleccionado'] = self.request.GET.get('grupo', 'todos')
        
        context['estadisticas_grupos'] = {}
        for grupo in context['grupos_disponibles']:
            count = Registro.objects.filter(user__groups=grupo).count()
            context['estadisticas_grupos'][grupo.name] = count
        
        context['total_usuarios'] = Registro.objects.count()

        # Años disponibles para el filtro (años con registros)
        from django.db.models.functions import ExtractYear
        anios = (
            Registro.objects.annotate(anio=ExtractYear('user__date_joined'))
            .values_list('anio', flat=True)
            .distinct()
            .order_by('-anio')
        )
        context['anios_disponibles'] = [a for a in anios if a]
        context['anio_seleccionado'] = self.request.GET.get('anio', '')
        context['fecha_desde'] = self.request.GET.get('fecha_desde', '')
        context['fecha_hasta'] = self.request.GET.get('fecha_hasta', '')
        context['orden'] = self.request.GET.get('orden', 'desc')

        return context


@login_required
def export_usuarios_excel(request):
    search_query = request.GET.get('search', '')
    registros = Registro.objects.filter(user__groups__name='Estudiantes')

    if search_query:
        registros = registros.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(carnet__icontains=search_query)
        ).distinct()

    context = {
        'registros': registros
    }
    excel_file = generate_excel(context)
    response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios_registrados.xlsx"'
    return response

@login_required
def export_matriculas_pdf(request):
    curso_academico_id = request.GET.get('curso_academico')
    curso_id = request.GET.get('curso')
    student_id = request.GET.get('student')

    matriculas = Matriculas.objects.all()

    if curso_academico_id:
        matriculas = matriculas.filter(course__curso_academico__id=curso_academico_id)
    if curso_id:
        matriculas = matriculas.filter(course__id=curso_id)
    if student_id:
        matriculas = matriculas.filter(student__id=student_id)

    context = {
        'matriculas': matriculas,
        'curso_academico': CursoAcademico.objects.get(id=curso_academico_id) if curso_academico_id else None,
    }
    return render_to_pdf('matriculas_pdf.html', context)

@login_required
def export_matriculas_excel(request):
    curso_academico_id = request.GET.get('curso_academico')
    curso_id = request.GET.get('curso')
    student_id = request.GET.get('student')

    # Aplicar los mismos filtros que MatriculasListView
    matriculas = Matriculas.objects.select_related(
        'student', 'course', 'curso_academico'
    ).all()

    if curso_academico_id:
        matriculas = matriculas.filter(curso_academico__id=curso_academico_id)
    if curso_id:
        matriculas = matriculas.filter(course__id=curso_id)
    if student_id:
        matriculas = matriculas.filter(student__id=student_id)

    matriculas = matriculas.order_by('course__name', 'student__first_name', 'student__last_name')

    # ── Estilos ───────────────────────────────────────────────────────────────
    header_font   = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill   = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    center        = Alignment(horizontal='center', vertical='center')
    left_align    = Alignment(horizontal='left',   vertical='center')
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    estado_fills = {
        'A':  PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Aprobado - verde
        'P':  PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),  # Activo - azul
        'BA': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Baja por Ausencia - gris
        'BL': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Baja por Licencia - gris
        'BI': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # Baja por Insuf. - amarillo
    }
    estado_fonts = {
        'A':  Font(name='Arial', bold=True, color='276221'),
        'P':  Font(name='Arial', bold=True, color='1F4E79'),
        'BA': Font(name='Arial', bold=True, color='595959'),
        'BL': Font(name='Arial', bold=True, color='595959'),
        'BI': Font(name='Arial', bold=True, color='7D5A00'),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía por defecto

    # Agrupar matrículas por curso
    cursos_vistos = {}
    for m in matriculas:
        cid = m.course.id
        if cid not in cursos_vistos:
            cursos_vistos[cid] = {'course': m.course, 'matriculas': []}
        cursos_vistos[cid]['matriculas'].append(m)

    if not cursos_vistos:
        ws = wb.create_sheet(title="Sin datos")
        ws['A1'] = "No se encontraron matrículas con los filtros seleccionados."
    else:
        HEADERS = ['Estudiante', 'Curso Académico', 'Fecha de Matrícula', 'Estado']

        for cid, grupo in cursos_vistos.items():
            course = grupo['course']
            mats   = grupo['matriculas']
            sheet_title = course.name[:31]

            ws = wb.create_sheet(title=sheet_title)

            # Fila 1: título del curso
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
            tc = ws.cell(row=1, column=1, value=f"Matrículas — {course.name}")
            tc.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
            tc.fill = PatternFill(start_color='001F4D', end_color='001F4D', fill_type='solid')
            tc.alignment = center
            ws.row_dimensions[1].height = 26

            # Fila 2: encabezados
            for col, h in enumerate(HEADERS, 1):
                c = ws.cell(row=2, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = thin_border

            # Filas de datos
            for row_num, m in enumerate(mats, 3):
                nombre = m.student.get_full_name() or m.student.username
                ca_nombre = m.curso_academico.nombre if m.curso_academico else '—'
                fecha = m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '—'
                estado_display = m.get_estado_display()

                valores = [nombre, ca_nombre, fecha, estado_display]
                for col, val in enumerate(valores, 1):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.alignment = left_align if col == 1 else center
                    c.border = thin_border

                # Color en la columna Estado
                estado_cell = ws.cell(row=row_num, column=4)
                if m.estado in estado_fills:
                    estado_cell.fill = estado_fills[m.estado]
                    estado_cell.font = estado_fonts[m.estado]

            # Anchos de columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 14

            ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="matriculas.xlsx"'
    return response

# Función auxiliar para generar PDF
def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

# Función auxiliar para generar Excel
def generate_excel(context_dict={}):
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    
    # Estilos para el Excel
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Obtener datos del contexto
    curso_academico = context_dict.get('curso_academico')
    cursos = context_dict.get('cursos', [])
    matriculas = context_dict.get('matriculas', [])
    calificaciones = context_dict.get('calificaciones', [])
    asistencias = context_dict.get('asistencias', [])
    registros = context_dict.get('registros', []) # Añadir registros al contexto
    
    # Hoja de información general
    if curso_academico:
        ws_info = wb.active
        ws_info.title = "Información General"
        ws_info['A1'] = f"Curso Académico: {curso_academico.nombre}"
        ws_info['A2'] = f"Activo: {'Sí' if curso_academico.activo else 'No'}"
        ws_info['A3'] = f"Archivado: {'Sí' if curso_academico.archivado else 'No'}"
        ws_info['A4'] = f"Fecha de Creación: {curso_academico.fecha_creacion}"
        # Crear hoja de matrículas como hoja adicional
        ws_matriculas = wb.create_sheet(title="Matrículas")
    else:
        # Si no hay curso_academico, usar la hoja activa para matrículas
        ws_matriculas = wb.active
        ws_matriculas.title = "Matrículas"
    
    # Hoja de cursos
    if cursos:
        ws_cursos = wb.create_sheet(title="Cursos")
        # Encabezados
        headers_cursos = ["Nombre del Curso", "Profesor", "Estado"]
        for col_num, header in enumerate(headers_cursos, 1):
            cell = ws_cursos.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, curso in enumerate(cursos, 2):
            ws_cursos.cell(row=row_num, column=1, value=curso.name)
            ws_cursos.cell(row=row_num, column=2, value=curso.teacher.get_full_name() or curso.teacher.username)
            ws_cursos.cell(row=row_num, column=3, value=curso.get_status_display())
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 4):
                ws_cursos.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_cursos.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_cursos.column_dimensions[column].width = adjusted_width
    
    # Encabezados para la hoja de matrículas
    headers = ["Estudiante", "Curso Académico", "Curso", "Fecha Matrícula", "Estado Matrícula"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_matriculas.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    if matriculas:
        # Datos
        for row_num, matricula in enumerate(matriculas, 2):
            ws_matriculas.cell(row=row_num, column=1, value=matricula.student.get_full_name() or matricula.student.username)
            ws_matriculas.cell(row=row_num, column=2, value=matricula.course.curso_academico.nombre if matricula.course and matricula.course.curso_academico else 'N/A')
            ws_matriculas.cell(row=row_num, column=3, value=matricula.course.name if matricula.course else 'N/A')
            ws_matriculas.cell(row=row_num, column=4, value=matricula.fecha_matricula.strftime('%d/%m/%Y') if matricula.fecha_matricula else 'N/A')
            ws_matriculas.cell(row=row_num, column=5, value=matricula.get_estado_display())
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 6):
                ws_matriculas.cell(row=row_num, column=col_num).border = border
    else:
        # Si no hay matrículas, agregar una fila indicándolo
        ws_matriculas.cell(row=2, column=1, value="No se encontraron matrículas con los filtros seleccionados")
        ws_matriculas.merge_cells('A2:E2')
        ws_matriculas.cell(row=2, column=1).alignment = Alignment(horizontal='center')
        ws_matriculas.cell(row=2, column=1).font = Font(italic=True, color='666666')
    
    # Ajustar ancho de columnas
    for col in ws_matriculas.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_matriculas.column_dimensions[column].width = adjusted_width

    
    # Hoja de calificaciones
    if calificaciones:
        ws_calificaciones = wb.create_sheet(title="Calificaciones")
        
        # Determinar el número máximo de notas individuales
        max_notas = 0
        for calificacion in calificaciones:
            num_notas = calificacion.notas.count()
            if num_notas > max_notas:
                max_notas = num_notas
        
        # Crear encabezados dinámicos
        headers = ["Estudiante", "Curso"]
        for i in range(1, max_notas + 1):
            headers.append(f"Nota {i}")
        headers.append("Promedio")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_calificaciones.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, calificacion in enumerate(calificaciones, 2):
            # Información básica
            ws_calificaciones.cell(row=row_num, column=1, value=calificacion.student.get_full_name() or calificacion.student.username)
            ws_calificaciones.cell(row=row_num, column=2, value=calificacion.course.name)
            
            # Notas individuales
            notas_individuales = list(calificacion.notas.all().order_by('fecha_creacion'))
            for i, nota in enumerate(notas_individuales, 3):
                ws_calificaciones.cell(row=row_num, column=i, value=nota.valor)
            
            # Rellenar con N/A las notas que no existen
            for i in range(len(notas_individuales) + 3, max_notas + 3):
                ws_calificaciones.cell(row=row_num, column=i, value='N/A')
            
            # Promedio
            ws_calificaciones.cell(row=row_num, column=max_notas + 3, value=calificacion.average if calificacion.average is not None else 'N/A')
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, max_notas + 4):
                ws_calificaciones.cell(row=row_num, column=col_num).border = border
        
    # Hoja de usuarios registrados
    if registros:
        ws_usuarios = wb.create_sheet(title="Usuarios Registrados")
        # Encabezados
        headers = [
            "Nombre", "Apellidos", "Email", "Nacionalidad", "Carnet ID", "Carnet Disponible", "Sexo",
            "Dirección", "Municipio", "Provincia", "Movil", "Grado Académico",
            "Ocupación", "Religioso", "Título", "Título Disponible", "Grupo", "Fecha de Registro"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws_usuarios.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, registro in enumerate(registros, 2):
            ws_usuarios.cell(row=row_num, column=1, value=registro.user.first_name)
            ws_usuarios.cell(row=row_num, column=2, value=registro.user.last_name)
            ws_usuarios.cell(row=row_num, column=3, value=registro.user.email)
            ws_usuarios.cell(row=row_num, column=4, value=registro.nacionalidad)
            ws_usuarios.cell(row=row_num, column=5, value=registro.carnet)
            ws_usuarios.cell(row=row_num, column=6, value="Sí" if registro.foto_carnet else "No")
            ws_usuarios.cell(row=row_num, column=7, value=registro.sexo)
            ws_usuarios.cell(row=row_num, column=8, value=registro.address)
            ws_usuarios.cell(row=row_num, column=9, value=registro.location)
            ws_usuarios.cell(row=row_num, column=10, value=registro.provincia)
            ws_usuarios.cell(row=row_num, column=11, value=registro.movil)
            ws_usuarios.cell(row=row_num, column=12, value=registro.get_grado_display())
            ws_usuarios.cell(row=row_num, column=13, value=registro.get_ocupacion_display())
            ws_usuarios.cell(row=row_num, column=14, value="Sí" if registro.es_religioso else "No")
            ws_usuarios.cell(row=row_num, column=15, value=registro.titulo)
            ws_usuarios.cell(row=row_num, column=16, value="Sí" if registro.foto_titulo else "No")
            ws_usuarios.cell(row=row_num, column=17, value=registro.user.groups.first().name if registro.user.groups.first() else '')
            ws_usuarios.cell(row=row_num, column=18, value=registro.user.date_joined.strftime("%d/%m/%Y"))
            
            for col_num in range(1, len(headers) + 1):
                ws_usuarios.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_usuarios.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_usuarios.column_dimensions[column].width = adjusted_width

    if calificaciones:
        # Ajustar ancho de columnas
        for col in ws_calificaciones.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_calificaciones.column_dimensions[column].width = adjusted_width

    # Hoja de asistencias
    if asistencias:
        ws_asistencias = wb.create_sheet(title="Asistencias")
        # Encabezados
        headers = ["Estudiante", "Curso", "Fecha", "Presente"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_asistencias.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_num, asistencia in enumerate(asistencias, 2):
            ws_asistencias.cell(row=row_num, column=1, value=asistencia.student.get_full_name() or asistencia.student.username)
            ws_asistencias.cell(row=row_num, column=2, value=asistencia.course.name)
            ws_asistencias.cell(row=row_num, column=3, value=asistencia.date.strftime('%d/%m/%Y') if asistencia.date else 'N/A')
            ws_asistencias.cell(row=row_num, column=4, value='Sí' if asistencia.presente else 'No')

            # Aplicar bordes a todas las celdas
            for col_num in range(1, 5):
                ws_asistencias.cell(row=row_num, column=col_num).border = border

        # Ajustar ancho de columnas
        for col in ws_asistencias.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_asistencias.column_dimensions[column].width = adjusted_width

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


# ── Adaptadores para datos archivados ─────────────────────────────────────────
# Estos objetos envuelven los modelos de datos_archivados y exponen la misma
# interfaz que los modelos de principal, de modo que el template y el generador
# de Excel no necesitan distinguir entre datos activos y archivados.


class _SemestreCursoFinalizadoAdapter:
    """
    Adapta un SemestreCurso de un curso finalizado (status='F', no archivado aún)
    para que el template lo trate igual que un SemestreCursoArchivado en el
    selector de semestres.
    Se usa una clase interna _CursoRef para imitar la interfaz de curso_archivado.
    """

    class _CursoRef:
        def __init__(self, curso):
            self.id_original = curso.id
            self.name = curso.name
            self.id = curso.id

    def __init__(self, semestre_curso):
        self._sc = semestre_curso
        self.id = f"activo_{semestre_curso.id}"   # prefijo para distinguirlo de archivados
        self.pk = self.id
        self.numero_semestre = semestre_curso.numero_semestre
        self.fecha_cierre = semestre_curso.fecha_cierre
        self.fecha_creacion = semestre_curso.fecha_creacion
        self.curso_archivado = self._CursoRef(semestre_curso.curso)
        self.es_finalizado_no_archivado = True   # bandera para el template si se necesita

class _UsuarioArchivadoAdapter:
    """Adapta UsuarioArchivado para que se comporte como User en el template."""
    def __init__(self, ua):
        self._ua = ua
        self.id = ua.id
        self.pk = ua.id
        self.username = ua.username
        self.first_name = ua.first_name
        self.last_name = ua.last_name
        self.email = ua.email

    def get_full_name(self):
        nombre = f"{self._ua.first_name} {self._ua.last_name}".strip()
        return nombre or self._ua.username

    def __str__(self):
        return self.get_full_name()


class _CursoArchivadoAdapter:
    """Adapta CursoArchivado para que se comporte como Curso en el template."""
    def __init__(self, ca):
        self._ca = ca
        self.id = ca.id
        self.pk = ca.id
        self.name = ca.name
        self.description = ca.description
        self.area = ca.area
        self.tipo = ca.tipo
        self.class_quantity = ca.class_quantity
        self.status = ca.status
        self.enrollment_deadline = ca.enrollment_deadline
        self.start_date = ca.start_date
        # Exponer teacher como un objeto con get_full_name / username
        self.teacher = _TeacherProxy(ca)
        # Exponer curso_academico para compatibilidad con generate_excel
        self.curso_academico = ca.curso_academico

    def get_dynamic_status(self):
        return self._ca.status

    def get_dynamic_status_display(self):
        return dict(self._ca.STATUS_CHOICES).get(self._ca.status, self._ca.status)

    def get_status_display(self):
        return self.get_dynamic_status_display()

    @property
    def semestre_activo_numero(self):
        # Buscar el semestre activo archivado para este curso
        from datos_archivados.models import SemestreCursoArchivado
        semestre = (
            SemestreCursoArchivado.objects
            .filter(curso_archivado=self._ca)
            .order_by('-activo', '-numero_semestre')
            .first()
        )
        return semestre.numero_semestre if semestre else 1

    def __str__(self):
        return self.name


class _TeacherProxy:
    """Proxy mínimo para el profesor de un CursoArchivado."""
    def __init__(self, ca):
        self._ca = ca
        self.username = ca.teacher_name

    def get_full_name(self):
        if self._ca.teacher_actual:
            nombre = self._ca.teacher_actual.get_full_name()
            return nombre or self._ca.teacher_actual.username
        return self._ca.teacher_name

    def __str__(self):
        return self.get_full_name()


class _MatriculaArchivadaAdapter:
    """Adapta MatriculaArchivada para que se comporte como Matriculas en el template."""
    ESTADO_CHOICES = {
        'P': 'Activo', 'A': 'Aprobado',
        'BA': 'Baja por Ausencia', 'BL': 'Baja por Licencia', 'BI': 'Baja por Insuficiencia Académica',
    }

    def __init__(self, ma):
        self._ma = ma
        self.id = ma.id
        self.pk = ma.id
        self.activo = ma.activo
        self.fecha_matricula = ma.fecha_matricula
        self.estado = ma.estado
        self.student = _UsuarioArchivadoAdapter(ma.student)
        self.course = _CursoArchivadoAdapter(ma.course)
        # Exponer curso_academico para compatibilidad con templates que lo usan
        self.curso_academico = ma.course.curso_academico if ma.course else None

    def get_estado_display(self):
        return self.ESTADO_CHOICES.get(self._ma.estado, self._ma.estado)

    def __str__(self):
        return str(self._ma)


class _NotaArchivadaAdapter:
    """Adapta NotaIndividualArchivada para que se comporte como NotaIndividual."""
    def __init__(self, nota):
        self.id = nota.id
        self.valor = nota.valor
        self.fecha_creacion = nota.fecha_creacion

    def __str__(self):
        return str(self.valor)


class _NotasManager:
    """Simula el manager .notas de Calificaciones para datos archivados."""
    def __init__(self, notas_archivadas_qs):
        self._notas = [_NotaArchivadaAdapter(n) for n in notas_archivadas_qs]

    def all(self):
        return _NotasList(self._notas)

    def count(self):
        return len(self._notas)


class _NotasList(list):
    """Lista de notas que soporta .order_by() para compatibilidad con generate_excel."""
    def order_by(self, *args):
        return self  # ya viene ordenada por fecha_creacion desde el queryset


class _CalificacionArchivadaAdapter:
    """Adapta CalificacionArchivada para que se comporte como Calificaciones en el template."""
    def __init__(self, ca):
        self._ca = ca
        self.id = ca.id
        self.pk = ca.id
        self.average = ca.average
        self.student = _UsuarioArchivadoAdapter(ca.student)
        self.course = _CursoArchivadoAdapter(ca.course)
        # Exponer .notas con la misma interfaz que el manager de NotaIndividual
        self.notas = _NotasManager(ca.notas_archivadas.all())
        # Exponer curso_academico para compatibilidad con templates
        self.curso_academico = ca.course.curso_academico if ca.course else None

    def __str__(self):
        return str(self._ca)


class _AsistenciaArchivadaAdapter:
    """Adapta AsistenciaArchivada para que se comporte como Asistencia en el template."""
    def __init__(self, aa):
        self._aa = aa
        self.id = aa.id
        self.pk = aa.id
        self.presente = aa.presente
        self.date = aa.date
        self.student = _UsuarioArchivadoAdapter(aa.student)
        self.course = _CursoArchivadoAdapter(aa.course)

    def __str__(self):
        return str(self._aa)


def _estudiantes_de_archivado(ca_archivado):
    """Devuelve los UsuarioArchivado únicos matriculados en un CursoAcademicoArchivado."""
    from datos_archivados.models import MatriculaArchivada
    ids = (
        MatriculaArchivada.objects
        .filter(course__curso_academico=ca_archivado)
        .values_list('student_id', flat=True)
        .distinct()
    )
    from datos_archivados.models import UsuarioArchivado
    return UsuarioArchivado.objects.filter(id__in=ids)


class CursoAcademicoDetailView(DetailView):
    model = CursoAcademico
    template_name = 'curso_academico_detail.html'
    context_object_name = 'curso_academico'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico = self.get_object()

        # Obtener los filtros de la URL
        curso_id = self.request.GET.get('curso')
        estudiante_id = self.request.GET.get('estudiante')
        semestre_id = self.request.GET.get('semestre')
        # Descartar valores inválidos (ej. prefijos desconocidos)
        if semestre_id and not str(semestre_id).startswith('activo_') and not str(semestre_id).isdigit() and str(semestre_id) != 'todos':
            semestre_id = None
        estado_curso = self.request.GET.get('estado_curso')
        estado_matricula = self.request.GET.get('estado_matricula')

        # Páginas activas por sección
        page_m = self.request.GET.get('page_m', 1)
        page_c = self.request.GET.get('page_c', 1)
        page_a = self.request.GET.get('page_a', 1)
        page_cur = self.request.GET.get('page_cur', 1)
        per_page = 10

        # ── Determinar fuente de datos ────────────────────────────────────────
        if curso_academico.archivado:
            context.update(
                self._context_desde_archivados(
                    curso_academico, curso_id, estudiante_id,
                    page_m, page_c, page_a, page_cur, per_page,
                    semestre_id=semestre_id,
                    estado_curso=estado_curso,
                    estado_matricula=estado_matricula,
                )
            )
        else:
            context.update(
                self._context_desde_principal(
                    curso_academico, curso_id, estudiante_id,
                    page_m, page_c, page_a, page_cur, per_page,
                    semestre_id=semestre_id,
                    estado_curso=estado_curso,
                    estado_matricula=estado_matricula,
                )
            )

        # Tab activa
        context['active_tab'] = self.request.GET.get('tab', 'cursos')
        context['filter_params'] = {
            'curso': curso_id or '',
            'estudiante': estudiante_id or '',
            'semestre': semestre_id or '',
            'estado_curso': estado_curso or '',
            'estado_matricula': estado_matricula or '',
        }
        context['todos_los_cursos_academicos'] = CursoAcademico.objects.all().order_by('-fecha_creacion')

        # ── Pestaña Semestres: lista de todos los semestres del CA ────────────
        context.update(self._context_semestres_tab(curso_academico))

        return context

    def _context_semestres_tab(self, curso_academico):
        """
        Construye la lista completa de semestres para la pestaña 'Semestres'.
        Combina SemestreCurso de la tabla principal con SemestreCursoArchivado,
        evitando duplicados por id_original.
        Soporta filtro por estado y paginación (page_sem).
        """
        from principal.models import SemestreCurso, Matriculas
        from datos_archivados.models import SemestreCursoArchivado, MatriculaArchivada

        estado_semestre = self.request.GET.get('estado_semestre', '')
        numero_semestre_filtro = self.request.GET.get('numero_semestre', '')
        page_sem = self.request.GET.get('page_sem', 1)
        per_page = 10

        semestres_lista = []

        # 1. Semestres archivados
        archivados = SemestreCursoArchivado.objects.filter(
            curso_archivado__id_original__in=Curso.objects.filter(
                curso_academico=curso_academico
            ).values_list('id', flat=True)
        ).select_related('curso_archivado').order_by('curso_archivado__name', 'numero_semestre')

        ids_archivados = set()
        for s in archivados:
            ids_archivados.add(s.id_original)
            # Contar por curso_archivado ya que semestre_archivado FK puede ser None
            num_matriculas = MatriculaArchivada.objects.filter(
                course=s.curso_archivado
            ).count()
            semestres_lista.append({
                'curso_nombre': s.curso_archivado.name,
                'numero': s.numero_semestre,
                'estado': 'F',
                'estado_display': 'Finalizado',
                'fecha_inicio': s.fecha_inicio,
                'fecha_cierre': s.fecha_cierre,
                'num_matriculas': num_matriculas,
                'es_archivado': True,
                'semestre_id': s.id,
            })

        # 2. SemestreCurso actuales (incluyendo cerrados y activos)
        actuales = SemestreCurso.objects.filter(
            curso__curso_academico=curso_academico,
        ).exclude(
            id__in=ids_archivados,
        ).select_related('curso').order_by('curso__name', 'numero_semestre')

        for s in actuales:
            if s.activo and s.curso.status != 'F':
                estado = s.curso.get_dynamic_status()
                estado_display = s.curso.get_dynamic_status_display()
            else:
                estado = 'F'
                estado_display = 'Finalizado'

            # Si no hay fecha_cierre pero el curso está finalizado,
            # usar fecha_actualizacion del curso como fecha de cierre de facto.
            fecha_cierre = s.fecha_cierre
            if not fecha_cierre and s.curso.status == 'F':
                fecha_cierre = s.curso.fecha_actualizacion.date() if s.curso.fecha_actualizacion else None

            num_matriculas = Matriculas.objects.filter(semestre=s).count()
            semestres_lista.append({
                'curso_nombre': s.curso.name,
                'numero': s.numero_semestre,
                'estado': estado,
                'estado_display': estado_display,
                'fecha_inicio': s.fecha_inicio,
                'fecha_cierre': fecha_cierre,
                'num_matriculas': num_matriculas,
                'es_archivado': False,
                'semestre_id': f'activo_{s.id}',
            })

        # Ordenar: por nombre de curso, luego número de semestre
        semestres_lista.sort(key=lambda x: (x['curso_nombre'], x['numero']))

        # Guardar lista completa para extraer números únicos de semestre
        semestres_lista_completa = semestres_lista[:]
        numeros_semestre = sorted({s['numero'] for s in semestres_lista_completa})

        # Filtrar por estado si se especificó
        if estado_semestre:
            semestres_lista = [s for s in semestres_lista if s['estado'] == estado_semestre]

        # Filtrar por número de semestre si se especificó
        if numero_semestre_filtro:
            try:
                num = int(numero_semestre_filtro)
                semestres_lista = [s for s in semestres_lista if s['numero'] == num]
            except ValueError:
                pass

        # Paginar
        paginator = Paginator(semestres_lista, per_page)
        semestres_page = paginator.get_page(page_sem)

        return {
            'semestres_tab': semestres_page,
            'semestres_tab_total': len(semestres_lista),
            'estado_semestre': estado_semestre,
            'numero_semestre_filtro': numero_semestre_filtro,
            'numeros_semestre_disponibles': numeros_semestre,
        }

    # ── Fuente: gestión académica (curso activo / no archivado) ───────────────
    def _context_desde_principal(
        self, curso_academico, curso_id, estudiante_id,
        page_m, page_c, page_a, page_cur, per_page,
        semestre_id=None,
        estado_curso=None,
        estado_matricula=None,
    ):
        from principal.models import SemestreCurso
        from datos_archivados.models import SemestreCursoArchivado, MatriculaArchivada, CalificacionArchivada, AsistenciaArchivada

        # ── Semestres disponibles para este CA ────────────────────────────────
        # 1. Semestres de cursos ya archivados en datos_archivados
        semestres_archivados_qs = list(
            SemestreCursoArchivado.objects.filter(
                curso_archivado__id_original__in=Curso.objects.filter(
                    curso_academico=curso_academico
                ).values_list('id', flat=True)
            ).select_related('curso_archivado').order_by('numero_semestre')
        )
        # IDs de SemestreCurso (id_original) que ya tienen copia en datos_archivados
        ids_semestres_ya_archivados = {s.id_original for s in semestres_archivados_qs}

        # 2. Semestres cerrados (activo=False + fecha_cierre) de cursos NO archivados
        #    que aún no tienen copia en datos_archivados.
        #    Un semestre cerrado = activo=False y fecha_cierre definida.
        # 3. También incluir semestres activos de cursos con status='F' (finalizados),
        #    ya que esos cursos no deben aparecer en la vista "semestre actual".
        semestres_finalizados_no_archivados = [
            _SemestreCursoFinalizadoAdapter(sc)
            for sc in SemestreCurso.objects.filter(
                curso__curso_academico=curso_academico,
            ).filter(
                # Semestre cerrado (cualquier curso) O semestre de curso finalizado
                Q(activo=False, fecha_cierre__isnull=False) |
                Q(curso__status='F')
            ).exclude(
                id__in=ids_semestres_ya_archivados,
            ).select_related('curso').order_by('curso__name', 'numero_semestre')
        ]

        semestres_archivados = semestres_archivados_qs + semestres_finalizados_no_archivados

        # Si se seleccionó un semestre, mostrar sus datos
        semestre_seleccionado = None
        semestre_finalizado_no_archivado = None  # SemestreCurso de curso finalizado

        # Ignorar valores no numéricos que no sean el prefijo 'activo_' ni 'todos'
        if semestre_id and not str(semestre_id).startswith('activo_') and not str(semestre_id).isdigit() and str(semestre_id) != 'todos':
            semestre_id = None

        mostrar_todos = (str(semestre_id) == 'todos') if semestre_id else False

        if semestre_id and not mostrar_todos:
            if str(semestre_id).startswith('activo_'):
                # Semestre de curso finalizado no archivado
                real_id = str(semestre_id).replace('activo_', '')
                try:
                    sc = SemestreCurso.objects.select_related('curso').get(pk=real_id)
                    semestre_seleccionado = _SemestreCursoFinalizadoAdapter(sc)
                    semestre_finalizado_no_archivado = sc
                except SemestreCurso.DoesNotExist:
                    semestre_seleccionado = None
            else:
                try:
                    semestre_seleccionado = SemestreCursoArchivado.objects.get(pk=semestre_id)
                except SemestreCursoArchivado.DoesNotExist:
                    semestre_seleccionado = None

        if semestre_seleccionado:
            if semestre_finalizado_no_archivado:
                # ── Semestre de curso FINALIZADO pero no archivado ─────────────
                # Los datos siguen en la tabla principal; filtramos por semestre y curso.
                curso_finalizado = semestre_finalizado_no_archivado.curso

                cursos_qs = Curso.objects.filter(
                    curso_academico=curso_academico,
                    id=curso_finalizado.id,
                )

                matriculas_qs = Matriculas.objects.filter(
                    semestre=semestre_finalizado_no_archivado,
                ).select_related('student', 'course')
                if estudiante_id:
                    matriculas_qs = matriculas_qs.filter(student_id=estudiante_id)

                calificaciones_qs = Calificaciones.objects.filter(
                    semestre=semestre_finalizado_no_archivado,
                ).select_related('student', 'course').prefetch_related('notas')
                if estudiante_id:
                    calificaciones_qs = calificaciones_qs.filter(student_id=estudiante_id)

                asistencias_matriculas_qs = Matriculas.objects.filter(
                    semestre=semestre_finalizado_no_archivado,
                ).select_related('student', 'course').order_by(
                    'student__first_name', 'student__last_name'
                )
                if estudiante_id:
                    asistencias_matriculas_qs = asistencias_matriculas_qs.filter(student_id=estudiante_id)

                paginator_cur = Paginator(list(cursos_qs), per_page)
                paginator_m   = Paginator(list(matriculas_qs), per_page)
                paginator_c   = Paginator(list(calificaciones_qs), per_page)
                paginator_a   = Paginator(list(asistencias_matriculas_qs), per_page)

                return {
                    'cursos': paginator_cur.get_page(page_cur),
                    'cursos_total': cursos_qs.count(),
                    'matriculas': paginator_m.get_page(page_m),
                    'matriculas_total': matriculas_qs.count(),
                    'calificaciones': paginator_c.get_page(page_c),
                    'calificaciones_total': calificaciones_qs.count(),
                    'asistencias': paginator_a.get_page(page_a),
                    'asistencias_total': asistencias_matriculas_qs.count(),
                    'cursos_disponibles': Curso.objects.filter(curso_academico=curso_academico).distinct(),
                    'estudiantes_disponibles': User.objects.filter(matriculas__curso_academico=curso_academico).distinct(),
                    'es_archivado': False,
                    'semestres_disponibles': semestres_archivados,
                    'semestre_seleccionado': semestre_seleccionado,
                    'viendo_semestre_archivado': True,
                }
            else:
                # ── Semestre ya archivado en datos_archivados ──────────────────
                curso_archivado = semestre_seleccionado.curso_archivado

                matriculas_qs = MatriculaArchivada.objects.filter(
                    course=curso_archivado
                ).select_related('student', 'course')
                if estudiante_id:
                    matriculas_qs = matriculas_qs.filter(student__id_original=estudiante_id)

                calificaciones_qs = CalificacionArchivada.objects.filter(
                    course=curso_archivado
                ).select_related('student', 'course').prefetch_related('notas_archivadas')
                if estudiante_id:
                    calificaciones_qs = calificaciones_qs.filter(student__id_original=estudiante_id)

                asistencias_matriculas_qs = MatriculaArchivada.objects.filter(
                    course=curso_archivado
                ).select_related('student', 'course').order_by(
                    'student__first_name', 'student__last_name'
                )
                if estudiante_id:
                    asistencias_matriculas_qs = asistencias_matriculas_qs.filter(student__id_original=estudiante_id)

                matriculas_adaptadas = [_MatriculaArchivadaAdapter(m) for m in matriculas_qs]
                calificaciones_adaptadas = [_CalificacionArchivadaAdapter(c) for c in calificaciones_qs]
                asistencias_adaptadas = [_MatriculaArchivadaAdapter(m) for m in asistencias_matriculas_qs]

                # Cursos: solo el curso del semestre seleccionado
                cursos_qs = Curso.objects.filter(
                    curso_academico=curso_academico,
                    id=curso_archivado.id_original,
                )

                paginator_cur = Paginator(list(cursos_qs), per_page)
                paginator_m   = Paginator(matriculas_adaptadas, per_page)
                paginator_c   = Paginator(calificaciones_adaptadas, per_page)
                paginator_a   = Paginator(asistencias_adaptadas, per_page)

                return {
                    'cursos': paginator_cur.get_page(page_cur),
                    'cursos_total': cursos_qs.count(),
                    'matriculas': paginator_m.get_page(page_m),
                    'matriculas_total': len(matriculas_adaptadas),
                    'calificaciones': paginator_c.get_page(page_c),
                    'calificaciones_total': len(calificaciones_adaptadas),
                    'asistencias': paginator_a.get_page(page_a),
                    'asistencias_total': len(asistencias_adaptadas),
                    'cursos_disponibles': Curso.objects.filter(curso_academico=curso_academico).distinct(),
                    'estudiantes_disponibles': User.objects.filter(matriculas__curso_academico=curso_academico).distinct(),
                    'es_archivado': False,
                    'semestres_disponibles': semestres_archivados,
                    'semestre_seleccionado': semestre_seleccionado,
                    'viendo_semestre_archivado': True,
                }

        # ── "Todos": mostrar todos los cursos sin filtrar por semestre ────────
        if mostrar_todos:
            cursos_qs = Curso.objects.filter(
                curso_academico=curso_academico
            ).distinct()
            if curso_id:
                cursos_qs = cursos_qs.filter(id=curso_id)
            if estado_curso:
                cursos_qs = [c for c in cursos_qs if c.get_dynamic_status() == estado_curso]

            matriculas_qs = Matriculas.objects.filter(
                curso_academico=curso_academico,
            ).select_related('student', 'course')
            if curso_id:
                matriculas_qs = matriculas_qs.filter(course_id=curso_id)
            if estudiante_id:
                matriculas_qs = matriculas_qs.filter(student_id=estudiante_id)
            if estado_matricula:
                matriculas_qs = matriculas_qs.filter(estado=estado_matricula)
            if estado_curso:
                ids_cursos = [c.id for c in Curso.objects.filter(curso_academico=curso_academico) if c.get_dynamic_status() == estado_curso]
                matriculas_qs = matriculas_qs.filter(course_id__in=ids_cursos)

            calificaciones_qs = Calificaciones.objects.filter(
                curso_academico=curso_academico,
            ).select_related('student', 'course').prefetch_related('notas')
            if curso_id:
                calificaciones_qs = calificaciones_qs.filter(course_id=curso_id)
            if estudiante_id:
                calificaciones_qs = calificaciones_qs.filter(student_id=estudiante_id)
            if estado_curso:
                ids_cursos = [c.id for c in Curso.objects.filter(curso_academico=curso_academico) if c.get_dynamic_status() == estado_curso]
                calificaciones_qs = calificaciones_qs.filter(course_id__in=ids_cursos)

            asistencias_matriculas_qs = Matriculas.objects.filter(
                curso_academico=curso_academico,
            ).select_related('student', 'course').order_by(
                'student__first_name', 'student__last_name', 'student__username'
            )
            if curso_id:
                asistencias_matriculas_qs = asistencias_matriculas_qs.filter(course_id=curso_id)
            if estudiante_id:
                asistencias_matriculas_qs = asistencias_matriculas_qs.filter(student_id=estudiante_id)
            if estado_matricula:
                asistencias_matriculas_qs = asistencias_matriculas_qs.filter(estado=estado_matricula)
            if estado_curso:
                ids_cursos = [c.id for c in Curso.objects.filter(curso_academico=curso_academico) if c.get_dynamic_status() == estado_curso]
                asistencias_matriculas_qs = asistencias_matriculas_qs.filter(course_id__in=ids_cursos)

            cursos_list = list(cursos_qs)
            paginator_cur = Paginator(cursos_list, per_page)
            paginator_m   = Paginator(matriculas_qs, per_page)
            paginator_c   = Paginator(calificaciones_qs, per_page)
            paginator_a   = Paginator(asistencias_matriculas_qs, per_page)

            return {
                'cursos': paginator_cur.get_page(page_cur),
                'cursos_total': len(cursos_list),
                'matriculas': paginator_m.get_page(page_m),
                'matriculas_total': matriculas_qs.count(),
                'calificaciones': paginator_c.get_page(page_c),
                'calificaciones_total': calificaciones_qs.count(),
                'asistencias': paginator_a.get_page(page_a),
                'asistencias_total': asistencias_matriculas_qs.count(),
                'cursos_disponibles': Curso.objects.filter(curso_academico=curso_academico).distinct(),
                'estudiantes_disponibles': User.objects.filter(matriculas__curso_academico=curso_academico).distinct(),
                'es_archivado': False,
                'semestres_disponibles': semestres_archivados,
                'semestre_seleccionado': None,
                'viendo_semestre_archivado': False,
            }

        # ── Semestre activo (solo cursos no finalizados, semestre activo) ─────
        # Los semestres cerrados y los de cursos finalizados se consultan
        # desde el selector de semestres.
        semestres_activos_ids = SemestreCurso.objects.filter(
            curso__curso_academico=curso_academico,
            activo=True,
        ).exclude(
            curso__status='F',
        ).values_list('id', flat=True)

        cursos_qs = Curso.objects.filter(
            curso_academico=curso_academico,
            semestres__activo=True,
        ).exclude(status='F').distinct()
        if curso_id:
            cursos_qs = cursos_qs.filter(id=curso_id)
        if estado_curso:
            cursos_qs = [c for c in cursos_qs if c.get_dynamic_status() == estado_curso]

        matriculas_qs = Matriculas.objects.filter(
            semestre__in=semestres_activos_ids,
        ).select_related('student', 'course')
        if curso_id:
            matriculas_qs = matriculas_qs.filter(course_id=curso_id)
        if estudiante_id:
            matriculas_qs = matriculas_qs.filter(student_id=estudiante_id)
        if estado_matricula:
            matriculas_qs = matriculas_qs.filter(estado=estado_matricula)
        if estado_curso:
            ids_cursos = [c.id for c in Curso.objects.filter(curso_academico=curso_academico) if c.get_dynamic_status() == estado_curso]
            matriculas_qs = matriculas_qs.filter(course_id__in=ids_cursos)

        calificaciones_qs = Calificaciones.objects.filter(
            semestre__in=semestres_activos_ids,
        ).select_related('student', 'course').prefetch_related('notas')
        if curso_id:
            calificaciones_qs = calificaciones_qs.filter(course_id=curso_id)
        if estudiante_id:
            calificaciones_qs = calificaciones_qs.filter(student_id=estudiante_id)
        if estado_curso:
            ids_cursos = [c.id for c in Curso.objects.filter(curso_academico=curso_academico) if c.get_dynamic_status() == estado_curso]
            calificaciones_qs = calificaciones_qs.filter(course_id__in=ids_cursos)

        asistencias_matriculas_qs = Matriculas.objects.filter(
            semestre__in=semestres_activos_ids,
        ).select_related('student', 'course').order_by(
            'student__first_name', 'student__last_name', 'student__username'
        )
        if curso_id:
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(course_id=curso_id)
        if estudiante_id:
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(student_id=estudiante_id)
        if estado_matricula:
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(estado=estado_matricula)
        if estado_curso:
            ids_cursos = [c.id for c in Curso.objects.filter(curso_academico=curso_academico) if c.get_dynamic_status() == estado_curso]
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(course_id__in=ids_cursos)

        cursos_list = list(cursos_qs)
        paginator_cur = Paginator(cursos_list, per_page)
        paginator_m   = Paginator(matriculas_qs, per_page)
        paginator_c   = Paginator(calificaciones_qs, per_page)
        paginator_a   = Paginator(asistencias_matriculas_qs, per_page)

        return {
            'cursos': paginator_cur.get_page(page_cur),
            'cursos_total': len(cursos_list),
            'matriculas': paginator_m.get_page(page_m),
            'matriculas_total': matriculas_qs.count(),
            'calificaciones': paginator_c.get_page(page_c),
            'calificaciones_total': calificaciones_qs.count(),
            'asistencias': paginator_a.get_page(page_a),
            'asistencias_total': asistencias_matriculas_qs.count(),
            'cursos_disponibles': Curso.objects.filter(
                curso_academico=curso_academico
            ).distinct(),
            'estudiantes_disponibles': User.objects.filter(
                matriculas__curso_academico=curso_academico
            ).distinct(),
            'es_archivado': False,
            'semestres_disponibles': semestres_archivados,
            'semestre_seleccionado': None,
            'viendo_semestre_archivado': False,
        }

    # ── Fuente: datos archivados (curso archivado) ────────────────────────────
    def _context_desde_archivados(
        self, curso_academico, curso_id, estudiante_id,
        page_m, page_c, page_a, page_cur, per_page,
        semestre_id=None,
        estado_curso=None,
        estado_matricula=None,
    ):
        from datos_archivados.models import (
            CursoAcademicoArchivado,
            CursoArchivado,
            MatriculaArchivada,
            CalificacionArchivada,
            AsistenciaArchivada,
        )

        # Buscar el registro archivado correspondiente
        ca_archivado = CursoAcademicoArchivado.objects.filter(
            id_original=curso_academico.pk
        ).first()

        if ca_archivado is None:
            # El curso está marcado como archivado pero aún no tiene datos en
            # datos_archivados (p.ej. archivado manualmente sin pasar por el servicio).
            # Devolver contexto vacío con indicador para que el template lo informe.
            return {
                'cursos': Paginator([], per_page).get_page(1),
                'cursos_total': 0,
                'matriculas': Paginator([], per_page).get_page(1),
                'matriculas_total': 0,
                'calificaciones': Paginator([], per_page).get_page(1),
                'calificaciones_total': 0,
                'asistencias': Paginator([], per_page).get_page(1),
                'asistencias_total': 0,
                'cursos_disponibles': [],
                'estudiantes_disponibles': [],
                'es_archivado': True,
                'sin_datos_archivados': True,
            }

        # ── Cursos ────────────────────────────────────────────────────────────
        cursos_qs = CursoArchivado.objects.filter(
            curso_academico=ca_archivado
        ).select_related('teacher_actual')

        # ── Filtro por semestre ───────────────────────────────────────────────
        from datos_archivados.models import SemestreCursoArchivado
        semestres_archivados = SemestreCursoArchivado.objects.filter(
            curso_archivado__curso_academico=ca_archivado
        ).select_related('curso_archivado').order_by('curso_archivado__name', 'numero_semestre')

        semestre_seleccionado = None
        if semestre_id:
            try:
                semestre_seleccionado = SemestreCursoArchivado.objects.get(pk=semestre_id)
                # Filtrar solo el curso de ese semestre
                cursos_qs = cursos_qs.filter(id=semestre_seleccionado.curso_archivado.id)
            except SemestreCursoArchivado.DoesNotExist:
                semestre_seleccionado = None

        if curso_id:
            cursos_qs = cursos_qs.filter(id=curso_id)

        # Adaptar CursoArchivado para que el template pueda usar los mismos
        # atributos que usa con Curso (name, teacher.get_full_name, get_dynamic_status_display)
        if estado_curso:
            cursos_adaptados = [_CursoArchivadoAdapter(c) for c in cursos_qs if c.status == estado_curso]
        else:
            cursos_adaptados = [_CursoArchivadoAdapter(c) for c in cursos_qs]

        # ── Matrículas ────────────────────────────────────────────────────────
        matriculas_qs = MatriculaArchivada.objects.filter(
            course__curso_academico=ca_archivado
        ).select_related('student', 'course', 'semestre_archivado')
        if semestre_seleccionado:
            matriculas_qs = matriculas_qs.filter(semestre_archivado=semestre_seleccionado)
        if curso_id:
            matriculas_qs = matriculas_qs.filter(course_id=curso_id)
        if estudiante_id:
            matriculas_qs = matriculas_qs.filter(student_id=estudiante_id)
        if estado_matricula:
            matriculas_qs = matriculas_qs.filter(estado=estado_matricula)
        if estado_curso:
            ids_cursos_arch = [c.id for c in CursoArchivado.objects.filter(curso_academico=ca_archivado, status=estado_curso)]
            matriculas_qs = matriculas_qs.filter(course_id__in=ids_cursos_arch)

        matriculas_adaptadas = [_MatriculaArchivadaAdapter(m) for m in matriculas_qs]

        # ── Calificaciones ────────────────────────────────────────────────────
        calificaciones_qs = CalificacionArchivada.objects.filter(
            course__curso_academico=ca_archivado
        ).select_related('student', 'course', 'semestre_archivado').prefetch_related('notas_archivadas')
        if semestre_seleccionado:
            calificaciones_qs = calificaciones_qs.filter(semestre_archivado=semestre_seleccionado)
        if curso_id:
            calificaciones_qs = calificaciones_qs.filter(course_id=curso_id)
        if estudiante_id:
            calificaciones_qs = calificaciones_qs.filter(student_id=estudiante_id)
        if estado_curso:
            ids_cursos_arch = [c.id for c in CursoArchivado.objects.filter(curso_academico=ca_archivado, status=estado_curso)]
            calificaciones_qs = calificaciones_qs.filter(course_id__in=ids_cursos_arch)

        calificaciones_adaptadas = [_CalificacionArchivadaAdapter(c) for c in calificaciones_qs]

        # ── Asistencias ───────────────────────────────────────────────────────
        asistencias_matriculas_qs = MatriculaArchivada.objects.filter(
            course__curso_academico=ca_archivado
        ).select_related('student', 'course', 'semestre_archivado').order_by(
            'student__first_name', 'student__last_name', 'student__username'
        )
        if semestre_seleccionado:
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(semestre_archivado=semestre_seleccionado)
        if curso_id:
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(course_id=curso_id)
        if estudiante_id:
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(student_id=estudiante_id)
        if estado_matricula:
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(estado=estado_matricula)
        if estado_curso:
            ids_cursos_arch = [c.id for c in CursoArchivado.objects.filter(curso_academico=ca_archivado, status=estado_curso)]
            asistencias_matriculas_qs = asistencias_matriculas_qs.filter(course_id__in=ids_cursos_arch)

        asistencias_adaptadas = [_MatriculaArchivadaAdapter(m) for m in asistencias_matriculas_qs]

        # ── Selectores de filtro ──────────────────────────────────────────────
        cursos_disponibles = [
            _CursoArchivadoAdapter(c)
            for c in CursoArchivado.objects.filter(curso_academico=ca_archivado)
        ]
        estudiantes_disponibles = [
            _UsuarioArchivadoAdapter(u)
            for u in _estudiantes_de_archivado(ca_archivado)
        ]

        paginator_cur = Paginator(cursos_adaptados, per_page)
        paginator_m = Paginator(matriculas_adaptadas, per_page)
        paginator_c = Paginator(calificaciones_adaptadas, per_page)
        paginator_a = Paginator(asistencias_adaptadas, per_page)

        return {
            'cursos': paginator_cur.get_page(page_cur),
            'cursos_total': len(cursos_adaptados),
            'matriculas': paginator_m.get_page(page_m),
            'matriculas_total': len(matriculas_adaptadas),
            'calificaciones': paginator_c.get_page(page_c),
            'calificaciones_total': len(calificaciones_adaptadas),
            'asistencias': paginator_a.get_page(page_a),
            'asistencias_total': len(asistencias_adaptadas),
            'cursos_disponibles': cursos_disponibles,
            'estudiantes_disponibles': estudiantes_disponibles,
            'es_archivado': True,
            'sin_datos_archivados': False,
            'semestres_disponibles': semestres_archivados,
            'semestre_seleccionado': semestre_seleccionado,
            'viendo_semestre_archivado': semestre_seleccionado is not None,
        }
        
    def render_to_response(self, context, **response_kwargs):
        # Verificar si se solicita PDF
        if 'pdf' in self.request.GET:
            # Añadir fecha actual al contexto para el PDF
            context['now'] = datetime.now()
            # Crear un template para PDF basado en el mismo contexto
            pdf = render_to_pdf('curso_academico_pdf.html', context)
            if pdf:
                response = HttpResponse(pdf, content_type='application/pdf')
                filename = f"curso_academico_{context['curso_academico'].nombre}.pdf"
                content = f"attachment; filename={filename}"
                response['Content-Disposition'] = content
                return response
        # Verificar si se solicita Excel
        elif 'excel' in self.request.GET:
            # Construir contexto para Excel con asistencias reales (no matrículas)
            excel_context = dict(context)
            curso_academico = context['curso_academico']
            curso_id     = self.request.GET.get('curso')
            estudiante_id = self.request.GET.get('estudiante')

            if curso_academico.archivado:
                from datos_archivados.models import (
                    CursoAcademicoArchivado, AsistenciaArchivada
                )
                ca_archivado = CursoAcademicoArchivado.objects.filter(
                    id_original=curso_academico.pk
                ).first()
                if ca_archivado:
                    asistencias_qs = AsistenciaArchivada.objects.filter(
                        course__curso_academico=ca_archivado
                    ).select_related('student', 'course')
                    if curso_id:
                        asistencias_qs = asistencias_qs.filter(course_id=curso_id)
                    if estudiante_id:
                        asistencias_qs = asistencias_qs.filter(student_id=estudiante_id)
                    excel_context['asistencias'] = [
                        _AsistenciaArchivadaAdapter(a) for a in asistencias_qs
                    ]
                else:
                    excel_context['asistencias'] = []
            else:
                asistencias_qs = Asistencia.objects.filter(
                    course__matriculas__curso_academico=curso_academico
                ).distinct().select_related('student', 'course')
                if curso_id:
                    asistencias_qs = asistencias_qs.filter(course_id=curso_id)
                if estudiante_id:
                    asistencias_qs = asistencias_qs.filter(student_id=estudiante_id)
                excel_context['asistencias'] = asistencias_qs

            excel_file = generate_excel(excel_context)
            if excel_file:
                response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f"curso_academico_{context['curso_academico'].nombre}.xlsx"
                response['Content-Disposition'] = f'attachment; filename={filename}'
                return response
        # Si no se solicita PDF ni Excel, renderizar normalmente
        return super().render_to_response(context, **response_kwargs)


class BaseContextMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        group_name = None
        if user.is_authenticated:
            group = Group.objects.filter(user=user).first()
            if group:
                group_name = group.name
            context['group_name'] = group_name
        return context


class HomeView(DocumentsCourseMixin, BaseContextMixin, TemplateView):
    template_name = 'home.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            # Incluir todos los estados activos: inscripción, plazo terminado y en progreso.
            # Se excluye 'F' (Finalizado) porque esos cursos ya no son relevantes en la home.
            courses = Curso.objects.filter(
                curso_academico=curso_academico_activo,
                status__in=['I', 'IT', 'P']
            )
        else:
            courses = Curso.objects.none()
        
        # Obtener todos los formularios de aplicación existentes
        formularios = FormularioAplicacion.objects.all()
        formularios_por_curso = {f.curso_id: f for f in formularios}
        
        # Asignar los formularios a los cursos
        for curso in courses:
            if curso.id in formularios_por_curso:
                curso.formulario_aplicacion = formularios_por_curso[curso.id]
            else:
                curso.formulario_aplicacion = None
        
        # Group courses into chunks of four for the carousel (compatibilidad con otros templates)
        grouped_courses = [courses[i:i + 4] for i in range(0, len(courses), 4)]
        context['grouped_courses'] = grouped_courses

        # Lista plana de cursos activos para el nuevo template
        context['cursos_activos'] = list(courses)
        
        # Grupos de 3 para el carrusel de la home
        courses_list = list(courses)
        context['cursos_grupos_3'] = [courses_list[i:i+3] for i in range(0, len(courses_list), 3)]
        
        # Obtener las noticias publicadas más recientes
        noticias = Noticia.objects.filter(estado='publicado').order_by('-fecha_publicacion')[:8]
        
        # Agrupar noticias en chunks de 4 para el carousel (compatibilidad con otros templates)
        grouped_noticias = [noticias[i:i + 4] for i in range(0, len(noticias), 4)]
        context['grouped_noticias'] = grouped_noticias

        # Lista plana de noticias publicadas para el nuevo template
        context['noticias_publicadas'] = list(noticias)
        
        # Grupos de 3 para el carrusel de noticias de la home
        noticias_list = list(noticias)
        context['noticias_grupos_3'] = [noticias_list[i:i+3] for i in range(0, len(noticias_list), 3)]

        # Categorías que tienen al menos una noticia publicada
        categorias = Categoria.objects.filter(
            noticias__estado='publicado'
        ).distinct().order_by('nombre')
        context['categorias_con_noticias'] = categorias

        student = self.request.user if self.request.user.is_authenticated else None
        
        # Crear conjuntos de IDs de cursos con solicitudes pendientes y rechazadas
        cursos_con_solicitudes_pendientes = set()
        cursos_con_solicitudes_rechazadas = set()
        
        if student:
            # Obtener todos los IDs de cursos con solicitudes pendientes
            cursos_con_solicitudes_pendientes = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student,
                    estado='pendiente'
                ).values_list('curso_id', flat=True)
            )
            
            # Obtener todos los IDs de cursos con solicitudes rechazadas
            cursos_con_solicitudes_rechazadas = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student,
                    estado='rechazada'
                ).values_list('curso_id', flat=True)
            )
            
            print(f"DEBUG: Home - Cursos con solicitudes pendientes: {cursos_con_solicitudes_pendientes}")
            print(f"DEBUG: Home - Cursos con solicitudes rechazadas: {cursos_con_solicitudes_rechazadas}")

        for item in courses:
            if student:
                # Verificar si el estudiante está matriculado
                item.is_enrolled = Matriculas.objects.filter(
                    course=item, 
                    student=student
                ).exists()
                
                # Verificar si el estudiante tiene una solicitud pendiente para este curso
                item.tiene_solicitud_pendiente = item.id in cursos_con_solicitudes_pendientes
                
                # Verificar si el estudiante tiene una solicitud rechazada para este curso
                item.tiene_solicitud_rechazada = item.id in cursos_con_solicitudes_rechazadas
                
                if item.tiene_solicitud_pendiente:
                    print(f"DEBUG: Home - Curso {item.name} (ID: {item.id}) tiene solicitud pendiente")
                if item.tiene_solicitud_rechazada:
                    print(f"DEBUG: Home - Curso {item.name} (ID: {item.id}) tiene solicitud rechazada")
            else:
                item.is_enrolled = False
                item.tiene_solicitud_pendiente = False
                item.tiene_solicitud_rechazada = False
            
            # Calcular el conteo de inscripciones
            enrollment_count = Matriculas.objects.filter(course=item).count()
            item.enrollment_count = enrollment_count
            
            # Ya no necesitamos este código porque los formularios se cargan en get_context_data

        context['courses'] = courses
        return context
        return context



class ListadoCursosView(BaseContextMixin, TemplateView):
    template_name = 'cursos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            courses = Curso.objects.filter(curso_academico=curso_academico_activo)
        else:
            courses = Curso.objects.none()

        # Filtrar por área si viene el parámetro GET
        area = self.request.GET.get('area')
        if area:
            courses = courses.filter(area=area)

        # Filtrar por tipo si viene el parámetro GET
        tipo = self.request.GET.get('tipo')
        if tipo:
            courses = courses.filter(tipo=tipo)

        student = self.request.user if self.request.user.is_authenticated else None

        cursos_con_solicitudes_pendientes = set()
        cursos_con_solicitudes_rechazadas = set()
        if student and student.is_authenticated:
            cursos_con_solicitudes_pendientes = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student, estado='pendiente'
                ).values_list('curso_id', flat=True)
            )
            cursos_con_solicitudes_rechazadas = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student, estado='rechazada'
                ).values_list('curso_id', flat=True)
            )

        formularios = FormularioAplicacion.objects.all()
        formularios_por_curso = {f.curso_id: f for f in formularios}

        courses_list = list(courses)
        for item in courses_list:
            if student and student.is_authenticated:
                item.is_enrolled = Matriculas.objects.filter(
                    course=item, student=student).exists()
                item.tiene_solicitud_pendiente = item.id in cursos_con_solicitudes_pendientes
                item.tiene_solicitud_rechazada = item.id in cursos_con_solicitudes_rechazadas
            else:
                item.is_enrolled = False
                item.tiene_solicitud_pendiente = False
                item.tiene_solicitud_rechazada = False

            item.enrollment_count = Matriculas.objects.filter(course=item).count()
            item.formulario_aplicacion = formularios_por_curso.get(item.id)

        # Paginación: 8 tarjetas por página
        paginator = Paginator(courses_list, 8)
        page_number = self.request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        context['courses'] = page_obj
        context['page_obj'] = page_obj
        context['is_paginated'] = paginator.num_pages > 1
        context['area_seleccionada'] = area or ''
        context['tipo_seleccionado'] = tipo or ''
        context['filtro_servidor'] = True  # indica que el filtrado lo hace el servidor
        context['curso_academico_activo'] = curso_academico_activo

        user = self.request.user
        if user.is_authenticated:
            group = Group.objects.filter(user=user).first()
            context['group_name'] = group.name if group else None
            # Agregar info de documentos a cada curso (equivalente a DocumentsCourseMixin)
            for course in page_obj:
                from course_documents.mixins import DocumentsCourseMixin as _DCM
                _dcm = _DCM()
                _dcm.request = self.request
                _dcm._add_course_document_info(course, user)
        else:
            context['group_name'] = None

        return context


# para cerrar sesion


def logout_view(request):
    logout(request)
    return redirect('principal:home')


# pagina de Registro
""" class RegisterView(View):
    def get(self, request):
        data = {
            'form': RegisterForm()
        }
        return render(request, 'registration/registro.html', data)

    def post(self, request):
        user_creation_form=RegisterForm(data=request.POST)
        if user_creation_form.is_valid():
            user_creation_form.save()  
            user = authenticate(username=user_creation_form.cleaned_data['username'],
                                password=user_creation_form.cleaned_data['password'])
            login(request, user)
            return redirect('principal:home')
        data = {
            'form':user_creation_form
        }   
        return render(request, 'registration/registro.html', data) """


import random

def registro(request):
    if request.method == 'POST':
        user_creation_form = CustomUserCreationForm(
            data=request.POST, files=request.FILES)

        if user_creation_form.is_valid():
            # Generar código aleatorio de 4 dígitos
            verification_code = str(random.randint(1000, 9999))
            # Almacenar datos temporales en la sesión
            request.session['verification_code'] = verification_code
            
            # Convertir cleaned_data a un formato serializable
            form_data = {}
            for key, value in user_creation_form.cleaned_data.items():
                if hasattr(value, 'read'):  # Es un archivo
                    # No almacenar archivos en la sesión
                    continue
                else:
                    form_data[key] = value
            
            request.session['user_form_data'] = form_data
            
            # Almacenar archivos temporalmente en el sistema de archivos
            import tempfile
            import os
            temp_files = {}
            for key, file in request.FILES.items():
                if file:
                    # Crear archivo temporal
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f'_{file.name}')
                    for chunk in file.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    temp_files[key] = {
                        'path': temp_file.name,
                        'name': file.name,
                        'content_type': file.content_type,
                        'size': file.size
                    }
            
            request.session['temp_files'] = temp_files

            # Enviar email
            email_text = 'Bienvenido al Centro Fray Bartolome de las Casas, para completar su registro ingrese el siguiente codigo : ' + verification_code
            try:
                send_mail(
                    'Código de Verificación - Centro Fray Bartolome de las Casas',
                    email_text,
                    settings.DEFAULT_FROM_EMAIL,
                    [user_creation_form.cleaned_data['email']],
                    fail_silently=False,
                )
                # Redirigir a la página de verificación
                return redirect('principal:verify_email')
            except Exception as e:
                print(f"Error al enviar email: {str(e)}")
                messages.error(request, 'Error al enviar el código de verificación. Por favor, intente nuevamente más tarde.')
        else:
            # Mostrar solo errores específicos como mensajes, excepto email y carnet
            for field, errors in user_creation_form.errors.items():
                for error in errors:
                    # No mostrar errores de email y carnet como mensajes
                    if field not in ['email', 'carnet']:
                        if field == 'password2' and 'password_mismatch' in error:
                            messages.error(request, 'Las contraseñas no coinciden. Por favor, asegúrese de escribir la misma contraseña en ambos campos.')
                        else:
                            # No mostrar errores como mensajes para que aparezcan solo en los campos
                            pass
            print(f"Errores en el formulario: {user_creation_form.errors}")
    else:
        user_creation_form = CustomUserCreationForm()

    data = {
        'form': user_creation_form
    }
    return render(request, 'registration/registro.html', data)

def verify_email(request):
    if 'verification_code' not in request.session or 'user_form_data' not in request.session:
        messages.error(request, 'La sesión ha expirado. Por favor, inicie el proceso de registro nuevamente.')
        return redirect('principal:registro')
    
    if request.method == 'POST':
        code = request.POST.get('code')
        if code == request.session.get('verification_code'):
            user_form_data = request.session.get('user_form_data')
            temp_files_data = request.session.get('temp_files', {})
            
            # Recrear archivos desde archivos temporales
            from django.core.files.uploadedfile import SimpleUploadedFile
            import os
            
            files_dict = {}
            for key, file_info in temp_files_data.items():
                if os.path.exists(file_info['path']):
                    with open(file_info['path'], 'rb') as f:
                        files_dict[key] = SimpleUploadedFile(
                            file_info['name'],
                            f.read(),
                            content_type=file_info['content_type']
                        )
            
            user_creation_form = CustomUserCreationForm(data=user_form_data, files=files_dict)
            if user_creation_form.is_valid():
                user = user_creation_form.save(commit=True)
                messages.success(request, f"Usuario {user.username} creado correctamente")
                
                # Limpiar archivos temporales
                for file_info in temp_files_data.values():
                    if os.path.exists(file_info['path']):
                        os.unlink(file_info['path'])
                
                # Limpiar sesión
                del request.session['verification_code']
                del request.session['user_form_data']
                if 'temp_files' in request.session:
                    del request.session['temp_files']

                # Enviar correo de confirmación de registro
                confirmation_subject = 'Registro Exitoso - Centro Fray Bartolome de las Casas'
                confirmation_message = f'Usted se ha registrado satisfactoriamente. Su Nombre de Usuario es: {user.username}'
                try:
                    send_mail(
                        confirmation_subject,
                        confirmation_message,
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    print(f"Error al enviar correo de confirmación: {str(e)}")
                    # Considerar si se debe mostrar un mensaje de error al usuario o simplemente loguear

                return redirect('login')
            else:
                # Limpiar archivos temporales en caso de error
                for file_info in temp_files_data.values():
                    if os.path.exists(file_info['path']):
                        os.unlink(file_info['path'])
                
                # Recopilar errores específicos para mostrar en modal
                error_messages = []
                for field, errors in user_creation_form.errors.items():
                    for error in errors:
                        error_messages.append(error)
                
                if error_messages:
                    error_message = ' '.join(error_messages)
                else:
                    error_message = 'Error al crear el usuario. Por favor, intente nuevamente.'
        else:
            error_message = 'Código incorrecto. Por favor, intente nuevamente.'

        return render(request, 'registration/verify_email.html', {'error': error_message})

    return render(request, 'registration/verify_email.html')

def registro_resend_code(request):
    """
    Vista para reenviar el código de verificación de registro.
    """
    if 'user_form_data' not in request.session or 'verification_code' not in request.session:
        messages.error(request, 'La sesión ha expirado. Por favor, inicie el proceso de registro nuevamente.')
        return redirect('principal:registro')
    
    user_form_data = request.session.get('user_form_data')
    email = user_form_data.get('email')
    
    if not email:
        messages.error(request, 'No se encontró el correo electrónico. Por favor, inicie el proceso nuevamente.')
        return redirect('principal:registro')
    
    # Generar nuevo código aleatorio de 4 dígitos
    verification_code = str(random.randint(1000, 9999))
    request.session['verification_code'] = verification_code
    
    # Enviar email con el nuevo código
    email_text = 'Bienvenido al Centro Fray Bartolome de las Casas, para completar su registro ingrese el siguiente codigo : ' + verification_code
    try:
        send_mail(
            'Código de Verificación - Centro Fray Bartolome de las Casas',
            email_text,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        messages.success(request, 'Se ha reenviado un nuevo código de verificación a su correo electrónico.')
    except Exception as e:
        print(f"Error al enviar email: {str(e)}")
        messages.error(request, 'Error al reenviar el código de verificación. Por favor, intente nuevamente más tarde.')
    
    return redirect('principal:verify_email')

# Vista para manejar la redirección después del login

class LoginRedirectView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            # Verificar si el usuario fue creado automáticamente desde datos archivados
            if request.session.get('usuario_creado_automaticamente'):
                fuente = request.session.get('usuario_creado_desde', 'datos_archivados')
                
                if fuente == 'datos_dinamicos':
                    messages.success(
                        request, 
                        f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                        'Su cuenta ha sido creada automáticamente desde los datos archivados. '
                        'Ahora puede acceder a todos los servicios del sistema. '
                        'Se le ha enviado un email con los detalles de su cuenta.'
                    )
                else:
                    messages.success(
                        request, 
                        f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                        'Su cuenta ha sido reactivada automáticamente desde los datos archivados. '
                        'Ahora puede acceder a todos los servicios del sistema. '
                        'Se le ha enviado un email con los detalles de su cuenta.'
                    )
                
                # Limpiar las variables de sesión
                del request.session['usuario_creado_automaticamente']
                if 'usuario_creado_desde' in request.session:
                    del request.session['usuario_creado_desde']
            
            # Redirección según el grupo del usuario
            if user.groups.filter(name='Editor').exists():
                return redirect('blog:panel_editores')

            if user.groups.filter(name__in=['Profesores', 'Administración', 'Secretaría']).exists():
                return redirect('principal:profile')

            if user.groups.filter(name='Estudiantes').exists():
                # Ir al perfil solo si tiene al menos una matrícula aprobada
                tiene_matricula = Matriculas.objects.filter(student=user).exists()
                if tiene_matricula:
                    return redirect('principal:profile')

            return redirect('principal:cursos')
        return redirect('principal:home') # Redirige a home si no está autenticado (aunque LoginRequiredMixin ya lo manejaría)



    # Pagina de Perfil


class ProfileView(DocumentsProfileMixin, BaseContextMixin, TemplateView):
    template_name = 'profile/profile.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Obtener el primer grupo del usuario de forma segura
        user_group = user.groups.first()
        group_name = user_group.name if user_group else None
        
        # Asegurar que group_name esté en el contexto (BaseContextMixin ya lo hace, pero por seguridad)
        context['group_name'] = group_name

        if group_name == 'Profesores':
            # Obtener todos los cursos asignados al profesor del curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                assigned_courses = Curso.objects.filter(teacher=user, curso_academico=curso_academico_activo)
            else:
                assigned_courses = Curso.objects.none()
            context['assigned_courses'] = assigned_courses
            
            # Obtener las solicitudes de inscripción pendientes para los cursos del profesor con paginación
            # Solo mostrar solicitudes de cursos en etapa de inscripción (I o IT)
            pending_solicitudes = SolicitudInscripcion.objects.filter(
                curso__teacher=user,
                estado='pendiente',
                curso__status__in=['I', 'IT']  # Solo cursos en inscripción
            ).order_by('-fecha_solicitud')[:5]  # Limitar a 5 solicitudes
            context['pending_solicitudes'] = pending_solicitudes
        elif group_name == 'Estudiantes':
            # Obtener los cursos en los que el estudiante está inscrito y que pertenecen al curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                # Obtener todas las matrículas del estudiante en el curso académico activo de una sola consulta
                matriculas_dict = {
                    m.course_id: m
                    for m in Matriculas.objects.filter(
                        student=user,
                        curso_academico=curso_academico_activo
                    )
                }

                enrolled_courses = Curso.objects.filter(
                    id__in=matriculas_dict.keys(),
                    curso_academico=curso_academico_activo
                )

                # Separar cursos por estado de solicitud
                approved_courses = []
                pending_courses = []

                # Para cada curso inscrito, obtener información adicional sobre solicitudes
                for course in enrolled_courses:
                    # Estados que implican baja/inactividad del estudiante en el curso
                    ESTADOS_INACTIVOS = {'BA', 'BL', 'BI'}

                    # Asignar estado de matrícula directamente desde el dict
                    matricula = matriculas_dict.get(course.id)
                    if matricula:
                        course.matricula_activa = matricula.estado not in ESTADOS_INACTIVOS
                        course.matricula_estado = matricula.estado
                        course.matricula_estado_display = matricula.get_estado_display()
                    else:
                        course.matricula_activa = True
                        course.matricula_estado = 'P'
                        course.matricula_estado_display = ''

                    # Verificar si hay una solicitud de inscripción para este curso
                    try:
                        solicitud = SolicitudInscripcion.objects.get(
                            estudiante=user,
                            curso=course
                        )
                        course.solicitud_estado = solicitud.estado
                        course.fecha_revision = solicitud.fecha_revision
                        course.revisado_por = solicitud.revisado_por

                        # Separar por estado solo para cursos en inscripción
                        if course.status in ['I', 'IT'] and solicitud.estado == 'pendiente':
                            pending_courses.append(course)
                        else:
                            approved_courses.append(course)
                    except SolicitudInscripcion.DoesNotExist:
                        course.solicitud_estado = None
                        course.fecha_revision = None
                        course.revisado_por = None
                        approved_courses.append(course)

                    # Agregar indicador de contenido nuevo si course_documents está disponible
                    if ContentIndicatorService:
                        course.has_new_content_indicator = ContentIndicatorService.has_new_content(course, user)
                    else:
                        course.has_new_content_indicator = False

                context['enrolled_courses'] = approved_courses
                context['pending_courses'] = pending_courses

                # IDs de cursos que tienen al menos una evaluación publicada
                # y conteo de evaluaciones pendientes (sin intento) por curso
                from evaluaciones.models import Evaluacion as EvaluacionModel, IntentoEvaluacion
                from django.utils import timezone as tz_now

                evaluaciones_publicadas = EvaluacionModel.objects.filter(
                    curso__in=[c.id for c in approved_courses],
                    estado='publicada',
                ).select_related('curso')

                cursos_con_evaluaciones = set()
                evaluaciones_pendientes_por_curso = {}  # {curso_id: count}

                ahora = tz_now.now()
                intentos_del_estudiante = set(
                    IntentoEvaluacion.objects.filter(
                        estudiante=user,
                        evaluacion__in=evaluaciones_publicadas,
                    ).values_list('evaluacion_id', flat=True)
                )

                for ev in evaluaciones_publicadas:
                    cursos_con_evaluaciones.add(ev.curso_id)
                    cerrada = ev.fecha_limite is not None and ahora > ev.fecha_limite
                    if not cerrada and ev.pk not in intentos_del_estudiante:
                        evaluaciones_pendientes_por_curso[ev.curso_id] = (
                            evaluaciones_pendientes_por_curso.get(ev.curso_id, 0) + 1
                        )

                context['cursos_con_evaluaciones'] = cursos_con_evaluaciones
                context['evaluaciones_pendientes_por_curso'] = evaluaciones_pendientes_por_curso
            else:
                context['enrolled_courses'] = Curso.objects.none()
                context['pending_courses'] = Curso.objects.none()
        elif group_name in ['Administración', 'Secretaría']:
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                all_courses = Curso.objects.filter(curso_academico=curso_academico_activo)
            else:
                all_courses = Curso.objects.none()
            context['all_courses'] = all_courses
            context['curso_academico_activo'] = curso_academico_activo
            # Todos los cursos académicos archivados, para que la secretaria
            # pueda acceder a sus detalles históricos desde el perfil
            context['cursos_academicos_archivados'] = CursoAcademico.objects.filter(
                archivado=True
            ).order_by('-fecha_creacion')
        
        return context

# Vista de los Cursos


class MatriculasListView(BaseContextMixin, ListView):
    model = Matriculas
    template_name = 'matriculas_list.html'
    context_object_name = 'matriculas'
    paginate_by = 10

    def get_queryset(self):
        from datos_archivados.models import MatriculaArchivada, SemestreCursoArchivado
        semestre_id = self.request.GET.get('semestre')
        if semestre_id:
            # Devolver lista vacía; los datos archivados se manejan en get_context_data
            return Matriculas.objects.none()

        queryset = Matriculas.objects.select_related('student', 'course', 'curso_academico')

        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(curso_academico__id=curso_academico_id)

        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        student_id = self.request.GET.get('student')
        if student_id:
            queryset = queryset.filter(student__id=student_id)

        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        estado_curso = self.request.GET.get('estado_curso')
        if estado_curso:
            # Filtrar por estado dinámico: obtener IDs de cursos con ese estado
            ids_cursos = [
                c.id for c in Curso.objects.filter(
                    curso_academico__id=curso_academico_id
                ) if c.get_dynamic_status() == estado_curso
            ] if curso_academico_id else [
                c.id for c in Curso.objects.all()
                if c.get_dynamic_status() == estado_curso
            ]
            queryset = queryset.filter(course__id__in=ids_cursos)

        return queryset

    def get_context_data(self, **kwargs):
        from datos_archivados.models import SemestreCursoArchivado, MatriculaArchivada
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')

        # Semestres archivados disponibles (de todos los CAs o del CA seleccionado)
        curso_academico_id = self.request.GET.get('curso_academico')
        semestres_qs = SemestreCursoArchivado.objects.select_related('curso_archivado').order_by(
            'curso_archivado__curso_academico', 'numero_semestre'
        )
        if curso_academico_id:
            semestres_qs = semestres_qs.filter(
                curso_archivado__id_original__in=Curso.objects.filter(
                    curso_academico__id=curso_academico_id
                ).values_list('id', flat=True)
            )
        context['semestres_disponibles'] = semestres_qs

        semestre_id = self.request.GET.get('semestre')
        context['semestre_seleccionado'] = None
        context['viendo_semestre_archivado'] = False

        if semestre_id:
            try:
                semestre = SemestreCursoArchivado.objects.get(pk=semestre_id)
                context['semestre_seleccionado'] = semestre
                context['viendo_semestre_archivado'] = True

                # Cargar matrículas archivadas del semestre
                matriculas_arch_qs = MatriculaArchivada.objects.filter(
                    course=semestre.curso_archivado
                ).select_related('student', 'course')

                student_id = self.request.GET.get('student')
                if student_id:
                    matriculas_arch_qs = matriculas_arch_qs.filter(student__id_original=student_id)

                matriculas_adaptadas = [_MatriculaArchivadaAdapter(m) for m in matriculas_arch_qs]

                # Paginar manualmente
                paginator = Paginator(matriculas_adaptadas, self.paginate_by)
                page_number = self.request.GET.get('page', 1)
                context['matriculas'] = paginator.get_page(page_number)
                context['page_obj'] = context['matriculas']
                context['is_paginated'] = context['matriculas'].has_other_pages()
            except SemestreCursoArchivado.DoesNotExist:
                pass

        return context


@login_required
@require_POST
def cambiar_estado_matricula(request, matricula_id):
    """Permite a Secretaría cambiar el estado de una matrícula."""
    if not request.user.groups.filter(name='Secretaría').exists():
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()
    matricula = get_object_or_404(Matriculas, pk=matricula_id)
    nuevo_estado = request.POST.get('estado')
    estados_validos = [c[0] for c in Matriculas.ESTADO_CHOICES]
    if nuevo_estado in estados_validos:
        matricula.estado = nuevo_estado
        matricula.save()
        messages.success(request, f'Estado actualizado a "{matricula.get_estado_display()}".')
    else:
        messages.error(request, 'Estado no válido.')
    # Redirigir de vuelta con los mismos filtros
    return redirect(request.META.get('HTTP_REFERER', reverse('principal:matriculas')))


# Vistas para Calificaciones
class CalificacionesListView(BaseContextMixin, ListView):
    model = Calificaciones
    template_name = 'calificaciones_list.html'
    context_object_name = 'calificaciones'
    paginate_by = 10

    def get_queryset(self):
        semestre_id = self.request.GET.get('semestre')
        if semestre_id:
            return Calificaciones.objects.none()

        # Si se pide mostrar sin calificaciones, el queryset base se maneja en get_context_data
        mostrar_sin_cal = self.request.GET.get('mostrar_sin_calificaciones') == '1'
        if mostrar_sin_cal:
            return Calificaciones.objects.none()

        queryset = Calificaciones.objects.select_related('student', 'course', 'curso_academico')

        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(curso_academico__id=curso_academico_id)

        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        student_id = self.request.GET.get('student') or self.request.GET.get('estudiante')
        if student_id:
            queryset = queryset.filter(student__id=student_id)

        return queryset

    def get_context_data(self, **kwargs):
        from datos_archivados.models import SemestreCursoArchivado, CalificacionArchivada
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')

        curso_academico_id = self.request.GET.get('curso_academico')
        semestres_qs = SemestreCursoArchivado.objects.select_related('curso_archivado').order_by(
            'curso_archivado__curso_academico', 'numero_semestre'
        )
        if curso_academico_id:
            semestres_qs = semestres_qs.filter(
                curso_archivado__id_original__in=Curso.objects.filter(
                    curso_academico__id=curso_academico_id
                ).values_list('id', flat=True)
            )
        context['semestres_disponibles'] = semestres_qs

        semestre_id = self.request.GET.get('semestre')
        context['semestre_seleccionado'] = None
        context['viendo_semestre_archivado'] = False
        mostrar_sin_cal = self.request.GET.get('mostrar_sin_calificaciones') == '1'
        context['mostrar_sin_calificaciones'] = mostrar_sin_cal

        if semestre_id:
            try:
                semestre = SemestreCursoArchivado.objects.get(pk=semestre_id)
                context['semestre_seleccionado'] = semestre
                context['viendo_semestre_archivado'] = True

                calificaciones_arch_qs = CalificacionArchivada.objects.filter(
                    course=semestre.curso_archivado
                ).select_related('student', 'course').prefetch_related('notas_archivadas')

                student_id = self.request.GET.get('student') or self.request.GET.get('estudiante')
                if student_id:
                    calificaciones_arch_qs = calificaciones_arch_qs.filter(student__id_original=student_id)

                calificaciones_adaptadas = [_CalificacionArchivadaAdapter(c) for c in calificaciones_arch_qs]

                paginator = Paginator(calificaciones_adaptadas, self.paginate_by)
                page_number = self.request.GET.get('page', 1)
                context['calificaciones'] = paginator.get_page(page_number)
                context['page_obj'] = context['calificaciones']
                context['is_paginated'] = context['calificaciones'].has_other_pages()
            except SemestreCursoArchivado.DoesNotExist:
                pass

        elif mostrar_sin_cal:
            # Construir lista de estudiantes activos sin calificaciones
            matriculas_qs = Matriculas.objects.filter(
                estado='P', activo=True
            ).select_related('student', 'course', 'curso_academico')

            if curso_academico_id:
                matriculas_qs = matriculas_qs.filter(curso_academico__id=curso_academico_id)

            curso_id = self.request.GET.get('curso')
            if curso_id:
                matriculas_qs = matriculas_qs.filter(course__id=curso_id)

            student_id = self.request.GET.get('student') or self.request.GET.get('estudiante')
            if student_id:
                matriculas_qs = matriculas_qs.filter(student__id=student_id)

            # Excluir los que ya tienen calificación
            estudiantes_con_cal = Calificaciones.objects.values_list('student_id', 'course_id')
            pares_con_cal = set(estudiantes_con_cal)

            # Crear objetos pseudo-calificacion para el template
            class _EstudianteSinCalificacion:
                def __init__(self, matricula):
                    self.student = matricula.student
                    self.course = matricula.course
                    self.curso_academico = matricula.curso_academico
                    self.average = None

            sin_cal = [
                _EstudianteSinCalificacion(m)
                for m in matriculas_qs
                if (m.student_id, m.course_id) not in pares_con_cal
            ]

            paginator = Paginator(sin_cal, self.paginate_by)
            page_number = self.request.GET.get('page', 1)
            context['calificaciones'] = paginator.get_page(page_number)
            context['page_obj'] = context['calificaciones']
            context['is_paginated'] = context['calificaciones'].has_other_pages()

        return context


class StudentCourseAttendanceView(BaseContextMixin, ListView):
    model = Asistencia
    template_name = 'student_asistencias.html'
    context_object_name = 'asistencias'

    def dispatch(self, request, *args, **kwargs):
        # Bloquear acceso si la matrícula está en estado de baja
        course_id = self.kwargs.get('course_id')
        student_id = self.kwargs.get('student_id')
        if course_id and student_id:
            matricula = Matriculas.objects.filter(
                student_id=student_id,
                course_id=course_id
            ).first()
            if matricula and matricula.estado in ('BA', 'BL', 'BI'):
                if request.user.is_authenticated and request.user.id == int(student_id):
                    messages.error(request, 'No tienes acceso a este curso porque tu matrícula está inactiva.')
                    return redirect('principal:profile')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        queryset = queryset.filter(student__id=student_id, course__id=course_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        context['student'] = User.objects.get(id=student_id)
        context['course'] = Curso.objects.get(id=course_id)
        
        # Calculate attendance statistics
        asistencias = context['asistencias']
        total_classes = asistencias.count()
        present_count = asistencias.filter(presente=True).count()
        
        context['total_classes'] = total_classes
        context['present_count'] = present_count
        
        # Calculate attendance percentage
        if total_classes > 0:
            context['attendance_percentage'] = round((present_count / total_classes) * 100, 1)
        else:
            context['attendance_percentage'] = 0
            
        return context


class CalificacionDetalleEstudianteView(BaseContextMixin, TemplateView):
    """Vista de detalle de calificaciones de un estudiante en un curso (para secretaria)."""
    template_name = 'calificacion_detalle_estudiante.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id  = self.kwargs['course_id']

        student = get_object_or_404(User, id=student_id)
        course  = get_object_or_404(Curso, id=course_id)

        calificacion = Calificaciones.objects.filter(
            student=student, course=course
        ).prefetch_related('notas').first()

        notas = calificacion.notas.all().order_by('fecha_creacion') if calificacion else []

        matricula = Matriculas.objects.filter(
            student=student, course=course
        ).select_related('curso_academico').order_by(
            '-curso_academico__activo', '-curso_academico__id'
        ).first()

        context['student']      = student
        context['course']       = course
        context['calificacion'] = calificacion
        context['notas']        = notas
        context['matricula']    = matricula
        return context


class CalificacionDetalleArchivadoView(BaseContextMixin, TemplateView):
    """Vista de detalle de calificaciones archivadas de un estudiante en un curso."""
    template_name = 'calificacion_detalle_estudiante.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datos_archivados.models import (
            UsuarioArchivado, CalificacionArchivada, MatriculaArchivada
        )

        student_id = self.kwargs['student_id']
        course_id  = self.kwargs['course_id']

        student = get_object_or_404(UsuarioArchivado, id=student_id)
        calificacion_arch = CalificacionArchivada.objects.filter(
            student=student, course_id=course_id
        ).prefetch_related('notas_archivadas').first()

        notas = []
        average = None
        if calificacion_arch:
            notas = [_NotaArchivadaAdapter(n) for n in
                     calificacion_arch.notas_archivadas.all().order_by('fecha_creacion')]
            average = calificacion_arch.average

        # Adaptamos para que el template reutilice la misma estructura
        student_adapter = _UsuarioArchivadoAdapter(student)

        # Buscar el curso archivado para el adaptador
        from datos_archivados.models import CursoArchivado
        course_arch = get_object_or_404(CursoArchivado, id=course_id)
        course_adapter = _CursoArchivadoAdapter(course_arch)

        # Crear un objeto calificacion compatible con el template
        class _CalificacionProxy:
            def __init__(self, notas_list, avg):
                self.average = avg
                self._notas = notas_list
                self.notas = type('NM', (), {
                    'all': lambda s: notas_list,
                    'count': lambda s: len(notas_list),
                })()

        calificacion_proxy = _CalificacionProxy(notas, average) if calificacion_arch else None

        context['student']      = student_adapter
        context['course']       = course_adapter
        context['calificacion'] = calificacion_proxy
        context['notas']        = notas
        context['matricula']    = None
        context['es_archivado'] = True
        return context


@login_required
def export_calificacion_pdf(request, student_id, course_id):
    """Exporta el detalle de calificaciones de un estudiante en un curso como PDF."""
    student      = get_object_or_404(User, id=student_id)
    course       = get_object_or_404(Curso, id=course_id)
    calificacion = Calificaciones.objects.filter(
        student=student, course=course
    ).prefetch_related('notas').first()
    notas = list(calificacion.notas.all().order_by('fecha_creacion')) if calificacion else []

    context = {
        'student':      student,
        'course':       course,
        'calificacion': calificacion,
        'notas':        notas,
    }
    return render_to_pdf('calificacion_detalle_pdf.html', context)


@login_required
def export_calificaciones_excel(request):
    """
    Exporta calificaciones a Excel según los filtros activos.
    Una hoja por curso. Columnas: Estudiante, Nota 1, Nota 2, ..., Promedio.
    """
    curso_academico_id = request.GET.get('curso_academico')
    curso_id           = request.GET.get('curso')
    estudiante_id      = request.GET.get('estudiante')

    # Mismos filtros que CalificacionesListView
    calificaciones = Calificaciones.objects.select_related(
        'student', 'course', 'curso_academico'
    ).prefetch_related('notas')

    if curso_academico_id:
        calificaciones = calificaciones.filter(curso_academico__id=curso_academico_id)
    if curso_id:
        calificaciones = calificaciones.filter(course__id=curso_id)
    if estudiante_id:
        calificaciones = calificaciones.filter(student__id=estudiante_id)

    calificaciones = calificaciones.order_by(
        'course__name', 'student__first_name', 'student__last_name', 'student__username'
    )

    # ── Estilos ───────────────────────────────────────────────────────────────
    header_font    = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill    = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    nota_fill      = PatternFill(start_color='1F5C99', end_color='1F5C99', fill_type='solid')
    aprobado_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    reprobado_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    na_fill        = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    aprobado_font  = Font(name='Arial', bold=True, color='276221')
    reprobado_font = Font(name='Arial', bold=True, color='9C0006')
    na_font        = Font(name='Arial', color='595959')
    center         = Alignment(horizontal='center', vertical='center')
    left_align     = Alignment(horizontal='left',   vertical='center')
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Agrupar por curso
    cursos_vistos = {}
    for cal in calificaciones:
        cid = cal.course.id
        if cid not in cursos_vistos:
            cursos_vistos[cid] = {'course': cal.course, 'calificaciones': []}
        cursos_vistos[cid]['calificaciones'].append(cal)

    if not cursos_vistos:
        ws = wb.create_sheet(title="Sin datos")
        ws['A1'] = "No se encontraron calificaciones con los filtros seleccionados."
    else:
        for cid, grupo in cursos_vistos.items():
            course = grupo['course']
            cals   = grupo['calificaciones']
            sheet_title = course.name[:31]

            # Determinar el máximo de notas en este curso
            max_notas = max((cal.notas.count() for cal in cals), default=0)

            ws = wb.create_sheet(title=sheet_title)

            # Fila 1: título
            total_cols = 2 + max_notas + 1   # Estudiante + CursoAcad + notas + Promedio
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            tc = ws.cell(row=1, column=1, value=f"Calificaciones — {course.name}")
            tc.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
            tc.fill = PatternFill(start_color='001F4D', end_color='001F4D', fill_type='solid')
            tc.alignment = center
            ws.row_dimensions[1].height = 26

            # Fila 2: encabezados
            headers_fijos = ['Estudiante', 'Curso Académico']
            for col, h in enumerate(headers_fijos, 1):
                c = ws.cell(row=2, column=col, value=h)
                c.font = header_font; c.fill = header_fill
                c.alignment = center; c.border = thin_border

            for i in range(max_notas):
                c = ws.cell(row=2, column=3 + i, value=f'Nota {i + 1}')
                c.font = header_font; c.fill = nota_fill
                c.alignment = center; c.border = thin_border

            col_prom = 3 + max_notas
            c = ws.cell(row=2, column=col_prom, value='Promedio')
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = thin_border

            # Filas de datos
            for row_num, cal in enumerate(cals, 3):
                nombre    = cal.student.get_full_name() or cal.student.username
                ca_nombre = cal.curso_academico.nombre if cal.curso_academico else '—'
                notas     = list(cal.notas.all().order_by('fecha_creacion'))

                c = ws.cell(row=row_num, column=1, value=nombre)
                c.alignment = left_align; c.border = thin_border

                c = ws.cell(row=row_num, column=2, value=ca_nombre)
                c.alignment = center; c.border = thin_border

                # Notas individuales
                for i in range(max_notas):
                    c = ws.cell(row=row_num, column=3 + i)
                    c.border = thin_border
                    c.alignment = center
                    if i < len(notas):
                        val = float(notas[i].valor)
                        c.value = val
                        if val >= 6:
                            c.fill = aprobado_fill
                            c.font = aprobado_font
                        else:
                            c.fill = reprobado_fill
                            c.font = reprobado_font
                    else:
                        c.value = '—'
                        c.fill = na_fill
                        c.font = na_font

                # Promedio
                c = ws.cell(row=row_num, column=col_prom)
                c.border = thin_border; c.alignment = center
                if cal.average is not None:
                    prom = float(cal.average)
                    c.value = round(prom, 1)
                    if prom >= 6:
                        c.fill = aprobado_fill
                        c.font = Font(name='Arial', bold=True, color='276221')
                    else:
                        c.fill = reprobado_fill
                        c.font = Font(name='Arial', bold=True, color='9C0006')
                else:
                    c.value = 'N/A'
                    c.fill = na_fill
                    c.font = na_font

            # Anchos de columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 22
            for i in range(max_notas):
                ws.column_dimensions[
                    ws.cell(row=2, column=3 + i).column_letter
                ].width = 10
            ws.column_dimensions[
                ws.cell(row=2, column=col_prom).column_letter
            ].width = 12

            ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="calificaciones.xlsx"'
    return response


# Vistas para Asistencias
class AsistenciasListView(BaseContextMixin, ListView):
    model = Matriculas
    template_name = 'asistencias_list.html'
    context_object_name = 'matriculas'
    paginate_by = 10

    def get_queryset(self):
        semestre_id = self.request.GET.get('semestre')
        if semestre_id:
            return Matriculas.objects.none()

        queryset = Matriculas.objects.select_related('student', 'course', 'curso_academico').filter(activo=True)

        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(curso_academico__id=curso_academico_id)

        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        estudiante_id = self.request.GET.get('estudiante')
        if estudiante_id:
            queryset = queryset.filter(student__id=estudiante_id)

        return queryset.order_by('student__first_name', 'student__last_name', 'student__username')

    def get_context_data(self, **kwargs):
        from datos_archivados.models import SemestreCursoArchivado, MatriculaArchivada
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')
        context['selected_curso_academico'] = self.request.GET.get('curso_academico')
        context['selected_curso'] = self.request.GET.get('curso')
        context['selected_estudiante'] = self.request.GET.get('estudiante')

        curso_academico_id = self.request.GET.get('curso_academico')
        semestres_qs = SemestreCursoArchivado.objects.select_related('curso_archivado').order_by(
            'curso_archivado__curso_academico', 'numero_semestre'
        )
        if curso_academico_id:
            semestres_qs = semestres_qs.filter(
                curso_archivado__id_original__in=Curso.objects.filter(
                    curso_academico__id=curso_academico_id
                ).values_list('id', flat=True)
            )
        context['semestres_disponibles'] = semestres_qs

        semestre_id = self.request.GET.get('semestre')
        context['semestre_seleccionado'] = None
        context['viendo_semestre_archivado'] = False

        if semestre_id:
            try:
                semestre = SemestreCursoArchivado.objects.get(pk=semestre_id)
                context['semestre_seleccionado'] = semestre
                context['viendo_semestre_archivado'] = True

                matriculas_arch_qs = MatriculaArchivada.objects.filter(
                    course=semestre.curso_archivado
                ).select_related('student', 'course').order_by(
                    'student__first_name', 'student__last_name'
                )

                estudiante_id = self.request.GET.get('estudiante')
                if estudiante_id:
                    matriculas_arch_qs = matriculas_arch_qs.filter(student__id_original=estudiante_id)

                matriculas_adaptadas = [_MatriculaArchivadaAdapter(m) for m in matriculas_arch_qs]

                paginator = Paginator(matriculas_adaptadas, self.paginate_by)
                page_number = self.request.GET.get('page', 1)
                context['matriculas'] = paginator.get_page(page_number)
                context['page_obj'] = context['matriculas']
                context['is_paginated'] = context['matriculas'].has_other_pages()
            except SemestreCursoArchivado.DoesNotExist:
                pass

        return context


class AsistenciaDetalleEstudianteView(BaseContextMixin, TemplateView):
    """Vista de detalle de asistencias de un estudiante en un curso (para secretaria)."""
    template_name = 'asistencias_detalle_estudiante.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']

        student = get_object_or_404(User, id=student_id)
        course = get_object_or_404(Curso, id=course_id)

        asistencias = Asistencia.objects.filter(
            student=student, course=course
        ).order_by('date')

        total = asistencias.count()
        presentes = asistencias.filter(presente=True).count()
        ausentes = total - presentes
        porcentaje = round((presentes / total) * 100, 2) if total > 0 else 0

        context['student'] = student
        context['course'] = course
        context['asistencias'] = asistencias
        context['total'] = total
        context['presentes'] = presentes
        context['ausentes'] = ausentes
        context['porcentaje'] = porcentaje
        return context


class AsistenciaDetalleArchivadoView(BaseContextMixin, TemplateView):
    """Vista de detalle de asistencias archivadas de un estudiante en un curso."""
    template_name = 'asistencias_detalle_archivado.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datos_archivados.models import (
            UsuarioArchivado, MatriculaArchivada, AsistenciaArchivada
        )

        student_id = self.kwargs['student_id']
        course_id  = self.kwargs['course_id']

        student = get_object_or_404(UsuarioArchivado, id=student_id)
        matricula = get_object_or_404(
            MatriculaArchivada,
            student=student, course_id=course_id
        )
        course = matricula.course

        asistencias = AsistenciaArchivada.objects.filter(
            student=student, course=course
        ).order_by('date')

        total     = asistencias.count()
        presentes = asistencias.filter(presente=True).count()
        ausentes  = total - presentes
        porcentaje = round((presentes / total) * 100, 2) if total > 0 else 0

        context['student']    = _UsuarioArchivadoAdapter(student)
        context['course']     = _CursoArchivadoAdapter(course)
        context['asistencias'] = [_AsistenciaArchivadaAdapter(a) for a in asistencias]
        context['total']      = total
        context['presentes']  = presentes
        context['ausentes']   = ausentes
        context['porcentaje'] = porcentaje
        return context


@login_required
def export_asistencia_detalle_pdf(request, student_id, course_id):
    """Exporta el detalle de asistencias de un estudiante en un curso como PDF."""
    student = get_object_or_404(User, id=student_id)
    course  = get_object_or_404(Curso, id=course_id)

    asistencias = Asistencia.objects.filter(
        student=student, course=course
    ).order_by('date')

    total    = asistencias.count()
    presentes = asistencias.filter(presente=True).count()
    ausentes  = total - presentes
    porcentaje = round((presentes / total) * 100, 2) if total > 0 else 0

    context = {
        'student':    student,
        'course':     course,
        'asistencias': asistencias,
        'total':      total,
        'presentes':  presentes,
        'ausentes':   ausentes,
        'porcentaje': porcentaje,
    }
    return render_to_pdf('asistencia_detalle_pdf.html', context)


@login_required
def export_asistencias_curso_excel(request, course_id):
    """
    Exporta a Excel una tabla con todos los estudiantes del curso,
    sus asistencias por fecha y su porcentaje de asistencia.
    """
    course = get_object_or_404(Curso, id=course_id)

    # Obtener matrículas activas del curso
    matriculas = Matriculas.objects.filter(
        course=course, activo=True
    ).select_related('student').order_by('student__first_name', 'student__last_name', 'student__username')

    # Obtener todas las fechas únicas con asistencias registradas, ordenadas
    fechas = list(
        Asistencia.objects.filter(course=course)
        .values_list('date', flat=True)
        .distinct()
        .order_by('date')
    )

    # Estilos
    header_font       = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill       = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    subheader_fill    = PatternFill(start_color='1F5C99', end_color='1F5C99', fill_type='solid')
    presente_fill     = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ausente_fill      = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    pct_ok_fill       = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    pct_warn_fill     = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    pct_bad_fill      = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    center            = Alignment(horizontal='center', vertical='center')
    left              = Alignment(horizontal='left',   vertical='center')
    thin_border       = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    # ── Fila 1: título del curso ──────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fechas) + 4)
    title_cell = ws.cell(row=1, column=1, value=f"Asistencias — {course.name}")
    title_cell.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='001F4D', end_color='001F4D', fill_type='solid')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Fila 2: encabezados ───────────────────────────────────────────────────
    headers_fijos = ['Estudiante', 'Total Clases', 'Presentes', 'Ausentes']
    col_offset = len(headers_fijos) + 1   # columna donde empiezan las fechas

    for col, h in enumerate(headers_fijos, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    for i, fecha in enumerate(fechas):
        c = ws.cell(row=2, column=col_offset + i, value=fecha.strftime('%d/%m/%Y'))
        c.font = header_font
        c.fill = subheader_fill
        c.alignment = center
        c.border = thin_border

    # columna % al final
    col_pct = col_offset + len(fechas)
    c = ws.cell(row=2, column=col_pct, value='% Asistencia')
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

    # ── Filas de datos ────────────────────────────────────────────────────────
    for row_num, matricula in enumerate(matriculas, 3):
        student = matricula.student
        nombre = student.get_full_name() or student.username

        # Asistencias de este estudiante indexadas por fecha
        asistencias_qs = Asistencia.objects.filter(student=student, course=course)
        asistencias_dict = {a.date: a.presente for a in asistencias_qs}

        total   = len(fechas)
        present = sum(1 for f in fechas if asistencias_dict.get(f) is True)
        absent  = total - present
        pct     = round((present / total) * 100, 1) if total > 0 else 0.0

        # Nombre
        c = ws.cell(row=row_num, column=1, value=nombre)
        c.alignment = left
        c.border = thin_border

        # Total / Presentes / Ausentes
        for col, val in zip([2, 3, 4], [total, present, absent]):
            c = ws.cell(row=row_num, column=col, value=val)
            c.alignment = center
            c.border = thin_border

        # Columnas de fechas
        for i, fecha in enumerate(fechas):
            presente = asistencias_dict.get(fecha)
            if presente is True:
                texto = '✓'
                fill  = presente_fill
            elif presente is False:
                texto = '✗'
                fill  = ausente_fill
            else:
                texto = '—'
                fill  = None
            c = ws.cell(row=row_num, column=col_offset + i, value=texto)
            c.alignment = center
            c.border = thin_border
            if fill:
                c.fill = fill

        # % Asistencia
        c = ws.cell(row=row_num, column=col_pct, value=f'{pct}%')
        c.alignment = center
        c.border = thin_border
        if pct >= 75:
            c.fill = pct_ok_fill
            c.font = Font(name='Arial', bold=True, color='276221')
        elif pct >= 50:
            c.fill = pct_warn_fill
            c.font = Font(name='Arial', bold=True, color='7D5A00')
        else:
            c.fill = pct_bad_fill
            c.font = Font(name='Arial', bold=True, color='9C0006')

    # ── Ajustar anchos de columna ─────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30   # Nombre
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 14
    for i in range(len(fechas)):
        col_letter = ws.cell(row=2, column=col_offset + i).column_letter
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions[ws.cell(row=2, column=col_pct).column_letter].width = 14

    # Congelar la primera columna y la fila de encabezados
    ws.freeze_panes = 'B3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"asistencias_{course.name.replace(' ', '_')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_asistencias_excel(request):
    """
    Exporta a Excel las asistencias según los filtros activos de la página
    asistencias_list (curso_academico, curso, estudiante).
    Genera una hoja por cada curso presente en los resultados.
    """
    # ── Aplicar los mismos filtros que AsistenciasListView ────────────────────
    matriculas = Matriculas.objects.select_related(
        'student', 'course', 'curso_academico'
    ).filter(activo=True)

    curso_academico_id = request.GET.get('curso_academico')
    if curso_academico_id:
        matriculas = matriculas.filter(curso_academico__id=curso_academico_id)

    curso_id = request.GET.get('curso')
    if curso_id:
        matriculas = matriculas.filter(course__id=curso_id)

    estudiante_id = request.GET.get('estudiante')
    if estudiante_id:
        matriculas = matriculas.filter(student__id=estudiante_id)

    matriculas = matriculas.order_by(
        'course__name', 'student__first_name', 'student__last_name', 'student__username'
    )

    # ── Estilos ───────────────────────────────────────────────────────────────
    header_font    = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill    = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    subheader_fill = PatternFill(start_color='1F5C99', end_color='1F5C99', fill_type='solid')
    presente_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ausente_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    pct_ok_fill    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    pct_warn_fill  = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    pct_bad_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    center         = Alignment(horizontal='center', vertical='center')
    left_align     = Alignment(horizontal='left',   vertical='center')
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # eliminar hoja vacía por defecto

    # ── Agrupar matrículas por curso ──────────────────────────────────────────
    from itertools import groupby
    from operator import attrgetter

    cursos_vistos = {}
    for matricula in matriculas:
        cid = matricula.course.id
        if cid not in cursos_vistos:
            cursos_vistos[cid] = {'course': matricula.course, 'matriculas': []}
        cursos_vistos[cid]['matriculas'].append(matricula)

    if not cursos_vistos:
        # Sin datos: devolver Excel con mensaje
        ws = wb.create_sheet(title="Sin datos")
        ws['A1'] = "No se encontraron estudiantes con los filtros seleccionados."
    else:
        for cid, grupo in cursos_vistos.items():
            course      = grupo['course']
            mats        = grupo['matriculas']
            sheet_title = course.name[:31]   # Excel limita a 31 chars

            # Fechas únicas del curso ordenadas
            fechas = list(
                Asistencia.objects.filter(course=course)
                .values_list('date', flat=True)
                .distinct()
                .order_by('date')
            )

            ws = wb.create_sheet(title=sheet_title)
            col_offset = 5   # Estudiante + Total + Presentes + Ausentes + (empieza en col 5)

            # Fila 1: título
            total_cols = col_offset + len(fechas)   # última col = %
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            tc = ws.cell(row=1, column=1, value=f"Asistencias — {course.name}")
            tc.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
            tc.fill = PatternFill(start_color='001F4D', end_color='001F4D', fill_type='solid')
            tc.alignment = center
            ws.row_dimensions[1].height = 26

            # Fila 2: encabezados fijos
            for col, h in enumerate(['Estudiante', 'Total Clases', 'Presentes', 'Ausentes'], 1):
                c = ws.cell(row=2, column=col, value=h)
                c.font = header_font; c.fill = header_fill
                c.alignment = center; c.border = thin_border

            # Encabezados de fechas
            for i, fecha in enumerate(fechas):
                c = ws.cell(row=2, column=col_offset + i, value=fecha.strftime('%d/%m/%Y'))
                c.font = header_font; c.fill = subheader_fill
                c.alignment = center; c.border = thin_border

            # Columna % al final
            col_pct = col_offset + len(fechas)
            c = ws.cell(row=2, column=col_pct, value='% Asistencia')
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = thin_border

            # Filas de datos
            for row_num, matricula in enumerate(mats, 3):
                student = matricula.student
                nombre  = student.get_full_name() or student.username

                asistencias_dict = {
                    a.date: a.presente
                    for a in Asistencia.objects.filter(student=student, course=course)
                }

                total   = len(fechas)
                present = sum(1 for f in fechas if asistencias_dict.get(f) is True)
                absent  = total - present
                pct     = round((present / total) * 100, 1) if total > 0 else 0.0

                c = ws.cell(row=row_num, column=1, value=nombre)
                c.alignment = left_align; c.border = thin_border

                for col, val in zip([2, 3, 4], [total, present, absent]):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.alignment = center; c.border = thin_border

                for i, fecha in enumerate(fechas):
                    presente = asistencias_dict.get(fecha)
                    if presente is True:
                        texto, fill = '✓', presente_fill
                    elif presente is False:
                        texto, fill = '✗', ausente_fill
                    else:
                        texto, fill = '—', None
                    c = ws.cell(row=row_num, column=col_offset + i, value=texto)
                    c.alignment = center; c.border = thin_border
                    if fill:
                        c.fill = fill

                c = ws.cell(row=row_num, column=col_pct, value=f'{pct}%')
                c.alignment = center; c.border = thin_border
                if pct >= 75:
                    c.fill = pct_ok_fill
                    c.font = Font(name='Arial', bold=True, color='276221')
                elif pct >= 50:
                    c.fill = pct_warn_fill
                    c.font = Font(name='Arial', bold=True, color='7D5A00')
                else:
                    c.fill = pct_bad_fill
                    c.font = Font(name='Arial', bold=True, color='9C0006')

            # Anchos de columna
            ws.column_dimensions['A'].width = 30
            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 14
            for i in range(len(fechas)):
                ws.column_dimensions[
                    ws.cell(row=2, column=col_offset + i).column_letter
                ].width = 12
            ws.column_dimensions[
                ws.cell(row=2, column=col_pct).column_letter
            ].width = 14

            ws.freeze_panes = 'B3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="asistencias.xlsx"'
    return response


class StudentCourseNotesView(BaseContextMixin, ListView):
    model = Calificaciones
    template_name = 'student_notes.html'  # New template for student notes
    context_object_name = 'calificaciones'

    def dispatch(self, request, *args, **kwargs):
        # Bloquear acceso si la matrícula está en estado de baja
        course_id = self.kwargs.get('course_id')
        student_id = self.kwargs.get('student_id')
        if course_id and student_id:
            matricula = Matriculas.objects.filter(
                student_id=student_id,
                course_id=course_id
            ).first()
            if matricula and matricula.estado in ('BA', 'BL', 'BI'):
                if request.user.is_authenticated and request.user.id == int(student_id):
                    messages.error(request, 'No tienes acceso a este curso porque tu matrícula está inactiva.')
                    return redirect('principal:profile')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        # Filter grades for the specific student and course
        return Calificaciones.objects.filter(student__id=student_id, course__id=course_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']

        student = User.objects.get(id=student_id)
        course = Curso.objects.get(id=course_id)

        context['student'] = student
        context['course'] = course

        print(f"[DEBUG] StudentCourseNotesView - student_id: {student_id}, course_id: {course_id}")

        # Get all Calificaciones for the student and course
        calificaciones_for_student_course = Calificaciones.objects.filter(
            student=student,
            course=course
        )
        print(f"[DEBUG] Calificaciones found: {calificaciones_for_student_course.count()}")

        all_notes = []
        total_score = 0
        num_grades = 0

        for calificacion in calificaciones_for_student_course:
            print(f"[DEBUG] Processing Calificacion ID: {calificacion.id}")
            for nota in calificacion.notas.all():
                print(f"[DEBUG] Adding Nota: {nota.valor} (ID: {nota.id})")
                all_notes.append(nota)
                if nota.valor is not None:
                    total_score += nota.valor
                    num_grades += 1

        if num_grades > 0:
            average_score = total_score / num_grades
        else:
            average_score = 0

        context['all_notes'] = all_notes
        context['average_score'] = average_score
        return context


class CoursesView(BaseContextMixin, TemplateView):
    template_name = 'cursos.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            courses = Curso.objects.filter(curso_academico=curso_academico_activo)
        else:
            courses = Curso.objects.none()

        # Filtrar por área y tipo si vienen como parámetros GET
        area = self.request.GET.get('area')
        tipo = self.request.GET.get('tipo')
        if area:
            courses = courses.filter(area=area)
        if tipo:
            courses = courses.filter(tipo=tipo)

        student = self.request.user if self.request.user.is_authenticated else None

        # Obtener solicitudes pendientes y rechazadas del estudiante de una sola vez
        cursos_con_solicitudes_pendientes = set()
        cursos_con_solicitudes_rechazadas = set()
        if student and student.is_authenticated:
            cursos_con_solicitudes_pendientes = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student, estado='pendiente'
                ).values_list('curso_id', flat=True)
            )
            cursos_con_solicitudes_rechazadas = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student, estado='rechazada'
                ).values_list('curso_id', flat=True)
            )

        # Obtener formularios de aplicación de una sola vez
        formularios = FormularioAplicacion.objects.all()
        formularios_por_curso = {f.curso_id: f for f in formularios}

        courses_list = list(courses)
        for item in courses_list:
            if student and student.is_authenticated:
                item.is_enrolled = Matriculas.objects.filter(
                    course=item, student=student).exists()
                item.tiene_solicitud_pendiente  = item.id in cursos_con_solicitudes_pendientes
                item.tiene_solicitud_rechazada  = item.id in cursos_con_solicitudes_rechazadas
            else:
                item.is_enrolled = False
                item.tiene_solicitud_pendiente  = False
                item.tiene_solicitud_rechazada  = False

            item.enrollment_count = Matriculas.objects.filter(course=item).count()
            item.formulario_aplicacion = formularios_por_curso.get(item.id)

        # Paginación: 8 tarjetas por página
        paginator   = Paginator(courses_list, 8)
        page_number = self.request.GET.get('page', 1)
        page_obj    = paginator.get_page(page_number)

        context['courses']      = page_obj
        context['page_obj']     = page_obj
        context['is_paginated'] = paginator.num_pages > 1
        context['area_seleccionada'] = area or ''
        context['tipo_seleccionado'] = tipo or ''
        context['filtro_servidor'] = True
        context['curso_academico_activo'] = curso_academico_activo

        user = self.request.user
        if user.is_authenticated:
            group = Group.objects.filter(user=user).first()
            context['group_name'] = group.name if group else None
        else:
            context['group_name'] = None

        return context

# Crear nuevo Curso


class CourseCreateView(LoginRequiredMixin, CreateView):
    model = Curso
    form_class = CourseForm
    template_name = 'create_course.html'
    success_url = reverse_lazy('principal:cursos')

    def dispatch(self, request, *args, **kwargs):
        if not CursoAcademico.objects.filter(activo=True).exists():
            messages.error(
                request,
                'No hay un curso académico activo. Debe activar un curso académico antes de crear cursos.'
            )
            return redirect('principal:cursos')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # Asigna el curso académico activo al curso
        active_academic_course = CursoAcademico.objects.filter(activo=True).first()
        if active_academic_course:
            form.instance.curso_academico = active_academic_course
        messages.success(self.request, 'El Curso se guardo correctamente')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(
            self.request, 'Ha ocurrido un ERROR al guardar el Curso')
        return self.render_to_response(self.get_context_data(form=form))

# Vista para inscribirse a un curso


@login_required
def inscribirse_curso(request, curso_id):
    # Obtener el curso
    curso = Curso.objects.get(id=curso_id)
    estudiante = request.user

    # Verificar si ya está inscrito
    inscripcion_existente = Matriculas.objects.filter(
        course=curso, student=estudiante).exists()

    if not inscripcion_existente:
        # Obtener el curso académico activo
        curso_academico = CursoAcademico.objects.filter(activo=True).first()
        
        if not curso_academico:
            messages.error(request, 'No hay un curso académico activo configurado. Contacte al administrador.')
            return redirect('principal:cursos')
            
        # Crear nueva matrícula asignada al curso académico activo
        matricula = Matriculas(
            course=curso, 
            student=estudiante, 
            activo=True,
            curso_academico=curso_academico,
            estado='P'  # Estado inicial: Pendiente
        )
        matricula.save()
        messages.success(
            request, f'Te has inscrito exitosamente al curso {curso.name} para el año académico {curso_academico.nombre}')
    else:
        messages.info(request, 'Ya estás inscrito en este curso')

    return redirect('principal:cursos')

# Vista para editar un curso


class CourseUpdateView(BaseContextMixin, UpdateView):
    model = Curso
    form_class = CourseForm
    template_name = 'create_course.html'  # Reutilizamos el mismo template
    success_url = reverse_lazy('principal:cursos')

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        # Verificar si el curso tiene un formulario de aplicación
        try:
            formulario = FormularioAplicacion.objects.filter(curso=obj).first()
            if formulario:
                obj.formulario_aplicacion = formulario
        except Exception:
            pass
        return obj

    @override
    def form_valid(self, form):
        messages.success(self.request, 'El Curso se actualizó correctamente')
        return super().form_valid(form)

    @override
    def form_invalid(self, form):
        messages.error(
            self.request, 'Ha ocurrido un ERROR al actualizar el Curso')
        return self.render_to_response(self.get_context_data(form=form))


# Vista para eliminar un curso
@login_required
def eliminar_curso(request, curso_id):
    # Verificar si el usuario pertenece al grupo 'Secretaría'
    if request.user.groups.filter(name='Secretaría').exists():
        try:
            # Obtener el curso
            curso = Curso.objects.get(id=curso_id)
            nombre_curso = curso.name

            # Eliminar el curso
            curso.delete()
            messages.success(
                request, f'El curso {nombre_curso} ha sido eliminado correctamente')
        except Curso.DoesNotExist:
            messages.error(request, 'El curso no existe')
    else:
        messages.error(request, 'No tienes permisos para eliminar cursos')

    return redirect('principal:cursos')

# Mostrar lista de alumnos y notas a los profesores


class StudentListNotasView(BaseContextMixin, ListView):
    model = Matriculas
    template_name = 'student_list_notas.html'
    context_object_name = 'matriculas'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'student',
            'course',
            'course__teacher',
            'calificaciones'
        )

        # Verificar si se está accediendo desde la URL con course_id
        course_id = self.kwargs.get('course_id')
        if course_id:
            queryset = queryset.filter(course__id=course_id)
            return queryset

        # Si no hay course_id en la URL, usar los filtros normales
        search_query = self.request.GET.get('search_query')
        course_filter = self.request.GET.get('course')
        teacher_filter = self.request.GET.get('teacher')

        if search_query:
            queryset = queryset.filter(
                Q(student__username__icontains=search_query) |
                Q(student__first_name__icontains=search_query) |
                Q(student__last_name__icontains=search_query)
            )


        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        course_id = self.kwargs.get('course_id')
        if course_id:
            course = get_object_or_404(Curso, id=course_id)
            context['course'] = course

            # Obtener el curso académico activo (solo para contexto)
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            context['curso_academico'] = curso_academico_activo

            # Obtener todas las matrículas del curso para mostrar inactivos transparentes.
            # Un estudiante es "activo en el semestre" solo si estado == 'P'.
            # Cualquier baja (BA, BL, BI) o aprobado (A) se trata como inactivo.
            search_query = self.request.GET.get('search_query', '').strip()
            all_enrollments = Matriculas.objects.filter(
                course=course,
                curso_academico=curso_academico_activo,
            ).select_related('student').order_by(
                # Activos (estado P) primero
                Case(When(estado='P', then=0), default=1, output_field=IntegerField()),
                'student__first_name', 'student__last_name'
            )

            if search_query:
                all_enrollments = all_enrollments.filter(
                    Q(student__first_name__icontains=search_query) |
                    Q(student__last_name__icontains=search_query) |
                    Q(student__username__icontains=search_query)
                )

            student_data = []
            max_notas = 0
            for enrollment in all_enrollments:
                student = enrollment.student
                # Activo en semestre = estado 'P' (Activo). Cualquier otro estado es inactivo.
                es_activo = enrollment.estado == 'P'
                calificacion = Calificaciones.objects.filter(
                    course=course,
                    student=student,
                    curso_academico=curso_academico_activo,
                ).first()

                all_notes = []
                if calificacion:
                    all_notes = list(calificacion.notas.all().order_by('fecha_creacion'))
                    # Solo contar notas para el ancho de columnas si el estudiante está activo
                    if es_activo:
                        num_notas = len(all_notes)
                        if num_notas > max_notas:
                            max_notas = num_notas

                student_data.append({
                    'calificacion_id': calificacion.id if calificacion else None,
                    'name': student.get_full_name() or student.username,
                    'notas': all_notes,
                    'average': calificacion.average if calificacion else None,
                    'matricula_id': enrollment.id,
                    'student_id': student.id,
                    'matricula_activa': es_activo,
                    'estado_display': enrollment.get_estado_display(),
                })
            # Pre-calcular notas faltantes por estudiante para el template
            for data in student_data:
                data['notas_faltantes'] = range(max_notas - len(data['notas']))

            context['student_data'] = student_data
            context['max_notas'] = max_notas
            context['nota_range'] = range(1, max_notas + 1)
        else:
            context['courses'] = CursoAcademico.objects.all()
            context['teachers'] = User.objects.filter(groups__name='Docente')

        return context
# Agregar Notas de los estudiantes

class AddNotaView(LoginRequiredMixin, View):

    def _get_calificacion(self, matricula):
        """Obtiene o crea el objeto Calificaciones para la matrícula dada."""
        # Primero intentar con curso_academico exacto
        calificacion = Calificaciones.objects.filter(
            course=matricula.course,
            student=matricula.student,
            curso_academico=matricula.curso_academico,
        ).first()

        if not calificacion:
            # Fallback: buscar cualquier calificación existente para este estudiante/curso
            calificacion = Calificaciones.objects.filter(
                course=matricula.course,
                student=matricula.student,
            ).first()

        if not calificacion:
            # Crear nueva
            calificacion = Calificaciones.objects.create(
                course=matricula.course,
                student=matricula.student,
                curso_academico=matricula.curso_academico,
                matricula=matricula,
            )
        else:
            # Asegurar que la FK matricula esté enlazada
            if calificacion.matricula_id != matricula.pk:
                calificacion.matricula = matricula
                calificacion.save(update_fields=['matricula'])

        return calificacion

    def _redirect_after_save(self, request, matricula):
        """Redirige al origen correcto: next param, detalle secretaria, o lista profesor."""
        next_url = request.POST.get('next') or request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect(
            'principal:calificacion_detalle_estudiante',
            student_id=matricula.student.id,
            course_id=matricula.course.id,
        )

    def get(self, request, matricula_id):
        matricula = get_object_or_404(Matriculas, id=matricula_id)

        # Bloquear acceso si la matrícula no está activa en el semestre (estado != 'P')
        if matricula.estado != 'P':
            messages.error(request, f'Este estudiante está inactivo ({matricula.get_estado_display()}). No se pueden agregar ni editar notas.')
            return redirect('principal:calificacion_detalle_estudiante', student_id=matricula.student.id, course_id=matricula.course.id)

        calificacion = Calificaciones.objects.filter(
            course=matricula.course,
            student=matricula.student,
            curso_academico=matricula.curso_academico,
        ).prefetch_related('notas').first()

        # Fallback: buscar sin filtro de curso_academico
        if not calificacion:
            calificacion = Calificaciones.objects.filter(
                course=matricula.course,
                student=matricula.student,
            ).prefetch_related('notas').first()

        notas_existentes = list(calificacion.notas.order_by('id')) if calificacion else []

        context = {
            'matricula': matricula,
            'notas_existentes': notas_existentes,
            'next': request.GET.get('next', ''),
        }
        return render(request, 'add_nota.html', context)

    def post(self, request, matricula_id):
        matricula = get_object_or_404(Matriculas, id=matricula_id)

        # Bloquear si la matrícula no está activa en el semestre (estado != 'P')
        if matricula.estado != 'P':
            messages.error(request, f'Este estudiante está inactivo ({matricula.get_estado_display()}). No se pueden agregar ni editar notas.')
            return redirect('principal:calificacion_detalle_estudiante', student_id=matricula.student.id, course_id=matricula.course.id)

        # ── Recopilar notas enviadas ──────────────────────────────────────────
        # El template envía: nota_id[], nota_valor[], nota_delete[]
        ids     = request.POST.getlist('nota_id')
        valores = request.POST.getlist('nota_valor')
        deletes = request.POST.getlist('nota_delete')  # lista de ids a eliminar

        errores = []
        filas_validas = []

        for nota_id, valor_str in zip(ids, valores):
            nota_id = nota_id.strip()
            valor_str = valor_str.strip()

            # Si está marcada para eliminar, la procesamos aparte
            if nota_id in deletes:
                continue

            # Fila vacía (nueva sin valor) → ignorar
            if not nota_id and not valor_str:
                continue

            # Validar valor
            if valor_str:
                try:
                    valor = float(valor_str.replace(',', '.'))
                    if not (0 <= valor <= 10):
                        errores.append(f'El valor {valor_str} está fuera del rango 0-10.')
                        continue
                except ValueError:
                    errores.append(f'"{valor_str}" no es un número válido.')
                    continue
            else:
                valor = None

            filas_validas.append({'id': nota_id, 'valor': valor})

        if errores:
            calificacion = Calificaciones.objects.filter(
                course=matricula.course,
                student=matricula.student,
                curso_academico=matricula.curso_academico,
            ).prefetch_related('notas').first()
            notas_existentes = list(calificacion.notas.order_by('id')) if calificacion else []
            for e in errores:
                messages.error(request, e)
            return render(request, 'add_nota.html', {
                'matricula': matricula,
                'notas_existentes': notas_existentes,
                'next': request.POST.get('next', ''),
            })

        # ── Obtener/crear Calificaciones ──────────────────────────────────────
        calificacion = self._get_calificacion(matricula)

        # ── Eliminar notas marcadas ───────────────────────────────────────────
        if deletes:
            NotaIndividual.objects.filter(
                calificacion=calificacion,
                id__in=[d for d in deletes if d],
            ).delete()

        # ── Guardar / actualizar notas válidas ────────────────────────────────
        for fila in filas_validas:
            if fila['valor'] is None:
                continue
            if fila['id']:
                # Actualizar existente
                NotaIndividual.objects.filter(
                    id=fila['id'], calificacion=calificacion
                ).update(valor=fila['valor'])
            else:
                # Crear nueva
                NotaIndividual.objects.create(
                    calificacion=calificacion,
                    valor=fila['valor'],
                )

        # Recalcular promedio
        calificacion.save()

        messages.success(request, 'Calificaciones guardadas correctamente.')
        return self._redirect_after_save(request, matricula)

#esto es de la ia
""" def crear_cursos(request):
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES) # Asegúrate de incluir request.FILES aquí
        if form.is_valid():
            form.save()
            # Por ejemplo, puedes añadir un mensaje de éxito y redirigir
            # messages.success(request, 'Curso creado exitosamente!')
            # return redirect('nombre_de_tu_url_de_cursos')
    else:
        form = CourseForm()
    return render(request, 'create_course.html', {'form': form})


def editar_curso(request, course_id):
    course = get_object_or_404(Curso, id=course_id)
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, instance=course) # Asegúrate de incluir request.FILES aquí
        if form.is_valid():
            form.save()
            # Por ejemplo, puedes añadir un mensaje de éxito y redirigir
            # messages.success(request, 'Curso actualizado exitosamente!')
            # return redirect('nombre_de_tu_url_de_cursos')
    else:
        form = CourseForm(instance=course)
    return render(request, 'edit_course.html', {'form': form, 'course': course})

 """

#vistas para historicos

def historico_alumno(request, student_id):
        matriculas = Matriculas.objects.filter(student_id=student_id).select_related('curso_academico')
        return render(request, 'historico.html', {'matriculas': matriculas})


# Agregando asistencias

class AsistenciaView(BaseContextMixin, TemplateView):
    template_name = 'asistencias.html'
    
    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_id = kwargs['course_id']
        course = get_object_or_404(Curso, id=course_id)
        
        # Obtener el curso académico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        
        # Filtrar asistencias — incluir todos los estudiantes del curso académico activo
        asistencias = Asistencia.objects.filter(
            course=course,
            student__matriculas__curso_academico=curso_academico_activo,
        ).select_related('student', 'course').order_by('-date').distinct()
        
        # Todas las matrículas del curso: activos (estado='P') primero, luego inactivos
        matriculas_qs = Matriculas.objects.filter(
            course=course,
            curso_academico=curso_academico_activo
        ).select_related('student').order_by(
            Case(When(estado='P', then=0), default=1, output_field=IntegerField()),
            'student__first_name', 'student__last_name'
        )
        # Anotar cada matrícula con flag de activo en semestre para el template
        matriculas = list(matriculas_qs)
        for m in matriculas:
            m.es_activo_semestre = (m.estado == 'P')
        
        # Filtrar por fecha si se proporciona en la solicitud
        fecha_filtro = self.request.GET.get('fecha')
        if fecha_filtro:
            asistencias = asistencias.filter(date=fecha_filtro)
        
        # Calcular la cantidad de asistencias registradas (fechas únicas)
        asistencias_registradas = Asistencia.objects.filter(course=course).values('date').distinct().count()
        
        # Obtener la cantidad total de clases del curso
        cantidad_total_clases = course.class_quantity
        
        # Calcular las clases restantes
        clases_restantes = cantidad_total_clases - asistencias_registradas

        context['course'] = course
        context['asistencias'] = asistencias
        context['matriculas'] = matriculas
        context['curso_academico'] = curso_academico_activo
        context['cantidad_total_clases'] = cantidad_total_clases
        context['asistencias_registradas'] = asistencias_registradas
        context['clases_restantes'] = clases_restantes
        return context


class AddAsistenciaView(LoginRequiredMixin, TemplateView):
    template_name='add_asistencias.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_id = kwargs['course_id']
        course = Curso.objects.get(id=course_id)
        # Todas las matrículas: activos (estado='P') primero, inactivos al final
        matriculas_qs = Matriculas.objects.filter(course=course).select_related('student').order_by(
            Case(When(estado='P', then=0), default=1, output_field=IntegerField()),
            'student__first_name', 'student__last_name'
        )
        # Anotar flag de activo en semestre para el template
        matriculas = list(matriculas_qs)
        for m in matriculas:
            m.es_activo_semestre = (m.estado == 'P')
        asistencias = Asistencia.objects.order_by('date')

        asistencias_registradas = Asistencia.objects.filter(course=course).values('date').distinct().count()
        clases_restantes = course.class_quantity - asistencias_registradas

        context['course'] = course
        context['matriculas'] = matriculas
        context['asistencias'] = asistencias
        context['today'] = date.today()
        context['clases_restantes'] = clases_restantes
        context['modo_modificar'] = clases_restantes <= 0
        return context

    def post(self, request, course_id):
        course = Curso.objects.get(id=course_id)
        # Solo procesar matrículas con estado 'P' (activo en el semestre)
        matriculas = Matriculas.objects.filter(course=course, estado='P')

        if request.method == 'POST':
            date_str = request.POST.get('date')

            # Calcular clases restantes
            asistencias_registradas = Asistencia.objects.filter(course=course).values('date').distinct().count()
            clases_restantes = course.class_quantity - asistencias_registradas

            # Verificar si la fecha ya tiene asistencias registradas
            fecha_ya_existe = Asistencia.objects.filter(course=course, date=date_str).exists()

            # Si no quedan clases y la fecha es nueva, bloquear
            if clases_restantes <= 0 and not fecha_ya_existe:
                messages.error(request, "No quedan clases disponibles. Solo puedes modificar asistencias de fechas ya registradas.")
                return redirect('principal:add_asistencias', course_id=course_id)

            for matricula in matriculas:
                absent = request.POST.get('asistencia_' + str(matricula.id))
                # Buscar si ya existe un registro de asistencia para este estudiante en esta fecha
                asistencia, created = Asistencia.objects.get_or_create(
                    student=matricula.student,
                    course=course,
                    date=date_str,
                    defaults={'presente': not bool(absent)}
                )
                # Si el registro ya existía, actualizar el estado de presente
                if not created:
                    asistencia.presente = not bool(absent)
                    asistencia.save()

        # Redirigir a la misma página para mostrar las asistencias actualizadas
        return redirect('principal:asistencias', course_id=course_id)


# Eliminar asistencia
def eliminar_asistencia(request, asistencia_id):
    # Obtener la asistencia o devolver 404 si no existe
    asistencia = get_object_or_404(Asistencia, id=asistencia_id)
    
    # Guardar el ID del curso antes de eliminar la asistencia
    course_id = asistencia.course.id
    
    # Eliminar la asistencia
    asistencia.delete()
    
    # Redirigir a la página de asistencias del curso
    return redirect('principal:asistencias', course_id=course_id)


def add_asistencias(request, course_id):
    course = get_object_or_404(Curso, id=course_id)
    # Solo estudiantes con estado 'P' (activo en el semestre)
    matriculas = Matriculas.objects.filter(course=course, estado='P')

    if request.method == 'POST':
        date_str = request.POST.get('date')
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return redirect('principal:add_asistencias', course_id=course.id)

        # Eliminar asistencias existentes para esta fecha y curso
        Asistencia.objects.filter(course=course, date=attendance_date).delete()

        for matricula in matriculas:
            is_absent = request.POST.get(f'asistencia_{matricula.id}')
            presente = not bool(is_absent) # Si está marcado, significa que está ausente, por lo tanto, no presente

            Asistencia.objects.create(
                course=course,
                student=matricula.student,
                date=attendance_date,
                presente=presente
            )
        messages.success(request, "Asistencias guardadas correctamente.")
        return redirect('principal:asistencias', course_id=course.id) # Redirige a la página de asistencias del curso
    
    today = date.today()
    context = {
        'course': course,
        'matriculas': matriculas,
        'today': today,
    }
    return render(request, 'add_asistencias.html', context)


@login_required
def undo_last_asistencia(request, course_id):
    course = get_object_or_404(Curso, id=course_id)

    # Encontrar el registro de asistencia más reciente para este curso
    # Ordenar por fecha y luego por ID para obtener el último registro creado
    latest_attendance = Asistencia.objects.filter(course=course).order_by('-date', '-id').first()

    if latest_attendance:
        # Obtener información del registro antes de eliminarlo
        student_name = latest_attendance.student.get_full_name() or latest_attendance.student.username
        attendance_date = latest_attendance.date.strftime('%d-%m-%Y')
        status = "Presente" if latest_attendance.presente else "Ausente"
        
        # Eliminar únicamente este registro específico
        latest_attendance.delete()
        
        messages.success(request, f"Se deshizo la última asistencia: {student_name} - {status} ({attendance_date})")
    else:
        messages.info(request, "No hay asistencias registradas para deshacer en este curso.")

    return redirect('principal:asistencias', course_id=course.id)

    

# Vistas para el sistema de formularios de aplicación a cursos

class SecretariaRequiredMixin(UserPassesTestMixin):
    """
    Mixin que verifica que el usuario pertenezca al grupo Secretaría.
    """
    def test_func(self):
        return self.request.user.groups.filter(name='Secretaría').exists()

class ProfesorRequiredMixin(UserPassesTestMixin):
    """
    Mixin que verifica que el usuario pertenezca al grupo Profesores.
    """
    def test_func(self):
        return self.request.user.groups.filter(name='Profesores').exists()

class FormularioAplicacionListView(LoginRequiredMixin, SecretariaRequiredMixin, ListView):
    """
    Vista para listar los formularios de aplicación creados por el grupo secretaría.
    """
    model = FormularioAplicacion
    template_name = 'formularios/formulario_list.html'
    context_object_name = 'formularios'

    def get_queryset(self):
        return FormularioAplicacion.objects.all().order_by('-fecha_modificacion')

class FormularioAplicacionCreateView(LoginRequiredMixin, SecretariaRequiredMixin, CreateView):
    """
    Vista para crear un nuevo formulario de aplicación.
    """
    model = FormularioAplicacion
    form_class = FormularioAplicacionForm
    template_name = 'formularios/formulario_form.html'
    
    def get_form(self, form_class=None):
        """
        Personaliza el formulario dinámicamente según si hay curso preseleccionado.
        """
        form = super().get_form(form_class)
        
        # Si hay un curso preseleccionado, ocultar el campo curso del formulario
        curso_id = self.request.GET.get('curso_id')
        if curso_id:
            # Eliminar el campo curso del formulario ya que está preseleccionado
            if 'curso' in form.fields:
                del form.fields['curso']
        
        return form
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Formulario de Aplicación'
        
        # Verificar si se recibió un curso_id en la URL
        curso_id = self.request.GET.get('curso_id')
        if curso_id:
            try:
                curso = Curso.objects.get(id=curso_id)
                context['curso_preseleccionado'] = curso
                return context
            except Curso.DoesNotExist:
                pass
        
        # Obtener los cursos que no tienen formulario de aplicación
        cursos_sin_formulario = Curso.objects.filter(
            curso_academico__activo=True
        ).exclude(
            id__in=FormularioAplicacion.objects.values_list('curso_id', flat=True)
        )
        context['cursos'] = cursos_sin_formulario
        
        return context
        
    def get_initial(self):
        """
        Establece valores iniciales para el formulario.
        """
        initial = super().get_initial()
        initial['descripcion'] = "Por favor, conteste responsablemente todas las preguntas que le hacemos a continuación, eso ayudará a los profesores a una mejor organización del curso. Muchas gracias."
        return initial
    
    def form_valid(self, form):
        curso_id = self.request.POST.get('curso')
        if curso_id:
            curso = get_object_or_404(Curso, id=curso_id)
            form.instance.curso = curso
            response = super().form_valid(form)
            
            # Imprimir información de depuración
            print(f"DEBUG: Formulario creado para curso {curso.name} (ID: {curso.id})")
            print(f"DEBUG: ID del formulario: {self.object.id}")
            print(f"DEBUG: Verificando relación: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
            
            # Limpiar la caché de la sesión para forzar una recarga de los datos
            if 'cursos_con_formularios' in self.request.session:
                del self.request.session['cursos_con_formularios']
            
            return response
        else:
            messages.error(self.request, 'Debe seleccionar un curso.')
            return self.form_invalid(form)
    
    def get_success_url(self):
        # Añadir parámetro para indicar que viene de la creación del formulario
        return reverse('principal:formulario_preguntas', kwargs={'pk': self.object.pk}) + '?from_create=1'

class FormularioAplicacionUpdateView(LoginRequiredMixin, SecretariaRequiredMixin, UpdateView):
    """
    Vista para editar un formulario de aplicación existente.
    """
    model = FormularioAplicacion
    form_class = FormularioAplicacionForm
    template_name = 'formularios/formulario_form.html'
    
    def get_object(self, queryset=None):
        """
        Obtiene el objeto que se va a editar y maneja posibles errores.
        """
        try:
            obj = super().get_object(queryset)
            return obj
        except Exception as e:
            # Registrar el error para depuración
            print(f"Error al obtener el objeto FormularioAplicacion: {e}")
            # Redirigir a la lista de formularios con un mensaje de error
            messages.error(self.request, f"No se pudo encontrar el formulario solicitado. Error: {e}")
            return None
    
    def get(self, request, *args, **kwargs):
        """
        Maneja la solicitud GET y redirige si no se encuentra el objeto.
        """
        self.object = self.get_object()
        if self.object is None:
            return redirect('principal:formulario_list')
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Formulario de Aplicación'
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Formulario de aplicación actualizado correctamente.')
        return response
    
    def get_success_url(self):
        return reverse('principal:formulario_preguntas', kwargs={'pk': self.object.pk})

class FormularioPreguntasView(LoginRequiredMixin, SecretariaRequiredMixin, UpdateView):
    """
    Vista para gestionar las preguntas de un formulario de aplicación.
    """
    model = FormularioAplicacion
    template_name = 'formularios/formulario_preguntas.html'
    fields = []  # No necesitamos campos para esta vista
    
    def get(self, request, *args, **kwargs):
        # Limpiar todos los mensajes existentes para evitar duplicados
        storage = messages.get_messages(request)
        # Consumir todos los mensajes para limpiarlos
        for _ in storage:
            pass
        
        # Si viene de la creación del formulario, mostrar un solo mensaje
        if request.GET.get('from_create'):
            messages.success(request, 'Formulario de aplicación creado correctamente.')
        
        response = super().get(request, *args, **kwargs)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        formulario = self.get_object()
        
        if self.request.POST:
            context['pregunta_formset'] = PreguntaFormularioFormSet(
                self.request.POST, instance=formulario
            )
        else:
            formset = PreguntaFormularioFormSet(instance=formulario)
            
            # Si viene el parámetro add_new, agregar una nueva pregunta con el orden correcto
            if self.request.GET.get('add_new') == '1':
                # Calcular el siguiente orden disponible
                max_orden = formulario.preguntas.aggregate(Max('orden'))['orden__max']
                siguiente_orden = (max_orden + 1) if max_orden is not None else 0
                
                # Crear una nueva instancia de pregunta con el orden correcto
                nueva_pregunta = PreguntaFormulario(formulario=formulario, orden=siguiente_orden)
                
                # Agregar la nueva pregunta a las instancias del formset
                formset.extra = 1
                # Crear un nuevo formset que incluya la nueva pregunta con valores por defecto
                initial_data = [{'orden': siguiente_orden, 'requerida': True}]
                formset = PreguntaFormularioFormSet(instance=formulario, initial=initial_data)
            
            context['pregunta_formset'] = formset
        
        context['formulario'] = formulario
        context['add_new'] = self.request.GET.get('add_new') == '1'
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        pregunta_formset = context['pregunta_formset']
        
        # Imprimir datos del formulario para depuración
        print("\n=== DATOS DEL FORMULARIO RECIBIDOS ===")
        for key, value in self.request.POST.items():
            print(f"{key}: '{value}'")
        print("=====================================\n")
        
        if pregunta_formset.is_valid():
            # Guardar las preguntas
            preguntas = pregunta_formset.save(commit=True)
            print(f"Preguntas guardadas: {preguntas}")
            
            # Asegurarse de que el curso tenga el atributo tiene_formulario
            formulario = self.object
            curso = formulario.curso
            print(f"DEBUG: Formulario {formulario.id} asociado al curso {curso.name} (ID: {curso.id})")
            print(f"DEBUG: Verificando relación después de guardar preguntas: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
            
            # Limpiar la caché de la sesión para forzar una recarga de los datos
            if 'cursos_con_formularios' in self.request.session:
                del self.request.session['cursos_con_formularios']
            
            # Si se está redirigiendo a las opciones, buscar la última pregunta creada
            if self.request.POST.get('redirect_to_options') or self.request.POST.get('save_and_continue'):
                # Obtener la última pregunta creada para este formulario
                ultima_pregunta = self.object.preguntas.order_by('-id').first()
                if ultima_pregunta:
                    # No mostrar mensaje aquí, lo mostraremos en la vista de opciones
                    # Imprimir información de depuración
                    print(f"Redirigiendo a opciones de pregunta {ultima_pregunta.id}: {ultima_pregunta.texto}")
                    # Redirigir directamente a la página de opciones de la pregunta con parámetro
                    return redirect(reverse('principal:pregunta_opciones', kwargs={'pk': ultima_pregunta.pk}) + '?from_redirect=1')
            
            # Limpiar todos los mensajes existentes antes de añadir uno nuevo
            storage = messages.get_messages(self.request)
            for _ in storage:
                pass  # Consumir todos los mensajes
            
            # Añadir un solo mensaje de éxito
            messages.success(self.request, 'Preguntas guardadas correctamente.')
            
            # Redirigir a la página de cursos
            return redirect(reverse('principal:cursos'))
        else:
            print("\n=== ERRORES DEL FORMSET ===")
            for i, form_errors in enumerate(pregunta_formset.errors):
                print(f"Formulario {i}: {form_errors}")
            print("=========================\n")
            return self.render_to_response(self.get_context_data(form=form))
    
    def get_success_url(self):
        # Redirigir a la página de cursos
        return reverse('principal:cursos')

class PreguntaOpcionesView(LoginRequiredMixin, SecretariaRequiredMixin, UpdateView):
    """
    Vista para gestionar las opciones de respuesta de una pregunta.
    """
    model = PreguntaFormulario
    template_name = 'formularios/pregunta_opciones.html'
    fields = []  # No necesitamos campos para esta vista
    
    def get(self, request, *args, **kwargs):
        # Limpiar todos los mensajes existentes para evitar duplicados
        storage = messages.get_messages(request)
        # Consumir todos los mensajes para limpiarlos
        for _ in storage:
            pass
        
        # Si viene de la redirección de una pregunta guardada, mostrar un mensaje
        if request.GET.get('from_redirect'):
            messages.success(request, 'Pregunta guardada correctamente. Ahora puedes agregar opciones de respuesta.')
        
        response = super().get(request, *args, **kwargs)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pregunta = self.get_object()
        
        if self.request.POST:
            context['opcion_formset'] = OpcionRespuestaFormSet(
                self.request.POST, instance=pregunta
            )
        else:
            # Crear el formset con la instancia de la pregunta
            # El formset ya está configurado para añadir una fila extra (extra=1)
            context['opcion_formset'] = OpcionRespuestaFormSet(instance=pregunta)
        
        context['pregunta'] = pregunta
        context['formulario'] = pregunta.formulario
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pregunta = self.object
        
        # Imprimir información de depuración detallada
        print("\n=== DATOS DEL FORMULARIO RECIBIDOS ===")
        for key, value in request.POST.items():
            print(f"{key}: '{value}'")
        print("=====================================\n")
        
        # Procesar los datos del formulario manualmente
        try:
            # Obtener el número total de opciones
            total_opciones = int(request.POST.get('total_opciones', 0))
            print(f"Total de opciones: {total_opciones}")
            
            # Procesar cada opción
            for i in range(total_opciones):
                opcion_id = request.POST.get(f'opcion_id_{i}', '')
                texto = request.POST.get(f'texto_{i}', '')
                orden = request.POST.get(f'orden_{i}', i)
                eliminar = request.POST.get(f'eliminar_{i}', '') == 'on'
                
                print(f"Opción {i}: id='{opcion_id}', texto='{texto}', orden={orden}, eliminar={eliminar}")
                
                # Si la opción está marcada para eliminar, eliminarla
                if eliminar and opcion_id:
                    try:
                        opcion = OpcionRespuesta.objects.get(id=opcion_id)
                        opcion.delete()
                        print(f"  - Opción eliminada: {opcion}")
                        continue
                    except OpcionRespuesta.DoesNotExist:
                        pass
                
                # Si la opción ya existe, actualizarla
                if opcion_id:
                    try:
                        opcion = OpcionRespuesta.objects.get(id=opcion_id)
                        opcion.texto = texto
                        opcion.orden = int(orden) if orden else i
                        opcion.save()
                        print(f"  - Opción actualizada: {opcion}")
                    except OpcionRespuesta.DoesNotExist:
                        pass
                # Si es una nueva opción, crearla
                else:
                    opcion = OpcionRespuesta(
                        pregunta=pregunta,
                        texto=texto,
                        orden=int(orden) if orden else i
                    )
                    opcion.save()
                    print(f"  - Opción creada: {opcion}")
            
            # Limpiar todos los mensajes existentes antes de añadir uno nuevo
            storage = messages.get_messages(self.request)
            for _ in storage:
                pass  # Consumir todos los mensajes
            
            messages.success(self.request, 'Opciones de respuesta guardadas correctamente.')
            return redirect(self.get_success_url())
        except Exception as e:
            print(f"Error al procesar el formulario: {e}")
            messages.error(self.request, f'Error al guardar las opciones: {e}')
            return self.get(request, *args, **kwargs)
    

    
    def form_invalid(self, opcion_formset):
        # Mostrar errores específicos
        print("\n=== ERRORES DEL FORMSET ===")
        for i, form_errors in enumerate(opcion_formset.errors):
            print(f"Formulario {i}: {form_errors}")
            for field, errors in form_errors.items():
                for error in errors:
                    messages.error(self.request, f"Error en formulario {i}, campo {field}: {error}")
        print("=========================\n")
        
        # Preparar el contexto para renderizar la respuesta
        context = self.get_context_data()
        context['opcion_formset'] = opcion_formset
        return self.render_to_response(context)
    
    def get_success_url(self):
        return reverse('principal:formulario_preguntas', kwargs={'pk': self.object.formulario.pk})

# Vistas para los estudiantes

@login_required
def aplicar_curso(request, curso_id):
    """
    Vista para que un estudiante aplique a un curso mediante un formulario dinámico.
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar si el curso tiene un formulario de aplicación
    try:
        # Intentar obtener el formulario directamente
        formulario = FormularioAplicacion.objects.get(curso_id=curso_id)
        if not formulario:
            messages.error(request, 'Este curso no tiene un formulario de aplicación disponible.')
            return redirect('principal:cursos')
    except FormularioAplicacion.DoesNotExist:
        messages.error(request, 'Este curso no tiene un formulario de aplicación disponible.')
        return redirect('principal:cursos')
    
    # Verificar si el estudiante ya ha aplicado a este curso
    solicitud_existente = SolicitudInscripcion.objects.filter(
        curso=curso,
        estudiante=request.user
    ).first()
    
    if solicitud_existente:
        messages.info(request, 'Ya has aplicado a este curso. Tu solicitud está en proceso de revisión.')
        return redirect('principal:cursos')
    
    # Verificar si el estudiante ya está matriculado en este curso
    matricula_existente = Matriculas.objects.filter(
        course=curso,
        student=request.user
    ).exists()
    
    if matricula_existente:
        messages.info(request, 'Ya estás matriculado en este curso.')
        return redirect('principal:cursos')
    
    # Obtener las preguntas del formulario
    preguntas = formulario.preguntas.all().order_by('orden')
    
    if request.method == 'POST':
        # Validar que todas las preguntas tengan respuesta
        preguntas_sin_responder = []
        for pregunta in preguntas:
            if pregunta.tipo == 'seleccion_multiple':
                if not request.POST.getlist(f'pregunta_{pregunta.id}'):
                    preguntas_sin_responder.append(pregunta)
            elif pregunta.tipo == 'seleccion_unica':
                if not request.POST.get(f'pregunta_{pregunta.id}'):
                    preguntas_sin_responder.append(pregunta)
            elif pregunta.tipo == 'escritura_libre':
                if not request.POST.get(f'pregunta_{pregunta.id}', '').strip():
                    preguntas_sin_responder.append(pregunta)

        if preguntas_sin_responder:
            cantidad = len(preguntas_sin_responder)
            if cantidad == 1:
                msg = 'Hay 1 pregunta obligatoria sin responder. Por favor completa todas las preguntas requeridas.'
            else:
                msg = f'Hay {cantidad} preguntas obligatorias sin responder. Por favor completa todas las preguntas requeridas.'
            messages.error(request, msg)
            formularios_preguntas = []
            for pregunta in preguntas:
                form = RespuestaEstudianteForm(pregunta=pregunta)
                formularios_preguntas.append((pregunta, form))
            context = {
                'curso': curso,
                'formulario': formulario,
                'formularios_preguntas': formularios_preguntas
            }
            return render(request, 'formularios/aplicar_curso.html', context)

        try:
            # Crear la solicitud de inscripción
            solicitud = SolicitudInscripcion.objects.create(
                curso=curso,
                estudiante=request.user,
                formulario=formulario,
                estado='pendiente'  # Asegurarse de que el estado sea 'pendiente'
            )
            
            # Mensaje de depuración
            print(f"DEBUG: Solicitud creada - ID: {solicitud.id}, Estado: {solicitud.estado}, Curso: {curso.name}, Estudiante: {request.user.username}")
            
            # Verificar que la solicitud se haya creado correctamente
            solicitud_verificada = SolicitudInscripcion.objects.filter(
                id=solicitud.id
            ).first()
            
            if solicitud_verificada:
                print(f"DEBUG: Solicitud verificada - ID: {solicitud_verificada.id}, Estado: {solicitud_verificada.estado}")
            else:
                print("ERROR: No se pudo verificar la solicitud creada")
        except Exception as e:
            print(f"ERROR al crear la solicitud: {str(e)}")
            raise
        
        # Procesar las respuestas
        for pregunta in preguntas:
            respuesta = RespuestaEstudiante.objects.create(
                solicitud=solicitud,
                pregunta=pregunta
            )
            
            # Obtener las opciones seleccionadas
            if pregunta.tipo == 'seleccion_multiple':
                opcion_ids = request.POST.getlist(f'pregunta_{pregunta.id}')
                for opcion_id in opcion_ids:
                    opcion = get_object_or_404(OpcionRespuesta, id=opcion_id)
                    respuesta.opciones_seleccionadas.add(opcion)
            elif pregunta.tipo == 'seleccion_unica':
                opcion_id = request.POST.get(f'pregunta_{pregunta.id}')
                if opcion_id:
                    opcion = get_object_or_404(OpcionRespuesta, id=opcion_id)
                    respuesta.opciones_seleccionadas.add(opcion)
            elif pregunta.tipo == 'escritura_libre':
                # Para preguntas de escritura libre, creamos una opción de respuesta con el texto ingresado
                texto_respuesta = request.POST.get(f'pregunta_{pregunta.id}', '')
                if texto_respuesta:
                    # Crear una opción de respuesta para almacenar el texto
                    opcion = OpcionRespuesta.objects.create(
                        pregunta=pregunta,
                        texto=texto_respuesta,
                        orden=0
                    )
                    respuesta.opciones_seleccionadas.add(opcion)
        
        # En lugar de redirigir a la lista de cursos, redirigimos a una página de confirmación
        # que luego redirigirá automáticamente a la lista de cursos
        request.session['solicitud_enviada_curso_id'] = curso_id
        return redirect('principal:solicitud_enviada', curso_id=curso_id)
    
    # Crear formularios dinámicos para cada pregunta
    formularios_preguntas = []
    for pregunta in preguntas:
        form = RespuestaEstudianteForm(pregunta=pregunta)
        formularios_preguntas.append((pregunta, form))
    
    context = {
        'curso': curso,
        'formulario': formulario,
        'formularios_preguntas': formularios_preguntas
    }
    
    return render(request, 'formularios/aplicar_curso.html', context)

# Vista para mostrar la página de confirmación después de enviar una solicitud
@login_required
def solicitud_enviada(request, curso_id):
    """
    Vista para mostrar una página de confirmación después de enviar una solicitud.
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar si el estudiante realmente envió una solicitud para este curso
    solicitud_existente = SolicitudInscripcion.objects.filter(
        curso=curso,
        estudiante=request.user
    ).exists()
    
    if not solicitud_existente:
        # Si no existe una solicitud, redirigir a la lista de cursos
        return redirect('principal:cursos')
    
    # Imprimir mensaje de depuración
    print(f"DEBUG: Mostrando página de confirmación para solicitud del curso {curso.name} (ID: {curso.id})")
    
    return render(request, 'formularios/solicitud_enviada.html', {'curso': curso})

# Vista para mostrar el reglamento del curso
def reglamento_curso(request):
    """
    Vista para mostrar el reglamento del curso.
    """
    return render(request, 'formularios/reglamento_curso.html')

def reglamento_general(request):
    """
    Vista para mostrar el reglamento general del centro.
    Si existe un ReglamentoGeneral dinámico lo muestra; si no, muestra el estático.
    """
    reglamento = ReglamentoGeneral.objects.prefetch_related('articulos').first()
    if reglamento and (reglamento.introduccion.strip() or reglamento.articulos.exists()):
        articulos = reglamento.articulos.order_by('orden', 'fecha_creacion')
        context = {
            'reglamento': reglamento,
            'articulos': articulos,
            'ano_anterior': timezone.now().year - 1,
            'ano_actual': timezone.now().year,
            'dinamico': True,
        }
    else:
        context = {
            'reglamento': None,
            'articulos': [],
            'ano_anterior': timezone.now().year - 1,
            'ano_actual': timezone.now().year,
            'dinamico': False,
        }
    return render(request, 'registration/reglamento_general.html', context)

def exportar_reglamento_general_pdf(request):
    """
    Vista para exportar el reglamento general del centro en PDF.
    Si existe un ReglamentoGeneral dinámico lo usa; si no, usa el contenido estático.
    """
    reglamento = ReglamentoGeneral.objects.prefetch_related('articulos').first()
    dinamico = reglamento and (reglamento.introduccion.strip() or reglamento.articulos.exists())

    context = {
        'fecha_generacion': timezone.now(),
        'ano_anterior': timezone.now().year - 1,
        'ano_actual': timezone.now().year,
        'reglamento': reglamento if dinamico else None,
        'articulos': reglamento.articulos.order_by('orden', 'fecha_creacion') if dinamico else [],
        'dinamico': bool(dinamico),
    }
    
    # Generar el PDF usando la plantilla específica para PDF
    template = get_template('registration/reglamento_general_pdf.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reglamento_General_CFBC.pdf"'
        return response
    
    return HttpResponse('Error al generar el PDF', status=500)


# ---------------------------------------------------------------------------
# Vistas para el Reglamento del Curso (Requisitos 1.1–1.6, 2.x, 3.x, 4.x)
# ---------------------------------------------------------------------------

class ReglamentoCursoCreateView(LoginRequiredMixin, SecretariaRequiredMixin, CreateView):
    """
    Vista para crear un nuevo ReglamentoCurso para un curso específico.
    Requisitos: 1.1, 1.2, 1.4, 1.6, 2.2, 2.5
    """
    template_name = 'formularios/reglamento_curso_form.html'
    model = ReglamentoCurso
    form_class = ReglamentoCursoForm

    def get_curso(self):
        return get_object_or_404(Curso, pk=self.kwargs['curso_id'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso = self.get_curso()
        context['curso'] = curso
        if self.request.method == 'POST':
            context['form'] = ReglamentoCursoForm(self.request.POST)
            context['articulo_formset'] = ArticuloReglamentoFormSet(self.request.POST)
        else:
            context['form'] = ReglamentoCursoForm()
            context['articulo_formset'] = ArticuloReglamentoFormSet()
        return context

    def post(self, request, *args, **kwargs):
        curso = self.get_curso()
        form = ReglamentoCursoForm(request.POST)
        formset = ArticuloReglamentoFormSet(request.POST)

        # Informar al formset el valor de introduccion para la validación cruzada
        formset.set_introduccion(request.POST.get('introduccion', ''))

        if form.is_valid() and formset.is_valid():
            reglamento = form.save(commit=False)
            reglamento.curso = curso
            reglamento.save()

            formset.instance = reglamento
            formset.save()

            messages.success(request, f'Reglamento del curso "{curso.name}" creado exitosamente.')
            return redirect(reverse('principal:cursos'))

        # Re-render con errores
        context = {
            'curso': curso,
            'form': form,
            'articulo_formset': formset,
        }
        return render(request, self.template_name, context)


class ReglamentoCursoUpdateView(LoginRequiredMixin, SecretariaRequiredMixin, UpdateView):
    """
    Vista para editar un ReglamentoCurso existente.
    Requisitos: 1.3, 1.5, 1.6, 2.8, 2.9
    """
    template_name = 'formularios/reglamento_curso_form.html'
    model = ReglamentoCurso
    form_class = ReglamentoCursoForm

    def get_reglamento(self):
        return get_object_or_404(ReglamentoCurso, curso__pk=self.kwargs['curso_id'])

    def get_object(self, queryset=None):
        return self.get_reglamento()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        reglamento = self.get_reglamento()
        curso = reglamento.curso
        context['curso'] = curso
        if self.request.method == 'POST':
            context['form'] = ReglamentoCursoForm(self.request.POST, instance=reglamento)
            context['articulo_formset'] = ArticuloReglamentoFormSet(self.request.POST, instance=reglamento)
        else:
            context['form'] = ReglamentoCursoForm(instance=reglamento)
            context['articulo_formset'] = ArticuloReglamentoFormSet(instance=reglamento)
        return context

    def post(self, request, *args, **kwargs):
        reglamento = self.get_reglamento()
        curso = reglamento.curso
        form = ReglamentoCursoForm(request.POST, instance=reglamento)
        formset = ArticuloReglamentoFormSet(request.POST, instance=reglamento)

        # Informar al formset el valor de introduccion para la validación cruzada
        formset.set_introduccion(request.POST.get('introduccion', ''))

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()

            messages.success(request, f'Reglamento del curso "{curso.name}" actualizado exitosamente.')
            return redirect(reverse('principal:cursos'))

        # Re-render con errores
        context = {
            'curso': curso,
            'form': form,
            'articulo_formset': formset,
        }
        return render(request, self.template_name, context)


def reglamento_curso_detalle(request, curso_id):
    """
    Vista pública para que los estudiantes vean el reglamento de un curso específico.
    Si el curso no tiene reglamento dinámico, muestra el reglamento estático de respaldo.
    Requisitos: 3.1, 3.2, 3.3, 3.4, 3.5, 3.6, 5.3
    """
    curso = get_object_or_404(Curso, pk=curso_id)
    reglamento = getattr(curso, 'reglamento_curso', None)

    if reglamento is not None:
        articulos = reglamento.articulos.order_by('orden', 'fecha_creacion')
        context = {
            'curso': curso,
            'reglamento': reglamento,
            'articulos': articulos,
            'ano_actual': timezone.now().year,
        }
        return render(request, 'formularios/reglamento_curso_detalle.html', context)

    # Fallback: reglamento estático genérico (Req. 3.5)
    return render(request, 'formularios/reglamento_curso.html', {'curso': curso})


def exportar_reglamento_curso_pdf(request, curso_id):
    """
    Vista para exportar el reglamento de un curso específico en PDF.
    Requisitos: 4.1, 4.2, 4.3, 4.4, 4.5
    """
    curso = get_object_or_404(Curso, pk=curso_id)
    reglamento = get_object_or_404(ReglamentoCurso, curso=curso)
    articulos = reglamento.articulos.order_by('orden', 'fecha_creacion')

    context = {
        'curso': curso,
        'reglamento': reglamento,
        'articulos': articulos,
        'ano_actual': timezone.now().year,
    }

    template = get_template('formularios/reglamento_curso_pdf.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode('UTF-8')), result)

    if not pdf.err:
        # Sanitizar el nombre del curso para el nombre del archivo (Req. 4.4)
        nombre_normalizado = unicodedata.normalize('NFKD', curso.name)
        nombre_ascii = nombre_normalizado.encode('ascii', 'ignore').decode('ascii')
        nombre_sanitizado = nombre_ascii.replace(' ', '_')
        # Eliminar caracteres no seguros para cabeceras HTTP / nombres de archivo
        nombre_sanitizado = ''.join(
            c for c in nombre_sanitizado
            if c.isalnum() or c in ('_', '-', '.')
        )
        nombre_archivo = f'Reglamento_{nombre_sanitizado}.pdf'

        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response

    return HttpResponse('Error al generar el PDF del reglamento', status=500)


# Vistas para los profesores

class SolicitudesInscripcionListView(LoginRequiredMixin, ProfesorRequiredMixin, ListView):
    """
    Vista para que un profesor vea las solicitudes de inscripción a sus cursos.
    """
    model = SolicitudInscripcion
    template_name = 'formularios/solicitudes_list.html'
    context_object_name = 'solicitudes'
    paginate_by = 10  # Mostrar 10 solicitudes por página
    
    def get_queryset(self):
        # Obtener el filtro de estado desde los parámetros GET
        estado_filtro = self.request.GET.get('estado', 'pendiente')
        
        # Obtener las solicitudes de los cursos que imparte el profesor filtradas por estado
        # Solo mostrar solicitudes de cursos en etapa de inscripción (I o IT)
        return SolicitudInscripcion.objects.filter(
            curso__teacher=self.request.user,
            estado=estado_filtro,
            curso__status__in=['I', 'IT']  # Solo cursos en inscripción
        ).order_by('-fecha_solicitud')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener el estado actual del filtro
        estado_filtro = self.request.GET.get('estado', 'pendiente')
        context['estado_filtro'] = estado_filtro
        
        # Agregar estadísticas de solicitudes (contadores)
        # Solo contar solicitudes de cursos en etapa de inscripción (I o IT)
        context['total_pendientes'] = SolicitudInscripcion.objects.filter(
            curso__teacher=self.request.user,
            estado='pendiente',
            curso__status__in=['I', 'IT']
        ).count()
        
        context['total_aprobadas'] = SolicitudInscripcion.objects.filter(
            curso__teacher=self.request.user,
            estado='aprobada',
            curso__status__in=['I', 'IT']
        ).count()
        
        context['total_rechazadas'] = SolicitudInscripcion.objects.filter(
            curso__teacher=self.request.user,
            estado='rechazada',
            curso__status__in=['I', 'IT']
        ).count()
        
        return context

class SolicitudInscripcionDetailView(LoginRequiredMixin, ProfesorRequiredMixin, DetailView):
    """
    Vista para que un profesor vea el detalle de una solicitud de inscripción.
    """
    model = SolicitudInscripcion
    template_name = 'formularios/solicitud_detail.html'
    context_object_name = 'solicitud'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        solicitud = self.get_object()
        
        # Verificar que el profesor sea el profesor del curso
        if solicitud.curso.teacher != self.request.user:
            raise PermissionDenied
        
        # Obtener las respuestas del estudiante
        respuestas = solicitud.respuestas.all().select_related('pregunta').prefetch_related('opciones_seleccionadas')
        context['respuestas'] = respuestas
        
        # Obtener la siguiente solicitud pendiente
        siguiente_solicitud = SolicitudInscripcion.objects.filter(
            curso__teacher=self.request.user,
            estado='pendiente',
            fecha_solicitud__gt=solicitud.fecha_solicitud
        ).order_by('fecha_solicitud').first()
        
        # Si no hay siguiente con fecha mayor, buscar la primera pendiente
        if not siguiente_solicitud:
            siguiente_solicitud = SolicitudInscripcion.objects.filter(
                curso__teacher=self.request.user,
                estado='pendiente'
            ).exclude(id=solicitud.id).order_by('fecha_solicitud').first()
        
        context['siguiente_solicitud'] = siguiente_solicitud
        
        return context

@login_required
def aprobar_solicitud(request, pk):
    """
    Vista para que un profesor apruebe una solicitud de inscripción.
    """
    solicitud = get_object_or_404(SolicitudInscripcion, pk=pk)
    
    # Verificar que el profesor sea el profesor del curso
    if solicitud.curso.teacher != request.user:
        raise PermissionDenied
    
    # Aprobar la solicitud
    matricula = solicitud.aprobar(request.user)
    
    # Enviar correo de confirmación al estudiante
    try:
        nombre_estudiante = solicitud.estudiante.get_full_name() or solicitud.estudiante.username
        nombre_curso = solicitud.curso.name
        email_estudiante = solicitud.estudiante.email
        
        if email_estudiante:
            asunto = f'¡Enhorabuena! Su aplicación al curso {nombre_curso} ha sido aprobada'
            mensaje = f'''¡Enhorabuena! Su aplicación al curso "{nombre_curso}" ha sido aprobada.

Ya puede acceder al curso y comenzar con las actividades académicas.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            send_mail(
                asunto,
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [email_estudiante],
                fail_silently=False,
            )
            print(f"Correo enviado exitosamente a {email_estudiante} para el curso {nombre_curso}")
        else:
            print(f"No se pudo enviar correo: el estudiante {nombre_estudiante} no tiene email registrado")
    except Exception as e:
        print(f"Error al enviar correo de aprobación: {str(e)}")
        # No interrumpimos el proceso si falla el envío del correo
    
    # Agregar un solo mensaje de éxito
    messages.success(request, f'La solicitud de {solicitud.estudiante.get_full_name() or solicitud.estudiante.username} ha sido aprobada.')
    
    # Redirigir a la misma página de detalle de la solicitud
    return redirect('principal:solicitud_detail', pk=solicitud.id)

@login_required
def rechazar_solicitud(request, pk):
    """
    Vista para que un profesor rechace una solicitud de inscripción.
    """
    solicitud = get_object_or_404(SolicitudInscripcion, pk=pk)
    
    # Verificar que el profesor sea el profesor del curso
    if solicitud.curso.teacher != request.user:
        raise PermissionDenied
    
    # Rechazar la solicitud
    solicitud.rechazar(request.user)
    
    # Enviar correo de notificación al estudiante
    try:
        nombre_estudiante = solicitud.estudiante.get_full_name() or solicitud.estudiante.username
        nombre_curso = solicitud.curso.name
        email_estudiante = solicitud.estudiante.email
        
        if email_estudiante:
            asunto = f'Su aplicación al curso {nombre_curso} ha sido denegada'
            mensaje = f'''Lo sentimos! Su aplicación al curso "{nombre_curso}" ha sido denegada.

Le recomendamos revisar los requisitos del curso y considerar aplicar en futuras convocatorias.

Si tiene alguna pregunta, no dude en contactarnos.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            send_mail(
                asunto,
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [email_estudiante],
                fail_silently=False,
            )
            print(f"Correo de denegación enviado exitosamente a {email_estudiante} para el curso {nombre_curso}")
        else:
            print(f"No se pudo enviar correo: el estudiante {nombre_estudiante} no tiene email registrado")
    except Exception as e:
        print(f"Error al enviar correo de denegación: {str(e)}")
        # No interrumpimos el proceso si falla el envío del correo
    
    messages.success(request, f'La solicitud de {solicitud.estudiante.get_full_name() or solicitud.estudiante.username} ha sido rechazada.')
    
    # Redirigir a la misma página de detalle de la solicitud
    return redirect('principal:solicitud_detail', pk=solicitud.id)

@login_required
def deshacer_aprobacion_solicitud(request, pk):
    """
    Vista para que un profesor deshaga la aprobación de una solicitud de inscripción.
    """
    solicitud = get_object_or_404(SolicitudInscripcion, pk=pk)
    
    # Verificar que el profesor sea el profesor del curso
    if solicitud.curso.teacher != request.user:
        raise PermissionDenied
    
    # Verificar que la solicitud esté aprobada
    if solicitud.estado != 'aprobada':
        messages.error(request, 'Esta solicitud no está aprobada.')
        return redirect('principal:solicitud_detail', pk=solicitud.id)
    
    # Buscar y eliminar la matrícula asociada
    try:
        from principal.models import Matriculas
        matricula = Matriculas.objects.get(
            student=solicitud.estudiante,
            course=solicitud.curso
        )
        matricula.delete()
        print(f"Matrícula eliminada para {solicitud.estudiante.username} en {solicitud.curso.name}")
    except Matriculas.DoesNotExist:
        print(f"No se encontró matrícula para {solicitud.estudiante.username} en {solicitud.curso.name}")
    except Exception as e:
        print(f"Error al eliminar matrícula: {str(e)}")
    
    # Cambiar el estado de la solicitud a pendiente
    solicitud.estado = 'pendiente'
    solicitud.fecha_revision = None
    solicitud.revisado_por = None
    solicitud.save()
    
    messages.success(request, f'Se ha deshecho la aprobación de la solicitud de {solicitud.estudiante.get_full_name() or solicitud.estudiante.username}. La solicitud vuelve a estar pendiente.')
    
    # Redirigir a la misma página de detalle de la solicitud
    return redirect('principal:solicitud_detail', pk=solicitud.id)

@login_required
def deshacer_rechazo_solicitud(request, pk):
    """
    Vista para que un profesor deshaga el rechazo de una solicitud de inscripción.
    """
    solicitud = get_object_or_404(SolicitudInscripcion, pk=pk)
    
    # Verificar que el profesor sea el profesor del curso
    if solicitud.curso.teacher != request.user:
        raise PermissionDenied
    
    # Verificar que la solicitud esté rechazada
    if solicitud.estado != 'rechazada':
        messages.error(request, 'Esta solicitud no está rechazada.')
        return redirect('principal:solicitud_detail', pk=solicitud.id)
    
    # Cambiar el estado de la solicitud a pendiente
    solicitud.estado = 'pendiente'
    solicitud.fecha_revision = None
    solicitud.revisado_por = None
    solicitud.save()
    
    messages.success(request, f'Se ha deshecho el rechazo de la solicitud de {solicitud.estudiante.get_full_name() or solicitud.estudiante.username}. La solicitud vuelve a estar pendiente.')
    
    # Redirigir a la misma página de detalle de la solicitud
    return redirect('principal:solicitud_detail', pk=solicitud.id)

@login_required
def exportar_solicitudes_excel(request):
    """
    Vista para exportar las solicitudes de inscripción a Excel según el filtro de estado.
    """
    # Verificar permisos - solo profesores y secretarías pueden exportar
    if not request.user.groups.filter(name__in=['Profesores', 'Secretaría']).exists():
        messages.error(request, 'No tienes permisos para exportar solicitudes.')
        return redirect('principal:solicitudes_list')
    
    # Obtener el filtro de estado
    estado_filtro = request.GET.get('estado', 'pendiente')
    
    # Obtener las solicitudes filtradas
    # Solo exportar solicitudes de cursos en etapa de inscripción (I o IT)
    solicitudes = SolicitudInscripcion.objects.filter(
        curso__teacher=request.user,
        estado=estado_filtro,
        curso__status__in=['I', 'IT']  # Solo cursos en inscripción
    ).select_related('estudiante', 'curso', 'revisado_por').order_by('-fecha_solicitud')
    
    if not solicitudes.exists():
        messages.warning(request, 'No hay solicitudes para exportar con el filtro seleccionado.')
        return redirect('principal:solicitudes_list')
    
    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Solicitudes {estado_filtro.capitalize()}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'N°',
        'Estudiante',
        'Email',
        'Curso',
        'Estado',
        'Fecha Solicitud',
        'Fecha Revisión',
        'Revisado Por'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row_num, solicitud in enumerate(solicitudes, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = border
        ws.cell(row=row_num, column=2, value=solicitud.estudiante.get_full_name() or solicitud.estudiante.username).border = border
        ws.cell(row=row_num, column=3, value=solicitud.estudiante.email).border = border
        ws.cell(row=row_num, column=4, value=solicitud.curso.name).border = border
        ws.cell(row=row_num, column=5, value=solicitud.get_estado_display()).border = border
        ws.cell(row=row_num, column=6, value=solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row_num, column=7, value=solicitud.fecha_revision.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_revision else 'N/A').border = border
        ws.cell(row=row_num, column=8, value=solicitud.revisado_por.get_full_name() if solicitud.revisado_por else 'N/A').border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'solicitudes_{estado_filtro}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def guardar_pregunta_y_redirigir(request, formulario_id):
    """
    Vista para guardar una pregunta y redirigir a la página de opciones.
    """
    # Verificar que el usuario pertenezca al grupo 'Secretaría'
    if not request.user.groups.filter(name='Secretaría').exists():
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('principal:formulario_list')
    
    # Obtener el formulario
    formulario = get_object_or_404(FormularioAplicacion, pk=formulario_id)
    
    if request.method == 'POST':
        # Crear una nueva pregunta
        requerida_value = request.POST.get('requerida', 'True')
        requerida = requerida_value.lower() == 'true' if isinstance(requerida_value, str) else bool(requerida_value)
        
        # Calcular el siguiente orden disponible
        max_orden = formulario.preguntas.aggregate(models.Max('orden'))['orden__max']
        siguiente_orden = (max_orden + 1) if max_orden is not None else 0
        
        # Imprimir información de depuración
        print(f"Guardando pregunta para formulario {formulario_id}")
        print(f"Texto: {request.POST.get('texto', '')}")
        print(f"Tipo: {request.POST.get('tipo', 'seleccion_multiple')}")
        print(f"Requerida: {requerida}")
        print(f"Orden calculado automáticamente: {siguiente_orden}")
        
        pregunta = PreguntaFormulario(
            formulario=formulario,
            texto=request.POST.get('texto', ''),
            tipo=request.POST.get('tipo', 'seleccion_multiple'),
            requerida=requerida,
            orden=siguiente_orden
        )
        pregunta.save()
        
        # Imprimir información de la pregunta guardada
        print(f"Pregunta guardada con ID: {pregunta.pk}")
        
        # Limpiar la caché de la sesión para forzar una recarga de los datos
        if 'cursos_con_formularios' in request.session:
            del request.session['cursos_con_formularios']
        
        # No mostramos mensaje aquí para evitar duplicación, ya que la vista FormularioPreguntasView ya muestra un mensaje
        
        # Usar una redirección con JavaScript
        from django.http import HttpResponse
        redirect_url = reverse('principal:pregunta_opciones', kwargs={'pk': pregunta.pk}) + '?from_redirect=1'
        return HttpResponse(f"""
            <html>
                <head>
                    <title>Redirigiendo...</title>
                    <script>
                        window.location.href = "{redirect_url}";
                    </script>
                </head>
                <body>
                    <p>Redirigiendo a la página de opciones de respuesta...</p>
                    <p>Si no eres redirigido automáticamente, <a href="{redirect_url}">haz clic aquí</a>.</p>
                </body>
            </html>
        """)
    
    # Si no es POST, redirigir a la página de preguntas del formulario
    return redirect('principal:formulario_preguntas', pk=formulario_id)

@login_required
def eliminar_formulario(request, pk):
    """
    Vista para eliminar un formulario de aplicación.
    """
    # Verificar que el usuario pertenezca al grupo Secretaría
    if not request.user.groups.filter(name='Secretaría').exists():
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('principal:cursos')
    
    formulario = get_object_or_404(FormularioAplicacion, pk=pk)
    curso = formulario.curso
    curso_id = curso.id  # Guardar el ID del curso antes de eliminar el formulario
    
    # Imprimir información de depuración antes de eliminar
    print(f"DEBUG: Eliminando formulario {pk} del curso {curso.name} (ID: {curso_id})")
    print(f"DEBUG: Verificando relación antes de eliminar: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
    
    # Eliminar el formulario
    formulario.delete()
    
    # Verificar que se haya eliminado correctamente
    print(f"DEBUG: Verificando relación después de eliminar: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
    
    # Limpiar la caché de la sesión para forzar una recarga de los datos
    if 'cursos_con_formularios' in request.session:
        del request.session['cursos_con_formularios']
    
    messages.success(request, 'El formulario de aplicación ha sido eliminado correctamente.')
    return redirect('principal:cursos')

# Vistas para recuperación de contraseña
import random
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.contrib.auth.models import User
from accounts.models import Registro

def password_reset_request(request):
    """
    Vista para solicitar el restablecimiento de contraseña.
    El usuario ingresa su correo electrónico y se le envía un código de verificación.
    """
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            # Generar código aleatorio de 4 dígitos
            verification_code = str(random.randint(1000, 9999))
            # Almacenar datos temporales en la sesión
            request.session['reset_verification_code'] = verification_code
            request.session['reset_user_id'] = user.id
            
            # Enviar email con el código de verificación
            email_text = f'Para restablecer su contraseña en el Centro Fray Bartolome de las Casas, ingrese el siguiente código: {verification_code}'
            try:
                send_mail(
                    'Código para Restablecer Contraseña - Centro Fray Bartolome de las Casas',
                    email_text,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                messages.success(request, 'Se ha enviado un código de verificación a su correo electrónico.')
                return redirect('principal:password_reset_verify')
            except Exception as e:
                print(f"Error al enviar email: {str(e)}")
                messages.error(request, 'Error al enviar el código de verificación. Por favor, intente nuevamente más tarde.')
        except User.DoesNotExist:
            messages.error(request, 'No existe una cuenta con ese correo electrónico.')
    
    return render(request, 'registration/password_reset_request.html')

def password_reset_verify(request):
    """
    Vista para verificar el código enviado al correo electrónico.
    """
    if 'reset_verification_code' not in request.session or 'reset_user_id' not in request.session:
        messages.error(request, 'La sesión ha expirado. Por favor, inicie el proceso nuevamente.')
        return redirect('principal:password_reset_request')
    
    if request.method == 'POST':
        code = request.POST.get('code')
        if code == request.session.get('reset_verification_code'):
            return redirect('principal:password_reset_confirm')
        else:
            messages.error(request, 'El código ingresado no es válido. Por favor, intente nuevamente.')
    
    return render(request, 'registration/password_reset_verify.html')

def password_reset_resend_code(request):
    """
    Vista para reenviar el código de verificación de recuperación de contraseña.
    """
    if 'reset_user_id' not in request.session:
        messages.error(request, 'La sesión ha expirado. Por favor, inicie el proceso nuevamente.')
        return redirect('principal:password_reset_request')
    
    try:
        user = User.objects.get(id=request.session.get('reset_user_id'))
        
        # Generar nuevo código aleatorio de 4 dígitos
        verification_code = str(random.randint(1000, 9999))
        request.session['reset_verification_code'] = verification_code
        
        # Enviar email con el nuevo código
        email_text = f'Para restablecer su contraseña en el Centro Fray Bartolome de las Casas, ingrese el siguiente código: {verification_code}'
        try:
            send_mail(
                'Código para Restablecer Contraseña - Centro Fray Bartolome de las Casas',
                email_text,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            messages.success(request, 'Se ha reenviado un nuevo código de verificación a su correo electrónico.')
        except Exception as e:
            print(f"Error al enviar email: {str(e)}")
            messages.error(request, 'Error al reenviar el código de verificación. Por favor, intente nuevamente más tarde.')
    except User.DoesNotExist:
        messages.error(request, 'Ha ocurrido un error. Por favor, inicie el proceso nuevamente.')
        return redirect('principal:password_reset_request')
    
    return redirect('principal:password_reset_verify')

def password_reset_confirm(request):
    """
    Vista para establecer la nueva contraseña después de verificar el código.
    """
    if 'reset_user_id' not in request.session:
        messages.error(request, 'La sesión ha expirado. Por favor, inicie el proceso nuevamente.')
        return redirect('principal:password_reset_request')
    
    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        if password1 != password2:
            messages.error(request, 'Las contraseñas no coinciden. Por favor, inténtelo nuevamente.')
            return render(request, 'registration/password_reset_confirm.html')
        
        try:
            user = User.objects.get(id=request.session.get('reset_user_id'))
            user.password = make_password(password1)
            user.save()
            
            # Enviar correo de confirmación de cambio de contraseña
            try:
                nombre_usuario = user.get_full_name() or user.username
                email_usuario = user.email
                
                if email_usuario:
                    asunto = 'Su contraseña ha sido cambiada satisfactoriamente'
                    mensaje = f'''Estimado/a {nombre_usuario},

Su contraseña ha sido cambiada satisfactoriamente en el Centro Fray Bartolomé de las Casas.

Si usted no realizó este cambio, por favor contacte inmediatamente con el administrador del sistema.

Fecha y hora del cambio: {timezone.now().strftime('%d/%m/%Y a las %H:%M')}

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
                    
                    send_mail(
                        asunto,
                        mensaje,
                        settings.DEFAULT_FROM_EMAIL,
                        [email_usuario],
                        fail_silently=False,
                    )
                    print(f"Correo de confirmación de cambio de contraseña enviado a {email_usuario}")
                else:
                    print(f"No se pudo enviar correo: el usuario {nombre_usuario} no tiene email registrado")
            except Exception as e:
                print(f"Error al enviar correo de confirmación de cambio de contraseña: {str(e)}")
                # No interrumpimos el proceso si falla el envío del correo
            
            # Limpiar datos de sesión
            del request.session['reset_verification_code']
            del request.session['reset_user_id']
            
            messages.success(request, 'Su contraseña ha sido restablecida exitosamente. Ahora puede iniciar sesión con su nueva contraseña.')
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, 'Ha ocurrido un error. Por favor, inicie el proceso nuevamente.')
            return redirect('principal:password_reset_request')
    
    return render(request, 'registration/password_reset_confirm.html')

# Vistas para validación AJAX en tiempo real
def validate_username(request):
    """Validar si el username ya existe"""
    if request.method == 'GET':
        username = request.GET.get('username', '').strip()
        if username:
            exists = User.objects.filter(username=username).exists()
            return JsonResponse({
                'exists': exists,
                'message': f"El nombre de usuario '{username}' ya existe." if exists else "Nombre de usuario disponible."
            })
    return JsonResponse({'exists': False, 'message': ''})

def validate_email(request):
    """Validar si el email ya existe"""
    if request.method == 'GET':
        email = request.GET.get('email', '').strip()
        if email:
            exists = User.objects.filter(email=email).exists()
            return JsonResponse({
                'exists': exists,
                'message': f"El correo '{email}' ya está registrado." if exists else "Correo disponible."
            })
    return JsonResponse({'exists': False, 'message': ''})

def validate_carnet(request):
    """Validar si el carnet ya existe"""
    if request.method == 'GET':
        carnet = request.GET.get('carnet', '').strip()
        if carnet:
            exists = Registro.objects.filter(carnet=carnet).exists()
            return JsonResponse({
                'exists': exists,
                'message': f"El carnet '{carnet}' ya está registrado." if exists else "Carnet disponible."
            })
    return JsonResponse({'exists': False, 'message': ''})

@login_required
def exportar_solicitud_excel(request, solicitud_id):
    """
    Vista para exportar los datos de una solicitud individual a Excel.
    Incluye información del estudiante, solicitud y respuestas del formulario.
    """
    # Verificar permisos - solo profesores y secretarías pueden exportar
    if not request.user.groups.filter(name__in=['Profesores', 'Secretaría']).exists():
        messages.error(request, 'No tienes permisos para exportar esta solicitud.')
        return redirect('principal:solicitudes_list')
    
    # Obtener la solicitud
    solicitud = get_object_or_404(SolicitudInscripcion, pk=solicitud_id)
    
    # Obtener las respuestas del estudiante
    respuestas = RespuestaEstudiante.objects.filter(solicitud=solicitud).order_by('pregunta__orden')
    
    # Obtener información adicional del estudiante si existe
    registro_estudiante = None
    try:
        registro_estudiante = Registro.objects.get(user=solicitud.estudiante)
    except Registro.DoesNotExist:
        pass
    
    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitud de Inscripción"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, color="000000")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal
    ws.merge_cells('A1:D1')
    ws['A1'] = f"SOLICITUD DE INSCRIPCIÓN - {solicitud.curso.name}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = border
    
    row = 3
    
    # Información del Estudiante
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "INFORMACIÓN DEL ESTUDIANTE"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].border = border
    row += 1
    
    # Datos del estudiante
    estudiante_data = [
        ("Nombre Completo", solicitud.estudiante.get_full_name() or solicitud.estudiante.username),
        ("Nombre de Usuario", solicitud.estudiante.username),
        ("Correo Electrónico", solicitud.estudiante.email),
    ]
    
    # Agregar datos adicionales si existe el registro
    if registro_estudiante:
        estudiante_data.extend([
            ("Carnet", registro_estudiante.carnet or "No especificado"),
            ("Nacionalidad", registro_estudiante.nacionalidad or "No especificada"),
            ("Sexo", registro_estudiante.get_sexo_display() if registro_estudiante.sexo else "No especificado"),
            ("Teléfono", registro_estudiante.telephone or "No especificado"),
            ("Móvil", registro_estudiante.movil or "No especificado"),
            ("Dirección", registro_estudiante.address or "No especificada"),
            ("Municipio", registro_estudiante.location or "No especificado"),
            ("Provincia", registro_estudiante.provincia or "No especificada"),
            ("Grado Académico", registro_estudiante.get_grado_display() if registro_estudiante.grado else "No especificado"),
            ("Ocupación", registro_estudiante.get_ocupacion_display() if registro_estudiante.ocupacion else "No especificada"),
            ("Título", registro_estudiante.titulo or "No especificado"),
        ])
    
    for campo, valor in estudiante_data:
        ws[f'A{row}'] = campo
        ws[f'B{row}'] = valor
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 1
    
    # Información de la Solicitud
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "INFORMACIÓN DE LA SOLICITUD"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].border = border
    row += 1
    
    solicitud_data = [
        ("Curso", solicitud.curso.name),
        ("Profesor", solicitud.curso.teacher.get_full_name() or solicitud.curso.teacher.username),
        ("Estado", solicitud.get_estado_display()),
        ("Fecha de Solicitud", solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M')),
    ]
    
    if solicitud.fecha_revision:
        solicitud_data.append(("Fecha de Revisión", solicitud.fecha_revision.strftime('%d/%m/%Y %H:%M')))
    
    if solicitud.revisado_por:
        solicitud_data.append(("Revisado por", solicitud.revisado_por.get_full_name() or solicitud.revisado_por.username))
    
    for campo, valor in solicitud_data:
        ws[f'A{row}'] = campo
        ws[f'B{row}'] = valor
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 1
    
    # Respuestas del Formulario
    if respuestas:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"RESPUESTAS DEL FORMULARIO: {solicitud.formulario.titulo}"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].border = border
        row += 1
        
        for i, respuesta in enumerate(respuestas, 1):
            # Pregunta
            ws[f'A{row}'] = f"Pregunta {i}"
            ws[f'B{row}'] = respuesta.pregunta.texto
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            # Tipo de pregunta
            ws[f'A{row}'] = "Tipo"
            ws[f'B{row}'] = respuesta.pregunta.get_tipo_display()
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            # Respuesta
            ws[f'A{row}'] = "Respuesta"
            opciones = respuesta.opciones_seleccionadas.all()
            if opciones:
                if respuesta.pregunta.tipo == 'escritura_libre':
                    # Para escritura libre, mostrar el texto completo
                    respuesta_texto = '\n'.join([opcion.texto for opcion in opciones])
                else:
                    # Para opciones múltiples o únicas, mostrar como lista
                    respuesta_texto = ', '.join([opcion.texto for opcion in opciones])
                ws[f'B{row}'] = respuesta_texto
            else:
                ws[f'B{row}'] = "Sin respuesta"
            
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 2
    else:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "NO HAY RESPUESTAS REGISTRADAS"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo
    estudiante_nombre = solicitud.estudiante.get_full_name() or solicitud.estudiante.username
    curso_nombre = solicitud.curso.name
    filename = f'Solicitud_{estudiante_nombre}_{curso_nombre}_{solicitud.fecha_solicitud.strftime("%Y%m%d")}.xlsx'
    filename = filename.replace(' ', '_').replace('/', '_').replace('\\', '_')
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar el archivo en la respuesta
    wb.save(response)
    
    return response
@login_required
@login_required
def obtener_historial_usuario(request, user_id):
    """
    Vista AJAX para obtener el historial COMPLETO de un usuario.
    Retorna datos de TODAS las 11 tablas históricas de Docencia.
    """
    from historial.models import (
        HistoricalArea,
        HistoricalCourseCategory,
        HistoricalCourseInformation,
        HistoricalCourseInformationAdminTeachers,
        HistoricalEnrollmentApplication,
        HistoricalAccountNumber,
        HistoricalEnrollmentPay,
        HistoricalSubjectInformation,
        HistoricalEdition,
        HistoricalEnrollment,
        HistoricalApplication,
        HistoricalClass,
        HistoricalClassStudentView,
    )
    
    # Verificar que el usuario sea secretaria
    if not request.user.groups.filter(name='Secretaría').exists():
        return JsonResponse({'error': 'No tiene permisos para ver esta información'}, status=403)
    
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    
    # Estructura completa de datos históricos
    historial_data = {
        'usuario': {
            'nombre': user.get_full_name() or user.username,
            'email': user.email,
            'username': user.username,
        },
        'aplicaciones': [],
        'matriculas': [],
        'solicitudes_inscripcion': [],
        'cuentas_bancarias': [],
        'pagos': [],
        'cursos_como_profesor': [],
        'cursos_administrados': [],
        'ediciones': [],
        'asignaturas': [],
        'areas': [],
        'categorias': [],
        'clases': []
    }
    
    # 1. APLICACIONES (Docencia_application)
    aplicaciones, usuario_ids_originales = _buscar_aplicaciones_historicas(user)
    usuario_id_original = usuario_ids_originales[0] if usuario_ids_originales else None
    for app in aplicaciones:
        historial_data['aplicaciones'].append({
            'id': app.id,
            'curso': app.curso.nombre if app.curso else 'N/A',
            'curso_codigo': app.curso.codigo if app.curso else 'N/A',
            'area': app.curso.area.nombre if app.curso and app.curso.area else 'N/A',
            'categoria': app.curso.categoria.nombre if app.curso and app.curso.categoria else 'N/A',
            'edicion': app.edicion.nombre if app.edicion else 'N/A',
            'edicion_fecha_inicio': app.edicion.fecha_inicio.strftime('%d/%m/%Y') if app.edicion and app.edicion.fecha_inicio else 'N/A',
            'edicion_fecha_fin': app.edicion.fecha_fin.strftime('%d/%m/%Y') if app.edicion and app.edicion.fecha_fin else 'N/A',
            'fecha_solicitud': app.fecha_solicitud.strftime('%d/%m/%Y') if app.fecha_solicitud else 'N/A',
            'estado': app.estado or 'N/A',
            'beca': 'Sí' if app.beca else 'No',
            'pagado': 'Sí' if app.pagado else 'No',
            'nota_primaria': app.nota_primaria,
            'nota_secundaria': app.nota_secundaria,
            'nota_final': app.nota_final,
            'nota_extra': app.nota_extra,
            'comentarios': app.comentarios or 'Sin comentarios',
            'tabla_origen': app.tabla_origen,
            'fecha_consolidacion': app.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })
    
    # 2. MATRÍCULAS (Docencia_enrollment)
    matriculas = _buscar_matriculas_historicas(user, usuario_ids_originales)
    for mat in matriculas:
        historial_data['matriculas'].append({
            'id': mat.id,
            'curso': mat.edicion.curso.nombre if mat.edicion and mat.edicion.curso else (mat.curso.nombre if mat.curso else 'N/A'),
            'edicion': mat.edicion.nombre if mat.edicion else 'N/A',
            'fecha_inscripcion': mat.fecha_inscripcion.strftime('%d/%m/%Y %H:%M'),
            'estado': mat.estado or 'N/A',
            'ausencias': mat.ausencias,
            'intento': mat.intento,
            'nota_primaria': mat.nota_primaria if hasattr(mat, 'nota_primaria') else 'N/A',
            'nota_secundaria': mat.nota_secundaria if hasattr(mat, 'nota_secundaria') else 'N/A',
            'nota_final': mat.nota_final,
            'nota_extra': mat.nota_extra if hasattr(mat, 'nota_extra') else 'N/A',
            'slug': mat.slug if hasattr(mat, 'slug') else 'N/A',
            'tabla_origen': mat.tabla_origen,
            'fecha_consolidacion': mat.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })
    
    # 3. SOLICITUDES DE INSCRIPCIÓN (Docencia_enrollmentapplication)
    solicitudes = _buscar_solicitudes_historicas(user, usuario_ids_originales)
    for sol in solicitudes:
        historial_data['solicitudes_inscripcion'].append({
            'id': sol.id,
            'curso': sol.curso.nombre if sol.curso else 'N/A',
            'curso_codigo': sol.curso.codigo if sol.curso else 'N/A',
            'fecha_solicitud': sol.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            'estado': sol.estado or 'N/A',
            'tabla_origen': sol.tabla_origen,
            'fecha_consolidacion': sol.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })
    
    # 4. CUENTAS BANCARIAS (Docencia_accountnumber)
    # user_id apunta directamente a auth_user
    from datos_archivados.models import DatoArchivadoDinamico
    if usuario_ids_originales:
        cuentas = HistoricalAccountNumber.objects.filter(
            id_original__in=DatoArchivadoDinamico.objects.filter(
                tabla_origen='Docencia_accountnumber',
                datos_originales__user_id__in=usuario_ids_originales
            ).values_list('id_original', flat=True)
        ) | HistoricalAccountNumber.objects.filter(usuario=user)
        cuentas = cuentas.distinct()
    else:
        cuentas = HistoricalAccountNumber.objects.filter(usuario=user)
    for cuenta in cuentas:
        historial_data['cuentas_bancarias'].append({
            'id': cuenta.id,
            'numero_cuenta': cuenta.numero_cuenta,
            'banco': cuenta.banco or 'N/A',
            'tabla_origen': cuenta.tabla_origen,
            'fecha_consolidacion': cuenta.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })

    # 5. PAGOS (Docencia_enrollmentpay)
    # app_id apunta al ID original de Docencia_application en la BD antigua
    # que está guardado en dato_archivado.datos_originales['id']
    aplicaciones_ids_originales = []
    for app in aplicaciones:
        if app.dato_archivado:
            orig_id = app.dato_archivado.datos_originales.get('id')
            if orig_id:
                aplicaciones_ids_originales.append(orig_id)
        elif app.id_original:
            aplicaciones_ids_originales.append(app.id_original)

    if aplicaciones_ids_originales:
        pagos_data = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_enrollmentpay',
            datos_originales__app_id__in=aplicaciones_ids_originales
        )
        for pago_dato in pagos_data:
            datos = pago_dato.datos_originales
            historial_data['pagos'].append({
                'id': datos.get('id'),
                'monto': datos.get('monto', datos.get('amount', 'N/A')),
                'fecha': datos.get('datepub', datos.get('date', 'N/A')),
                'metodo': datos.get('transfernumber', datos.get('method', 'N/A')),
                'referencia': datos.get('cardnumber_id', datos.get('reference', 'N/A')),
                'aceptado': 'Sí' if datos.get('accept') else 'No',
                'tabla_origen': 'Docencia_enrollmentpay',
            })
    
    # 6. CURSOS COMO PROFESOR (Docencia_courseinformation_adminteachers)
    cursos_profesor = HistoricalCourseInformationAdminTeachers.objects.filter(
        profesor=user
    ).select_related('curso', 'curso__area', 'curso__categoria')
    for cp in cursos_profesor:
        historial_data['cursos_como_profesor'].append({
            'id': cp.id,
            'curso': cp.curso.nombre if cp.curso else 'N/A',
            'curso_codigo': cp.curso.codigo if cp.curso else 'N/A',
            'area': cp.curso.area.nombre if cp.curso and cp.curso.area else 'N/A',
            'categoria': cp.curso.categoria.nombre if cp.curso and cp.curso.categoria else 'N/A',
            'tabla_origen': cp.tabla_origen,
            'fecha_consolidacion': cp.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })
    
    # 7. CURSOS ADMINISTRADOS - Obtener cursos donde el usuario es admin
    cursos_admin = HistoricalCourseInformation.objects.filter(
        admin_teachers__profesor=user
    ).distinct()
    for curso in cursos_admin:
        historial_data['cursos_administrados'].append({
            'id': curso.id,
            'nombre': curso.nombre,
            'codigo': curso.codigo if hasattr(curso, 'codigo') else 'N/A',
            'descripcion': curso.descripcion if hasattr(curso, 'descripcion') else 'N/A',
            'area': curso.area.nombre if curso.area else 'N/A',
            'categoria': curso.categoria.nombre if curso.categoria else 'N/A',
            'tabla_origen': curso.tabla_origen,
            'fecha_consolidacion': curso.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })
    
    # 8. EDICIONES - Ediciones de cursos donde el usuario participó
    ediciones_ids = set()
    for app in aplicaciones:
        if app.edicion:
            ediciones_ids.add(app.edicion.id)
    for mat in matriculas:
        if mat.edicion:
            ediciones_ids.add(mat.edicion.id)
    
    if ediciones_ids:
        ediciones = HistoricalEdition.objects.filter(id__in=ediciones_ids).select_related('curso')
        for ed in ediciones:
            historial_data['ediciones'].append({
                'id': ed.id,
                'nombre': ed.nombre,
                'curso': ed.curso.nombre if ed.curso else 'N/A',
                'fecha_inicio': ed.fecha_inicio.strftime('%d/%m/%Y') if ed.fecha_inicio else 'N/A',
                'fecha_fin': ed.fecha_fin.strftime('%d/%m/%Y') if ed.fecha_fin else 'N/A',
                'tabla_origen': ed.tabla_origen,
                'fecha_consolidacion': ed.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })
    
    # 9. ASIGNATURAS - Asignaturas de los cursos del usuario
    cursos_ids = set()
    for app in aplicaciones:
        if app.curso:
            cursos_ids.add(app.curso.id)
    
    if cursos_ids:
        asignaturas = HistoricalSubjectInformation.objects.filter(
            curso_id__in=cursos_ids
        ).select_related('curso')
        for asig in asignaturas:
            historial_data['asignaturas'].append({
                'id': asig.id,
                'nombre': asig.nombre,
                'codigo': asig.codigo if hasattr(asig, 'codigo') else 'N/A',
                'curso': asig.curso.nombre if asig.curso else 'N/A',
                'descripcion': asig.descripcion if hasattr(asig, 'descripcion') else 'N/A',
                'tabla_origen': asig.tabla_origen,
                'fecha_consolidacion': asig.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })
    
    # 10. ÁREAS - Áreas de los cursos del usuario
    areas_ids = set()
    for app in aplicaciones:
        if app.curso and app.curso.area:
            areas_ids.add(app.curso.area.id)
    
    if areas_ids:
        areas = HistoricalArea.objects.filter(id__in=areas_ids)
        for area in areas:
            historial_data['areas'].append({
                'id': area.id,
                'nombre': area.nombre,
                'codigo': area.codigo if hasattr(area, 'codigo') else 'N/A',
                'descripcion': area.descripcion or 'Sin descripción',
                'tabla_origen': area.tabla_origen,
                'fecha_consolidacion': area.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })
    
    # 11. CATEGORÍAS - Categorías de los cursos del usuario
    categorias_ids = set()
    for app in aplicaciones:
        if app.curso and app.curso.categoria:
            categorias_ids.add(app.curso.categoria.id)
    for curso in cursos_admin:
        if curso.categoria:
            categorias_ids.add(curso.categoria.id)
    
    if categorias_ids:
        categorias = HistoricalCourseCategory.objects.filter(id__in=categorias_ids)
        for cat in categorias:
            historial_data['categorias'].append({
                'id': cat.id,
                'nombre': cat.nombre,
                'codigo': cat.codigo if hasattr(cat, 'codigo') else 'N/A',
                'descripcion': cat.descripcion or 'Sin descripción',
                'precio': str(cat.precio) if hasattr(cat, 'precio') and cat.precio else 'N/A',
                'es_servicio': 'Sí' if hasattr(cat, 'es_servicio') and cat.es_servicio else 'No',
                'registro_abierto': 'Sí' if hasattr(cat, 'registro_abierto') and cat.registro_abierto else 'No',
                'tabla_origen': cat.tabla_origen,
                'fecha_consolidacion': cat.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })
    

    # 12. CLASES (Docencia_class) - clases donde el usuario tiene student views
    clases_ids = set()
    for app in aplicaciones:
        for csv in app.class_views.select_related('class_field').all():
            clases_ids.add(csv.class_field.id)

    if clases_ids:
        clases = HistoricalClass.objects.filter(id__in=clases_ids).select_related('subject')
        for clase in clases:
            historial_data['clases'].append({
                'id': clase.id,
                'nombre': clase.name,
                'asignatura': clase.subject.nombre if clase.subject else 'N/A',
                'fecha_carga': clase.uploaddate.strftime('%d/%m/%Y') if clase.uploaddate else 'N/A',
                'fecha_publicacion': clase.datepub.strftime('%d/%m/%Y %H:%M') if clase.datepub else 'N/A',
                'fecha_fin': clase.dateend.strftime('%d/%m/%Y %H:%M') if clase.dateend else 'N/A',
                'slug': clase.slug,
                'fecha_consolidacion': clase.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })

    return JsonResponse(historial_data)


def _buscar_aplicaciones_historicas(user):
    """
    Busca HistoricalApplication para un usuario cubriendo todos los casos.

    La cadena real es:
        auth_user.id -> Docencia_studentpersonalinformation.user_id
                     -> Docencia_application.student_id

    También cubre el caso donde se guardó user_id directo, y FK directa.
    """
    from datos_archivados.models import DatoArchivadoDinamico
    from historial.models import HistoricalApplication
    from django.db.models import Q

    # Paso 1: encontrar el/los IDs del usuario en auth_user archivado
    datos_user = DatoArchivadoDinamico.objects.filter(
        tabla_origen='auth_user',
        datos_originales__username=user.username
    )
    if not datos_user.exists() and user.email:
        datos_user = DatoArchivadoDinamico.objects.filter(
            tabla_origen='auth_user',
            datos_originales__email=user.email
        )
    usuario_ids_originales = [
        d.datos_originales.get('id')
        for d in datos_user
        if d.datos_originales.get('id')
    ]

    if not usuario_ids_originales:
        return HistoricalApplication.objects.filter(usuario=user).select_related(
            'curso', 'edicion', 'curso__area', 'curso__categoria'
        ), []

    # Paso 2: encontrar los IDs de Docencia_studentpersonalinformation para este usuario
    student_info_ids = list(
        DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_studentpersonalinformation',
            datos_originales__user_id__in=usuario_ids_originales
        ).values_list('id_original', flat=True)
    )

    # Paso 3: buscar aplicaciones por student_id (cadena correcta),
    # por user_id directo (algunos registros lo usan), y por FK directa
    q = Q(usuario=user)

    if student_info_ids:
        ids_por_student_id = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_application',
            datos_originales__student_id__in=student_info_ids
        ).values_list('id_original', flat=True)
        q |= Q(id_original__in=ids_por_student_id)

    # También cubrir si algún registro usó user_id directo de auth_user
    ids_por_user_id = DatoArchivadoDinamico.objects.filter(
        tabla_origen='Docencia_application',
        datos_originales__user_id__in=usuario_ids_originales
    ).values_list('id_original', flat=True)
    q |= Q(id_original__in=ids_por_user_id)

    aplicaciones = HistoricalApplication.objects.filter(q).distinct().select_related(
        'curso', 'edicion', 'curso__area', 'curso__categoria'
    )
    return aplicaciones, usuario_ids_originales


def _buscar_matriculas_historicas(user, usuario_ids_originales):
    """
    Busca HistoricalEnrollment.
    Cadena: auth_user -> studentpersonalinformation.user_id -> enrollment.student_id
    """
    from datos_archivados.models import DatoArchivadoDinamico
    from historial.models import HistoricalEnrollment
    from django.db.models import Q

    if not usuario_ids_originales:
        return HistoricalEnrollment.objects.filter(usuario=user).select_related(
            'edicion', 'edicion__curso', 'curso'
        )

    student_info_ids = list(
        DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_studentpersonalinformation',
            datos_originales__user_id__in=usuario_ids_originales
        ).values_list('id_original', flat=True)
    )

    q = Q(usuario=user)
    if student_info_ids:
        ids_por_student_id = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_enrollment',
            datos_originales__student_id__in=student_info_ids
        ).values_list('id_original', flat=True)
        q |= Q(id_original__in=ids_por_student_id)

    ids_por_user_id = DatoArchivadoDinamico.objects.filter(
        tabla_origen='Docencia_enrollment',
        datos_originales__user_id__in=usuario_ids_originales
    ).values_list('id_original', flat=True)
    q |= Q(id_original__in=ids_por_user_id)

    return HistoricalEnrollment.objects.filter(q).distinct().select_related(
        'edicion', 'edicion__curso', 'curso'
    )


def _buscar_solicitudes_historicas(user, usuario_ids_originales):
    """Busca HistoricalEnrollmentApplication cubriendo user_id y student_id."""
    from datos_archivados.models import DatoArchivadoDinamico
    from historial.models import HistoricalEnrollmentApplication
    from django.db.models import Q

    if usuario_ids_originales:
        ids_por_user_id = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_enrollmentapplication',
            datos_originales__user_id__in=usuario_ids_originales
        ).values_list('id_original', flat=True)
        ids_por_student_id = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_enrollmentapplication',
            datos_originales__student_id__in=usuario_ids_originales
        ).values_list('id_original', flat=True)
        # También buscar por nombre completo (campo legacy)
        ids_por_nombre = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_enrollmentapplication',
            datos_originales__name=user.get_full_name()
        ).values_list('id_original', flat=True) if user.get_full_name() else []
        todos_ids = set(list(ids_por_user_id) + list(ids_por_student_id) + list(ids_por_nombre))
        return HistoricalEnrollmentApplication.objects.filter(
            Q(id_original__in=todos_ids) | Q(usuario=user)
        ).distinct().select_related('curso')
    return HistoricalEnrollmentApplication.objects.filter(usuario=user).select_related('curso')


@login_required
def ver_detalles_historial_usuario(request, user_id):
    """
    Vista para mostrar el historial completo de un usuario en una página completa.
    Similar a obtener_historial_usuario pero renderiza un template en lugar de retornar JSON.
    """
    from historial.models import (
        HistoricalArea,
        HistoricalCourseCategory,
        HistoricalCourseInformation,
        HistoricalCourseInformationAdminTeachers,
        HistoricalEnrollmentApplication,
        HistoricalAccountNumber,
        HistoricalEnrollmentPay,
        HistoricalSubjectInformation,
        HistoricalEdition,
        HistoricalEnrollment,
        HistoricalApplication,
        HistoricalClass,
        HistoricalClassStudentView,
    )

    # Verificar que el usuario sea secretaria
    if not request.user.groups.filter(name='Secretaría').exists():
        messages.error(request, 'No tiene permisos para ver esta información')
        return redirect('principal:profile')

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, 'Usuario no encontrado')
        return redirect('principal:usuarios_registrados')

    # Estructura completa de datos históricos
    historial_data = {
        'usuario': {
            'nombre': user.get_full_name() or user.username,
            'email': user.email,
            'username': user.username,
        },
        'aplicaciones': [],
        'matriculas': [],
        'solicitudes_inscripcion': [],
        'cuentas_bancarias': [],
        'pagos': [],
        'cursos_como_profesor': [],
        'cursos_administrados': [],
        'ediciones': [],
        'asignaturas': [],
        'areas': [],
        'categorias': [],
        'clases': []
    }

    # 1. APLICACIONES (Docencia_application)
    aplicaciones, usuario_ids_originales = _buscar_aplicaciones_historicas(user)
    usuario_id_original = usuario_ids_originales[0] if usuario_ids_originales else None
    for app in aplicaciones:
        historial_data['aplicaciones'].append({
            'id': app.id,
            'curso': app.curso.nombre if app.curso else 'N/A',
            'curso_codigo': app.curso.codigo if app.curso else 'N/A',
            'area': app.curso.area.nombre if app.curso and app.curso.area else 'N/A',
            'categoria': app.curso.categoria.nombre if app.curso and app.curso.categoria else 'N/A',
            'edicion': app.edicion.nombre if app.edicion else 'N/A',
            'edicion_fecha_inicio': app.edicion.fecha_inicio.strftime('%d/%m/%Y') if app.edicion and app.edicion.fecha_inicio else 'N/A',
            'edicion_fecha_fin': app.edicion.fecha_fin.strftime('%d/%m/%Y') if app.edicion and app.edicion.fecha_fin else 'N/A',
            'fecha_solicitud': app.fecha_solicitud.strftime('%d/%m/%Y') if app.fecha_solicitud else 'N/A',
            'estado': app.estado or 'N/A',
            'beca': 'Sí' if app.beca else 'No',
            'pagado': 'Sí' if app.pagado else 'No',
            'nota_primaria': app.nota_primaria,
            'nota_secundaria': app.nota_secundaria,
            'nota_final': app.nota_final,
            'nota_extra': app.nota_extra,
            'comentarios': app.comentarios or 'Sin comentarios',
            'tabla_origen': app.tabla_origen,
            'fecha_consolidacion': app.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })

    # 2. MATRÍCULAS (Docencia_enrollment)
    matriculas = _buscar_matriculas_historicas(user, usuario_ids_originales)
    for mat in matriculas:
        historial_data['matriculas'].append({
            'id': mat.id,
            'curso': mat.edicion.curso.nombre if mat.edicion and mat.edicion.curso else (mat.curso.nombre if mat.curso else 'N/A'),
            'edicion': mat.edicion.nombre if mat.edicion else 'N/A',
            'fecha_inscripcion': mat.fecha_inscripcion.strftime('%d/%m/%Y %H:%M'),
            'estado': mat.estado or 'N/A',
            'ausencias': mat.ausencias,
            'intento': mat.intento,
            'nota_primaria': mat.nota_primaria if hasattr(mat, 'nota_primaria') else 'N/A',
            'nota_secundaria': mat.nota_secundaria if hasattr(mat, 'nota_secundaria') else 'N/A',
            'nota_final': mat.nota_final,
            'nota_extra': mat.nota_extra if hasattr(mat, 'nota_extra') else 'N/A',
            'slug': mat.slug if hasattr(mat, 'slug') else 'N/A',
            'tabla_origen': mat.tabla_origen,
            'fecha_consolidacion': mat.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })

    # 3. SOLICITUDES DE INSCRIPCIÓN (Docencia_enrollmentapplication)
    solicitudes = _buscar_solicitudes_historicas(user, usuario_ids_originales)
    for sol in solicitudes:
        historial_data['solicitudes_inscripcion'].append({
            'id': sol.id,
            'curso': sol.curso.nombre if sol.curso else 'N/A',
            'curso_codigo': sol.curso.codigo if sol.curso else 'N/A',
            'fecha_solicitud': sol.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            'estado': sol.estado or 'N/A',
            'tabla_origen': sol.tabla_origen,
            'fecha_consolidacion': sol.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })

    # 4. CUENTAS BANCARIAS (Docencia_accountnumber)
    # user_id apunta directamente a auth_user
    from datos_archivados.models import DatoArchivadoDinamico
    if usuario_ids_originales:
        cuentas = (
            HistoricalAccountNumber.objects.filter(
                id_original__in=DatoArchivadoDinamico.objects.filter(
                    tabla_origen='Docencia_accountnumber',
                    datos_originales__user_id__in=usuario_ids_originales
                ).values_list('id_original', flat=True)
            ) | HistoricalAccountNumber.objects.filter(usuario=user)
        ).distinct()
    else:
        cuentas = HistoricalAccountNumber.objects.filter(usuario=user)
    for cuenta in cuentas:
        historial_data['cuentas_bancarias'].append({
            'id': cuenta.id,
            'numero_cuenta': cuenta.numero_cuenta,
            'banco': cuenta.banco or 'N/A',
            'tabla_origen': cuenta.tabla_origen,
            'fecha_consolidacion': cuenta.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })

    # 5. PAGOS (Docencia_enrollmentpay)
    # app_id apunta al ID original de Docencia_application en la BD antigua
    aplicaciones_ids_originales = []
    for app in aplicaciones:
        if app.dato_archivado:
            orig_id = app.dato_archivado.datos_originales.get('id')
            if orig_id:
                aplicaciones_ids_originales.append(orig_id)
        elif app.id_original:
            aplicaciones_ids_originales.append(app.id_original)

    if aplicaciones_ids_originales:
        pagos_data = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_enrollmentpay',
            datos_originales__app_id__in=aplicaciones_ids_originales
        )
        for pago_dato in pagos_data:
            datos = pago_dato.datos_originales
            historial_data['pagos'].append({
                'id': datos.get('id'),
                'monto': datos.get('monto', datos.get('amount', 'N/A')),
                'fecha': datos.get('datepub', datos.get('date', 'N/A')),
                'metodo': datos.get('transfernumber', datos.get('method', 'N/A')),
                'referencia': datos.get('cardnumber_id', datos.get('reference', 'N/A')),
                'aceptado': 'Sí' if datos.get('accept') else 'No',
                'tabla_origen': 'Docencia_enrollmentpay',
            })

    # 6. CURSOS COMO PROFESOR (Docencia_courseinformation_adminteachers)
    cursos_profesor = HistoricalCourseInformationAdminTeachers.objects.filter(
        profesor=user
    ).select_related('curso', 'curso__area', 'curso__categoria')
    for cp in cursos_profesor:
        historial_data['cursos_como_profesor'].append({
            'id': cp.id,
            'curso': cp.curso.nombre if cp.curso else 'N/A',
            'curso_codigo': cp.curso.codigo if cp.curso else 'N/A',
            'area': cp.curso.area.nombre if cp.curso and cp.curso.area else 'N/A',
            'categoria': cp.curso.categoria.nombre if cp.curso and cp.curso.categoria else 'N/A',
            'tabla_origen': cp.tabla_origen,
            'fecha_consolidacion': cp.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })

    # 7. CURSOS ADMINISTRADOS
    cursos_admin = HistoricalCourseInformation.objects.filter(
        admin_teachers__profesor=user
    ).distinct()
    for curso in cursos_admin:
        historial_data['cursos_administrados'].append({
            'id': curso.id,
            'nombre': curso.nombre,
            'codigo': curso.codigo if hasattr(curso, 'codigo') else 'N/A',
            'descripcion': curso.descripcion if hasattr(curso, 'descripcion') else 'N/A',
            'area': curso.area.nombre if curso.area else 'N/A',
            'categoria': curso.categoria.nombre if curso.categoria else 'N/A',
            'tabla_origen': curso.tabla_origen,
            'fecha_consolidacion': curso.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
        })

    # 8. EDICIONES
    ediciones_ids = set()
    for app in aplicaciones:
        if app.edicion:
            ediciones_ids.add(app.edicion.id)
    for mat in matriculas:
        if mat.edicion:
            ediciones_ids.add(mat.edicion.id)

    if ediciones_ids:
        ediciones = HistoricalEdition.objects.filter(id__in=ediciones_ids).select_related('curso')
        for ed in ediciones:
            historial_data['ediciones'].append({
                'id': ed.id,
                'nombre': ed.nombre,
                'curso': ed.curso.nombre if ed.curso else 'N/A',
                'fecha_inicio': ed.fecha_inicio.strftime('%d/%m/%Y') if ed.fecha_inicio else 'N/A',
                'fecha_fin': ed.fecha_fin.strftime('%d/%m/%Y') if ed.fecha_fin else 'N/A',
                'tabla_origen': ed.tabla_origen,
                'fecha_consolidacion': ed.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })

    # 9. ASIGNATURAS
    cursos_ids = set()
    for app in aplicaciones:
        if app.curso:
            cursos_ids.add(app.curso.id)

    if cursos_ids:
        asignaturas = HistoricalSubjectInformation.objects.filter(
            curso_id__in=cursos_ids
        ).select_related('curso')
        for asig in asignaturas:
            historial_data['asignaturas'].append({
                'id': asig.id,
                'nombre': asig.nombre,
                'codigo': asig.codigo if hasattr(asig, 'codigo') else 'N/A',
                'curso': asig.curso.nombre if asig.curso else 'N/A',
                'descripcion': asig.descripcion if hasattr(asig, 'descripcion') else 'N/A',
                'tabla_origen': asig.tabla_origen,
                'fecha_consolidacion': asig.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })

    # 10. ÁREAS
    areas_ids = set()
    for app in aplicaciones:
        if app.curso and app.curso.area:
            areas_ids.add(app.curso.area.id)

    if areas_ids:
        areas = HistoricalArea.objects.filter(id__in=areas_ids)
        for area in areas:
            historial_data['areas'].append({
                'id': area.id,
                'nombre': area.nombre,
                'codigo': area.codigo if hasattr(area, 'codigo') else 'N/A',
                'descripcion': area.descripcion or 'Sin descripción',
                'tabla_origen': area.tabla_origen,
                'fecha_consolidacion': area.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })

    # 11. CATEGORÍAS
    categorias_ids = set()
    for app in aplicaciones:
        if app.curso and app.curso.categoria:
            categorias_ids.add(app.curso.categoria.id)
    for curso in cursos_admin:
        if curso.categoria:
            categorias_ids.add(curso.categoria.id)

    if categorias_ids:
        categorias = HistoricalCourseCategory.objects.filter(id__in=categorias_ids)
        for cat in categorias:
            historial_data['categorias'].append({
                'id': cat.id,
                'nombre': cat.nombre,
                'codigo': cat.codigo if hasattr(cat, 'codigo') else 'N/A',
                'descripcion': cat.descripcion or 'Sin descripción',
                'precio': str(cat.precio) if hasattr(cat, 'precio') and cat.precio else 'N/A',
                'es_servicio': 'Sí' if hasattr(cat, 'es_servicio') and cat.es_servicio else 'No',
                'registro_abierto': 'Sí' if hasattr(cat, 'registro_abierto') and cat.registro_abierto else 'No',
                'tabla_origen': cat.tabla_origen,
                'fecha_consolidacion': cat.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })


    # 12. CLASES (Docencia_class) - clases donde el usuario tiene student views
    clases_ids = set()
    for app in aplicaciones:
        for csv in app.class_views.select_related('class_field').all():
            clases_ids.add(csv.class_field.id)

    if clases_ids:
        clases = HistoricalClass.objects.filter(id__in=clases_ids).select_related('subject')
        for clase in clases:
            historial_data['clases'].append({
                'id': clase.id,
                'nombre': clase.name,
                'asignatura': clase.subject.nombre if clase.subject else 'N/A',
                'fecha_carga': clase.uploaddate.strftime('%d/%m/%Y') if clase.uploaddate else 'N/A',
                'fecha_publicacion': clase.datepub.strftime('%d/%m/%Y %H:%M') if clase.datepub else 'N/A',
                'fecha_fin': clase.dateend.strftime('%d/%m/%Y %H:%M') if clase.dateend else 'N/A',
                'slug': clase.slug,
                'fecha_consolidacion': clase.fecha_consolidacion.strftime('%d/%m/%Y %H:%M'),
            })

    context = {
        'historial': historial_data,
        'usuario_historial': user,
    }

    return render(request, 'detalles_historial_usuario.html', context)


@login_required
def exportar_detalles_historial_pdf(request, user_id):
    """
    Vista para exportar el historial completo de un usuario a PDF.
    """
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from io import BytesIO
    from historial.models import (
        HistoricalArea,
        HistoricalCourseCategory,
        HistoricalCourseInformation,
        HistoricalCourseInformationAdminTeachers,
        HistoricalEnrollmentApplication,
        HistoricalAccountNumber,
        HistoricalEnrollmentPay,
        HistoricalSubjectInformation,
        HistoricalEdition,
        HistoricalEnrollment,
        HistoricalApplication,
        HistoricalClass,
        HistoricalClassStudentView,
    )

    # Verificar que el usuario sea secretaria
    if not request.user.groups.filter(name='Secretaría').exists():
        messages.error(request, 'No tiene permisos para exportar esta información')
        return redirect('principal:profile')

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, 'Usuario no encontrado')
        return redirect('principal:usuarios_registrados')

    # Obtener los mismos datos que en ver_detalles_historial_usuario
    historial_data = {
        'usuario': {
            'nombre': user.get_full_name() or user.username,
            'email': user.email,
            'username': user.username,
        },
        'aplicaciones': [],
        'matriculas': [],
        'solicitudes_inscripcion': [],
        'cuentas_bancarias': [],
        'pagos': [],
        'cursos_como_profesor': [],
        'cursos_administrados': [],
        'ediciones': [],
        'asignaturas': [],
        'areas': [],
        'categorias': [],
        'clases': []
    }

    # 1. APLICACIONES
    aplicaciones, usuario_ids_originales = _buscar_aplicaciones_historicas(user)
    usuario_id_original = usuario_ids_originales[0] if usuario_ids_originales else None
    for app in aplicaciones:
        historial_data['aplicaciones'].append({
            'curso': app.curso.nombre if app.curso else 'N/A',
            'curso_codigo': app.curso.codigo if app.curso else 'N/A',
            'area': app.curso.area.nombre if app.curso and app.curso.area else 'N/A',
            'categoria': app.curso.categoria.nombre if app.curso and app.curso.categoria else 'N/A',
            'edicion': app.edicion.nombre if app.edicion else 'N/A',
            'edicion_fecha_inicio': app.edicion.fecha_inicio.strftime('%d/%m/%Y') if app.edicion and app.edicion.fecha_inicio else 'N/A',
            'edicion_fecha_fin': app.edicion.fecha_fin.strftime('%d/%m/%Y') if app.edicion and app.edicion.fecha_fin else 'N/A',
            'fecha_solicitud': app.fecha_solicitud.strftime('%d/%m/%Y') if app.fecha_solicitud else 'N/A',
            'estado': app.estado or 'N/A',
            'beca': 'Sí' if app.beca else 'No',
            'pagado': 'Sí' if app.pagado else 'No',
            'nota_primaria': app.nota_primaria,
            'nota_secundaria': app.nota_secundaria,
            'nota_final': app.nota_final,
            'nota_extra': app.nota_extra,
            'comentarios': app.comentarios or 'Sin comentarios',
        })

    # 2. MATRÍCULAS
    matriculas = _buscar_matriculas_historicas(user, usuario_ids_originales)
    for mat in matriculas:
        historial_data['matriculas'].append({
            'curso': mat.edicion.curso.nombre if mat.edicion and mat.edicion.curso else (mat.curso.nombre if mat.curso else 'N/A'),
            'edicion': mat.edicion.nombre if mat.edicion else 'N/A',
            'fecha_inscripcion': mat.fecha_inscripcion.strftime('%d/%m/%Y %H:%M'),
            'estado': mat.estado or 'N/A',
            'ausencias': mat.ausencias,
            'intento': mat.intento,
            'nota_primaria': mat.nota_primaria if hasattr(mat, 'nota_primaria') else 'N/A',
            'nota_secundaria': mat.nota_secundaria if hasattr(mat, 'nota_secundaria') else 'N/A',
            'nota_final': mat.nota_final,
            'nota_extra': mat.nota_extra if hasattr(mat, 'nota_extra') else 'N/A',
        })

    # 3. SOLICITUDES DE INSCRIPCIÓN
    solicitudes = _buscar_solicitudes_historicas(user, usuario_ids_originales)
    for sol in solicitudes:
        historial_data['solicitudes_inscripcion'].append({
            'curso': sol.curso.nombre if sol.curso else 'N/A',
            'curso_codigo': sol.curso.codigo if sol.curso else 'N/A',
            'fecha_solicitud': sol.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            'estado': sol.estado or 'N/A',
        })

    # 4. CUENTAS BANCARIAS
    from datos_archivados.models import DatoArchivadoDinamico
    if usuario_ids_originales:
        cuentas = (
            HistoricalAccountNumber.objects.filter(
                id_original__in=DatoArchivadoDinamico.objects.filter(
                    tabla_origen='Docencia_accountnumber',
                    datos_originales__user_id__in=usuario_ids_originales
                ).values_list('id_original', flat=True)
            ) | HistoricalAccountNumber.objects.filter(usuario=user)
        ).distinct()
    else:
        cuentas = HistoricalAccountNumber.objects.filter(usuario=user)
    for cuenta in cuentas:
        historial_data['cuentas_bancarias'].append({
            'numero_cuenta': cuenta.numero_cuenta,
            'banco': cuenta.banco or 'N/A',
        })

    # 5. PAGOS
    aplicaciones_ids_originales = []
    for app in aplicaciones:
        if app.dato_archivado:
            orig_id = app.dato_archivado.datos_originales.get('id')
            if orig_id:
                aplicaciones_ids_originales.append(orig_id)
        elif app.id_original:
            aplicaciones_ids_originales.append(app.id_original)

    if aplicaciones_ids_originales:
        pagos_data = DatoArchivadoDinamico.objects.filter(
            tabla_origen='Docencia_enrollmentpay',
            datos_originales__app_id__in=aplicaciones_ids_originales
        )
        for pago_dato in pagos_data:
            datos = pago_dato.datos_originales
            historial_data['pagos'].append({
                'monto': datos.get('amount', 'N/A'),
                'fecha': datos.get('date', 'N/A'),
                'metodo': datos.get('method', 'N/A'),
                'referencia': datos.get('reference', 'N/A'),
            })

    # 6. CURSOS COMO PROFESOR
    cursos_profesor = HistoricalCourseInformationAdminTeachers.objects.filter(
        profesor=user
    ).select_related('curso', 'curso__area', 'curso__categoria')
    for cp in cursos_profesor:
        historial_data['cursos_como_profesor'].append({
            'curso': cp.curso.nombre if cp.curso else 'N/A',
            'curso_codigo': cp.curso.codigo if cp.curso else 'N/A',
            'area': cp.curso.area.nombre if cp.curso and cp.curso.area else 'N/A',
            'categoria': cp.curso.categoria.nombre if cp.curso and cp.curso.categoria else 'N/A',
        })

    # 7. CURSOS ADMINISTRADOS
    cursos_admin = HistoricalCourseInformation.objects.filter(
        admin_teachers__profesor=user
    ).distinct()
    for curso in cursos_admin:
        historial_data['cursos_administrados'].append({
            'nombre': curso.nombre,
            'codigo': curso.codigo if hasattr(curso, 'codigo') else 'N/A',
            'descripcion': curso.descripcion if hasattr(curso, 'descripcion') else 'N/A',
            'area': curso.area.nombre if curso.area else 'N/A',
            'categoria': curso.categoria.nombre if curso.categoria else 'N/A',
        })

    # 8. EDICIONES
    ediciones_ids = set()
    for app in aplicaciones:
        if app.edicion:
            ediciones_ids.add(app.edicion.id)
    for mat in matriculas:
        if mat.edicion:
            ediciones_ids.add(mat.edicion.id)

    if ediciones_ids:
        ediciones = HistoricalEdition.objects.filter(id__in=ediciones_ids).select_related('curso')
        for ed in ediciones:
            historial_data['ediciones'].append({
                'nombre': ed.nombre,
                'curso': ed.curso.nombre if ed.curso else 'N/A',
                'fecha_inicio': ed.fecha_inicio.strftime('%d/%m/%Y') if ed.fecha_inicio else 'N/A',
                'fecha_fin': ed.fecha_fin.strftime('%d/%m/%Y') if ed.fecha_fin else 'N/A',
            })

    # 9. ASIGNATURAS
    cursos_ids = set()
    for app in aplicaciones:
        if app.curso:
            cursos_ids.add(app.curso.id)

    if cursos_ids:
        asignaturas = HistoricalSubjectInformation.objects.filter(
            curso_id__in=cursos_ids
        ).select_related('curso')
        for asig in asignaturas:
            historial_data['asignaturas'].append({
                'nombre': asig.nombre,
                'codigo': asig.codigo if hasattr(asig, 'codigo') else 'N/A',
                'curso': asig.curso.nombre if asig.curso else 'N/A',
                'descripcion': asig.descripcion if hasattr(asig, 'descripcion') else 'N/A',
            })

    # 10. ÁREAS
    areas_ids = set()
    for app in aplicaciones:
        if app.curso and app.curso.area:
            areas_ids.add(app.curso.area.id)

    if areas_ids:
        areas = HistoricalArea.objects.filter(id__in=areas_ids)
        for area in areas:
            historial_data['areas'].append({
                'nombre': area.nombre,
                'codigo': area.codigo if hasattr(area, 'codigo') else 'N/A',
                'descripcion': area.descripcion or 'Sin descripción',
            })

    # 11. CATEGORÍAS
    categorias_ids = set()
    for app in aplicaciones:
        if app.curso and app.curso.categoria:
            categorias_ids.add(app.curso.categoria.id)
    for curso in cursos_admin:
        if curso.categoria:
            categorias_ids.add(curso.categoria.id)

    if categorias_ids:
        categorias = HistoricalCourseCategory.objects.filter(id__in=categorias_ids)
        for cat in categorias:
            historial_data['categorias'].append({
                'nombre': cat.nombre,
                'codigo': cat.codigo if hasattr(cat, 'codigo') else 'N/A',
                'descripcion': cat.descripcion or 'Sin descripción',
                'precio': str(cat.precio) if hasattr(cat, 'precio') and cat.precio else 'N/A',
                'es_servicio': 'Sí' if hasattr(cat, 'es_servicio') and cat.es_servicio else 'No',
                'registro_abierto': 'Sí' if hasattr(cat, 'registro_abierto') and cat.registro_abierto else 'No',
            })

    # Renderizar el template HTML para PDF
    context = {
        'historial': historial_data,
        'usuario_historial': user,
    }
    
    html_string = render_to_string('detalles_historial_usuario_pdf.html', context)
    
    # Crear el PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f'historial_{user.username}_{user.id}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    messages.error(request, 'Error al generar el PDF')
    return redirect('principal:detalles_historial_usuario', user_id=user_id)


# ── API AJAX para el admin: verificar si hay curso académico activo ───────────

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def api_hay_curso_activo(request):
    """
    Endpoint AJAX usado por el admin para verificar si existe algún
    CursoAcademico activo antes de restaurar uno archivado.

    Retorna JSON:
        { "hay_activo": true/false, "nombre": "2024-2025" | null }
    """
    from django.http import JsonResponse
    curso_activo = CursoAcademico.objects.filter(activo=True).first()
    if curso_activo:
        return JsonResponse({'hay_activo': True, 'nombre': curso_activo.nombre})
    return JsonResponse({'hay_activo': False, 'nombre': None})


# ---------------------------------------------------------------------------
# Vistas para el Reglamento General del Centro (gestionado por Secretaría)
# ---------------------------------------------------------------------------

class ReglamentoGeneralEditView(LoginRequiredMixin, SecretariaRequiredMixin, View):
    """
    Vista única para crear o editar el ReglamentoGeneral (singleton).
    Si ya existe lo edita; si no, lo crea.
    """
    template_name = 'registration/reglamento_general_form.html'

    def _get_or_none(self):
        return ReglamentoGeneral.objects.first()

    def get(self, request, *args, **kwargs):
        reglamento = self._get_or_none()
        form = ReglamentoGeneralForm(instance=reglamento)
        formset = ArticuloReglamentoGeneralFormSet(instance=reglamento)
        return render(request, self.template_name, {
            'form': form,
            'articulo_formset': formset,
            'reglamento': reglamento,
        })

    def post(self, request, *args, **kwargs):
        reglamento = self._get_or_none()
        form = ReglamentoGeneralForm(request.POST, instance=reglamento)
        formset = ArticuloReglamentoGeneralFormSet(request.POST, instance=reglamento)
        formset.set_introduccion(request.POST.get('introduccion', ''))

        if form.is_valid() and formset.is_valid():
            reglamento = form.save()
            formset.instance = reglamento
            formset.save()
            messages.success(request, 'Reglamento General guardado exitosamente.')
            return redirect(reverse('principal:reglamento_general'))

        return render(request, self.template_name, {
            'form': form,
            'articulo_formset': formset,
            'reglamento': reglamento,
        })


# ---------------------------------------------------------------------------
# Vista AJAX: Terminar Semestre
# ---------------------------------------------------------------------------

@login_required
@require_POST
def terminar_semestre_view(request, curso_id):
    """
    Endpoint AJAX para la acción "Terminar Semestre" o "Finalizar Curso".
    Solo accesible para el grupo Secretaría.

    Body JSON esperado:
      - accion: 'terminar_semestre' | 'finalizar_curso'
      - enrollment_deadline: 'YYYY-MM-DD'  (solo para terminar_semestre)
      - start_date: 'YYYY-MM-DD'           (solo para terminar_semestre)

    Retorna JsonResponse con {success, message} o {success: false, error}.
    """
    import json
    from django.http import JsonResponse
    from principal.semestre_service import terminar_semestre, SemestreError
    from principal.models import Curso

    # Verificar permisos: solo Secretaría
    if not request.user.groups.filter(name='Secretaría').exists():
        return JsonResponse(
            {'success': False, 'error': 'No tienes permisos para realizar esta acción.'},
            status=403
        )

    try:
        curso = Curso.objects.get(pk=curso_id)
    except Curso.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Curso no encontrado.'}, status=404)

    # Parsear body JSON
    try:
        body = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        body = {}

    accion = body.get('accion', 'terminar_semestre')

    # ── Acción: Finalizar Curso ───────────────────────────────────────────────
    if accion == 'finalizar_curso':
        try:
            curso.status = 'F'
            curso.save(update_fields=['status'])
            return JsonResponse({
                'success': True,
                'message': f"El curso '{curso.name}' ha sido finalizado.",
            })
        except Exception as e:
            return JsonResponse(
                {'success': False, 'error': f'Error al finalizar el curso: {str(e)}'},
                status=500
            )

    # ── Acción: Terminar Semestre ─────────────────────────────────────────────
    enrollment_deadline = body.get('enrollment_deadline', '').strip()
    start_date = body.get('start_date', '').strip()
    mantener_documentos = body.get('mantener_documentos', True)  # True = conservar, False = eliminar

    if not enrollment_deadline or not start_date:
        return JsonResponse(
            {'success': False, 'error': 'Debe proporcionar la fecha límite de inscripción y la fecha de inicio.'},
            status=400
        )

    # Validar formato de fechas
    from datetime import date
    try:
        from datetime import datetime
        enrollment_deadline_date = datetime.strptime(enrollment_deadline, '%Y-%m-%d').date()
        start_date_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse(
            {'success': False, 'error': 'Formato de fecha inválido. Use YYYY-MM-DD.'},
            status=400
        )

    try:
        contadores = terminar_semestre(curso)
        # Actualizar las fechas del curso para el nuevo semestre
        curso.enrollment_deadline = enrollment_deadline_date
        curso.start_date = start_date_date
        curso.save(update_fields=['enrollment_deadline', 'start_date'])

        # Eliminar documentos del profesor si el usuario lo solicitó
        if not mantener_documentos:
            from course_documents.models import DocumentFolder, CourseDocument
            from django.core.files.storage import default_storage
            carpetas = list(DocumentFolder.objects.filter(curso=curso))
            docs_eliminados = 0
            for carpeta in carpetas:
                for doc in carpeta.documents.all():
                    if doc.file and default_storage.exists(doc.file.name):
                        default_storage.delete(doc.file.name)
                    doc.delete()
                    docs_eliminados += 1
                carpeta.delete()
            logger.info(
                f"Documentos eliminados al cambiar semestre de '{curso.name}': "
                f"{docs_eliminados} documentos en {len(carpetas)} carpetas."
            )

        mensaje = (
            f"Semestre {contadores['semestre_num']} terminado exitosamente. "
            f"Se archivaron: {contadores['matriculas']} matrículas, "
            f"{contadores['calificaciones']} calificaciones, "
            f"{contadores['asistencias']} asistencias."
        )
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'contadores': contadores,
        })
    except SemestreError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse(
            {'success': False, 'error': f'Error inesperado: {str(e)}'},
            status=500
        )


@login_required
def revertir_semestre_view(request, curso_id):
    """
    Endpoint AJAX para revertir el semestre actual de un curso y restaurar el anterior.

    GET: Devuelve la lista de semestres archivados disponibles para el curso.
    POST: Ejecuta la reversión del semestre indicado.

    Condiciones:
      - Solo accesible para Secretaría.
      - El curso debe estar en estado 'I' (inscripción) o 'IT' (plazo terminado).
      - Si está en 'P' (en progreso), devuelve error.
      - Elimina el semestre actual (matrículas, solicitudes, etc.) y restaura el archivado.
      - Envía correo a los estudiantes que habían aplicado.
    """
    import json
    from django.http import JsonResponse
    from principal.models import Curso, SemestreCurso, SolicitudInscripcion, Matriculas, Calificaciones, NotaIndividual, Asistencia
    from datos_archivados.models import SemestreCursoArchivado, MatriculaArchivada, CalificacionArchivada, NotaIndividualArchivada, AsistenciaArchivada, CursoArchivado

    if not request.user.groups.filter(name='Secretaría').exists():
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)

    try:
        curso = Curso.objects.get(pk=curso_id)
    except Curso.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Curso no encontrado.'}, status=404)

    # GET: devolver semestres archivados disponibles
    if request.method == 'GET':
        semestres = SemestreCursoArchivado.objects.filter(
            curso_archivado__id_original=curso_id
        ).order_by('-numero_semestre').values('id', 'numero_semestre', 'fecha_cierre', 'fecha_inicio')
        lista = []
        for s in semestres:
            lista.append({
                'id': s['id'],
                'numero_semestre': s['numero_semestre'],
                'fecha_cierre': s['fecha_cierre'].strftime('%d/%m/%Y') if s['fecha_cierre'] else '—',
                'fecha_inicio': s['fecha_inicio'].strftime('%d/%m/%Y') if s['fecha_inicio'] else '—',
            })
        return JsonResponse({'semestres': lista})

    # POST: ejecutar reversión
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)

    # Verificar estado del curso
    estado_actual = curso.get_dynamic_status()
    if estado_actual == 'P':
        return JsonResponse({
            'success': False,
            'error': 'No se puede revertir un semestre mientras el curso está en progreso.',
            'en_progreso': True,
        }, status=400)

    if estado_actual not in ('I', 'IT'):
        return JsonResponse({
            'success': False,
            'error': f'Solo se puede revertir en estado de inscripción o plazo terminado. Estado actual: {curso.get_dynamic_status_display()}',
        }, status=400)

    try:
        body = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        body = {}

    semestre_archivado_id = body.get('semestre_archivado_id')
    if not semestre_archivado_id:
        return JsonResponse({'success': False, 'error': 'Debe indicar el semestre a restaurar.'}, status=400)

    try:
        semestre_archivado = SemestreCursoArchivado.objects.select_related('curso_archivado').get(
            pk=semestre_archivado_id,
            curso_archivado__id_original=curso_id,
        )
    except SemestreCursoArchivado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Semestre archivado no encontrado para este curso.'}, status=404)

    # Recopilar emails de estudiantes con solicitudes pendientes/aprobadas en el semestre actual
    solicitudes_actuales = SolicitudInscripcion.objects.filter(
        curso=curso,
        estado__in=('pendiente', 'aprobada'),
    ).select_related('estudiante')

    emails_notificar = []
    for sol in solicitudes_actuales:
        if sol.estudiante.email:
            emails_notificar.append((sol.estudiante.email, sol.estudiante.get_full_name() or sol.estudiante.username))

    try:
        from django.db import transaction
        with transaction.atomic():
            # 1. Eliminar datos del semestre actual del curso
            NotaIndividual.objects.filter(calificacion__course=curso).delete()
            Calificaciones.objects.filter(course=curso).delete()
            Asistencia.objects.filter(course=curso).delete()
            Matriculas.objects.filter(course=curso).delete()
            SolicitudInscripcion.objects.filter(curso=curso).delete()

            # 2. Desactivar el SemestreCurso actual
            SemestreCurso.objects.filter(curso=curso, activo=True).update(activo=False)

            # 3. Restaurar datos del semestre archivado
            curso_archivado = semestre_archivado.curso_archivado

            # Restaurar matrículas
            from django.contrib.auth.models import User
            matriculas_archivadas = MatriculaArchivada.objects.filter(
                course=curso_archivado
            ).select_related('student__usuario_actual')

            matriculas_map = {}
            for ma in matriculas_archivadas:
                student = ma.student.usuario_actual if ma.student.usuario_actual else User.objects.filter(pk=ma.student.id_original).first()
                if not student:
                    continue
                m, _ = Matriculas.objects.get_or_create(
                    course=curso,
                    student=student,
                    curso_academico=curso.curso_academico,
                    defaults={'activo': ma.activo, 'fecha_matricula': ma.fecha_matricula, 'estado': ma.estado},
                )
                matriculas_map[ma.pk] = m

            # Restaurar calificaciones y notas
            calificaciones_archivadas = CalificacionArchivada.objects.filter(
                course=curso_archivado
            ).select_related('student__usuario_actual', 'matricula').prefetch_related('notas_archivadas')

            for cala in calificaciones_archivadas:
                student = cala.student.usuario_actual if cala.student.usuario_actual else User.objects.filter(pk=cala.student.id_original).first()
                if not student:
                    continue
                matricula = matriculas_map.get(cala.matricula.pk) if cala.matricula else None
                cal, created = Calificaciones.objects.get_or_create(
                    course=curso,
                    student=student,
                    curso_academico=curso.curso_academico,
                    defaults={'matricula': matricula, 'average': cala.average},
                )
                if created:
                    for na in cala.notas_archivadas.all():
                        NotaIndividual.objects.create(calificacion=cal, valor=na.valor, fecha_creacion=na.fecha_creacion)

            # Restaurar asistencias
            asistencias_archivadas = AsistenciaArchivada.objects.filter(
                course=curso_archivado
            ).select_related('student__usuario_actual')

            for aa in asistencias_archivadas:
                student = aa.student.usuario_actual if aa.student.usuario_actual else User.objects.filter(pk=aa.student.id_original).first()
                if not student:
                    continue
                Asistencia.objects.get_or_create(course=curso, student=student, date=aa.date, defaults={'presente': aa.presente})

            # 4. Reactivar el SemestreCurso archivado (restaurar en principal)
            SemestreCurso.objects.filter(
                curso=curso,
                numero_semestre=semestre_archivado.numero_semestre,
            ).update(activo=True)

            # Si no existe, crearlo
            if not SemestreCurso.objects.filter(curso=curso, numero_semestre=semestre_archivado.numero_semestre).exists():
                SemestreCurso.objects.create(
                    curso=curso,
                    numero_semestre=semestre_archivado.numero_semestre,
                    activo=True,
                    curso_academico=curso.curso_academico,
                    fecha_inicio=semestre_archivado.fecha_inicio,
                    fecha_cierre=semestre_archivado.fecha_cierre,
                )

            # 5. Eliminar los datos archivados del semestre restaurado
            from datos_archivados.models import UsuarioArchivado
            cursos_archivados_ids = [curso_archivado.pk]
            NotaIndividualArchivada.objects.filter(calificacion__course__pk__in=cursos_archivados_ids).delete()
            CalificacionArchivada.objects.filter(course__pk__in=cursos_archivados_ids).delete()
            AsistenciaArchivada.objects.filter(course__pk__in=cursos_archivados_ids).delete()
            MatriculaArchivada.objects.filter(course__pk__in=cursos_archivados_ids).delete()
            semestre_archivado.delete()

            # 6. Restaurar estado del curso a 'P' (en progreso) y limpiar fechas
            Curso.objects.filter(pk=curso_id).update(
                status='P',
                enrollment_deadline=None,
                start_date=None,
            )
            curso.status = 'P'
            curso.enrollment_deadline = None
            curso.start_date = None

    except Exception as exc:
        import logging
        logging.getLogger(__name__).error(f'[revertir_semestre] Error: {exc}', exc_info=True)
        return JsonResponse({'success': False, 'error': f'Error al revertir el semestre: {str(exc)}'}, status=500)

    # 7. Enviar correos a estudiantes notificados (fuera de la transacción para no bloquearla)
    correos_enviados = 0
    for email, nombre in emails_notificar:
        try:
            send_mail(
                subject='Aviso: Tu solicitud de inscripción ha sido cancelada',
                message=(
                    f'Estimado/a {nombre},\n\n'
                    f'Te informamos que el semestre del curso "{curso.name}" ha sido restablecido '
                    f'al semestre anterior. Como consecuencia, tu solicitud de inscripción para el '
                    f'semestre actual no es válida y ha sido cancelada.\n\n'
                    f'Si tienes alguna duda, por favor contacta con la secretaría del centro.\n\n'
                    f'Atentamente,\nEl equipo del centro.'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )
            correos_enviados += 1
        except Exception:
            pass

    return JsonResponse({
        'success': True,
        'message': (
            f'Semestre {semestre_archivado.numero_semestre} restaurado correctamente. '
            f'Se notificó a {correos_enviados} estudiante(s) por correo.'
        ),
    })
