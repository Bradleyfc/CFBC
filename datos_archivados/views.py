from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.db.models import Count
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.mail import send_mail
from django.conf import settings
import json
import logging

# Configurar logger
logger = logging.getLogger(__name__)

def es_secretaria(user):
    """Verifica si el usuario pertenece al grupo Secretaría"""
    return user.groups.filter(name='Secretaría').exists()

def tiene_permisos_datos_archivados(user):
    """Verifica si el usuario tiene permisos para acceder a datos archivados (Administración o Admin)"""
    return (user.groups.filter(name='Administración').exists() or 
            user.is_superuser or 
            user.is_staff)

def permisos_datos_archivados_required(view_func):
    """Decorador personalizado que verifica permisos para datos archivados"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        
        return view_func(request, *args, **kwargs)
    return wrapper

class PermisosDataArchivadosRequiredMixin:
    """Mixin que verifica permisos para datos archivados (Secretaría o Admin)"""
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        
        return super().dispatch(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class DashboardArchivadosView(TemplateView):
    """Vista principal del dashboard de datos archivados"""
    template_name = 'datos_archivados/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Verificar permisos
        if not tiene_permisos_datos_archivados(self.request.user):
            context['sin_permisos'] = True
            return context
        
        # Importar modelos aquí para evitar problemas de importación
        try:
            from .models import DatoArchivadoDinamico, MigracionLog
            
            # Estadísticas actuales basadas en datos existentes
            total_datos_actuales = DatoArchivadoDinamico.objects.count()
            total_tablas_actuales = DatoArchivadoDinamico.objects.values('tabla_origen').distinct().count()
            ultima_migracion = MigracionLog.objects.first()
            
            # Calcular estadísticas dinámicas
            if total_datos_actuales > 0:
                # Hay datos: usar estadísticas reales
                tablas_inspeccionadas = ultima_migracion.tablas_inspeccionadas if ultima_migracion else 0
                tablas_con_datos = total_tablas_actuales  # Basado en datos actuales
                tablas_vacias = max(0, tablas_inspeccionadas - tablas_con_datos)  # Calculado dinámicamente
            else:
                # No hay datos: todo en cero
                tablas_inspeccionadas = 0
                tablas_con_datos = 0
                tablas_vacias = 0
            
            # Estadísticas básicas
            context.update({
                'total_datos_archivados': total_datos_actuales,
                'total_tablas_migradas': tablas_con_datos,
                'ultima_migracion': ultima_migracion,
                'tablas_inspeccionadas_actuales': tablas_inspeccionadas,
                'tablas_vacias_actuales': tablas_vacias,
            })
            
            # Distribución por tabla (primeras 10)
            context['datos_por_tabla'] = DatoArchivadoDinamico.objects.values(
                'tabla_origen'
            ).annotate(
                total=Count('id')
            ).order_by('-total')[:10]
            
            # Últimas migraciones (primeras 5)
            context['ultimas_migraciones'] = MigracionLog.objects.all()[:5]
            
        except Exception as e:
            # Si hay problemas con los modelos, usar valores por defecto
            context.update({
                'total_datos_archivados': 0,
                'total_tablas_migradas': 0,
                'ultima_migracion': None,
                'datos_por_tabla': [],
                'ultimas_migraciones': [],
            })
        
        return context

@login_required
def configurar_migracion_view(request):
    """Vista para configurar y ejecutar migración automática desde MariaDB"""
    # Verificar permisos
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    if request.method == 'POST':
        # Obtener datos de configuración
        host = request.POST.get('host')
        database = request.POST.get('database')
        user = request.POST.get('user')
        password = request.POST.get('password')
        port = request.POST.get('port', 3306)
        
        if not all([host, database, user, password]):
            messages.error(request, 'Todos los campos de conexión son obligatorios.')
            return render(request, 'datos_archivados/configurar_migracion.html')
        
        try:
            # Importar servicio aquí para evitar problemas de importación
            from .services import MigracionService
            from django.core.cache import cache
            
            # Limpiar cualquier flag de interrupción residual que pueda interferir
            cache.delete('combinacion_interrumpida')
            cache.delete('combinacion_estado_interrupcion')
            cache.delete('combinacion_interrumpida_info')
            cache.delete('combinacion_estado')
            logger.info("🧹 Cache de interrupción limpiado antes de iniciar migración")
            
            # Crear servicio de migración
            servicio = MigracionService(host, database, user, password, int(port))
            
            # Probar conexión
            if not servicio.conectar_mariadb():
                messages.error(request, 'No se pudo conectar a la base de datos MariaDB. Verifique los datos de conexión.')
                return render(request, 'datos_archivados/configurar_migracion.html')
            
            servicio.desconectar_mariadb()
            
            # Ejecutar migración automática en hilo separado
            def ejecutar_migracion_automatica():
                try:
                    servicio.inspeccionar_y_migrar_automaticamente(request.user)
                except Exception as e:
                    pass  # Error handling is in the service
            
            import threading
            thread = threading.Thread(target=ejecutar_migracion_automatica)
            thread.daemon = True
            thread.start()
            
            messages.success(request, 'Migración automática iniciada correctamente. El sistema inspeccionará la base de datos y migrará todos los datos automáticamente.')
            return redirect('datos_archivados:dashboard')
            
        except Exception as e:
            messages.error(request, f'Error al iniciar la migración: {str(e)}')
    
    return render(request, 'datos_archivados/configurar_migracion.html')

@login_required
def detalle_dato_archivado(request, pk):
    """Vista para mostrar el detalle completo de un dato archivado"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.shortcuts import get_object_or_404
        import json
        
        # Obtener el dato archivado
        dato = get_object_or_404(DatoArchivadoDinamico, pk=pk)
        
        # Preparar los datos para mostrar
        context = {
            'dato': dato,
            'datos_formateados': json.dumps(dato.datos_originales, indent=2, ensure_ascii=False),
            'estructura_formateada': json.dumps(
                json.loads(dato.estructura_tabla) if dato.estructura_tabla else {}, 
                indent=2, ensure_ascii=False
            ) if dato.estructura_tabla else None,
        }
        
        return render(request, 'datos_archivados/dato_detail.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar el detalle: {str(e)}')
        return redirect('datos_archivados:tablas_list')

@method_decorator(login_required, name='dispatch')
class TablasArchivadosListView(TemplateView):
    """Vista para listar las tablas archivadas disponibles"""
    template_name = 'datos_archivados/tablas_list.html'
    
    def dispatch(self, request, *args, **kwargs):
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            from .models import DatoArchivadoDinamico
            from django.db.models import Count, Max, Q
            
            # Obtener parámetros de búsqueda
            search_query = self.request.GET.get('search', '').strip()
            search_type = self.request.GET.get('search_type', 'global')
            
            # Query base
            queryset = DatoArchivadoDinamico.objects.all()
            
            # Aplicar filtros de búsqueda
            if search_query:
                if search_type == 'tabla':
                    # Buscar por nombre de tabla
                    queryset = queryset.filter(tabla_origen__icontains=search_query)
                elif search_type == 'contenido':
                    # Buscar en el contenido de los datos (JSON)
                    queryset = queryset.filter(
                        Q(datos_originales__icontains=search_query) |
                        Q(nombre_registro__icontains=search_query)
                    )
                elif search_type == 'global':
                    # Búsqueda global (tabla + contenido)
                    queryset = queryset.filter(
                        Q(tabla_origen__icontains=search_query) |
                        Q(datos_originales__icontains=search_query) |
                        Q(nombre_registro__icontains=search_query)
                    )
            
            # Obtener estadísticas por tabla con filtros aplicados
            tablas_stats = queryset.values('tabla_origen').annotate(
                total_registros=Count('id'),
                ultima_migracion=Max('fecha_migracion')
            ).order_by('tabla_origen')
            
            # Estadísticas generales
            total_registros_filtrados = queryset.count()
            total_registros_totales = DatoArchivadoDinamico.objects.count()
            
            # Obtener la fecha de la última migración
            ultima_migracion = DatoArchivadoDinamico.objects.aggregate(
                ultima_fecha=Max('fecha_migracion')
            )['ultima_fecha']
            
            context.update({
                'tablas_stats': tablas_stats,
                'total_tablas': tablas_stats.count(),
                'total_registros': total_registros_totales,
                'total_registros_filtrados': total_registros_filtrados,
                'search_query': search_query,
                'search_type': search_type,
                'is_filtered': bool(search_query),
                'ultima_migracion': ultima_migracion,
            })
            
        except Exception as e:
            context.update({
                'tablas_stats': [],
                'total_tablas': 0,
                'total_registros': 0,
                'total_registros_filtrados': 0,
                'search_query': '',
                'search_type': 'tabla',
                'is_filtered': False,
                'ultima_migracion': None,
            })
        
        return context

@method_decorator(login_required, name='dispatch')
class DatosArchivadosListView(TemplateView):
    """Vista para listar datos archivados de una tabla específica con paginación"""
    template_name = 'datos_archivados/datos_list.html'
    
    def dispatch(self, request, *args, **kwargs):
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            from .models import DatoArchivadoDinamico
            from django.core.paginator import Paginator
            
            # Obtener la tabla desde la URL
            tabla = kwargs.get('tabla')
            
            if not tabla:
                # Si no hay tabla, redirigir a la lista de tablas
                context['datos'] = []
                context['tabla_actual'] = None
                return context
            
            # Obtener datos de la tabla específica
            queryset = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla)
            
            # Filtro de búsqueda
            search = self.request.GET.get('search')
            if search:
                queryset = queryset.filter(nombre_registro__icontains=search)
            
            # Ordenamiento
            order_by = self.request.GET.get('order_by', 'fecha_migracion')
            order_direction = self.request.GET.get('order_direction', 'desc')
            
            # Validar campos de ordenamiento
            valid_order_fields = {
                'fecha_migracion': 'fecha_migracion',
                'id_original': 'id_original',
                'nombre': 'nombre_registro',
                'tabla': 'tabla_origen'
            }
            
            if order_by in valid_order_fields:
                order_field = valid_order_fields[order_by]
                if order_direction == 'desc':
                    order_field = f'-{order_field}'
                queryset = queryset.order_by(order_field)
            else:
                # Ordenamiento por defecto
                queryset = queryset.order_by('-fecha_migracion')
            
            # Configurar paginación - 100 registros por página
            paginator = Paginator(queryset, 100)
            page_number = self.request.GET.get('page')
            page_obj = paginator.get_page(page_number)
            
            # Para ordenamiento por nombre, necesitamos ordenar después de obtener los datos
            # porque nombre_registro puede ser None y necesitamos usar obtener_nombre_legible()
            if order_by == 'nombre':
                # Obtener todos los datos de la página actual
                datos_list = list(page_obj.object_list)
                datos_list.sort(
                    key=lambda x: (x.obtener_nombre_legible() or '').lower(),
                    reverse=(order_direction == 'desc')
                )
                # Crear un objeto similar a page_obj pero con datos ordenados
                context['datos'] = datos_list
                context['page_obj'] = page_obj
                context['is_paginated'] = page_obj.has_other_pages()
            else:
                context['datos'] = page_obj.object_list
                context['page_obj'] = page_obj
                context['is_paginated'] = page_obj.has_other_pages()
            
            context.update({
                'tabla_actual': tabla,
                'total_registros': paginator.count,
                'search_query': search or '',
                'order_by': order_by,
                'order_direction': order_direction,
                'paginator': paginator,
            })
            
        except Exception as e:
            context['datos'] = []
            context['tabla_actual'] = None
            context['total_registros'] = 0
            context['page_obj'] = None
            context['is_paginated'] = False
        
        return context

@login_required
def estado_migracion_ajax(request):
    """Vista AJAX para obtener el estado de la migración actual"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from .models import MigracionLog
        from datetime import datetime, timedelta
        from django.utils import timezone
        from django.core.cache import cache
        
        # Función helper para manejar valores None
        def safe_int(value):
            try:
                return int(value) if value is not None else 0
            except (ValueError, TypeError):
                return 0
        
        def safe_str(value):
            try:
                return str(value) if value is not None else ''
            except:
                return ''
        
        # Función helper para obtener atributos que pueden no existir
        def safe_getattr(obj, attr, default=0):
            try:
                return getattr(obj, attr, default) if hasattr(obj, attr) else default
            except:
                return default
        
        # Verificar si hay una migración con error reciente (últimos 30 minutos)
        try:
            hace_30_minutos = timezone.now() - timedelta(minutes=30)
            migracion_error = MigracionLog.objects.filter(
                fecha_inicio__gte=hace_30_minutos,
                estado='error'
            ).first()
        except Exception as e:
            logger.warning(f"Error buscando migraciones con error: {e}")
            migracion_error = None
        
        if migracion_error:
            # Hay una migración con error reciente
            try:
                total_migrados = (
                    safe_int(migracion_error.usuarios_migrados) +
                    safe_int(getattr(migracion_error, 'cursos_academicos_migrados', 0)) +
                    safe_int(getattr(migracion_error, 'cursos_migrados', 0)) +
                    safe_int(getattr(migracion_error, 'matriculas_migradas', 0)) +
                    safe_int(getattr(migracion_error, 'calificaciones_migradas', 0)) +
                    safe_int(getattr(migracion_error, 'notas_migradas', 0)) +
                    safe_int(getattr(migracion_error, 'asistencias_migradas', 0))
                )
                
                # Determinar tipo de error basado en el mensaje
                mensaje_error = safe_str(migracion_error.errores)
                tipo_error = 'unknown'
                
                if 'timeout' in mensaje_error.lower() or 'time' in mensaje_error.lower():
                    tipo_error = 'timeout'
                elif 'connection' in mensaje_error.lower() or 'conectar' in mensaje_error.lower():
                    tipo_error = 'connection'
                elif 'permission' in mensaje_error.lower() or 'permiso' in mensaje_error.lower():
                    tipo_error = 'permission'
                elif 'database' in mensaje_error.lower() or 'base de datos' in mensaje_error.lower():
                    tipo_error = 'database'
                
                data = {
                    'error_migracion': True,
                    'en_progreso': False,
                    'tipo': tipo_error,
                    'mensaje': mensaje_error,
                    'mensaje_tecnico': mensaje_error,
                    'fecha_inicio': migracion_error.fecha_inicio.isoformat() if migracion_error.fecha_inicio else None,
                    'fecha_error': migracion_error.fecha_fin.isoformat() if migracion_error.fecha_fin else timezone.now().isoformat(),
                    'total_migrados': total_migrados,
                    'registros_migrados': total_migrados,
                    'tablas_inspeccionadas': safe_int(safe_getattr(migracion_error, 'tablas_inspeccionadas', 0)),
                    'tablas_con_datos': safe_int(safe_getattr(migracion_error, 'tablas_con_datos', 0)),
                    'tablas_vacias': safe_int(safe_getattr(migracion_error, 'tablas_vacias', 0)),
                    'total_tablas': safe_int(safe_getattr(migracion_error, 'tablas_inspeccionadas', 0)),
                    'host_origen': safe_str(safe_getattr(migracion_error, 'host_origen', '')) or 'N/A',
                    'base_datos_origen': safe_str(safe_getattr(migracion_error, 'base_datos_origen', '')) or 'N/A',
                }
                
                # Calcular progreso alcanzado
                tablas_procesadas = data['tablas_con_datos'] + data['tablas_vacias']
                if data['tablas_inspeccionadas'] > 0:
                    data['progreso_porcentaje'] = int((tablas_procesadas / data['tablas_inspeccionadas']) * 100)
                else:
                    data['progreso_porcentaje'] = 0
                
                return JsonResponse(data)
            except Exception as e:
                logger.error(f"Error procesando migración con error: {e}")
        
        # Verificar progreso en cache
        try:
            progreso_cache = cache.get('migracion_progreso')
        except Exception as e:
            logger.warning(f"Error accediendo al cache: {e}")
            progreso_cache = None
        
        if progreso_cache and progreso_cache.get('en_progreso'):
            # Verificar si la migración en cache está atascada
            try:
                fecha_actualizacion = progreso_cache.get('fecha_actualizacion')
                if fecha_actualizacion:
                    ultima_actualizacion = datetime.fromisoformat(fecha_actualizacion.replace('Z', '+00:00'))
                    if timezone.is_naive(ultima_actualizacion):
                        ultima_actualizacion = timezone.make_aware(ultima_actualizacion)
                    
                    tiempo_sin_actualizacion = timezone.now() - ultima_actualizacion
                    
                    # Si han pasado más de 25 minutos sin actualización, considerar como error
                    if tiempo_sin_actualizacion.total_seconds() > 1500:  # 25 minutos
                        # Crear un log de error
                        try:
                            migracion_log = MigracionLog.objects.filter(
                                estado__in=['iniciada', 'en_progreso']
                            ).first()
                            
                            if migracion_log:
                                migracion_log.estado = 'error'
                                migracion_log.fecha_fin = timezone.now()
                                migracion_log.errores = f'Migración atascada por más de 25 minutos. Última actualización: {fecha_actualizacion}'
                                migracion_log.save()
                        except Exception as e:
                            logger.error(f"Error actualizando log de migración: {e}")
                        
                        # Limpiar cache
                        try:
                            cache.delete('migracion_progreso')
                        except:
                            pass
                        
                        # Retornar error
                        data = {
                            'error_migracion': True,
                            'en_progreso': False,
                            'tipo': 'timeout',
                            'mensaje': 'La migración se ha detenido por inactividad prolongada',
                            'mensaje_tecnico': f'Sin actualizaciones por {int(tiempo_sin_actualizacion.total_seconds() / 60)} minutos',
                            'tabla_actual': progreso_cache.get('tabla_actual', 'N/A'),
                            'progreso_porcentaje': progreso_cache.get('progreso_porcentaje', 0),
                            'fecha_inicio': progreso_cache.get('fecha_inicio'),
                            'fecha_error': timezone.now().isoformat(),
                            'tablas_con_datos': progreso_cache.get('tablas_con_datos', 0),
                            'tablas_vacias': progreso_cache.get('tablas_vacias', 0),
                            'total_tablas': progreso_cache.get('total_tablas', 0),
                            'registros_migrados': progreso_cache.get('registros_migrados', 0),
                            'host_origen': progreso_cache.get('host_origen', 'N/A'),
                            'base_datos_origen': progreso_cache.get('base_datos_origen', 'N/A'),
                        }
                        return JsonResponse(data)
                        
            except Exception as e:
                logger.warning(f"Error verificando tiempo de actualización: {e}")
            
            # Usar datos del cache para progreso en tiempo real
            try:
                data = {
                    'en_progreso': True,
                    'estado': f"Procesando tabla: {progreso_cache.get('tabla_actual', 'N/A')}",
                    'progreso_real': progreso_cache.get('progreso_porcentaje', 0),
                    'progreso_porcentaje': progreso_cache.get('progreso_porcentaje', 0),
                    'tabla_actual': progreso_cache.get('tabla_actual', 'N/A'),
                    'tabla_numero': progreso_cache.get('tabla_numero', 0),
                    'total_tablas': progreso_cache.get('total_tablas', 0),
                    'tablas_con_datos': progreso_cache.get('tablas_con_datos', 0),
                    'tablas_vacias': progreso_cache.get('tablas_vacias', 0),
                    'total_migrados': progreso_cache.get('registros_migrados', 0),
                    'registros_migrados': progreso_cache.get('registros_migrados', 0),
                    'fecha_actualizacion': progreso_cache.get('fecha_actualizacion'),
                    'fecha_inicio': progreso_cache.get('fecha_inicio', timezone.now().isoformat()),
                    'host_origen': progreso_cache.get('host_origen', 'Cache en tiempo real'),
                    'base_datos_origen': progreso_cache.get('base_datos_origen', 'Migración activa'),
                    # Campos adicionales para progreso en tiempo real de tabla
                    'registros_procesados_tabla': progreso_cache.get('registros_procesados_tabla', 0),
                    'total_registros_tabla': progreso_cache.get('total_registros_tabla', 0),
                    'porcentaje_tabla': progreso_cache.get('porcentaje_tabla', 0),
                    'registros_migrados_tabla': progreso_cache.get('registros_migrados_tabla', 0),
                    'estado_detalle': progreso_cache.get('estado_detalle', 'Procesando...'),
                }
                return JsonResponse(data)
            except Exception as e:
                logger.error(f"Error procesando datos del cache: {e}")
        
        # Buscar migración en progreso en la base de datos
        try:
            migracion_en_progreso = MigracionLog.objects.filter(
                estado__in=['iniciada', 'en_progreso']
            ).first()
        except Exception as e:
            logger.warning(f"Error buscando migración en progreso: {e}")
            migracion_en_progreso = None
        
        if migracion_en_progreso:
            # Verificar si la migración está atascada (más de 30 minutos sin cambios)
            try:
                tiempo_transcurrido = timezone.now() - migracion_en_progreso.fecha_inicio
                if tiempo_transcurrido.total_seconds() > 1800:  # 30 minutos
                    # Marcar como error por timeout
                    migracion_en_progreso.estado = 'error'
                    migracion_en_progreso.fecha_fin = timezone.now()
                    migracion_en_progreso.errores = f'Migración atascada por más de 30 minutos. Iniciada: {migracion_en_progreso.fecha_inicio}'
                    migracion_en_progreso.save()
                    
                    # Retornar como error
                    total_migrados = safe_int(migracion_en_progreso.usuarios_migrados)
                    
                    data = {
                        'error_migracion': True,
                        'en_progreso': False,
                        'tipo': 'timeout',
                        'mensaje': 'La migración ha excedido el tiempo límite de 30 minutos',
                        'mensaje_tecnico': f'Tiempo transcurrido: {int(tiempo_transcurrido.total_seconds() / 60)} minutos',
                        'fecha_inicio': migracion_en_progreso.fecha_inicio.isoformat(),
                        'fecha_error': timezone.now().isoformat(),
                        'total_migrados': total_migrados,
                        'registros_migrados': total_migrados,
                        'tablas_inspeccionadas': safe_int(safe_getattr(migracion_en_progreso, 'tablas_inspeccionadas', 0)),
                        'tablas_con_datos': safe_int(safe_getattr(migracion_en_progreso, 'tablas_con_datos', 0)),
                        'tablas_vacias': safe_int(safe_getattr(migracion_en_progreso, 'tablas_vacias', 0)),
                        'total_tablas': safe_int(safe_getattr(migracion_en_progreso, 'tablas_inspeccionadas', 0)),
                        'host_origen': safe_str(safe_getattr(migracion_en_progreso, 'host_origen', '')) or 'N/A',
                        'base_datos_origen': safe_str(safe_getattr(migracion_en_progreso, 'base_datos_origen', '')) or 'N/A',
                    }
                    
                    # Calcular progreso alcanzado
                    tablas_procesadas = data['tablas_con_datos'] + data['tablas_vacias']
                    if data['tablas_inspeccionadas'] > 0:
                        data['progreso_porcentaje'] = int((tablas_procesadas / data['tablas_inspeccionadas']) * 100)
                    else:
                        data['progreso_porcentaje'] = 0
                    
                    return JsonResponse(data)
            except Exception as e:
                logger.error(f"Error verificando timeout de migración: {e}")
            
            # Migración en progreso normal
            try:
                total_migrados = safe_int(migracion_en_progreso.usuarios_migrados)
                
                # Calcular progreso basado en tablas procesadas
                tablas_inspeccionadas = safe_int(safe_getattr(migracion_en_progreso, 'tablas_inspeccionadas', 0))
                tablas_procesadas = safe_int(safe_getattr(migracion_en_progreso, 'tablas_con_datos', 0)) + safe_int(safe_getattr(migracion_en_progreso, 'tablas_vacias', 0))
                progreso_real = int((tablas_procesadas / tablas_inspeccionadas) * 100) if tablas_inspeccionadas > 0 else 0
                
                data = {
                    'en_progreso': True,
                    'estado': safe_str(migracion_en_progreso.estado) or 'iniciada',
                    'progreso_real': min(progreso_real, 100),
                    'fecha_inicio': migracion_en_progreso.fecha_inicio.isoformat() if migracion_en_progreso.fecha_inicio else None,
                    'total_migrados': total_migrados,
                    'tablas_inspeccionadas': tablas_inspeccionadas,
                    'tablas_con_datos': safe_int(safe_getattr(migracion_en_progreso, 'tablas_con_datos', 0)),
                    'tablas_vacias': safe_int(safe_getattr(migracion_en_progreso, 'tablas_vacias', 0)),
                    'host_origen': safe_str(safe_getattr(migracion_en_progreso, 'host_origen', '')) or 'N/A',
                    'base_datos_origen': safe_str(safe_getattr(migracion_en_progreso, 'base_datos_origen', '')) or 'N/A',
                }
                return JsonResponse(data)
            except Exception as e:
                logger.error(f"Error procesando migración en progreso: {e}")
        
        # Verificar migración completada recientemente
        try:
            hace_10_minutos = timezone.now() - timedelta(minutes=10)
            migracion_reciente = MigracionLog.objects.filter(
                fecha_inicio__gte=hace_10_minutos,
                estado='completada'
            ).first()
            
            if migracion_reciente:
                total_migrados = safe_int(migracion_reciente.usuarios_migrados)
                
                data = {
                    'en_progreso': False,
                    'completada_recientemente': True,
                    'estado': safe_str(migracion_reciente.estado) or 'completada',
                    'fecha_inicio': migracion_reciente.fecha_inicio.isoformat() if migracion_reciente.fecha_inicio else None,
                    'fecha_fin': migracion_reciente.fecha_fin.isoformat() if migracion_reciente.fecha_fin else None,
                    'total_migrados': total_migrados,
                    'tablas_inspeccionadas': safe_int(safe_getattr(migracion_reciente, 'tablas_inspeccionadas', 0)),
                    'tablas_con_datos': safe_int(safe_getattr(migracion_reciente, 'tablas_con_datos', 0)),
                    'tablas_vacias': safe_int(safe_getattr(migracion_reciente, 'tablas_vacias', 0)),
                    'host_origen': safe_str(safe_getattr(migracion_reciente, 'host_origen', '')) or 'N/A',
                }
                return JsonResponse(data)
        except Exception as e:
            logger.warning(f"Error buscando migración reciente: {e}")
        
        # No hay migraciones
        data = {
            'en_progreso': False,
            'completada_recientemente': False,
            'estado': 'sin_migraciones',
            'mensaje': 'No hay migraciones registradas'
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        # Log del error para debugging
        logger.error(f"Error en estado_migracion_ajax: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Respuesta de error segura
        return JsonResponse({
            'error': 'Error interno del servidor',
            'en_progreso': False,
            'completada_recientemente': False,
            'estado': 'error',
            'mensaje': 'Error al obtener estado de migración'
        })

@login_required
def cancelar_migracion_ajax(request):
    """Vista AJAX para cancelar una migración en progreso"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        from .models import MigracionLog
        from django.utils import timezone
        from django.core.cache import cache
        
        # Buscar migración en progreso
        migracion_en_progreso = MigracionLog.objects.filter(
            estado__in=['iniciada', 'en_progreso']
        ).first()
        
        if migracion_en_progreso:
            # Marcar como cancelada
            migracion_en_progreso.estado = 'cancelada'
            migracion_en_progreso.fecha_fin = timezone.now()
            migracion_en_progreso.errores = f'Migración cancelada por el usuario {request.user.username}'
            migracion_en_progreso.save()
        
        # Limpiar cache de progreso
        cache.delete('migracion_progreso')
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Migración cancelada correctamente'
        })
        
    except Exception as e:
        logger.error(f"Error cancelando migración: {str(e)}")
        
        return JsonResponse({
            'error': 'Error al cancelar la migración',
            'mensaje': str(e)
        }, status=500)

@login_required
def continuar_migracion_ajax(request):
    """Vista AJAX para continuar una migración desde donde se quedó"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        from .models import MigracionLog
        from django.utils import timezone
        from django.core.cache import cache
        import json
        
        # Obtener datos de la solicitud
        data = json.loads(request.body)
        
        # Buscar la última migración con error
        migracion_error = MigracionLog.objects.filter(
            estado='error'
        ).first()
        
        if not migracion_error:
            return JsonResponse({
                'error': 'No se encontró una migración con error para continuar'
            }, status=404)
        
        # Obtener configuración de conexión desde la migración anterior
        host = migracion_error.host_origen
        database = migracion_error.base_datos_origen
        
        # Necesitamos las credenciales del usuario (deberían enviarse en la solicitud)
        user_db = data.get('user')
        password_db = data.get('password')
        port = data.get('port', 3306)
        
        if not all([user_db, password_db]):
            return JsonResponse({
                'error': 'Se requieren las credenciales de la base de datos para continuar'
            }, status=400)
        
        # Crear nueva migración marcando la anterior como "continuada"
        migracion_error.errores += f' | Continuada por {request.user.username} el {timezone.now()}'
        migracion_error.save()
        
        # Ejecutar migración en hilo separado
        def ejecutar_continuacion_migracion():
            try:
                from .services import MigracionService
                servicio = MigracionService(host, database, user_db, password_db, int(port))
                servicio.inspeccionar_y_migrar_automaticamente(request.user)
            except Exception as e:
                logger.error(f"Error en continuación de migración: {e}")
        
        import threading
        thread = threading.Thread(target=ejecutar_continuacion_migracion)
        thread.daemon = True
        thread.start()
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Migración reiniciada correctamente'
        })
        
    except Exception as e:
        logger.error(f"Error continuando migración: {str(e)}")
        
        return JsonResponse({
            'error': 'Error al continuar la migración',
            'mensaje': str(e)
        }, status=500)

@login_required
def limpiar_cache_migracion(request):
    """Vista AJAX para limpiar el cache de migración y permitir continuar"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        from django.core.cache import cache
        from .models import MigracionLog
        from django.utils import timezone
        
        # Limpiar cache de progreso de migración
        cache.delete('migracion_progreso')
        
        # Buscar migraciones en estado de error y marcarlas como "reiniciables"
        migraciones_error = MigracionLog.objects.filter(
            estado='error',
            fecha_inicio__gte=timezone.now() - timezone.timedelta(hours=2)  # Últimas 2 horas
        )
        
        for migracion in migraciones_error:
            # Agregar nota de que se limpió el cache
            if migracion.errores:
                migracion.errores += f' | Cache limpiado por {request.user.username} el {timezone.now()}'
            else:
                migracion.errores = f'Cache limpiado por {request.user.username} el {timezone.now()}'
            migracion.save()
        
        logger.info(f"Cache de migración limpiado por usuario {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Cache de migración limpiado correctamente. Puede intentar continuar la migración.'
        })
        
    except Exception as e:
        logger.error(f"Error limpiando cache de migración: {str(e)}")
        
        return JsonResponse({
            'error': 'Error al limpiar el cache de migración',
            'mensaje': str(e)
        }, status=500)

@login_required
def exportar_excel(request, pk):
    """Vista para exportar un dato archivado a Excel"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.shortcuts import get_object_or_404
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import json
        from datetime import datetime
        
        # Obtener el dato archivado
        dato = get_object_or_404(DatoArchivadoDinamico, pk=pk)
        
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Dato_{dato.tabla_origen}_{dato.id_original}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        info_font = Font(bold=True, color="000000")
        info_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        
        # Título principal
        ws.merge_cells('A1:C1')
        ws['A1'] = f"DETALLE DEL DATO ARCHIVADO"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Información general
        row = 3
        ws[f'A{row}'] = "INFORMACIÓN GENERAL"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        info_data = [
            ("Tabla de Origen", dato.tabla_origen),
            ("ID Original", dato.id_original),
            ("Fecha de Migración", dato.fecha_migracion.strftime('%d/%m/%Y %H:%M:%S')),
            ("Nombre del Registro", dato.obtener_nombre_legible()),
        ]
        
        for campo, valor in info_data:
            ws[f'A{row}'] = campo
            ws[f'A{row}'].font = info_font
            ws[f'A{row}'].fill = info_fill
            ws[f'B{row}'] = str(valor)
            row += 1
        
        # Datos del registro
        row += 1
        ws[f'A{row}'] = "DATOS DEL REGISTRO"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Campo"
        ws[f'B{row}'] = "Valor"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        
        # Agregar datos del registro
        for campo, valor in dato.datos_originales.items():
            row += 1
            ws[f'A{row}'] = str(campo)
            ws[f'B{row}'] = str(valor) if valor is not None else "NULL"
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"dato_archivado_{dato.tabla_origen}_{dato.id_original}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en response
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar a Excel: {str(e)}')
        return redirect('datos_archivados:dato_detail', pk=pk)

@login_required
def exportar_tabla_excel(request, tabla):
    """Vista para exportar todos los datos de una tabla a Excel"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Obtener todos los datos de la tabla
        datos = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla).order_by('id_original')
        
        if not datos.exists():
            messages.error(request, f'No hay datos para exportar de la tabla {tabla}.')
            return redirect('datos_archivados:datos_list', tabla=tabla)
        
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Tabla_{tabla}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16, color="000000")
        info_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título principal
        ws.merge_cells('A1:F1')
        ws['A1'] = f"DATOS ARCHIVADOS - TABLA: {tabla.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Información de la exportación
        row = 3
        ws[f'A{row}'] = f"Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{row}'].font = info_font
        row += 1
        ws[f'A{row}'] = f"Total de registros: {datos.count()}"
        ws[f'A{row}'].font = info_font
        row += 1
        ws[f'A{row}'] = f"Usuario: {request.user.username}"
        ws[f'A{row}'].font = info_font
        
        # Obtener todos los campos únicos de todos los registros
        todos_los_campos = set()
        for dato in datos:
            todos_los_campos.update(dato.datos_originales.keys())
        
        campos_ordenados = sorted(list(todos_los_campos))
        
        # Headers de la tabla
        row = 6
        ws[f'A{row}'] = "ID Sistema"
        ws[f'B{row}'] = "ID Original"
        ws[f'C{row}'] = "Fecha Migración"
        
        # Agregar headers de campos dinámicos
        col_start = 4  # Columna D
        for i, campo in enumerate(campos_ordenados):
            col_letter = get_column_letter(col_start + i)
            ws[f'{col_letter}{row}'] = campo
        
        # Aplicar estilo a headers
        for col in range(1, len(campos_ordenados) + 4):
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Agregar datos
        for dato in datos:
            row += 1
            ws[f'A{row}'] = dato.pk
            ws[f'B{row}'] = dato.id_original
            ws[f'C{row}'] = dato.fecha_migracion.strftime('%d/%m/%Y %H:%M')
            
            # Agregar datos de campos dinámicos
            for i, campo in enumerate(campos_ordenados):
                col_letter = get_column_letter(col_start + i)
                valor = dato.datos_originales.get(campo)
                ws[f'{col_letter}{row}'] = str(valor) if valor is not None else "NULL"
            
            # Aplicar bordes a toda la fila
            for col in range(1, len(campos_ordenados) + 4):
                col_letter = get_column_letter(col)
                ws[f'{col_letter}{row}'].border = border
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Máximo 30 caracteres
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)  # Mínimo 10
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"tabla_{tabla}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en response
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar tabla a Excel: {str(e)}')
        return redirect('datos_archivados:datos_list', tabla=tabla)

@login_required
def eliminar_tablas(request):
    """Vista para eliminar tablas seleccionadas y todos sus registros"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    if request.method == 'POST':
        try:
            from .models import DatoArchivadoDinamico
            
            # Obtener tablas a eliminar
            tablas = request.POST.getlist('tablas')
            
            if not tablas:
                messages.warning(request, 'No se seleccionaron tablas para eliminar.')
                return redirect('datos_archivados:tablas_list')
            
            # Contar registros antes de eliminar
            total_registros = 0
            for tabla in tablas:
                count = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla).count()
                total_registros += count
            
            # Eliminar registros de las tablas seleccionadas
            eliminados = 0
            for tabla in tablas:
                eliminados += DatoArchivadoDinamico.objects.filter(tabla_origen=tabla).delete()[0]
            
            messages.success(
                request, 
                f'Se eliminaron {len(tablas)} tabla(s) con un total de {eliminados} registros.'
            )
            
        except Exception as e:
            messages.error(request, f'Error al eliminar tablas: {str(e)}')
    
    return redirect('datos_archivados:tablas_list')

def reclamar_usuario_archivado(request):
    """Vista para que usuarios reclamen su cuenta archivada - Paso 1: Validar datos y enviar código"""
    from .forms import ReclamarUsuarioArchivadoForm
    from .models import CodigoVerificacionReclamacion, UsuarioArchivado
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    if request.method == 'POST':
        form = ReclamarUsuarioArchivadoForm(request.POST)
        if form.is_valid():
            try:
                # Obtener el usuario encontrado y su tipo
                usuario_encontrado = form.cleaned_data['usuario_encontrado']
                tipo_fuente = form.cleaned_data['tipo_fuente']
                
                # Extraer email y username según el tipo de fuente
                if tipo_fuente == 'usuario_archivado':
                    email = usuario_encontrado.email
                    username = usuario_encontrado.username
                    usuario_archivado_id = usuario_encontrado.id
                    dato_archivado_id = None
                else:  # dato_dinamico
                    datos = usuario_encontrado.datos_originales
                    email = datos.get('email')
                    username = datos.get('username')
                    usuario_archivado_id = None
                    dato_archivado_id = usuario_encontrado.id
                
                # Generar código de 4 dígitos
                codigo = ''.join(random.choices(string.digits, k=4))
                
                # Crear o actualizar código de verificación
                fecha_expiracion = timezone.now() + timedelta(minutes=15)
                
                # Eliminar códigos anteriores para este email
                CodigoVerificacionReclamacion.objects.filter(email=email).delete()
                
                # Crear nuevo código
                codigo_obj = CodigoVerificacionReclamacion.objects.create(
                    email=email,
                    codigo=codigo,
                    fecha_expiracion=fecha_expiracion
                )
                
                # Vincular con la fuente correspondiente
                if tipo_fuente == 'usuario_archivado':
                    codigo_obj.usuario_archivado = usuario_encontrado
                else:
                    codigo_obj.dato_archivado = usuario_encontrado
                codigo_obj.save()
                
                # Enviar email
                email_text = f'''Hola {username},

Se ha solicitado reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Para continuar con el proceso, ingrese el siguiente código de verificación:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
                
                send_mail(
                    'Código de Verificación - Reclamación de Cuenta Archivada',
                    email_text,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                
                # Guardar información en sesión para el siguiente paso
                request.session['reclamacion_tipo_fuente'] = tipo_fuente
                if usuario_archivado_id:
                    request.session['reclamacion_usuario_archivado_id'] = usuario_archivado_id
                if dato_archivado_id:
                    request.session['reclamacion_dato_archivado_id'] = dato_archivado_id
                request.session['reclamacion_email'] = email
                request.session['reclamacion_nueva_password'] = form.cleaned_data['nueva_password']
                
                messages.success(request, f'Se ha enviado un código de verificación a {email}')
                return redirect('datos_archivados:verificar_codigo_reclamacion_tradicional')
                
            except Exception as e:
                messages.error(request, f'Error al enviar el código de verificación: {str(e)}')
    else:
        form = ReclamarUsuarioArchivadoForm()
    
    return render(request, 'datos_archivados/reclamar_usuario.html', {'form': form})

def reenviar_codigo_reclamacion_tradicional(request):
    """Vista para reenviar el código de verificación"""
    from .models import CodigoVerificacionReclamacion, UsuarioArchivado, DatoArchivadoDinamico
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    # Verificar que hay una reclamación en proceso
    tipo_fuente = request.session.get('reclamacion_tipo_fuente')
    usuario_archivado_id = request.session.get('reclamacion_usuario_archivado_id')
    dato_archivado_id = request.session.get('reclamacion_dato_archivado_id')
    email = request.session.get('reclamacion_email')
    
    if not email or not tipo_fuente:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:reclamar_usuario')
    
    try:
        # Generar nuevo código de 4 dígitos
        codigo = ''.join(random.choices(string.digits, k=4))
        
        # Eliminar códigos anteriores para este email
        CodigoVerificacionReclamacion.objects.filter(email=email).delete()
        
        # Crear nuevo código
        fecha_expiracion = timezone.now() + timedelta(minutes=15)
        codigo_obj = CodigoVerificacionReclamacion.objects.create(
            email=email,
            codigo=codigo,
            fecha_expiracion=fecha_expiracion
        )
        
        # Vincular con la fuente correspondiente
        if tipo_fuente == 'usuario_archivado' and usuario_archivado_id:
            usuario_archivado = UsuarioArchivado.objects.get(id=usuario_archivado_id)
            codigo_obj.usuario_archivado = usuario_archivado
            username = usuario_archivado.username
        elif tipo_fuente == 'dato_dinamico' and dato_archivado_id:
            dato_archivado = DatoArchivadoDinamico.objects.get(id=dato_archivado_id)
            codigo_obj.dato_archivado = dato_archivado
            datos = dato_archivado.datos_originales
            username = datos.get('username', email.split('@')[0])
        else:
            username = email.split('@')[0]
        
        codigo_obj.save()
        
        # Enviar email
        email_text = f'''Hola {username},

Se ha solicitado un nuevo código de verificación para reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Su nuevo código de verificación es:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
        
        send_mail(
            'Nuevo Código de Verificación - Reclamación de Cuenta',
            email_text,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        
        messages.success(request, f'Se ha enviado un nuevo código de verificación a {email}')
        return redirect('datos_archivados:verificar_codigo_reclamacion_tradicional')
        
    except Exception as e:
        messages.error(request, f'Error al reenviar el código: {str(e)}')
        return redirect('datos_archivados:verificar_codigo_reclamacion_tradicional')

def verificar_codigo_reclamacion_tradicional(request):
    """Vista para verificar el código de 4 dígitos en reclamación tradicional"""
    from .models import CodigoVerificacionReclamacion, UsuarioArchivado, DatoArchivadoDinamico
    from django.contrib.auth import login
    from django.contrib.auth.models import User, Group
    
    # Verificar que hay una reclamación en proceso
    tipo_fuente = request.session.get('reclamacion_tipo_fuente')
    usuario_archivado_id = request.session.get('reclamacion_usuario_archivado_id')
    dato_archivado_id = request.session.get('reclamacion_dato_archivado_id')
    email = request.session.get('reclamacion_email')
    nueva_password = request.session.get('reclamacion_nueva_password')
    
    if not email or not nueva_password or not tipo_fuente:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:reclamar_usuario')
    
    if request.method == 'POST':
        codigo_ingresado = request.POST.get('code', '').strip()
        
        if not codigo_ingresado:
            messages.error(request, 'Debe ingresar el código de verificación.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
                'email': email
            })
        
        try:
            # Buscar código válido (el más reciente si hay múltiples)
            codigo_verificacion = CodigoVerificacionReclamacion.objects.filter(
                email=email,
                codigo=codigo_ingresado,
                usado=False
            ).order_by('-fecha_creacion').first()
            
            if not codigo_verificacion:
                raise CodigoVerificacionReclamacion.DoesNotExist()
            
            if codigo_verificacion.is_expired():
                messages.error(request, 'El código de verificación ha expirado. Solicite uno nuevo.')
                return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
                    'email': email
                })
            
            # NO marcar como usado todavía - esperar a que todo se complete
            
            # Obtener datos según el tipo de fuente
            if tipo_fuente == 'usuario_archivado':
                usuario_archivado = UsuarioArchivado.objects.get(id=usuario_archivado_id)
                username_original = usuario_archivado.username
                email_usuario = usuario_archivado.email
                first_name = usuario_archivado.first_name
                last_name = usuario_archivado.last_name
            else:  # dato_dinamico
                dato_archivado = DatoArchivadoDinamico.objects.get(id=dato_archivado_id)
                datos = dato_archivado.datos_originales
                username_original = datos.get('username', '')
                email_usuario = datos.get('email', '')
                first_name = datos.get('first_name', '')
                last_name = datos.get('last_name', '')
            
            # Generar username único con protección contra condiciones de carrera
            from django.db import IntegrityError
            
            user = None
            max_intentos = 100
            
            for intento in range(max_intentos):
                # Generar username candidato
                if intento == 0:
                    username_candidato = username_original
                else:
                    username_candidato = f"{username_original}{intento - 1}"
                
                # Verificar si ya existe
                if User.objects.filter(username=username_candidato).exists():
                    continue
                
                # Intentar crear el usuario
                try:
                    user = User.objects.create_user(
                        username=username_candidato,
                        email=email_usuario,
                        first_name=first_name,
                        last_name=last_name,
                        password=nueva_password,
                        is_active=True
                    )
                    username_final = username_candidato
                    break  # Usuario creado exitosamente
                except IntegrityError:
                    # El username fue creado por otro proceso entre la verificación y la creación
                    # Continuar con el siguiente número
                    continue
            
            if not user:
                messages.error(request, 'No se pudo generar un username único. Contacte al administrador.')
                return redirect('datos_archivados:reclamar_usuario')
            
            # Asignar al grupo Estudiantes
            try:
                grupo_estudiantes = Group.objects.get(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            except Group.DoesNotExist:
                grupo_estudiantes = Group.objects.create(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            
            # Vincular con la fuente correspondiente
            if tipo_fuente == 'usuario_archivado':
                usuario_archivado.usuario_actual = user
                usuario_archivado.save()
            
            # Enviar email de confirmación
            username_cambio = username_original != username_final
            
            if username_cambio:
                mensaje_username = f'''IMPORTANTE: Su nombre de usuario ha cambiado
Nombre de usuario original: {username_original}
Nuevo nombre de usuario: {username_final}

Esto se debe a que ya existía un usuario con el nombre "{username_original}" en el sistema.
Por favor, use "{username_final}" para futuros inicios de sesión.'''
            else:
                mensaje_username = f'Su nombre de usuario es: {username_final}'
            
            mensaje = f'''¡Bienvenido de vuelta al Centro Fray Bartolomé de las Casas!

Su cuenta ha sido reactivada exitosamente.

DETALLES DE SU CUENTA:
{mensaje_username}
Correo electrónico: {user.email}
Nombre completo: {user.get_full_name()}

Ahora puede acceder a todos los servicios del sistema usando sus credenciales.

Si tiene alguna pregunta o necesita ayuda, no dude en contactarnos.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            send_mail(
                'Cuenta Reactivada - Centro Fray Bartolomé de las Casas',
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            
            # Marcar código como usado SOLO después de que todo se completó exitosamente
            codigo_verificacion.usado = True
            codigo_verificacion.save()
            
            # Limpiar sesión
            if 'reclamacion_usuario_archivado_id' in request.session:
                del request.session['reclamacion_usuario_archivado_id']
            if 'reclamacion_dato_archivado_id' in request.session:
                del request.session['reclamacion_dato_archivado_id']
            if 'reclamacion_tipo_fuente' in request.session:
                del request.session['reclamacion_tipo_fuente']
            del request.session['reclamacion_email']
            del request.session['reclamacion_nueva_password']
            
            # Iniciar sesión automáticamente con backend específico
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            
            messages.success(
                request, 
                f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                'Su cuenta ha sido reactivada exitosamente. '
                'Se le ha enviado un email con los detalles de su cuenta.'
            )
            
            return redirect('principal:login_redirect')
            
        except CodigoVerificacionReclamacion.DoesNotExist:
            messages.error(request, 'Código de verificación incorrecto.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
                'email': email
            })
        except Exception as e:
            messages.error(request, f'Error al reactivar la cuenta: {str(e)}')
            return redirect('datos_archivados:reclamar_usuario')
    
    return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
        'email': email
    })

def buscar_usuario_archivado(request):
    """Vista para buscar usuarios archivados disponibles para reclamar"""
    from .forms import BuscarUsuarioArchivadoForm
    
    form = BuscarUsuarioArchivadoForm()
    usuarios_encontrados = []
    
    if request.method == 'GET' and 'busqueda' in request.GET:
        form = BuscarUsuarioArchivadoForm(request.GET)
        if form.is_valid():
            usuarios_encontrados = form.buscar_usuarios()
    
    context = {
        'form': form,
        'usuarios_encontrados': usuarios_encontrados,
    }
    
    return render(request, 'datos_archivados/buscar_usuario.html', context)

def iniciar_reclamacion_usuario(request, dato_id):
    """Vista para iniciar el proceso de reclamación enviando código de verificación"""
    from .models import DatoArchivadoDinamico, CodigoVerificacionReclamacion
    from django.shortcuts import get_object_or_404
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    # Obtener el dato archivado
    dato = get_object_or_404(DatoArchivadoDinamico, id=dato_id)
    datos = dato.datos_originales
    email = datos.get('email') or datos.get('correo')
    
    if not email:
        messages.error(request, 'No se encontró un email válido para este usuario archivado.')
        return redirect('datos_archivados:buscar_usuario')
    
    # Generar código de 4 dígitos
    codigo = ''.join(random.choices(string.digits, k=4))
    
    # Crear o actualizar código de verificación
    fecha_expiracion = timezone.now() + timedelta(minutes=15)  # Expira en 15 minutos
    
    # Eliminar códigos anteriores para este email
    CodigoVerificacionReclamacion.objects.filter(email=email).delete()
    
    # Crear nuevo código
    codigo_verificacion = CodigoVerificacionReclamacion.objects.create(
        email=email,
        codigo=codigo,
        dato_archivado=dato,
        fecha_expiracion=fecha_expiracion
    )
    
    # Enviar email
    username = datos.get('username') or datos.get('user') or 'Usuario'
    email_text = f'''Hola {username},

Se ha solicitado reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Para continuar con el proceso, ingrese el siguiente código de verificación:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
    
    try:
        send_mail(
            'Código de Verificación - Reclamación de Cuenta Archivada',
            email_text,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        
        # Guardar información en sesión para el siguiente paso
        request.session['reclamacion_dato_id'] = dato_id
        request.session['reclamacion_email'] = email
        
        messages.success(request, f'Se ha enviado un código de verificación a {email}')
        return redirect('datos_archivados:verificar_codigo_reclamacion')
        
    except Exception as e:
        messages.error(request, f'Error al enviar el código de verificación: {str(e)}')
        return redirect('datos_archivados:buscar_usuario')

def reenviar_codigo_reclamacion(request):
    """Vista para reenviar el código de verificación (búsqueda por email)"""
    from .models import CodigoVerificacionReclamacion, DatoArchivadoDinamico
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    # Verificar que hay una reclamación en proceso
    dato_id = request.session.get('reclamacion_dato_id')
    email = request.session.get('reclamacion_email')
    
    if not dato_id or not email:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:buscar_usuario')
    
    try:
        # Obtener el dato archivado
        dato = DatoArchivadoDinamico.objects.get(id=dato_id)
        datos = dato.datos_originales
        username = datos.get('username') or datos.get('user') or 'Usuario'
        
        # Generar nuevo código de 4 dígitos
        codigo = ''.join(random.choices(string.digits, k=4))
        
        # Eliminar códigos anteriores para este email
        CodigoVerificacionReclamacion.objects.filter(email=email).delete()
        
        # Crear nuevo código
        fecha_expiracion = timezone.now() + timedelta(minutes=15)
        CodigoVerificacionReclamacion.objects.create(
            email=email,
            codigo=codigo,
            dato_archivado=dato,
            fecha_expiracion=fecha_expiracion
        )
        
        # Enviar email
        email_text = f'''Hola {username},

Se ha solicitado un nuevo código de verificación para reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Su nuevo código de verificación es:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
        
        send_mail(
            'Nuevo Código de Verificación - Reclamación de Cuenta',
            email_text,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        
        messages.success(request, f'Se ha enviado un nuevo código de verificación a {email}')
        return redirect('datos_archivados:verificar_codigo_reclamacion')
        
    except Exception as e:
        messages.error(request, f'Error al reenviar el código: {str(e)}')
        return redirect('datos_archivados:verificar_codigo_reclamacion')

def verificar_codigo_reclamacion(request):
    """Vista para verificar el código de 4 dígitos"""
    from .models import CodigoVerificacionReclamacion
    
    # Verificar que hay una reclamación en proceso
    dato_id = request.session.get('reclamacion_dato_id')
    email = request.session.get('reclamacion_email')
    
    if not dato_id or not email:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:buscar_usuario')
    
    if request.method == 'POST':
        codigo_ingresado = request.POST.get('code', '').strip()
        
        if not codigo_ingresado:
            messages.error(request, 'Debe ingresar el código de verificación.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
                'email': email
            })
        
        try:
            # Buscar código válido (el más reciente si hay múltiples)
            codigo_verificacion = CodigoVerificacionReclamacion.objects.filter(
                email=email,
                codigo=codigo_ingresado,
                usado=False
            ).order_by('-fecha_creacion').first()
            
            if not codigo_verificacion:
                raise CodigoVerificacionReclamacion.DoesNotExist()
            
            if codigo_verificacion.is_expired():
                messages.error(request, 'El código de verificación ha expirado. Solicite uno nuevo.')
                return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
                    'email': email
                })
            
            # Marcar código como usado SOLO después de verificar que es válido
            codigo_verificacion.usado = True
            codigo_verificacion.save()
            
            # Redirigir a la página de reclamación
            messages.success(request, 'Código verificado correctamente.')
            return redirect('datos_archivados:reclamar_usuario_dinamico', dato_id=dato_id)
            
        except CodigoVerificacionReclamacion.DoesNotExist:
            messages.error(request, 'Código de verificación incorrecto.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
                'email': email
            })
    
    return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
        'email': email
    })

def reclamar_usuario_dinamico(request, dato_id):
    """Vista para reclamar un usuario desde datos archivados dinámicos (después de verificación)"""
    from .models import DatoArchivadoDinamico
    from django.contrib.auth.models import User, Group
    from django.contrib.auth import login
    from django.shortcuts import get_object_or_404
    from django.core.mail import send_mail
    from django.conf import settings
    
    # Verificar que el proceso de verificación se completó
    session_dato_id = request.session.get('reclamacion_dato_id')
    if not session_dato_id or int(session_dato_id) != dato_id:
        messages.error(request, 'Debe completar la verificación por email primero.')
        return redirect('datos_archivados:buscar_usuario')
    
    # Obtener el dato archivado
    dato = get_object_or_404(DatoArchivadoDinamico, id=dato_id)
    datos = dato.datos_originales
    
    if request.method == 'POST':
        # Obtener datos del formulario
        nueva_password = request.POST.get('nueva_password')
        confirmar_password = request.POST.get('confirmar_password')
        
        # Validaciones
        if not nueva_password or not confirmar_password:
            messages.error(request, 'Debe proporcionar una nueva contraseña.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        if nueva_password != confirmar_password:
            messages.error(request, 'Las contraseñas no coinciden.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        if len(nueva_password) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        # Extraer información del usuario
        username_original = datos.get('username') or datos.get('user') or f"usuario_{dato.id_original}"
        email = datos.get('email') or datos.get('correo') or ''
        first_name = datos.get('first_name') or datos.get('nombre') or datos.get('name') or ''
        last_name = datos.get('last_name') or datos.get('apellido') or datos.get('apellidos') or ''
        
        # Generar username único con protección contra condiciones de carrera
        from django.db import IntegrityError
        
        user = None
        max_intentos = 100
        
        for intento in range(max_intentos):
            # Generar username candidato
            if intento == 0:
                username_candidato = username_original
            else:
                username_candidato = f"{username_original}{intento - 1}"
            
            # Verificar si ya existe
            if User.objects.filter(username=username_candidato).exists():
                continue
            
            # Intentar crear el usuario
            try:
                user = User.objects.create_user(
                    username=username_candidato,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password=nueva_password,
                    is_active=True
                )
                username_final = username_candidato
                break  # Usuario creado exitosamente
            except IntegrityError:
                # El username fue creado por otro proceso entre la verificación y la creación
                # Continuar con el siguiente número
                continue
        
        if not user:
            messages.error(request, f'No se pudo generar un username único para {username_original}. Contacte al administrador.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        try:
            
            # Asignar al grupo Estudiantes por defecto
            try:
                grupo_estudiantes = Group.objects.get(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            except Group.DoesNotExist:
                # Crear el grupo si no existe
                grupo_estudiantes = Group.objects.create(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            
            # Enviar email de confirmación con detalles de la cuenta
            username_cambio = username_original != username_final
            
            if username_cambio:
                mensaje_username = f'''IMPORTANTE: Su nombre de usuario ha cambiado
Nombre de usuario original: {username_original}
Nuevo nombre de usuario: {username_final}

Esto se debe a que ya existía un usuario con el nombre "{username_original}" en el sistema.
Por favor, use "{username_final}" para futuros inicios de sesión.'''
            else:
                mensaje_username = f'Su nombre de usuario es: {username_final}'
            
            mensaje = f'''¡Bienvenido de vuelta al Centro Fray Bartolomé de las Casas!

Su cuenta ha sido reactivada exitosamente desde los datos archivados.

DETALLES DE SU CUENTA:
{mensaje_username}
Correo electrónico: {user.email}
Nombre completo: {user.get_full_name()}

Ahora puede acceder a todos los servicios del sistema usando sus credenciales.

Si tiene alguna pregunta o necesita ayuda, no dude en contactarnos.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            try:
                send_mail(
                    'Cuenta Reactivada - Centro Fray Bartolomé de las Casas',
                    mensaje,
                    settings.DEFAULT_FROM_EMAIL,
                    [user.email],
                    fail_silently=False,
                )
            except Exception as e:
                # No fallar si el email no se puede enviar
                pass
            
            # Limpiar sesión
            if 'reclamacion_dato_id' in request.session:
                del request.session['reclamacion_dato_id']
            if 'reclamacion_email' in request.session:
                del request.session['reclamacion_email']
            
            # Iniciar sesión automáticamente con backend específico
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            
            messages.success(
                request, 
                f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                'Su cuenta ha sido reactivada exitosamente desde los datos archivados. '
                'Se le ha enviado un email con los detalles de su cuenta.'
            )
            
            return redirect('principal:login_redirect')
            
        except Exception as e:
            messages.error(request, f'Error al crear la cuenta: {str(e)}')
    
    context = {
        'dato': dato,
        'datos': datos,
        'username': datos.get('username') or datos.get('user') or f"usuario_{dato.id_original}",
        'email': datos.get('email') or datos.get('correo') or '',
        'first_name': datos.get('first_name') or datos.get('nombre') or datos.get('name') or '',
        'last_name': datos.get('last_name') or datos.get('apellido') or datos.get('apellidos') or '',
    }
    
    return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', context)

@login_required
def buscar_datos_ajax(request):
    """Vista AJAX para búsqueda en tiempo real de datos archivados"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.db.models import Q, Count
        
        search_query = request.GET.get('q', '').strip()
        search_type = request.GET.get('type', 'global')
        limit = int(request.GET.get('limit', 10))
        
        if not search_query:
            return JsonResponse({
                'results': [],
                'total': 0,
                'query': search_query
            })
        
        # Construir query de búsqueda
        queryset = DatoArchivadoDinamico.objects.all()
        
        if search_type == 'tabla':
            queryset = queryset.filter(tabla_origen__icontains=search_query)
        elif search_type == 'contenido':
            queryset = queryset.filter(
                Q(datos_originales__icontains=search_query) |
                Q(nombre_registro__icontains=search_query)
            )
        else:  # global
            queryset = queryset.filter(
                Q(tabla_origen__icontains=search_query) |
                Q(datos_originales__icontains=search_query) |
                Q(nombre_registro__icontains=search_query)
            )
        
        # Obtener resultados agrupados por tabla
        tablas_resultados = queryset.values('tabla_origen').annotate(
            total=Count('id')
        ).order_by('-total', 'tabla_origen')[:limit]
        
        # Obtener algunos registros de ejemplo para cada tabla
        resultados = []
        for tabla in tablas_resultados:
            tabla_nombre = tabla['tabla_origen']
            total_registros = tabla['total']
            
            # Obtener algunos registros de ejemplo
            ejemplos = queryset.filter(tabla_origen=tabla_nombre)[:3]
            ejemplos_data = []
            
            for ejemplo in ejemplos:
                # Extraer información relevante del registro
                datos = ejemplo.datos_originales
                nombre = ejemplo.obtener_nombre_legible()
                
                # Buscar campos que contengan el término de búsqueda
                campos_coincidentes = []
                if search_type in ['contenido', 'global']:
                    for campo, valor in datos.items():
                        if search_query.lower() in str(valor).lower():
                            campos_coincidentes.append({
                                'campo': campo,
                                'valor': str(valor)[:100] + ('...' if len(str(valor)) > 100 else '')
                            })
                
                ejemplos_data.append({
                    'id': ejemplo.id,
                    'nombre': nombre,
                    'campos_coincidentes': campos_coincidentes[:3]  # Máximo 3 campos
                })
            
            resultados.append({
                'tabla': tabla_nombre,
                'total_registros': total_registros,
                'ejemplos': ejemplos_data
            })
        
        return JsonResponse({
            'results': resultados,
            'total': queryset.count(),
            'query': search_query,
            'search_type': search_type
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error en búsqueda: {str(e)}',
            'results': [],
            'total': 0
        })

@login_required
def buscar_en_tabla_ajax(request, tabla):
    """Vista AJAX para búsqueda en tiempo real dentro de una tabla específica"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.db.models import Q
        
        search_query = request.GET.get('q', '').strip()
        order_by = request.GET.get('order_by', 'fecha_migracion')
        order_direction = request.GET.get('order_direction', 'desc')
        limit = int(request.GET.get('limit', 20))
        
        # Query base para la tabla específica
        queryset = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla)
        
        # Aplicar filtro de búsqueda si existe
        if search_query:
            queryset = queryset.filter(
                Q(datos_originales__icontains=search_query) |
                Q(nombre_registro__icontains=search_query)
            )
        
        # Aplicar ordenamiento
        valid_order_fields = {
            'fecha_migracion': 'fecha_migracion',
            'id_original': 'id_original',
            'nombre': 'nombre_registro',
            'tabla': 'tabla_origen'
        }
        
        if order_by in valid_order_fields:
            order_field = valid_order_fields[order_by]
            if order_direction == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field)
        
        # Obtener resultados limitados
        resultados = queryset[:limit]
        
        # Para ordenamiento por nombre, necesitamos ordenar después
        if order_by == 'nombre':
            resultados_list = list(resultados)
            resultados_list.sort(
                key=lambda x: (x.obtener_nombre_legible() or '').lower(),
                reverse=(order_direction == 'desc')
            )
            resultados = resultados_list
        
        # Preparar datos para JSON
        datos_json = []
        for dato in resultados:
            # Resaltar términos de búsqueda en el nombre
            nombre = dato.obtener_nombre_legible()
            if search_query and nombre:
                # Simple highlighting (se puede mejorar)
                nombre_resaltado = nombre.replace(
                    search_query, 
                    f'<mark>{search_query}</mark>'
                )
            else:
                nombre_resaltado = nombre
            
            # Buscar campos que coincidan con la búsqueda
            campos_coincidentes = []
            if search_query:
                for campo, valor in dato.datos_originales.items():
                    if search_query.lower() in str(valor).lower():
                        valor_str = str(valor)
                        if len(valor_str) > 100:
                            valor_str = valor_str[:100] + '...'
                        campos_coincidentes.append({
                            'campo': campo,
                            'valor': valor_str
                        })
                        if len(campos_coincidentes) >= 3:  # Máximo 3 campos
                            break
            
            datos_json.append({
                'id': dato.pk,
                'id_original': dato.id_original,
                'nombre': nombre,
                'nombre_resaltado': nombre_resaltado,
                'fecha_migracion': dato.fecha_migracion.strftime('%d/%m/%Y %H:%M'),
                'campos_coincidentes': campos_coincidentes,
                'url_detalle': f'/datos-archivados/datos/{dato.pk}/'
            })
        
        return JsonResponse({
            'success': True,
            'resultados': datos_json,
            'total_encontrados': queryset.count(),
            'total_mostrados': len(datos_json),
            'query': search_query,
            'tabla': tabla,
            'order_by': order_by,
            'order_direction': order_direction
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en búsqueda: {str(e)}',
            'resultados': []
        })

@login_required

def agregar_campos_faltantes_a_modelo(modelo, campos_necesarios, logger):
    """
    Agrega campos faltantes a un modelo de Django dinámicamente usando ALTER TABLE
    """
    from django.db import connection
    from datetime import datetime, date
    
    campos_actuales = {f.name for f in modelo._meta.get_fields()}
    campos_agregados = []
    
    with connection.cursor() as cursor:
        tabla_nombre = modelo._meta.db_table
        
        for campo_nombre, valor_ejemplo in campos_necesarios.items():
            if campo_nombre not in campos_actuales and campo_nombre not in ['id', 'pk']:
                try:
                    # Determinar tipo de campo basado en el valor
                    if isinstance(valor_ejemplo, bool):
                        tipo_sql = 'BOOLEAN DEFAULT FALSE'
                    elif isinstance(valor_ejemplo, int):
                        tipo_sql = 'INTEGER NULL'
                    elif isinstance(valor_ejemplo, float):
                        tipo_sql = 'DOUBLE PRECISION NULL'
                    elif isinstance(valor_ejemplo, (datetime, date)):
                        tipo_sql = 'TIMESTAMP NULL'
                    elif isinstance(valor_ejemplo, str) and len(valor_ejemplo) > 255:
                        tipo_sql = 'TEXT NULL'
                    else:
                        # Por defecto, VARCHAR
                        tipo_sql = 'VARCHAR(255) NULL'
                    
                    # Ejecutar ALTER TABLE
                    sql = f'ALTER TABLE {tabla_nombre} ADD COLUMN IF NOT EXISTS "{campo_nombre}" {tipo_sql}'
                    cursor.execute(sql)
                    
                    campos_agregados.append(campo_nombre)
                    logger.info(f"✅ Campo agregado: {tabla_nombre}.{campo_nombre} ({tipo_sql})")
                    
                except Exception as e:
                    logger.warning(f"⚠️ No se pudo agregar campo {campo_nombre} a {tabla_nombre}: {str(e)}")
    
    return campos_agregados

def mapear_campos_ingles_espanol(datos_origen, logger=None):
    """
    Mapea campos de inglés a español y mantiene ambos
    Retorna un diccionario con campos mapeados + campos originales
    """
    # MAPEO DE CAMPOS: Traducir nombres de inglés a español
    mapeo_campos = {
        'nacionality': 'nacionalidad',
        'numberidentification': 'carnet',
        'phone': 'telephone',
        'cellphone': 'movil',
        'street': 'address',
        'city': 'location',
        'state': 'provincia',
        'degree': 'grado',
        'ocupation': 'ocupacion',
        'title': 'titulo',
        'gender': 'sexo',
        'name': 'first_name',
        'lastname': 'last_name',
        'email': 'email',  # Mantener email
        'photo': 'image',  # Mapear photo a image
        'resume': 'titulo',  # Resume puede ser título
        'isReligious': 'es_religioso', # Mapear isReligious a es_religioso
    }
    
    # Crear copia con todos los datos originales
    datos_mapeados = datos_origen.copy()
    campos_mapeados_count = 0
    
    # Aplicar mapeo: agregar campos traducidos
    for campo_origen, campo_destino in mapeo_campos.items():
        if campo_origen in datos_origen:
            datos_mapeados[campo_destino] = datos_origen[campo_origen]
            campos_mapeados_count += 1
            if logger:
                logger.info(f"🔄 Mapeando: {campo_origen} → {campo_destino} = {datos_origen[campo_origen]}")
    
    if logger and campos_mapeados_count > 0:
        logger.info(f"📋 Total de campos mapeados: {campos_mapeados_count}")
        logger.info(f"📋 Total de campos disponibles: {len(datos_mapeados)} (originales + mapeados)")
    
    return datos_mapeados

def copiar_todos_los_campos(objeto_destino, datos_origen, campos_excluir=None, logger=None):
    """
    Copia TODOS los campos de datos_origen al objeto_destino
    NOTA: Los campos deben existir previamente (creados en la fase de detección)
    """
    if campos_excluir is None:
        campos_excluir = ['id', 'pk']
    
    campos_copiados = 0
    campos_no_encontrados = []
    
    # Obtener los nombres de campos del modelo
    campos_modelo = {f.name for f in objeto_destino._meta.get_fields() if not f.many_to_many and not f.one_to_many}
    
    for campo_nombre, valor in datos_origen.items():
        if campo_nombre in campos_excluir:
            continue
        
        # Verificar si el campo existe en el modelo
        if campo_nombre not in campos_modelo:
            campos_no_encontrados.append(campo_nombre)
            if logger:
                logger.debug(f"⚠️ Campo '{campo_nombre}' no existe en el modelo {objeto_destino.__class__.__name__}")
            continue
        
        try:
            # Establecer el valor
            setattr(objeto_destino, campo_nombre, valor)
            campos_copiados += 1
            if logger:
                logger.debug(f"✅ Campo copiado: {campo_nombre} = {valor}")
        except Exception as e:
            if logger:
                logger.warning(f"❌ No se pudo copiar campo {campo_nombre}: {str(e)}")
    
    if logger and campos_copiados > 0:
        logger.info(f"📊 Total campos copiados: {campos_copiados}")
        if campos_no_encontrados:
            logger.debug(f"⚠️ Campos no encontrados: {', '.join(campos_no_encontrados[:5])}")
    
    return campos_copiados

@login_required
def combinar_datos_archivados(request):
    """Vista para combinar TODAS las tablas archivadas con las tablas activas de la base de datos
    
    Esta función:
    1. Detecta campos faltantes en las tablas actuales
    2. Crea esos campos automáticamente usando ALTER TABLE
    3. Copia TODA la información de las tablas archivadas
    """
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    # Ejecutar combinación en hilo separado para no bloquear la respuesta
    def ejecutar_combinacion():
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            logger.info("=== INICIANDO COMBINACIÓN REAL DE DATOS ===")
            
            from .models import DatoArchivadoDinamico
            from django.contrib.auth.models import User, Group
            from accounts.models import Registro
            from principal.models import (
                CursoAcademico, Curso, Matriculas, 
                Asistencia, Calificaciones, NotaIndividual
            )
            from django.db import transaction, IntegrityError
            from django.db.models import Q
            from datetime import datetime, date
            from decimal import Decimal
            from django.core.cache import cache
            from django.utils import timezone
            
            # Limpiar cualquier flag de interrupción residual al inicio
            cache.delete('combinacion_interrumpida')
            cache.delete('combinacion_estado_interrupcion')
            cache.delete('combinacion_interrumpida_info')
            logger.info("🧹 Cache de interrupción limpiado al inicio de combinación")
            
            # Inicializar progreso en cache
            def actualizar_progreso(paso_actual, pasos_completados, pasos_totales=11, **kwargs):
                # Verificar si la combinación ha sido interrumpida
                if cache.get('combinacion_interrumpida'):
                    logger.info("🛑 Combinación interrumpida por el usuario - deteniendo proceso")
                    cache.set('combinacion_estado', 'interrumpida', timeout=300)
                    raise InterruptedError("Combinación interrumpida por el usuario")
                
                # Calcular porcentaje de progreso real (sin limitación artificial)
                porcentaje_progreso = min(int((pasos_completados / pasos_totales) * 100), 100) if pasos_totales > 0 else 0
                
                progreso = {
                    'paso_actual': paso_actual,
                    'pasos_completados': pasos_completados,
                    'pasos_totales': pasos_totales,
                    'porcentaje_progreso': porcentaje_progreso,
                    'fecha_inicio': timezone.now().isoformat(),
                    'usuarios_combinados': kwargs.get('usuarios_combinados', 0),
                    'registros_combinados': kwargs.get('registros_combinados', 0),
                    'cursos_academicos_combinados': kwargs.get('cursos_academicos_combinados', 0),
                    'cursos_combinados': kwargs.get('cursos_combinados', 0),
                    'matriculas_combinadas': kwargs.get('matriculas_combinadas', 0),
                    'asistencias_combinadas': kwargs.get('asistencias_combinadas', 0),
                    'calificaciones_combinadas': kwargs.get('calificaciones_combinadas', 0),
                    'notas_combinadas': kwargs.get('notas_combinadas', 0),
                    'otras_tablas': kwargs.get('otras_tablas', 0),
                    'campos_agregados': kwargs.get('campos_agregados', 0),
                    'errores_encontrados': kwargs.get('errores_encontrados', 0),
                    'tiempo_transcurrido': kwargs.get('tiempo_transcurrido', ''),
                    'tipo_combinacion': 'completa'
                }
                cache.set('combinacion_en_progreso', progreso, timeout=1800)  # Aumentado a 30 minutos
                logger.info(f"Progreso actualizado: {paso_actual} ({pasos_completados}/{pasos_totales}) - {porcentaje_progreso}%")
                if kwargs.get('campos_agregados', 0) > 0:
                    logger.info(f"📊 Campos agregados hasta ahora: {kwargs.get('campos_agregados', 0)}")
            
            # Función para agregar campos faltantes dinámicamente
            def agregar_campos_faltantes(modelo, datos_ejemplo, logger=None):
                """Agrega campos faltantes al modelo usando ALTER TABLE"""
                from django.db import connection
                
                campos_actuales = {f.name for f in modelo._meta.get_fields()}
                campos_agregados = []
                
                with connection.cursor() as cursor:
                    tabla_nombre = modelo._meta.db_table
                    
                    for campo_nombre, valor_ejemplo in datos_ejemplo.items():
                        if campo_nombre not in campos_actuales and campo_nombre not in ['id', 'pk']:
                            try:
                                # Determinar tipo de campo basado en el valor
                                if isinstance(valor_ejemplo, bool):
                                    tipo_sql = 'BOOLEAN DEFAULT FALSE'
                                elif isinstance(valor_ejemplo, int):
                                    if valor_ejemplo > 2147483647:  # Valor muy grande
                                        tipo_sql = 'BIGINT NULL'
                                    else:
                                        tipo_sql = 'INTEGER NULL'
                                elif isinstance(valor_ejemplo, float):
                                    tipo_sql = 'DOUBLE PRECISION NULL'
                                elif isinstance(valor_ejemplo, (datetime, date)):
                                    tipo_sql = 'TIMESTAMP NULL'
                                elif isinstance(valor_ejemplo, str):
                                    if len(valor_ejemplo) > 255:
                                        tipo_sql = 'TEXT NULL'
                                    else:
                                        tipo_sql = 'VARCHAR(255) NULL'
                                else:
                                    # Por defecto, TEXT para valores complejos
                                    tipo_sql = 'TEXT NULL'
                                
                                # Ejecutar ALTER TABLE
                                sql = f'ALTER TABLE "{tabla_nombre}" ADD COLUMN IF NOT EXISTS "{campo_nombre}" {tipo_sql}'
                                cursor.execute(sql)
                                
                                campos_agregados.append(campo_nombre)
                                if logger:
                                    logger.info(f"✅ Campo agregado: {tabla_nombre}.{campo_nombre} ({tipo_sql})")
                                
                            except Exception as e:
                                if logger:
                                    logger.warning(f"⚠️ No se pudo agregar campo {campo_nombre} a {tabla_nombre}: {str(e)}")
                
                return campos_agregados
            
            # Función para mapear campos de inglés a español
            def mapear_campos_ingles_espanol(datos_origen, logger=None):
                """Mapea campos de inglés a español y mantiene ambos"""
                mapeo_campos = {
                    'nacionality': 'nacionalidad',
                    'numberidentification': 'carnet', 
                    'phone': 'telephone',
                    'cellphone': 'movil',
                    'street': 'address',
                    'city': 'location',
                    'state': 'provincia',
                    'degree': 'grado',
                    'ocupation': 'ocupacion',
                    'title': 'titulo',
                    'gender': 'sexo',
                    'photo': 'image',
                    'isReligious': 'es_religioso',
                    'name': 'first_name',  # Para usuarios
                    'lastname': 'last_name'  # Para usuarios
                }
                
                # Crear copia con todos los datos originales
                datos_mapeados = datos_origen.copy()
                campos_mapeados_count = 0
                
                # Aplicar mapeo: agregar campos traducidos
                for campo_origen, campo_destino in mapeo_campos.items():
                    if campo_origen in datos_origen:
                        datos_mapeados[campo_destino] = datos_origen[campo_origen]
                        campos_mapeados_count += 1
                        if logger:
                            logger.debug(f"🔄 Mapeando: {campo_origen} → {campo_destino} = {datos_origen[campo_origen]}")
                
                if logger and campos_mapeados_count > 0:
                    logger.info(f"📋 Total de campos mapeados: {campos_mapeados_count}")
                
                return datos_mapeados
            
            # Función para copiar campos dinámicamente (mejorada)
            def copiar_campos_dinamicos(objeto_destino, datos_origen, campos_excluir=None, logger=None):
                if campos_excluir is None:
                    campos_excluir = ['id', 'pk']
                
                campos_copiados = 0
                campos_no_encontrados = []
                
                for campo, valor in datos_origen.items():
                    if campo in campos_excluir:
                        continue
                    
                    try:
                        if hasattr(objeto_destino, campo):
                            # Convertir valores especiales
                            if valor == 'NULL' or valor == 'null':
                                valor = None
                            elif isinstance(valor, str) and valor.lower() in ['true', 'false']:
                                valor = valor.lower() == 'true'
                            elif campo in ['last_login', 'date_joined'] and valor:
                                # Manejar campos de fecha con timezone
                                from django.utils import timezone
                                from datetime import datetime
                                try:
                                    if isinstance(valor, str):
                                        # Parsear la fecha string
                                        fecha_naive = datetime.fromisoformat(valor.replace('Z', ''))
                                    elif isinstance(valor, datetime):
                                        fecha_naive = valor
                                    else:
                                        fecha_naive = None
                                    
                                    if fecha_naive and timezone.is_naive(fecha_naive):
                                        valor = timezone.make_aware(fecha_naive)
                                    elif fecha_naive:
                                        valor = fecha_naive
                                except Exception as e:
                                    if logger:
                                        logger.warning(f"Error convirtiendo fecha {campo}: {e}")
                                    valor = None
                            
                            setattr(objeto_destino, campo, valor)
                            campos_copiados += 1
                            if logger:
                                logger.debug(f"Campo copiado: {campo} = {valor}")
                        else:
                            campos_no_encontrados.append(campo)
                            if logger:
                                logger.debug(f"Campo no encontrado en modelo: {campo}")
                    except Exception as e:
                        if logger:
                            logger.warning(f"Error copiando campo {campo}: {e}")
                
                if logger and campos_no_encontrados:
                    logger.info(f"Campos no encontrados en modelo: {', '.join(campos_no_encontrados[:5])}")
                
                return campos_copiados
            
            # Contadores
            estadisticas = {
                'usuarios_combinados': 0,
                'registros_combinados': 0,
                'cursos_academicos_combinados': 0,
                'cursos_combinados': 0,
                'matriculas_combinadas': 0,
                'asistencias_combinadas': 0,
                'calificaciones_combinadas': 0,
                'notas_combinadas': 0,
                'otras_tablas': 0
            }
            
            # Mapeo de usuarios archivados a usuarios actuales
            mapeo_usuarios = {}
            
            # Inicializar progreso
            actualizar_progreso('Iniciando combinación real...', 0, **estadisticas)
            
            # PASO 1: COMBINAR USUARIOS
            actualizar_progreso('Combinando usuarios...', 1, **estadisticas)
            logger.info("=== Iniciando combinación de auth_user ===")
            
            datos_auth_user = DatoArchivadoDinamico.objects.filter(tabla_origen='auth_user')
            logger.info(f"Encontrados {datos_auth_user.count()} usuarios archivados")
            
            # PASO 1.1: DETECTAR Y AGREGAR CAMPOS FALTANTES AL MODELO USER
            if datos_auth_user.exists():
                logger.info("=== DETECTANDO CAMPOS FALTANTES EN auth_user ===")
                
                # Obtener una muestra de datos para detectar campos
                muestra_datos = datos_auth_user.first().datos_originales
                campos_agregados_user = agregar_campos_faltantes(User, muestra_datos, logger)
                
                if campos_agregados_user:
                    logger.info(f"✅ Se agregaron {len(campos_agregados_user)} campos nuevos al modelo User")
                    estadisticas['campos_agregados'] = estadisticas.get('campos_agregados', 0) + len(campos_agregados_user)
                else:
                    logger.info("ℹ️ No se necesitaron campos adicionales en el modelo User")
            
            with transaction.atomic():
                for dato in datos_auth_user:
                    try:
                        sid = transaction.savepoint()
                        datos = dato.datos_originales
                        username = datos.get('username', '')
                        email = datos.get('email', '')
                        id_original = dato.id_original
                        
                        if not username:
                            logger.warning(f"Usuario sin username, saltando: {datos}")
                            transaction.savepoint_rollback(sid)
                            continue
                        
                        # Buscar si el usuario ya existe
                        usuario_existente = User.objects.filter(username=username).first()
                        
                        if usuario_existente:
                            logger.info(f"Actualizando usuario existente: {username}")
                            # Actualizar usuario existente con TODOS los campos
                            campos_copiados = copiar_campos_dinamicos(usuario_existente, datos, 
                                                  campos_excluir=['id', 'pk', 'username'], 
                                                  logger=logger)
                            
                            logger.info(f"Se copiaron {campos_copiados} campos para usuario existente {username}")
                            
                            # Procesar contraseña
                            password_original = datos.get('password')
                            if password_original and password_original.strip():
                                if password_original.startswith(('pbkdf2_sha256', 'bcrypt', 'argon2', 'sha1', 'md5')):
                                    usuario_existente.password = password_original
                                    logger.info(f"Contraseña hasheada copiada para {username}")
                                else:
                                    usuario_existente.set_password(password_original)
                                    logger.info(f"Contraseña en texto plano hasheada para {username}")
                            
                            usuario_existente.save()
                            # USAR EL ID REAL DE LOS DATOS ORIGINALES (no el id_original del registro Django)
                            user_id_real = datos.get('id')  # Este es el ID real de la tabla auth_user
                            mapeo_usuarios[user_id_real] = usuario_existente
                            logger.info(f"✅ Usuario {username} mapeado con ID real: {user_id_real}")
                            estadisticas['usuarios_combinados'] += 1
                            
                        else:
                            logger.info(f"Creando nuevo usuario: {username}")
                            # Crear nuevo usuario con TODOS los campos
                            nuevo_usuario = User(username=username)
                            campos_copiados = copiar_campos_dinamicos(nuevo_usuario, datos, 
                                                  campos_excluir=['id', 'pk'], 
                                                  logger=logger)
                            
                            logger.info(f"Se copiaron {campos_copiados} campos para nuevo usuario {username}")
                            
                            # Procesar contraseña
                            password_original = datos.get('password')
                            if password_original and password_original.strip():
                                if password_original.startswith(('pbkdf2_sha256', 'bcrypt', 'argon2', 'sha1', 'md5')):
                                    nuevo_usuario.password = password_original
                                    logger.info(f"Contraseña hasheada asignada para {username}")
                                else:
                                    nuevo_usuario.set_password(password_original)
                                    logger.info(f"Contraseña en texto plano hasheada para {username}")
                            else:
                                nuevo_usuario.set_unusable_password()
                                logger.info(f"Contraseña no utilizable asignada para {username}")
                            
                            nuevo_usuario.save()
                            # USAR EL ID REAL DE LOS DATOS ORIGINALES (no el id_original del registro Django)
                            user_id_real = datos.get('id')  # Este es el ID real de la tabla auth_user
                            mapeo_usuarios[user_id_real] = nuevo_usuario
                            logger.info(f"✅ Usuario {username} mapeado con ID real: {user_id_real}")
                            estadisticas['usuarios_combinados'] += 1
                        
                        transaction.savepoint_commit(sid)
                        
                    except InterruptedError:
                        # Interrupción por el usuario - detener completamente el procesamiento
                        transaction.savepoint_rollback(sid)
                        logger.info("🛑 Procesamiento de usuarios interrumpido por el usuario")
                        raise  # Re-lanzar la excepción para detener toda la combinación
                    except Exception as e:
                        transaction.savepoint_rollback(sid)
                        error_msg = f"Error procesando usuario {username}: {str(e)}"
                        logger.error(error_msg)
                        continue
            
            # PASO 2: COMBINAR REGISTROS DE ESTUDIANTES/PROFESORES
            actualizar_progreso('Combinando registros de estudiantes y profesores...', 2, **estadisticas)
            logger.info("=== Combinando registros de estudiantes/profesores ===")
            
            # Obtener o crear grupos
            grupo_estudiantes, _ = Group.objects.get_or_create(name='Estudiantes')
            grupo_profesores, _ = Group.objects.get_or_create(name='Profesores')
            
            # PASO 2A: PROCESAR PROFESORES PRIMERO
            datos_profesores = DatoArchivadoDinamico.objects.filter(
                tabla_origen='Docencia_teacherpersonalinformation'
            )
            logger.info(f"Encontrados {datos_profesores.count()} profesores archivados")
            
            # PASO 2A.1: DETECTAR Y AGREGAR CAMPOS FALTANTES AL MODELO REGISTRO (PROFESORES)
            if datos_profesores.exists():
                logger.info("=== DETECTANDO CAMPOS FALTANTES EN accounts_registro (PROFESORES) ===")
                
                # Obtener una muestra de datos para detectar campos
                muestra_datos = datos_profesores.first().datos_originales
                
                # Mapear campos de inglés a español antes de agregar
                muestra_mapeada = mapear_campos_ingles_espanol(muestra_datos, logger)
                
                campos_agregados = agregar_campos_faltantes(Registro, muestra_mapeada, logger)
                
                if campos_agregados:
                    logger.info(f"✅ Se agregaron {len(campos_agregados)} campos nuevos al modelo Registro (profesores)")
                    estadisticas['campos_agregados'] = len(campos_agregados)
                else:
                    logger.info("ℹ️ No se necesitaron campos adicionales en el modelo Registro (profesores)")
            
            with transaction.atomic():
                for dato in datos_profesores:
                    try:
                        sid = transaction.savepoint()
                        datos = dato.datos_originales
                        user_id_original = datos.get('user_id')
                        
                        if not user_id_original:
                            logger.warning(f"Profesor sin user_id, saltando: {datos}")
                            transaction.savepoint_rollback(sid)
                            continue
                        
                        if user_id_original not in mapeo_usuarios:
                            logger.warning(f"Usuario no encontrado para profesor: user_id={user_id_original}")
                            logger.warning(f"IDs disponibles en mapeo: {list(mapeo_usuarios.keys())}")
                            transaction.savepoint_rollback(sid)
                            continue
                        
                        usuario = mapeo_usuarios[user_id_original]
                        logger.info(f"✅ Procesando PROFESOR: {usuario.username} (user_id={user_id_original})")
                        
                        # Buscar o crear registro
                        registro, created = Registro.objects.get_or_create(user=usuario)
                        
                        if created:
                            logger.info(f"Creado nuevo registro para profesor: {usuario.username}")
                        else:
                            logger.info(f"Actualizando registro existente para profesor: {usuario.username}")
                        
                        # Copiar TODOS los campos (ahora que existen en el modelo)
                        # Primero mapear los campos de inglés a español
                        datos_mapeados = mapear_campos_ingles_espanol(datos, logger)
                        
                        campos_copiados = copiar_campos_dinamicos(registro, datos_mapeados, 
                                              campos_excluir=['id', 'pk', 'user_id', 'user'], 
                                              logger=logger)
                        
                        logger.info(f"Se copiaron {campos_copiados} campos para profesor {usuario.username}")
                        
                        registro.save()
                        estadisticas['registros_combinados'] += 1
                        
                        # Asignar grupo de PROFESORES (limpiar otros grupos primero)
                        usuario.groups.clear()  # Limpiar grupos existentes
                        usuario.groups.add(grupo_profesores)
                        logger.info(f"✅ Usuario {usuario.username} asignado EXCLUSIVAMENTE al grupo PROFESORES")
                        
                        transaction.savepoint_commit(sid)
                        
                    except InterruptedError:
                        # Interrupción por el usuario - detener completamente el procesamiento
                        transaction.savepoint_rollback(sid)
                        logger.info("🛑 Procesamiento de profesores interrumpido por el usuario")
                        raise  # Re-lanzar la excepción para detener toda la combinación
                    except Exception as e:
                        transaction.savepoint_rollback(sid)
                        error_msg = f"Error procesando profesor: {str(e)}"
                        logger.error(error_msg)
                        continue
            
            # PASO 2B: PROCESAR ESTUDIANTES DESPUÉS
            datos_estudiantes = DatoArchivadoDinamico.objects.filter(
                tabla_origen='Docencia_studentpersonalinformation'
            )
            logger.info(f"Encontrados {datos_estudiantes.count()} estudiantes archivados")
            
            # PASO 2B.1: DETECTAR Y AGREGAR CAMPOS FALTANTES AL MODELO REGISTRO (ESTUDIANTES)
            if datos_estudiantes.exists():
                logger.info("=== DETECTANDO CAMPOS FALTANTES EN accounts_registro (ESTUDIANTES) ===")
                
                # Obtener una muestra de datos para detectar campos
                muestra_datos = datos_estudiantes.first().datos_originales
                
                # Mapear campos de inglés a español antes de agregar
                muestra_mapeada = mapear_campos_ingles_espanol(muestra_datos, logger)
                
                campos_agregados_estudiantes = agregar_campos_faltantes(Registro, muestra_mapeada, logger)
                
                if campos_agregados_estudiantes:
                    logger.info(f"✅ Se agregaron {len(campos_agregados_estudiantes)} campos adicionales al modelo Registro (estudiantes)")
                    estadisticas['campos_agregados'] = estadisticas.get('campos_agregados', 0) + len(campos_agregados_estudiantes)
                else:
                    logger.info("ℹ️ No se necesitaron campos adicionales en el modelo Registro (estudiantes)")
            
            with transaction.atomic():
                for dato in datos_estudiantes:
                    try:
                        sid = transaction.savepoint()
                        datos = dato.datos_originales
                        user_id_original = datos.get('user_id')
                        
                        if not user_id_original:
                            logger.warning(f"Estudiante sin user_id, saltando: {datos}")
                            transaction.savepoint_rollback(sid)
                            continue
                        
                        if user_id_original not in mapeo_usuarios:
                            logger.warning(f"Usuario no encontrado para estudiante: user_id={user_id_original}")
                            logger.warning(f"IDs disponibles en mapeo: {list(mapeo_usuarios.keys())}")
                            transaction.savepoint_rollback(sid)
                            continue
                        
                        usuario = mapeo_usuarios[user_id_original]
                        logger.info(f"✅ Procesando ESTUDIANTE: {usuario.username} (user_id={user_id_original})")
                        
                        # Buscar o crear registro
                        registro, created = Registro.objects.get_or_create(user=usuario)
                        
                        if created:
                            logger.info(f"Creado nuevo registro para estudiante: {usuario.username}")
                        else:
                            logger.info(f"Actualizando registro existente para estudiante: {usuario.username}")
                        
                        # Copiar TODOS los campos (ahora que existen en el modelo)
                        # Primero mapear los campos de inglés a español
                        datos_mapeados = mapear_campos_ingles_espanol(datos, logger)
                        
                        campos_copiados = copiar_campos_dinamicos(registro, datos_mapeados, 
                                              campos_excluir=['id', 'pk', 'user_id', 'user'], 
                                              logger=logger)
                        
                        logger.info(f"Se copiaron {campos_copiados} campos para estudiante {usuario.username}")
                        
                        registro.save()
                        estadisticas['registros_combinados'] += 1
                        
                        # Asignar grupo de ESTUDIANTES (limpiar otros grupos primero)
                        usuario.groups.clear()  # Limpiar grupos existentes
                        usuario.groups.add(grupo_estudiantes)
                        logger.info(f"✅ Usuario {usuario.username} asignado EXCLUSIVAMENTE al grupo ESTUDIANTES")
                        
                        transaction.savepoint_commit(sid)
                        
                    except InterruptedError:
                        # Interrupción por el usuario - detener completamente el procesamiento
                        transaction.savepoint_rollback(sid)
                        logger.info("🛑 Procesamiento de estudiantes interrumpido por el usuario")
                        raise  # Re-lanzar la excepción para detener toda la combinación
                    except Exception as e:
                        transaction.savepoint_rollback(sid)
                        error_msg = f"Error procesando estudiante: {str(e)}"
                        logger.error(error_msg)
                        continue
            
            # PASO 2C: PROCESAR REGISTROS EXISTENTES (accounts_registro)
            datos_registros_existentes = DatoArchivadoDinamico.objects.filter(
                tabla_origen='accounts_registro'
            )
            
            if datos_registros_existentes.exists():
                logger.info(f"Encontrados {datos_registros_existentes.count()} registros existentes archivados")
                
                with transaction.atomic():
                    for dato in datos_registros_existentes:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            user_id_original = datos.get('user_id')
                            
                            if not user_id_original:
                                logger.warning(f"Registro existente sin user_id, saltando: {datos}")
                                transaction.savepoint_rollback(sid)
                                continue
                            
                            if user_id_original not in mapeo_usuarios:
                                logger.warning(f"Usuario no encontrado para registro existente: user_id={user_id_original}")
                                transaction.savepoint_rollback(sid)
                                continue
                            
                            usuario = mapeo_usuarios[user_id_original]
                            logger.info(f"✅ Procesando registro existente: {usuario.username} (user_id={user_id_original})")
                            
                            # Buscar o crear registro
                            registro, created = Registro.objects.get_or_create(user=usuario)
                            
                            if created:
                                logger.info(f"Creado nuevo registro desde datos existentes: {usuario.username}")
                            else:
                                logger.info(f"Actualizando registro desde datos existentes: {usuario.username}")
                            
                            # Copiar TODOS los campos
                            campos_copiados = copiar_campos_dinamicos(registro, datos, 
                                                      campos_excluir=['id', 'pk', 'user_id', 'user'], 
                                                      logger=logger)
                            
                            logger.info(f"Se copiaron {campos_copiados} campos desde registro existente para {usuario.username}")
                            
                            registro.save()
                            estadisticas['registros_combinados'] += 1
                            
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            error_msg = f"Error procesando registro existente: {str(e)}"
                            logger.error(error_msg)
                            continue
            
            # PASO 3: COMBINAR GRUPOS DE USUARIOS
            actualizar_progreso('Combinando grupos de usuarios...', 3, **estadisticas)
            logger.info("=== Combinando grupos de usuarios ===")
            
            # Primero, procesar la tabla auth_group para crear los grupos con sus nombres reales
            datos_auth_group = DatoArchivadoDinamico.objects.filter(tabla_origen='auth_group')
            logger.info(f"Encontrados {datos_auth_group.count()} grupos archivados")
            
            mapeo_grupos = {}  # Para mapear IDs originales a grupos nuevos
            
            with transaction.atomic():
                for dato in datos_auth_group:
                    try:
                        sid = transaction.savepoint()
                        datos = dato.datos_originales
                        group_id_original = datos.get('id')
                        group_name = datos.get('name')
                        
                        if not group_name:
                            logger.warning(f"Grupo sin nombre, usando nombre genérico: Grupo_{group_id_original}")
                            group_name = f'Grupo_{group_id_original}'
                        
                        from django.contrib.auth.models import Group
                        
                        # Verificar si ya existe un grupo con este nombre
                        grupo_existente = Group.objects.filter(name=group_name).first()
                        
                        if grupo_existente:
                            # Si el grupo ya existe, usarlo (no crear duplicado)
                            mapeo_grupos[group_id_original] = grupo_existente
                            logger.info(f"Grupo existente reutilizado: {group_name} (ID original: {group_id_original})")
                        else:
                            # Si no existe, crearlo
                            grupo = Group.objects.create(name=group_name)
                            mapeo_grupos[group_id_original] = grupo
                            logger.info(f"Grupo creado: {group_name} (ID original: {group_id_original})")
                        
                        estadisticas['grupos_creados'] = estadisticas.get('grupos_creados', 0) + 1
                        transaction.savepoint_commit(sid)
                        
                    except Exception as e:
                        transaction.savepoint_rollback(sid)
                        logger.error(f"Error procesando grupo: {e}")
                        continue
            
            # Luego, procesar auth_user_groups para asignar usuarios a grupos
            datos_grupos = DatoArchivadoDinamico.objects.filter(tabla_origen='auth_user_groups')
            logger.info(f"Encontrados {datos_grupos.count()} relaciones usuario-grupo archivadas")
            
            with transaction.atomic():
                for dato in datos_grupos:
                    try:
                        sid = transaction.savepoint()
                        datos = dato.datos_originales
                        user_id_original = datos.get('user_id')
                        group_id_original = datos.get('group_id')
                        
                        if user_id_original in mapeo_usuarios and group_id_original in mapeo_grupos:
                            usuario = mapeo_usuarios[user_id_original]
                            grupo = mapeo_grupos[group_id_original]
                            
                            usuario.groups.add(grupo)
                            logger.info(f"Usuario {usuario.username} agregado al grupo {grupo.name}")
                            estadisticas['relaciones_usuario_grupo'] = estadisticas.get('relaciones_usuario_grupo', 0) + 1
                        else:
                            if user_id_original not in mapeo_usuarios:
                                logger.warning(f"Usuario con ID {user_id_original} no encontrado en mapeo")
                            if group_id_original not in mapeo_grupos:
                                logger.warning(f"Grupo con ID {group_id_original} no encontrado en mapeo")
                        
                        transaction.savepoint_commit(sid)
                        
                    except Exception as e:
                        transaction.savepoint_rollback(sid)
                        logger.error(f"Error procesando relación usuario-grupo: {e}")
                        continue
            
            # PASO 3.5: COMBINAR PERMISOS DIRECTOS DE USUARIOS
            actualizar_progreso('Combinando permisos directos de usuarios...', 3.5, **estadisticas)
            logger.info("=== Combinando permisos directos de usuarios ===")
            
            datos_user_permissions = DatoArchivadoDinamico.objects.filter(tabla_origen='auth_user_user_permissions')
            logger.info(f"Encontrados {datos_user_permissions.count()} permisos directos de usuario archivados")
            
            if datos_user_permissions.exists():
                with transaction.atomic():
                    for dato in datos_user_permissions:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            user_id_original = datos.get('user_id')
                            permission_id_original = datos.get('permission_id')
                            
                            if user_id_original in mapeo_usuarios:
                                usuario = mapeo_usuarios[user_id_original]
                                
                                # Buscar el permiso por ID (los permisos se crean automáticamente por Django)
                                from django.contrib.auth.models import Permission
                                try:
                                    # Intentar encontrar el permiso por ID original
                                    # Nota: Los IDs de permisos pueden cambiar entre sistemas
                                    # Por eso es mejor buscar por codename si está disponible
                                    permission = Permission.objects.filter(id=permission_id_original).first()
                                    
                                    if permission:
                                        usuario.user_permissions.add(permission)
                                        logger.info(f"Permiso {permission.codename} asignado directamente a usuario {usuario.username}")
                                        estadisticas['permisos_directos_asignados'] = estadisticas.get('permisos_directos_asignados', 0) + 1
                                    else:
                                        logger.warning(f"Permiso con ID {permission_id_original} no encontrado")
                                        
                                except Exception as e:
                                    logger.warning(f"Error asignando permiso directo: {e}")
                            else:
                                logger.warning(f"Usuario con ID {user_id_original} no encontrado en mapeo")
                            
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            logger.error(f"Error procesando permiso directo de usuario: {e}")
                            continue
            else:
                logger.info("No hay permisos directos de usuario para migrar")
            
            # PASO 4: COMBINAR CURSOS ACADÉMICOS
            actualizar_progreso('Combinando cursos académicos...', 4, **estadisticas)
            logger.info("=== Combinando cursos académicos ===")
            
            # Buscar tablas de cursos académicos
            tablas_cursos_academicos = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen__icontains='academicyear') |
                Q(tabla_origen__icontains='curso_academico') |
                Q(tabla_origen='Docencia_academicyear')
            ).values_list('tabla_origen', flat=True).distinct()
            
            for tabla_curso_academico in tablas_cursos_academicos:
                datos_cursos_academicos = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_curso_academico)
                logger.info(f"Procesando {datos_cursos_academicos.count()} cursos académicos de tabla {tabla_curso_academico}")
                
                with transaction.atomic():
                    for dato in datos_cursos_academicos:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            
                            # Crear o actualizar curso académico
                            curso_academico_data = {
                                'year': datos.get('year', datos.get('anio', datetime.now().year)),
                                'name': datos.get('name', datos.get('nombre', f'Curso Académico {datos.get("year", datetime.now().year)}')),
                                'activo': datos.get('activo', datos.get('active', False)),
                                'fecha_inicio': datos.get('fecha_inicio', datos.get('start_date')),
                                'fecha_fin': datos.get('fecha_fin', datos.get('end_date'))
                            }
                            
                            # Buscar curso académico existente por año
                            curso_academico_existente = CursoAcademico.objects.filter(year=curso_academico_data['year']).first()
                            
                            if curso_academico_existente:
                                # Actualizar curso académico existente
                                copiar_campos_dinamicos(curso_academico_existente, datos, 
                                                      campos_excluir=['id', 'pk'], 
                                                      logger=logger)
                                curso_academico_existente.save()
                                logger.info(f"Curso académico actualizado: {curso_academico_data['name']}")
                            else:
                                # Crear nuevo curso académico
                                nuevo_curso_academico = CursoAcademico(**curso_academico_data)
                                copiar_campos_dinamicos(nuevo_curso_academico, datos, 
                                                      campos_excluir=['id', 'pk', 'year', 'name', 'activo'], 
                                                      logger=logger)
                                nuevo_curso_academico.save()
                                logger.info(f"Curso académico creado: {curso_academico_data['name']}")
                            
                            estadisticas['cursos_academicos_combinados'] += 1
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            logger.error(f"Error procesando curso académico: {e}")
                            continue
            
            # PASO 5: COMBINAR CURSOS
            actualizar_progreso('Combinando cursos...', 5, **estadisticas)
            logger.info("=== Combinando cursos ===")
            
            # Buscar tablas de cursos (pueden tener diferentes nombres)
            tablas_cursos = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen__icontains='course') |
                Q(tabla_origen='Docencia_course') |
                Q(tabla_origen='principal_curso')
            ).values_list('tabla_origen', flat=True).distinct()
            
            # Mapeo de cursos archivados a cursos actuales (usando ID real de los datos)
            mapeo_cursos = {}
            
            for tabla_curso in tablas_cursos:
                datos_cursos = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_curso)
                logger.info(f"Procesando {datos_cursos.count()} cursos de tabla {tabla_curso}")
                
                with transaction.atomic():
                    for dato in datos_cursos:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            curso_id_real = datos.get('id')  # ID real de la tabla de cursos
                            
                            # Crear o actualizar curso
                            curso_data = {
                                'name': datos.get('name', datos.get('nombre', f'Curso_{curso_id_real}')),
                                'description': datos.get('description', datos.get('descripcion', '')),
                                'area': datos.get('area', 'general'),
                                'status': datos.get('status', datos.get('estado', 'F')),
                                'activo': datos.get('activo', datos.get('active', False))
                            }
                            
                            # Buscar curso existente por nombre
                            curso_existente = Curso.objects.filter(name=curso_data['name']).first()
                            
                            if curso_existente:
                                # Actualizar curso existente
                                copiar_campos_dinamicos(curso_existente, datos, 
                                                      campos_excluir=['id', 'pk'], 
                                                      logger=logger)
                                curso_existente.save()
                                mapeo_cursos[curso_id_real] = curso_existente
                                logger.info(f"✅ Curso actualizado: {curso_data['name']} (ID real: {curso_id_real})")
                            else:
                                # Crear nuevo curso
                                nuevo_curso = Curso(**curso_data)
                                copiar_campos_dinamicos(nuevo_curso, datos, 
                                                      campos_excluir=['id', 'pk', 'name'], 
                                                      logger=logger)
                                nuevo_curso.save()
                                mapeo_cursos[curso_id_real] = nuevo_curso
                                logger.info(f"✅ Curso creado: {curso_data['name']} (ID real: {curso_id_real})")
                            
                            estadisticas['cursos_combinados'] += 1
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            logger.error(f"Error procesando curso: {e}")
                            continue
            
            logger.info(f"📋 Mapeo de cursos creado: {len(mapeo_cursos)} cursos mapeados")
            
            # PASO 6: COMBINAR MATRÍCULAS
            actualizar_progreso('Combinando matrículas...', 6, **estadisticas)
            logger.info("=== Combinando matrículas ===")
            
            # Buscar tablas de matrículas
            tablas_matriculas = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen__icontains='matricula') |
                Q(tabla_origen__icontains='enrollment') |
                Q(tabla_origen='Docencia_enrollment') |
                Q(tabla_origen='principal_matriculas')
            ).values_list('tabla_origen', flat=True).distinct()
            
            for tabla_matricula in tablas_matriculas:
                datos_matriculas = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_matricula)
                logger.info(f"Procesando {datos_matriculas.count()} matrículas de tabla {tabla_matricula}")
                
                with transaction.atomic():
                    for dato in datos_matriculas:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            # USAR IDs REALES DE LOS DATOS, NO IDs DE DJANGO
                            student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                            course_id_real = datos.get('course_id')  # FK a course.id
                            
                            if student_id_real in mapeo_usuarios:
                                estudiante = mapeo_usuarios[student_id_real]
                                logger.info(f"✅ Estudiante encontrado: {estudiante.username} (ID real: {student_id_real})")
                                
                                # Buscar curso usando mapeo correcto
                                curso = None
                                if course_id_real and course_id_real in mapeo_cursos:
                                    curso = mapeo_cursos[course_id_real]
                                    logger.info(f"✅ Curso encontrado: {curso.name} (ID real: {course_id_real})")
                                else:
                                    # Buscar por nombre si no hay mapeo directo
                                    course_name = datos.get('course_name', datos.get('nombre_curso'))
                                    if course_name:
                                        curso = Curso.objects.filter(name=course_name).first()
                                        if curso:
                                            logger.info(f"✅ Curso encontrado por nombre: {curso.name}")
                                    
                                    if not curso:
                                        # Usar el primer curso disponible como fallback
                                        curso = Curso.objects.first()
                                        if curso:
                                            logger.warning(f"⚠️ Usando curso fallback: {curso.name}")
                                
                                # Buscar curso académico
                                curso_academico = CursoAcademico.objects.filter(activo=True).first()
                                if not curso_academico:
                                    curso_academico = CursoAcademico.objects.first()
                                
                                if curso and curso_academico:
                                    matricula, created = Matriculas.objects.get_or_create(
                                        student=estudiante,
                                        course=curso,
                                        curso_academico=curso_academico,
                                        defaults={
                                            'estado': datos.get('estado', datos.get('status', 'P')),
                                            'activo': datos.get('activo', datos.get('active', True))
                                        }
                                    )
                                    
                                    if not created:
                                        # Actualizar matrícula existente
                                        copiar_campos_dinamicos(matricula, datos, 
                                                              campos_excluir=['id', 'pk', 'student', 'course', 'curso_academico', 'student_id', 'course_id'], 
                                                              logger=logger)
                                        matricula.save()
                                    
                                    if created:
                                        logger.info(f"✅ Matrícula creada para {estudiante.username} en {curso.name}")
                                    else:
                                        logger.info(f"✅ Matrícula actualizada para {estudiante.username} en {curso.name}")
                                    
                                    estadisticas['matriculas_combinadas'] += 1
                                else:
                                    logger.warning(f"⚠️ No se pudo crear matrícula: curso={curso}, curso_academico={curso_academico}")
                            else:
                                logger.warning(f"⚠️ Usuario no encontrado para matrícula: student_id={student_id_real}")
                                logger.warning(f"IDs de usuarios disponibles: {list(mapeo_usuarios.keys())}")
                            
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            logger.error(f"Error procesando matrícula: {e}")
                            continue
            
            # PASO 7: COMBINAR ASISTENCIAS
            actualizar_progreso('Combinando asistencias...', 7, **estadisticas)
            logger.info("=== Combinando asistencias ===")
            
            # Buscar tablas de asistencias
            tablas_asistencias = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen__icontains='asistencia') |
                Q(tabla_origen__icontains='attendance') |
                Q(tabla_origen='Docencia_attendance') |
                Q(tabla_origen='principal_asistencia')
            ).values_list('tabla_origen', flat=True).distinct()
            
            for tabla_asistencia in tablas_asistencias:
                datos_asistencias = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_asistencia)
                logger.info(f"Procesando {datos_asistencias.count()} asistencias de tabla {tabla_asistencia}")
                
                with transaction.atomic():
                    for dato in datos_asistencias:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            # USAR IDs REALES DE LOS DATOS
                            student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                            course_id_real = datos.get('course_id')  # FK a course.id
                            
                            if student_id_real in mapeo_usuarios:
                                estudiante = mapeo_usuarios[student_id_real]
                                
                                # Buscar curso usando mapeo correcto
                                curso = None
                                if course_id_real and course_id_real in mapeo_cursos:
                                    curso = mapeo_cursos[course_id_real]
                                else:
                                    curso = Curso.objects.first()  # Fallback
                                
                                if curso:
                                    # Crear o actualizar asistencia
                                    fecha_asistencia = datos.get('fecha', datos.get('date', timezone.now().date()))
                                    
                                    asistencia, created = Asistencia.objects.get_or_create(
                                        student=estudiante,
                                        course=curso,
                                        fecha=fecha_asistencia,
                                        defaults={
                                            'presente': datos.get('presente', datos.get('present', True)),
                                            'justificada': datos.get('justificada', datos.get('justified', False))
                                        }
                                    )
                                    
                                    if not created:
                                        # Actualizar asistencia existente
                                        copiar_campos_dinamicos(asistencia, datos, 
                                                              campos_excluir=['id', 'pk', 'student', 'course', 'fecha', 'student_id', 'course_id'], 
                                                              logger=logger)
                                        asistencia.save()
                                    
                                    estadisticas['asistencias_combinadas'] += 1
                                    
                                    if created:
                                        logger.info(f"✅ Asistencia creada para {estudiante.username}")
                                    else:
                                        logger.info(f"✅ Asistencia actualizada para {estudiante.username}")
                            else:
                                logger.warning(f"⚠️ Usuario no encontrado para asistencia: student_id={student_id_real}")
                            
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            logger.error(f"Error procesando asistencia: {e}")
                            continue
            
            # PASO 8: COMBINAR CALIFICACIONES
            actualizar_progreso('Combinando calificaciones...', 8, **estadisticas)
            logger.info("=== Combinando calificaciones ===")
            
            # Buscar tablas de calificaciones
            tablas_calificaciones = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen__icontains='calificacion') |
                Q(tabla_origen__icontains='grade') |
                Q(tabla_origen__icontains='qualification') |
                Q(tabla_origen='Docencia_grade') |
                Q(tabla_origen='principal_calificaciones')
            ).values_list('tabla_origen', flat=True).distinct()
            
            for tabla_calificacion in tablas_calificaciones:
                datos_calificaciones = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_calificacion)
                logger.info(f"Procesando {datos_calificaciones.count()} calificaciones de tabla {tabla_calificacion}")
                
                with transaction.atomic():
                    for dato in datos_calificaciones:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            # USAR IDs REALES DE LOS DATOS
                            student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                            course_id_real = datos.get('course_id')  # FK a course.id
                            
                            if student_id_real in mapeo_usuarios:
                                estudiante = mapeo_usuarios[student_id_real]
                                
                                # Buscar curso usando mapeo correcto
                                curso = None
                                if course_id_real and course_id_real in mapeo_cursos:
                                    curso = mapeo_cursos[course_id_real]
                                else:
                                    curso = Curso.objects.first()  # Fallback
                                
                                if curso:
                                    # Crear o actualizar calificación
                                    calificacion_data = {
                                        'student': estudiante,
                                        'course': curso,
                                        'nota_final': datos.get('nota_final', datos.get('final_grade', 0)),
                                        'estado': datos.get('estado', datos.get('status', 'P')),
                                        'activo': datos.get('activo', datos.get('active', True))
                                    }
                                    
                                    calificacion, created = Calificaciones.objects.get_or_create(
                                        student=estudiante,
                                        course=curso,
                                        defaults=calificacion_data
                                    )
                                    
                                    if not created:
                                        # Actualizar calificación existente
                                        copiar_campos_dinamicos(calificacion, datos, 
                                                              campos_excluir=['id', 'pk', 'student', 'course', 'student_id', 'course_id'], 
                                                              logger=logger)
                                        calificacion.save()
                                    
                                    estadisticas['calificaciones_combinadas'] += 1
                                    
                                    if created:
                                        logger.info(f"✅ Calificación creada para {estudiante.username}")
                                    else:
                                        logger.info(f"✅ Calificación actualizada para {estudiante.username}")
                            else:
                                logger.warning(f"⚠️ Usuario no encontrado para calificación: student_id={student_id_real}")
                            
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            logger.error(f"Error procesando calificación: {e}")
                            continue
            
            # PASO 9: COMBINAR NOTAS INDIVIDUALES
            actualizar_progreso('Combinando notas individuales...', 9, **estadisticas)
            logger.info("=== Combinando notas individuales ===")
            
            # Buscar tablas de notas individuales
            tablas_notas = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen__icontains='nota') |
                Q(tabla_origen__icontains='individual') |
                Q(tabla_origen='Docencia_individualnote') |
                Q(tabla_origen='principal_notaindividual')
            ).values_list('tabla_origen', flat=True).distinct()
            
            for tabla_nota in tablas_notas:
                datos_notas = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_nota)
                logger.info(f"Procesando {datos_notas.count()} notas individuales de tabla {tabla_nota}")
                
                with transaction.atomic():
                    for dato in datos_notas:
                        try:
                            sid = transaction.savepoint()
                            datos = dato.datos_originales
                            # USAR IDs REALES DE LOS DATOS
                            student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                            course_id_real = datos.get('course_id')  # FK a course.id
                            
                            if student_id_real in mapeo_usuarios:
                                estudiante = mapeo_usuarios[student_id_real]
                                
                                # Buscar curso usando mapeo correcto
                                curso = None
                                if course_id_real and course_id_real in mapeo_cursos:
                                    curso = mapeo_cursos[course_id_real]
                                else:
                                    curso = Curso.objects.first()  # Fallback
                                
                                if curso:
                                    # Crear o actualizar nota individual
                                    nota_data = {
                                        'student': estudiante,
                                        'course': curso,
                                        'tipo_evaluacion': datos.get('tipo_evaluacion', datos.get('evaluation_type', 'examen')),
                                        'nota': datos.get('nota', datos.get('grade', 0)),
                                        'fecha': datos.get('fecha', datos.get('date', timezone.now().date())),
                                        'activo': datos.get('activo', datos.get('active', True))
                                    }
                                    
                                    # Buscar nota existente por criterios únicos
                                    nota_existente = NotaIndividual.objects.filter(
                                        student=estudiante,
                                        course=curso,
                                        tipo_evaluacion=nota_data['tipo_evaluacion'],
                                        fecha=nota_data['fecha']
                                    ).first()
                                    
                                    if nota_existente:
                                        # Actualizar nota existente
                                        copiar_campos_dinamicos(nota_existente, datos, 
                                                              campos_excluir=['id', 'pk', 'student', 'course', 'student_id', 'course_id'], 
                                                              logger=logger)
                                        nota_existente.save()
                                        logger.info(f"✅ Nota individual actualizada para {estudiante.username}")
                                    else:
                                        # Crear nueva nota
                                        nueva_nota = NotaIndividual(**nota_data)
                                        copiar_campos_dinamicos(nueva_nota, datos, 
                                                              campos_excluir=['id', 'pk', 'student', 'course', 'tipo_evaluacion', 'nota', 'fecha', 'student_id', 'course_id'], 
                                                              logger=logger)
                                        nueva_nota.save()
                                        logger.info(f"✅ Nota individual creada para {estudiante.username}")
                                    
                                    estadisticas['notas_combinadas'] += 1
                            else:
                                logger.warning(f"⚠️ Usuario no encontrado para nota individual: student_id={student_id_real}")
                            
                            transaction.savepoint_commit(sid)
                            
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            logger.error(f"Error procesando nota individual: {e}")
                            continue
            
            # PASO 10: PROCESAR OTRAS TABLAS RESTANTES
            actualizar_progreso('Procesando otras tablas...', 10, **estadisticas)
            logger.info("=== Procesando otras tablas restantes ===")
            
            # Obtener todas las tablas que no hemos procesado específicamente
            tablas_procesadas = {
                'auth_user', 'auth_user_groups', 'auth_group', 'auth_user_user_permissions',
                'Docencia_studentpersonalinformation', 'Docencia_teacherpersonalinformation', 'accounts_registro'
            }
            
            # Agregar tablas de cursos, matrículas, etc. que ya procesamos
            tablas_procesadas.update(tablas_cursos_academicos)
            tablas_procesadas.update(tablas_cursos)
            tablas_procesadas.update(tablas_matriculas)
            tablas_procesadas.update(tablas_asistencias)
            tablas_procesadas.update(tablas_calificaciones)
            tablas_procesadas.update(tablas_notas)
            
            # Obtener tablas restantes
            todas_las_tablas = DatoArchivadoDinamico.objects.values_list('tabla_origen', flat=True).distinct()
            tablas_restantes = [tabla for tabla in todas_las_tablas if tabla not in tablas_procesadas]
            
            logger.info(f"Tablas restantes por procesar: {tablas_restantes}")
            
            for tabla_restante in tablas_restantes:
                datos_restantes = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_restante)
                logger.info(f"Procesando {datos_restantes.count()} registros de tabla adicional: {tabla_restante}")
                
                # Para tablas no específicas, solo contamos los registros
                estadisticas['otras_tablas'] += datos_restantes.count()
                logger.info(f"Tabla {tabla_restante} registrada como procesada ({datos_restantes.count()} registros)")
            
            # PASO 11: ACTUALIZAR PROGRESO FINAL
            actualizar_progreso('Combinación completada exitosamente', 11, **estadisticas)
            
            # Marcar como completado
            resultado_final = {
                'fecha_inicio': timezone.now().isoformat(),
                'fecha_fin': timezone.now().isoformat(),
                **estadisticas,
                'campos_agregados': estadisticas.get('campos_agregados', 0)
            }
            cache.set('ultima_combinacion_completada', resultado_final, timeout=300)
            cache.delete('combinacion_en_progreso')
            
            logger.info("=== COMBINACIÓN REAL COMPLETADA EXITOSAMENTE ===")
            logger.info(f"Usuarios combinados: {estadisticas['usuarios_combinados']}")
            logger.info(f"Registros combinados: {estadisticas['registros_combinados']}")
            logger.info(f"Cursos académicos combinados: {estadisticas['cursos_academicos_combinados']}")
            logger.info(f"Cursos combinados: {estadisticas['cursos_combinados']}")
            logger.info(f"Matrículas combinadas: {estadisticas['matriculas_combinadas']}")
            logger.info(f"Asistencias combinadas: {estadisticas['asistencias_combinadas']}")
            logger.info(f"Calificaciones combinadas: {estadisticas['calificaciones_combinadas']}")
            logger.info(f"Notas individuales combinadas: {estadisticas['notas_combinadas']}")
            logger.info(f"Otras tablas procesadas: {estadisticas['otras_tablas']}")
                
        except InterruptedError as e:
            # Combinación interrumpida por el usuario
            logger.info(f"🛑 Combinación interrumpida: {str(e)}")
            
            cache.delete('combinacion_en_progreso')
            
            # Guardar estado de interrupción en cache
            interrupcion_info = {
                'estado': 'interrumpida',
                'mensaje': 'Combinación interrumpida por el usuario',
                'fecha_interrupcion': timezone.now().isoformat()
            }
            cache.set('combinacion_interrumpida_info', interrupcion_info, timeout=300)
            
        except Exception as e:
            # En caso de error, limpiar cache y registrar error
            logger.error(f"Error en combinación real: {str(e)}")
            logger.error(f"Traceback: ", exc_info=True)
            
            cache.delete('combinacion_en_progreso')
            
            # Guardar error en cache para mostrar en frontend
            error_info = {
                'estado': 'error',
                'mensaje': str(e),
                'fecha_error': timezone.now().isoformat()
            }
            cache.set('combinacion_error', error_info, timeout=300)
            raise
    
    # Ejecutar en hilo separado
    import threading
    import time
    
    def wrapper():
        # Pequeña pausa para asegurar que la respuesta se envíe primero
        time.sleep(0.1)
        ejecutar_combinacion()
    
    thread = threading.Thread(target=wrapper)
    thread.daemon = True
    thread.start()
    
    # Respuesta inmediata
    return JsonResponse({
        'success': True, 
        'message': 'Combinación iniciada. Puede seguir el progreso en tiempo real.'
    })
                        

@login_required
def estado_combinacion_ajax(request):
    """Vista AJAX para obtener el estado de la combinación actual"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from django.core.cache import cache
        from django.utils import timezone
        from datetime import timedelta
        
        # Buscar combinación en progreso
        progreso = cache.get('combinacion_en_progreso')
        
        # Verificar si la combinación fue interrumpida
        interrumpida = cache.get('combinacion_interrumpida')
        estado_interrupcion = cache.get('combinacion_estado_interrupcion')
        
        if interrumpida:
            # Limpiar cache de progreso si está interrumpida
            cache.delete('combinacion_en_progreso')
            
            data = {
                'en_progreso': False,
                'completada_recientemente': False,
                'estado': 'interrumpida',
                'progreso_real': 0,
                'mensaje': 'Combinación interrumpida por el usuario',
                'interrumpida_por': estado_interrupcion.get('interrumpida_por', 'Usuario') if estado_interrupcion else 'Usuario',
                'timestamp_interrupcion': estado_interrupcion.get('timestamp') if estado_interrupcion else timezone.now().isoformat()
            }
            return JsonResponse(data)
        
        if progreso:
            # Verificar si el progreso está "trabado" (sin cambios por mucho tiempo)
            fecha_inicio_str = progreso.get('fecha_inicio')
            if fecha_inicio_str:
                try:
                    from dateutil import parser
                    fecha_inicio = parser.parse(fecha_inicio_str)
                    tiempo_transcurrido = timezone.now() - fecha_inicio
                    
                    # Si han pasado más de 20 minutos, considerar el proceso trabado
                    if tiempo_transcurrido.total_seconds() > 1200:  # 20 minutos
                        logger.warning(f"Proceso de combinación trabado por {tiempo_transcurrido.total_seconds()} segundos")
                        # Limpiar cache y marcar como error
                        cache.delete('combinacion_en_progreso')
                        cache.set('combinacion_error', {
                            'mensaje': f'El proceso se trabó después de {int(tiempo_transcurrido.total_seconds()/60)} minutos. Esto puede deberse a un gran volumen de datos o problemas de conectividad.',
                            'fecha_error': timezone.now().isoformat(),
                            'tiempo_transcurrido': str(tiempo_transcurrido)
                        }, timeout=300)
                        progreso = None
                except Exception as e:
                    logger.error(f"Error verificando tiempo transcurrido: {e}")
        
        if progreso:
            # Calcular progreso real basado en pasos completados
            pasos_completados = progreso.get('pasos_completados', 0)
            pasos_totales = progreso.get('pasos_totales', 8)
            progreso_real = min(int((pasos_completados / pasos_totales) * 100), 100) if pasos_totales > 0 else 0
            
            data = {
                'en_progreso': True,
                'estado': progreso.get('paso_actual', 'Procesando...'),
                'progreso_real': progreso_real,
                'fecha_inicio': progreso.get('fecha_inicio'),
                'usuarios_combinados': progreso.get('usuarios_combinados', 0),
                'registros_combinados': progreso.get('registros_combinados', 0),
                'total_combinados': progreso.get('registros_combinados', 0),  # Alias para compatibilidad
                'cursos_academicos_combinados': progreso.get('cursos_academicos_combinados', 0),
                'cursos_combinados': progreso.get('cursos_combinados', 0),
                'matriculas_combinadas': progreso.get('matriculas_combinadas', 0),
                'asistencias_combinadas': progreso.get('asistencias_combinadas', 0),
                'calificaciones_combinadas': progreso.get('calificaciones_combinadas', 0),
                'notas_combinadas': progreso.get('notas_combinadas', 0),
                'otras_tablas': progreso.get('otras_tablas', 0),
                'pasos_completados': pasos_completados,
                'pasos_totales': pasos_totales,
                # Campos adicionales para progreso en tiempo real
                'tabla_actual_procesando': progreso.get('tabla_actual_procesando', ''),
                'registros_procesados_tabla': progreso.get('registros_procesados_tabla', 0),
                'total_registros_tabla': progreso.get('total_registros_tabla', 0),
                'porcentaje_tabla': progreso.get('porcentaje_tabla', 0),
            }
        else:
            # Verificar si hay una combinación completada recientemente (últimos 10 minutos)
            resultado_final = cache.get('ultima_combinacion_completada')
            error_info = cache.get('combinacion_error')
            
            if error_info:
                data = {
                    'en_progreso': False,
                    'completada_recientemente': False,
                    'estado': 'error',
                    'progreso_real': 0,
                    'mensaje': error_info.get('mensaje', 'Error desconocido'),
                    'fecha_error': error_info.get('fecha_error'),
                }
            elif resultado_final:
                data = {
                    'en_progreso': False,
                    'completada_recientemente': True,
                    'estado': 'Combinación completada exitosamente',
                    'progreso_real': 100,
                    'fecha_inicio': resultado_final.get('fecha_inicio'),
                    'fecha_fin': resultado_final.get('fecha_fin'),
                    'usuarios_combinados': resultado_final.get('usuarios_combinados', 0),
                    'registros_combinados': resultado_final.get('registros_combinados', 0),
                    'total_combinados': resultado_final.get('registros_combinados', 0),  # Alias para compatibilidad
                    'cursos_academicos_combinados': resultado_final.get('cursos_academicos_combinados', 0),
                    'cursos_combinados': resultado_final.get('cursos_combinados', 0),
                    'matriculas_combinadas': resultado_final.get('matriculas_combinadas', 0),
                    'asistencias_combinadas': resultado_final.get('asistencias_combinadas', 0),
                    'calificaciones_combinadas': resultado_final.get('calificaciones_combinadas', 0),
                    'notas_combinadas': resultado_final.get('notas_combinadas', 0),
                    'otras_tablas': resultado_final.get('otras_tablas', 0),
                    'campos_agregados': resultado_final.get('campos_agregados', 0),
                    # Campos de progreso final
                    'tabla_actual_procesando': 'Completada',
                    'registros_procesados_tabla': resultado_final.get('registros_combinados', 0),
                    'total_registros_tabla': resultado_final.get('registros_combinados', 0),
                    'porcentaje_tabla': 100,
                }
            else:
                data = {
                    'en_progreso': False,
                    'completada_recientemente': False,
                    'estado': 'sin_combinaciones',
                    'progreso_real': 0,
                    'mensaje': 'No hay combinaciones registradas'
                }
        
        return JsonResponse(data)
    except Exception as e:
        # Log del error para debugging
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error en estado_combinacion_ajax: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Respuesta de error segura
        return JsonResponse({
            'error': 'Error interno del servidor',
            'en_progreso': False,
            'completada_recientemente': False,
            'estado': 'error',
            'progreso_real': 0,
            'mensaje': 'Error al obtener estado de combinación'
        })


@login_required
def interrumpir_combinacion_ajax(request):
    """Vista AJAX para interrumpir la combinación en curso"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({
            'success': False,
            'error': 'No tienes permisos para interrumpir combinaciones'
        }, status=403)
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        import json
        import logging
        from django.utils import timezone
        
        logger = logging.getLogger(__name__)
        data = json.loads(request.body)
        tipo_combinacion = data.get('tipo', 'completa')
        
        # Limpiar el cache de combinación para detener el proceso
        from django.core.cache import cache
        
        # Marcar como interrumpida
        cache.set('combinacion_interrumpida', True, 300)  # 5 minutos
        cache.set('combinacion_estado_interrupcion', {
            'interrumpida_por': request.user.username,
            'timestamp': timezone.now().isoformat(),
            'tipo': tipo_combinacion
        }, 300)
        
        # Limpiar estados de progreso
        cache.delete('combinacion_en_progreso')
        cache.delete('combinacion_progreso')
        cache.delete('combinacion_estado')
        cache.delete('combinacion_completada')
        
        logger.info(f"🛑 Combinación {tipo_combinacion} interrumpida por usuario {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'mensaje': f'Combinación {tipo_combinacion} interrumpida exitosamente',
            'interrumpida_por': request.user.username,
            'timestamp': timezone.now().isoformat()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        }, status=400)
    except Exception as e:
        logger.error(f"❌ Error al interrumpir combinación: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }, status=500)


@login_required
def combinar_datos_seleccionadas(request):
    """Vista para combinar SOLO las tablas seleccionadas por el usuario
    
    Esta función permite al usuario elegir qué tablas específicas combinar
    en lugar de combinar todas las tablas archivadas
    """
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        # Obtener las tablas seleccionadas del request
        data = json.loads(request.body)
        tablas_seleccionadas = data.get('tablas_seleccionadas', [])
        
        if not tablas_seleccionadas:
            return JsonResponse({'success': False, 'error': 'No se seleccionaron tablas para combinar'})
        
        # Validar que las tablas existen
        from .models import DatoArchivadoDinamico
        tablas_disponibles = DatoArchivadoDinamico.objects.values_list('tabla_origen', flat=True).distinct()
        
        for tabla in tablas_seleccionadas:
            if tabla not in tablas_disponibles:
                return JsonResponse({'success': False, 'error': f'La tabla {tabla} no existe en los datos archivados'})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error procesando solicitud: {str(e)}'})
    
    # Ejecutar combinación en hilo separado para no bloquear la respuesta
    def ejecutar_combinacion_seleccionada():
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            logger.info(f"=== INICIANDO COMBINACIÓN SELECTIVA DE DATOS ===")
            logger.info(f"Tablas seleccionadas: {tablas_seleccionadas}")
            
            from .models import DatoArchivadoDinamico
            from django.contrib.auth.models import User, Group
            from accounts.models import Registro
            from principal.models import (
                CursoAcademico, Curso, Matriculas, 
                Asistencia, Calificaciones, NotaIndividual
            )
            from django.db import transaction, IntegrityError
            from django.db.models import Q
            from datetime import datetime, date
            from decimal import Decimal
            from django.core.cache import cache
            from django.utils import timezone
            
            # Limpiar cualquier flag de interrupción residual al inicio
            cache.delete('combinacion_interrumpida')
            cache.delete('combinacion_estado_interrupcion')
            cache.delete('combinacion_interrumpida_info')
            logger.info("🧹 Cache de interrupción limpiado al inicio de combinación selectiva")
            
            # Inicializar progreso en cache (usar clave diferente para no interferir)
            def actualizar_progreso_selectivo(paso_actual, pasos_completados, pasos_totales, **kwargs):
                # Verificar si la combinación ha sido interrumpida
                if cache.get('combinacion_interrumpida'):
                    logger.info("🛑 Combinación selectiva interrumpida por el usuario - deteniendo proceso")
                    cache.set('combinacion_estado', 'interrumpida', timeout=300)
                    raise InterruptedError("Combinación selectiva interrumpida por el usuario")
                
                progreso = {
                    'paso_actual': paso_actual,
                    'pasos_completados': pasos_completados,
                    'pasos_totales': pasos_totales,
                    'fecha_inicio': timezone.now().isoformat(),
                    'usuarios_combinados': kwargs.get('usuarios_combinados', 0),
                    'registros_combinados': kwargs.get('registros_combinados', 0),
                    'cursos_academicos_combinados': kwargs.get('cursos_academicos_combinados', 0),
                    'cursos_combinados': kwargs.get('cursos_combinados', 0),
                    'matriculas_combinadas': kwargs.get('matriculas_combinadas', 0),
                    'asistencias_combinadas': kwargs.get('asistencias_combinadas', 0),
                    'calificaciones_combinadas': kwargs.get('calificaciones_combinadas', 0),
                    'notas_combinadas': kwargs.get('notas_combinadas', 0),
                    'otras_tablas': kwargs.get('otras_tablas', 0),
                    'tablas_seleccionadas': tablas_seleccionadas,
                    'tipo_combinacion': 'selectiva'
                }
                cache.set('combinacion_en_progreso', progreso, timeout=300)  # 5 minutos
                logger.info(f"Progreso selectivo actualizado: {paso_actual} ({pasos_completados}/{pasos_totales})")
            
            # Función helper para actualizar progreso en tiempo real durante procesamiento de registros
            def actualizar_progreso_tiempo_real(tabla_actual, registros_procesados, total_registros, **kwargs):
                # Verificar interrupción
                if cache.get('combinacion_interrumpida'):
                    logger.info("🛑 Combinación selectiva interrumpida por el usuario - deteniendo proceso")
                    cache.set('combinacion_estado', 'interrumpida', timeout=300)
                    raise InterruptedError("Combinación selectiva interrumpida por el usuario")
                
                # Calcular porcentaje de la tabla actual
                porcentaje_tabla = int((registros_procesados / total_registros) * 100) if total_registros > 0 else 0
                
                # Crear progreso detallado
                progreso_detallado = {
                    'paso_actual': f'Procesando {tabla_actual}: {registros_procesados}/{total_registros} registros ({porcentaje_tabla}%)',
                    'pasos_completados': kwargs.get('pasos_completados', 0),
                    'pasos_totales': kwargs.get('pasos_totales', 8),
                    'fecha_inicio': kwargs.get('fecha_inicio', timezone.now().isoformat()),
                    'usuarios_combinados': kwargs.get('usuarios_combinados', 0),
                    'registros_combinados': kwargs.get('registros_combinados', 0),
                    'cursos_academicos_combinados': kwargs.get('cursos_academicos_combinados', 0),
                    'cursos_combinados': kwargs.get('cursos_combinados', 0),
                    'matriculas_combinadas': kwargs.get('matriculas_combinadas', 0),
                    'asistencias_combinadas': kwargs.get('asistencias_combinadas', 0),
                    'calificaciones_combinadas': kwargs.get('calificaciones_combinadas', 0),
                    'notas_combinadas': kwargs.get('notas_combinadas', 0),
                    'otras_tablas': kwargs.get('otras_tablas', 0),
                    'tablas_seleccionadas': tablas_seleccionadas,
                    'tipo_combinacion': 'selectiva',
                    # Campos específicos de progreso en tiempo real
                    'tabla_actual_procesando': tabla_actual,
                    'registros_procesados_tabla': registros_procesados,
                    'total_registros_tabla': total_registros,
                    'porcentaje_tabla': porcentaje_tabla,
                }
                
                cache.set('combinacion_en_progreso', progreso_detallado, timeout=300)
                logger.info(f"📊 Progreso tiempo real: {tabla_actual} {registros_procesados}/{total_registros} ({porcentaje_tabla}%) - Total combinados: {kwargs.get('registros_combinados', 0)}")
            
            # Función para mapear campos de inglés a español
            def mapear_campos_ingles_espanol(datos_origen, logger=None):
                """Mapea campos de inglés a español y mantiene ambos"""
                mapeo_campos = {
                    'nacionality': 'nacionalidad',
                    'numberidentification': 'carnet', 
                    'phone': 'telephone',
                    'cellphone': 'movil',
                    'street': 'address',
                    'city': 'location',
                    'state': 'provincia',
                    'degree': 'grado',
                    'ocupation': 'ocupacion',
                    'title': 'titulo',
                    'gender': 'sexo',
                    'photo': 'image',
                    'isReligious': 'es_religioso',
                    'name': 'first_name',  # Para usuarios
                    'lastname': 'last_name'  # Para usuarios
                }
                
                # Crear copia con todos los datos originales
                datos_mapeados = datos_origen.copy()
                campos_mapeados_count = 0
                
                # Aplicar mapeo: agregar campos traducidos
                for campo_origen, campo_destino in mapeo_campos.items():
                    if campo_origen in datos_origen:
                        datos_mapeados[campo_destino] = datos_origen[campo_origen]
                        campos_mapeados_count += 1
                        if logger:
                            logger.debug(f"🔄 Mapeando: {campo_origen} → {campo_destino} = {datos_origen[campo_origen]}")
                
                if logger and campos_mapeados_count > 0:
                    logger.info(f"📋 Total de campos mapeados: {campos_mapeados_count}")
                
                return datos_mapeados
            
            # Función para copiar campos dinámicamente (reutilizada)
            def copiar_campos_dinamicos(objeto_destino, datos_origen, campos_excluir=None, logger=None):
                if campos_excluir is None:
                    campos_excluir = ['id', 'pk']
                
                campos_copiados = 0
                for campo, valor in datos_origen.items():
                    if campo in campos_excluir:
                        continue
                    
                    try:
                        if hasattr(objeto_destino, campo):
                            # Convertir valores especiales
                            if valor == 'NULL' or valor == 'null':
                                valor = None
                            elif isinstance(valor, str) and valor.lower() in ['true', 'false']:
                                valor = valor.lower() == 'true'
                            elif campo in ['last_login', 'date_joined'] and valor:
                                # Manejar campos de fecha con timezone
                                from django.utils import timezone
                                from datetime import datetime
                                try:
                                    if isinstance(valor, str):
                                        # Parsear la fecha string
                                        fecha_naive = datetime.fromisoformat(valor.replace('Z', ''))
                                    elif isinstance(valor, datetime):
                                        fecha_naive = valor
                                    else:
                                        fecha_naive = None
                                    
                                    if fecha_naive and timezone.is_naive(fecha_naive):
                                        valor = timezone.make_aware(fecha_naive)
                                    elif fecha_naive:
                                        valor = fecha_naive
                                except Exception as e:
                                    if logger:
                                        logger.warning(f"Error convirtiendo fecha {campo}: {e}")
                                    valor = None
                            
                            setattr(objeto_destino, campo, valor)
                            campos_copiados += 1
                            if logger:
                                logger.debug(f"Campo copiado: {campo} = {valor}")
                    except Exception as e:
                        if logger:
                            logger.warning(f"Error copiando campo {campo}: {e}")
                
                return campos_copiados
            
            # Contadores
            estadisticas = {
                'usuarios_combinados': 0,
                'registros_combinados': 0,
                'cursos_academicos_combinados': 0,
                'cursos_combinados': 0,
                'matriculas_combinadas': 0,
                'asistencias_combinadas': 0,
                'calificaciones_combinadas': 0,
                'notas_combinadas': 0,
                'otras_tablas': 0
            }
            
            # Mapeo de usuarios archivados a usuarios actuales
            mapeo_usuarios = {}
            
            # Mapeo de grupos archivados a grupos actuales
            mapeo_grupos = {}
            
            # Variables para progreso en tiempo real
            fecha_inicio_proceso = timezone.now().isoformat()
            
            # Calcular pasos totales basado en tablas seleccionadas
            pasos_totales = len(tablas_seleccionadas) + 1  # +1 para finalización
            paso_actual = 0
            
            # Inicializar progreso
            actualizar_progreso_selectivo('Iniciando combinación selectiva...', 0, pasos_totales, **estadisticas)
            
            # ASEGURAR QUE auth_user SE PROCESE PRIMERO SI HAY TABLAS DEPENDIENTES
            tablas_dependientes = ['Docencia_studentpersonalinformation', 'Docencia_teacherpersonalinformation', 'accounts_registro']
            necesita_usuarios = any(tabla in tablas_dependientes for tabla in tablas_seleccionadas)
            
            # ASEGURAR QUE auth_group SE PROCESE PRIMERO SI HAY auth_user_groups
            necesita_grupos = 'auth_user_groups' in tablas_seleccionadas
            
            # FORZAR ORDEN CORRECTO DE PROCESAMIENTO SIEMPRE
            orden_correcto = [
                'auth_group',  # Grupos primero
                'auth_user',
                'auth_user_groups',  # Relaciones usuario-grupo después de usuarios y grupos
                'Docencia_teacherpersonalinformation',  # Profesores primero
                'Docencia_studentpersonalinformation',  # Estudiantes después
                'accounts_registro'
            ]
            
            # Crear lista ordenada con las tablas a procesar
            tablas_a_procesar = []
            
            # PASO 1: Si necesita usuarios pero auth_user no está seleccionado, agregarlo automáticamente
            if necesita_usuarios and 'auth_user' not in tablas_seleccionadas:
                tablas_a_procesar.append('auth_user')
                logger.info("✅ Se agregó auth_user automáticamente porque hay tablas que dependen de usuarios")
            
            # PASO 1.5: Si necesita grupos pero auth_group no está seleccionado, agregarlo automáticamente
            if necesita_grupos and 'auth_group' not in tablas_seleccionadas:
                tablas_a_procesar.append('auth_group')
                logger.info("✅ Se agregó auth_group automáticamente porque se seleccionó auth_user_groups")
            
            # PASO 2: Agregar tablas en el orden correcto SOLO si están seleccionadas
            for tabla_ordenada in orden_correcto:
                if tabla_ordenada in tablas_seleccionadas:
                    tablas_a_procesar.append(tabla_ordenada)
            
            # PASO 3: Agregar cualquier otra tabla que no esté en el orden predefinido
            for tabla in tablas_seleccionadas:
                if tabla not in orden_correcto:
                    tablas_a_procesar.append(tabla)
            
            # PASO 4: Recalcular pasos totales basado en las tablas finales a procesar
            pasos_totales = len(tablas_a_procesar) + 1  # +1 para finalización
            
            logger.info(f"📋 Tablas originalmente seleccionadas: {tablas_seleccionadas}")
            logger.info(f"🔄 Orden final de procesamiento: {tablas_a_procesar}")
            logger.info(f"📊 Pasos totales calculados: {pasos_totales}")
            
            # PROCESAR CADA TABLA EN EL ORDEN CORRECTO
            for tabla in tablas_a_procesar:
                paso_actual += 1
                actualizar_progreso_selectivo(f'Procesando tabla: {tabla}', paso_actual, pasos_totales, **estadisticas)
                logger.info(f"=== Procesando tabla: {tabla} ===")
                
                # Obtener datos de la tabla
                datos_tabla = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla)
                logger.info(f"Encontrados {datos_tabla.count()} registros en {tabla}")
                
                if tabla == 'auth_user':
                    # PROCESAR USUARIOS
                    total_usuarios = len(datos_tabla)
                    logger.info(f"📊 Iniciando procesamiento de {total_usuarios} usuarios")
                    
                    with transaction.atomic():
                        for i, dato in enumerate(datos_tabla, 1):
                            try:
                                sid = transaction.savepoint()
                                datos = dato.datos_originales
                                username = datos.get('username', '')
                                email = datos.get('email', '')
                                id_original = dato.id_original
                                
                                if not username:
                                    logger.warning(f"Usuario sin username, saltando: {datos}")
                                    transaction.savepoint_rollback(sid)
                                    continue
                                
                                # Buscar si el usuario ya existe
                                usuario_existente = User.objects.filter(username=username).first()
                                
                                if usuario_existente:
                                    logger.info(f"Actualizando usuario existente: {username}")
                                    # Actualizar usuario existente
                                    copiar_campos_dinamicos(usuario_existente, datos, 
                                                          campos_excluir=['id', 'pk', 'username'], 
                                                          logger=logger)
                                    
                                    # Procesar contraseña
                                    password_original = datos.get('password')
                                    if password_original and password_original.strip():
                                        if password_original.startswith(('pbkdf2_sha256', 'bcrypt', 'argon2', 'sha1', 'md5')):
                                            usuario_existente.password = password_original
                                            logger.info(f"Contraseña hasheada copiada para {username}")
                                        else:
                                            usuario_existente.set_password(password_original)
                                            logger.info(f"Contraseña en texto plano hasheada para {username}")
                                    
                                    usuario_existente.save()
                                    # USAR EL ID REAL DE LOS DATOS ORIGINALES
                                    user_id_real = datos.get('id')  # ID real de la tabla auth_user
                                    mapeo_usuarios[user_id_real] = usuario_existente
                                    logger.info(f"✅ Usuario {username} mapeado con ID real: {user_id_real}")
                                    estadisticas['usuarios_combinados'] += 1
                                    estadisticas['registros_combinados'] += 1
                                    
                                else:
                                    logger.info(f"Creando nuevo usuario: {username}")
                                    # Crear nuevo usuario
                                    nuevo_usuario = User(username=username)
                                    copiar_campos_dinamicos(nuevo_usuario, datos, 
                                                          campos_excluir=['id', 'pk'], 
                                                          logger=logger)
                                    
                                    # Procesar contraseña
                                    password_original = datos.get('password')
                                    if password_original and password_original.strip():
                                        if password_original.startswith(('pbkdf2_sha256', 'bcrypt', 'argon2', 'sha1', 'md5')):
                                            nuevo_usuario.password = password_original
                                            logger.info(f"Contraseña hasheada asignada para {username}")
                                        else:
                                            nuevo_usuario.set_password(password_original)
                                            logger.info(f"Contraseña en texto plano hasheada para {username}")
                                    else:
                                        nuevo_usuario.set_unusable_password()
                                
                                # Actualizar progreso en tiempo real cada 10 registros o al final
                                if i % 10 == 0 or i == total_usuarios:
                                    actualizar_progreso_tiempo_real(
                                        tabla_actual='auth_user',
                                        registros_procesados=i,
                                        total_registros=total_usuarios,
                                        pasos_completados=paso_actual,
                                        pasos_totales=pasos_totales,
                                        fecha_inicio=fecha_inicio_proceso,
                                        **estadisticas
                                    )
                                
                                nuevo_usuario.save()
                                # USAR EL ID REAL DE LOS DATOS ORIGINALES
                                user_id_real = datos.get('id')  # ID real de la tabla auth_user
                                mapeo_usuarios[user_id_real] = nuevo_usuario
                                logger.info(f"✅ Usuario {username} mapeado con ID real: {user_id_real}")
                                estadisticas['usuarios_combinados'] += 1
                                
                                transaction.savepoint_commit(sid)
                                
                            except InterruptedError:
                                # Interrupción por el usuario - detener completamente el procesamiento
                                transaction.savepoint_rollback(sid)
                                logger.info("🛑 Procesamiento selectivo de usuarios interrumpido por el usuario")
                                raise  # Re-lanzar la excepción para detener toda la combinación
                            except Exception as e:
                                transaction.savepoint_rollback(sid)
                                error_msg = f"Error procesando usuario {username}: {str(e)}"
                                logger.error(error_msg)
                                continue
                
                elif tabla == 'auth_group':
                    # PROCESAR GRUPOS
                    total_grupos = len(datos_tabla)
                    logger.info(f"📊 Iniciando procesamiento de {total_grupos} grupos")
                    
                    with transaction.atomic():
                        for i, dato in enumerate(datos_tabla, 1):
                            try:
                                sid = transaction.savepoint()
                                datos = dato.datos_originales
                                group_id_original = datos.get('id')
                                group_name = datos.get('name')
                                
                                if not group_name:
                                    logger.warning(f"Grupo sin nombre, usando nombre genérico: Grupo_{group_id_original}")
                                    group_name = f'Grupo_{group_id_original}'
                                
                                # Verificar si ya existe un grupo con este nombre
                                grupo_existente = Group.objects.filter(name=group_name).first()
                                
                                if grupo_existente:
                                    # Si el grupo ya existe, usarlo (no crear duplicado)
                                    mapeo_grupos[group_id_original] = grupo_existente
                                    logger.info(f"Grupo existente reutilizado: {group_name} (ID original: {group_id_original})")
                                else:
                                    # Si no existe, crearlo
                                    grupo = Group.objects.create(name=group_name)
                                    mapeo_grupos[group_id_original] = grupo
                                    logger.info(f"Grupo creado: {group_name} (ID original: {group_id_original})")
                                
                                estadisticas['otras_tablas'] += 1
                                
                                # Actualizar progreso en tiempo real cada 5 registros o al final
                                if i % 5 == 0 or i == total_grupos:
                                    actualizar_progreso_tiempo_real(
                                        tabla_actual='auth_group',
                                        registros_procesados=i,
                                        total_registros=total_grupos,
                                        pasos_completados=paso_actual,
                                        pasos_totales=pasos_totales,
                                        fecha_inicio=fecha_inicio_proceso,
                                        **estadisticas
                                    )
                                
                                transaction.savepoint_commit(sid)
                                
                            except InterruptedError:
                                # Interrupción por el usuario - detener completamente el procesamiento
                                transaction.savepoint_rollback(sid)
                                logger.info("🛑 Procesamiento selectivo de grupos interrumpido por el usuario")
                                raise  # Re-lanzar la excepción para detener toda la combinación
                            except Exception as e:
                                transaction.savepoint_rollback(sid)
                                error_msg = f"Error procesando grupo: {str(e)}"
                                logger.error(error_msg)
                                continue
                
                elif tabla == 'Docencia_teacherpersonalinformation':
                    # PASO 2A: PROCESAR PROFESORES PRIMERO
                    logger.info("Procesando información de profesores...")
                    grupo_profesores, _ = Group.objects.get_or_create(name='Profesores')
                    
                    total_profesores = len(datos_tabla)
                    logger.info(f"📊 Iniciando procesamiento de {total_profesores} profesores")
                    
                    with transaction.atomic():
                        for i, dato in enumerate(datos_tabla, 1):
                            try:
                                sid = transaction.savepoint()
                                datos = dato.datos_originales
                                user_id_original = datos.get('user_id')
                                
                                if not user_id_original or user_id_original not in mapeo_usuarios:
                                    logger.warning(f"Usuario no encontrado para profesor: user_id={user_id_original}")
                                    transaction.savepoint_rollback(sid)
                                    continue
                                
                                usuario = mapeo_usuarios[user_id_original]
                                
                                # Buscar o crear registro
                                registro, created = Registro.objects.get_or_create(user=usuario)
                                
                                if created:
                                    logger.info(f"Creado nuevo registro para profesor: {usuario.username}")
                                else:
                                    logger.info(f"Actualizando registro existente para profesor: {usuario.username}")
                                
                                # Aplicar mapeo de campos de inglés a español
                                datos_mapeados = mapear_campos_ingles_espanol(datos, logger)
                                
                                # Copiar todos los campos
                                copiar_campos_dinamicos(registro, datos_mapeados, 
                                                      campos_excluir=['id', 'pk', 'user_id', 'user'], 
                                                      logger=logger)
                                
                                registro.save()
                                estadisticas['registros_combinados'] += 1
                                
                                # Asignar al grupo de profesores (limpiar otros grupos primero)
                                usuario.groups.clear()  # Limpiar grupos existentes
                                usuario.groups.add(grupo_profesores)
                                logger.info(f"Profesor {usuario.username} asignado EXCLUSIVAMENTE al grupo Profesores")
                                
                                transaction.savepoint_commit(sid)
                                
                                # Actualizar progreso en tiempo real cada 5 registros o al final
                                if i % 5 == 0 or i == total_profesores:
                                    actualizar_progreso_tiempo_real(
                                        tabla_actual='Docencia_teacherpersonalinformation',
                                        registros_procesados=i,
                                        total_registros=total_profesores,
                                        pasos_completados=paso_actual,
                                        pasos_totales=pasos_totales,
                                        fecha_inicio=fecha_inicio_proceso,
                                        **estadisticas
                                    )
                                
                            except InterruptedError:
                                # Interrupción por el usuario - detener completamente el procesamiento
                                transaction.savepoint_rollback(sid)
                                logger.info("🛑 Procesamiento selectivo de profesores interrumpido por el usuario")
                                raise  # Re-lanzar la excepción para detener toda la combinación
                            except Exception as e:
                                transaction.savepoint_rollback(sid)
                                error_msg = f"Error procesando profesor: {str(e)}"
                                logger.error(error_msg)
                                continue
                
                elif tabla == 'Docencia_studentpersonalinformation':
                    # PASO 2B: PROCESAR ESTUDIANTES DESPUÉS
                    logger.info("Procesando información de estudiantes...")
                    grupo_estudiantes, _ = Group.objects.get_or_create(name='Estudiantes')
                    
                    total_estudiantes = len(datos_tabla)
                    logger.info(f"📊 Iniciando procesamiento de {total_estudiantes} estudiantes")
                    
                    with transaction.atomic():
                        for i, dato in enumerate(datos_tabla, 1):
                            try:
                                sid = transaction.savepoint()
                                datos = dato.datos_originales
                                user_id_original = datos.get('user_id')
                                
                                if not user_id_original or user_id_original not in mapeo_usuarios:
                                    logger.warning(f"Usuario no encontrado para estudiante: user_id={user_id_original}")
                                    transaction.savepoint_rollback(sid)
                                    continue
                                
                                usuario = mapeo_usuarios[user_id_original]
                                
                                # Buscar o crear registro
                                registro, created = Registro.objects.get_or_create(user=usuario)
                                
                                if created:
                                    logger.info(f"Creado nuevo registro para estudiante: {usuario.username}")
                                else:
                                    logger.info(f"Actualizando registro existente para estudiante: {usuario.username}")
                                
                                # Aplicar mapeo de campos de inglés a español
                                datos_mapeados = mapear_campos_ingles_espanol(datos, logger)
                                
                                # Copiar todos los campos
                                copiar_campos_dinamicos(registro, datos_mapeados, 
                                                      campos_excluir=['id', 'pk', 'user_id', 'user'], 
                                                      logger=logger)
                                
                                registro.save()
                                estadisticas['registros_combinados'] += 1
                                
                                # Asignar al grupo de estudiantes (limpiar otros grupos primero)
                                usuario.groups.clear()  # Limpiar grupos existentes
                                usuario.groups.add(grupo_estudiantes)
                                logger.info(f"Estudiante {usuario.username} asignado EXCLUSIVAMENTE al grupo Estudiantes")
                                
                                transaction.savepoint_commit(sid)
                                
                                # Actualizar progreso en tiempo real cada 10 registros o al final
                                if i % 10 == 0 or i == total_estudiantes:
                                    actualizar_progreso_tiempo_real(
                                        tabla_actual='Docencia_studentpersonalinformation',
                                        registros_procesados=i,
                                        total_registros=total_estudiantes,
                                        pasos_completados=paso_actual,
                                        pasos_totales=pasos_totales,
                                        fecha_inicio=fecha_inicio_proceso,
                                        **estadisticas
                                    )
                                
                            except InterruptedError:
                                # Interrupción por el usuario - detener completamente el procesamiento
                                transaction.savepoint_rollback(sid)
                                logger.info("🛑 Procesamiento selectivo de estudiantes interrumpido por el usuario")
                                raise  # Re-lanzar la excepción para detener toda la combinación
                            except Exception as e:
                                transaction.savepoint_rollback(sid)
                                error_msg = f"Error procesando estudiante: {str(e)}"
                                logger.error(error_msg)
                                continue
                
                elif tabla == 'accounts_registro':
                    # PASO 2C: PROCESAR REGISTROS EXISTENTES
                    logger.info("Procesando registros existentes...")
                    
                    with transaction.atomic():
                        for dato in datos_tabla:
                            try:
                                sid = transaction.savepoint()
                                datos = dato.datos_originales
                                user_id_original = datos.get('user_id')
                                
                                if not user_id_original or user_id_original not in mapeo_usuarios:
                                    logger.warning(f"Usuario no encontrado para registro: user_id={user_id_original}")
                                    transaction.savepoint_rollback(sid)
                                    continue
                                
                                usuario = mapeo_usuarios[user_id_original]
                                
                                # Buscar o crear registro
                                registro, created = Registro.objects.get_or_create(user=usuario)
                                
                                if created:
                                    logger.info(f"Creado nuevo registro para usuario: {usuario.username}")
                                else:
                                    logger.info(f"Actualizando registro existente para usuario: {usuario.username}")
                                
                                # Aplicar mapeo de campos de inglés a español
                                datos_mapeados = mapear_campos_ingles_espanol(datos, logger)
                                
                                # Copiar todos los campos
                                copiar_campos_dinamicos(registro, datos_mapeados, 
                                                      campos_excluir=['id', 'pk', 'user_id', 'user'], 
                                                      logger=logger)
                                
                                registro.save()
                                estadisticas['registros_combinados'] += 1
                                
                                transaction.savepoint_commit(sid)
                                
                            except Exception as e:
                                transaction.savepoint_rollback(sid)
                                error_msg = f"Error procesando registro: {str(e)}"
                                logger.error(error_msg)
                                continue
                
                elif tabla == 'auth_user_groups':
                    # PROCESAR GRUPOS DE USUARIOS (igual que función principal)
                    total_relaciones = len(datos_tabla)
                    logger.info(f"📊 Iniciando procesamiento de {total_relaciones} relaciones usuario-grupo")
                    
                    with transaction.atomic():
                        for i, dato in enumerate(datos_tabla, 1):
                            try:
                                sid = transaction.savepoint()
                                datos = dato.datos_originales
                                user_id_original = datos.get('user_id')
                                group_id_original = datos.get('group_id')
                                
                                if user_id_original in mapeo_usuarios:
                                    usuario = mapeo_usuarios[user_id_original]
                                    
                                    # USAR EL MAPEO DE GRUPOS CREADO ANTERIORMENTE
                                    if group_id_original in mapeo_grupos:
                                        grupo = mapeo_grupos[group_id_original]
                                        usuario.groups.add(grupo)
                                        logger.info(f"Usuario {usuario.username} agregado al grupo {grupo.name}")
                                        estadisticas['otras_tablas'] += 1
                                    else:
                                        logger.warning(f"Grupo con ID {group_id_original} no encontrado en mapeo")
                                else:
                                    logger.warning(f"Usuario con ID {user_id_original} no encontrado en mapeo")
                                
                                # Actualizar progreso en tiempo real cada 10 registros o al final
                                if i % 10 == 0 or i == total_relaciones:
                                    actualizar_progreso_tiempo_real(
                                        tabla_actual='auth_user_groups',
                                        registros_procesados=i,
                                        total_registros=total_relaciones,
                                        pasos_completados=paso_actual,
                                        pasos_totales=pasos_totales,
                                        fecha_inicio=fecha_inicio_proceso,
                                        **estadisticas
                                    )
                                
                                transaction.savepoint_commit(sid)
                                
                            except Exception as e:
                                transaction.savepoint_rollback(sid)
                                logger.error(f"Error procesando grupo de usuario: {e}")
                                continue
                
                else:
                    # PROCESAR OTRAS TABLAS CON EL MISMO MAPEO QUE LA FUNCIÓN PRINCIPAL
                    logger.info(f"Procesando tabla: {tabla}")
                    
                    # Mapeo de cursos (usando ID real de los datos)
                    mapeo_cursos = {}
                    
                    if (tabla.lower().find('academicyear') != -1 or 
                          tabla.lower().find('curso_academico') != -1 or 
                          tabla == 'Docencia_academicyear'):
                        # PROCESAR CURSOS ACADÉMICOS (igual que función principal)
                        from principal.models import CursoAcademico
                        
                        with transaction.atomic():
                            for dato in datos_tabla:
                                try:
                                    sid = transaction.savepoint()
                                    datos = dato.datos_originales
                                    
                                    # Crear o actualizar curso académico
                                    curso_academico_data = {
                                        'year': datos.get('year', datos.get('anio', datetime.now().year)),
                                        'name': datos.get('name', datos.get('nombre', f'Curso Académico {datos.get("year", datetime.now().year)}')),
                                        'activo': datos.get('activo', datos.get('active', False)),
                                        'fecha_inicio': datos.get('fecha_inicio', datos.get('start_date')),
                                        'fecha_fin': datos.get('fecha_fin', datos.get('end_date'))
                                    }
                                    
                                    # Buscar curso académico existente por año
                                    curso_academico_existente = CursoAcademico.objects.filter(year=curso_academico_data['year']).first()
                                    
                                    if curso_academico_existente:
                                        # Actualizar curso académico existente
                                        copiar_campos_dinamicos(curso_academico_existente, datos, 
                                                              campos_excluir=['id', 'pk'], 
                                                              logger=logger)
                                        curso_academico_existente.save()
                                        logger.info(f"Curso académico actualizado: {curso_academico_data['name']}")
                                    else:
                                        # Crear nuevo curso académico
                                        nuevo_curso_academico = CursoAcademico(**curso_academico_data)
                                        copiar_campos_dinamicos(nuevo_curso_academico, datos, 
                                                              campos_excluir=['id', 'pk', 'year', 'name', 'activo'], 
                                                              logger=logger)
                                        nuevo_curso_academico.save()
                                        logger.info(f"Curso académico creado: {curso_academico_data['name']}")
                                    
                                    estadisticas['cursos_academicos_combinados'] += 1
                                    transaction.savepoint_commit(sid)
                                    
                                except Exception as e:
                                    transaction.savepoint_rollback(sid)
                                    logger.error(f"Error procesando curso académico: {e}")
                                    continue
                    
                    elif (tabla.lower().find('course') != -1 or 
                          tabla == 'Docencia_course' or 
                          tabla == 'principal_curso'):
                        # PROCESAR CURSOS (igual que función principal)
                        from principal.models import Curso
                        
                        with transaction.atomic():
                            for dato in datos_tabla:
                                try:
                                    sid = transaction.savepoint()
                                    datos = dato.datos_originales
                                    curso_id_real = datos.get('id')  # ID real de la tabla de cursos
                                    
                                    # Crear o actualizar curso
                                    curso_data = {
                                        'name': datos.get('name', datos.get('nombre', f'Curso_{curso_id_real}')),
                                        'description': datos.get('description', datos.get('descripcion', '')),
                                        'area': datos.get('area', 'general'),
                                        'status': datos.get('status', datos.get('estado', 'F')),
                                        'activo': datos.get('activo', datos.get('active', False))
                                    }
                                    
                                    # Buscar curso existente por nombre
                                    curso_existente = Curso.objects.filter(name=curso_data['name']).first()
                                    
                                    if curso_existente:
                                        # Actualizar curso existente
                                        copiar_campos_dinamicos(curso_existente, datos, 
                                                              campos_excluir=['id', 'pk'], 
                                                              logger=logger)
                                        curso_existente.save()
                                        mapeo_cursos[curso_id_real] = curso_existente
                                        logger.info(f"✅ Curso actualizado: {curso_data['name']} (ID real: {curso_id_real})")
                                    else:
                                        # Crear nuevo curso
                                        nuevo_curso = Curso(**curso_data)
                                        copiar_campos_dinamicos(nuevo_curso, datos, 
                                                              campos_excluir=['id', 'pk', 'name'], 
                                                              logger=logger)
                                        nuevo_curso.save()
                                        mapeo_cursos[curso_id_real] = nuevo_curso
                                        logger.info(f"✅ Curso creado: {curso_data['name']} (ID real: {curso_id_real})")
                                    
                                    estadisticas['cursos_combinados'] += 1
                                    transaction.savepoint_commit(sid)
                                    
                                except Exception as e:
                                    transaction.savepoint_rollback(sid)
                                    logger.error(f"Error procesando curso: {e}")
                                    continue
                    
                    elif (tabla.lower().find('matricula') != -1 or 
                          tabla.lower().find('enrollment') != -1 or 
                          tabla == 'Docencia_enrollment' or 
                          tabla == 'principal_matriculas'):
                        # PROCESAR MATRÍCULAS (igual que función principal)
                        from principal.models import Matriculas, Curso, CursoAcademico
                        
                        with transaction.atomic():
                            for dato in datos_tabla:
                                try:
                                    sid = transaction.savepoint()
                                    datos = dato.datos_originales
                                    # USAR IDs REALES DE LOS DATOS, NO IDs DE DJANGO
                                    student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                                    course_id_real = datos.get('course_id')  # FK a course.id
                                    
                                    if student_id_real in mapeo_usuarios:
                                        estudiante = mapeo_usuarios[student_id_real]
                                        logger.info(f"✅ Estudiante encontrado: {estudiante.username} (ID real: {student_id_real})")
                                        
                                        # Buscar curso usando mapeo correcto
                                        curso = None
                                        if course_id_real and course_id_real in mapeo_cursos:
                                            curso = mapeo_cursos[course_id_real]
                                            logger.info(f"✅ Curso encontrado: {curso.name} (ID real: {course_id_real})")
                                        else:
                                            # Buscar por nombre si no hay mapeo directo
                                            course_name = datos.get('course_name', datos.get('nombre_curso'))
                                            if course_name:
                                                curso = Curso.objects.filter(name=course_name).first()
                                                if curso:
                                                    logger.info(f"✅ Curso encontrado por nombre: {curso.name}")
                                            
                                            if not curso:
                                                # Usar el primer curso disponible como fallback
                                                curso = Curso.objects.first()
                                                if curso:
                                                    logger.warning(f"⚠️ Usando curso fallback: {curso.name}")
                                        
                                        # Buscar curso académico
                                        curso_academico = CursoAcademico.objects.filter(activo=True).first()
                                        if not curso_academico:
                                            curso_academico = CursoAcademico.objects.first()
                                        
                                        if curso and curso_academico:
                                            matricula, created = Matriculas.objects.get_or_create(
                                                student=estudiante,
                                                course=curso,
                                                curso_academico=curso_academico,
                                                defaults={
                                                    'estado': datos.get('estado', datos.get('status', 'P')),
                                                    'activo': datos.get('activo', datos.get('active', True))
                                                }
                                            )
                                            
                                            if not created:
                                                # Actualizar matrícula existente
                                                copiar_campos_dinamicos(matricula, datos, 
                                                                      campos_excluir=['id', 'pk', 'student', 'course', 'curso_academico', 'student_id', 'course_id'], 
                                                                      logger=logger)
                                                matricula.save()
                                            
                                            if created:
                                                logger.info(f"✅ Matrícula creada para {estudiante.username} en {curso.name}")
                                            else:
                                                logger.info(f"✅ Matrícula actualizada para {estudiante.username} en {curso.name}")
                                            
                                            estadisticas['matriculas_combinadas'] += 1
                                        else:
                                            logger.warning(f"⚠️ No se pudo crear matrícula: curso={curso}, curso_academico={curso_academico}")
                                    else:
                                        logger.warning(f"⚠️ Usuario no encontrado para matrícula: student_id={student_id_real}")
                                        logger.warning(f"IDs de usuarios disponibles: {list(mapeo_usuarios.keys())}")
                                    
                                    transaction.savepoint_commit(sid)
                                    
                                except Exception as e:
                                    transaction.savepoint_rollback(sid)
                                    logger.error(f"Error procesando matrícula: {e}")
                                    continue
                    
                    elif (tabla.lower().find('asistencia') != -1 or 
                          tabla.lower().find('attendance') != -1 or 
                          tabla == 'Docencia_attendance' or 
                          tabla == 'principal_asistencia'):
                        # PROCESAR ASISTENCIAS (igual que función principal)
                        from principal.models import Asistencia, Curso
                        
                        with transaction.atomic():
                            for dato in datos_tabla:
                                try:
                                    sid = transaction.savepoint()
                                    datos = dato.datos_originales
                                    # USAR IDs REALES DE LOS DATOS
                                    student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                                    course_id_real = datos.get('course_id')  # FK a course.id
                                    
                                    if student_id_real in mapeo_usuarios:
                                        estudiante = mapeo_usuarios[student_id_real]
                                        
                                        # Buscar curso usando mapeo correcto
                                        curso = None
                                        if course_id_real and course_id_real in mapeo_cursos:
                                            curso = mapeo_cursos[course_id_real]
                                        else:
                                            curso = Curso.objects.first()  # Fallback
                                        
                                        if curso:
                                            # Crear o actualizar asistencia
                                            fecha_asistencia = datos.get('fecha', datos.get('date', timezone.now().date()))
                                            
                                            asistencia, created = Asistencia.objects.get_or_create(
                                                student=estudiante,
                                                course=curso,
                                                fecha=fecha_asistencia,
                                                defaults={
                                                    'presente': datos.get('presente', datos.get('present', True)),
                                                    'justificada': datos.get('justificada', datos.get('justified', False))
                                                }
                                            )
                                            
                                            if not created:
                                                # Actualizar asistencia existente
                                                copiar_campos_dinamicos(asistencia, datos, 
                                                                      campos_excluir=['id', 'pk', 'student', 'course', 'fecha', 'student_id', 'course_id'], 
                                                                      logger=logger)
                                                asistencia.save()
                                            
                                            estadisticas['asistencias_combinadas'] += 1
                                            
                                            if created:
                                                logger.info(f"✅ Asistencia creada para {estudiante.username}")
                                            else:
                                                logger.info(f"✅ Asistencia actualizada para {estudiante.username}")
                                    else:
                                        logger.warning(f"⚠️ Usuario no encontrado para asistencia: student_id={student_id_real}")
                                    
                                    transaction.savepoint_commit(sid)
                                    
                                except Exception as e:
                                    transaction.savepoint_rollback(sid)
                                    logger.error(f"Error procesando asistencia: {e}")
                                    continue
                    
                    elif (tabla.lower().find('calificacion') != -1 or 
                          tabla.lower().find('grade') != -1 or 
                          tabla.lower().find('qualification') != -1 or 
                          tabla == 'Docencia_grade' or 
                          tabla == 'principal_calificaciones'):
                        # PROCESAR CALIFICACIONES (igual que función principal)
                        from principal.models import Calificaciones, Curso
                        
                        with transaction.atomic():
                            for dato in datos_tabla:
                                try:
                                    sid = transaction.savepoint()
                                    datos = dato.datos_originales
                                    # USAR IDs REALES DE LOS DATOS
                                    student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                                    course_id_real = datos.get('course_id')  # FK a course.id
                                    
                                    if student_id_real in mapeo_usuarios:
                                        estudiante = mapeo_usuarios[student_id_real]
                                        
                                        # Buscar curso usando mapeo correcto
                                        curso = None
                                        if course_id_real and course_id_real in mapeo_cursos:
                                            curso = mapeo_cursos[course_id_real]
                                        else:
                                            curso = Curso.objects.first()  # Fallback
                                        
                                        if curso:
                                            # Crear o actualizar calificación
                                            calificacion_data = {
                                                'student': estudiante,
                                                'course': curso,
                                                'nota_final': datos.get('nota_final', datos.get('final_grade', 0)),
                                                'estado': datos.get('estado', datos.get('status', 'P')),
                                                'activo': datos.get('activo', datos.get('active', True))
                                            }
                                            
                                            calificacion, created = Calificaciones.objects.get_or_create(
                                                student=estudiante,
                                                course=curso,
                                                defaults=calificacion_data
                                            )
                                            
                                            if not created:
                                                # Actualizar calificación existente
                                                copiar_campos_dinamicos(calificacion, datos, 
                                                                      campos_excluir=['id', 'pk', 'student', 'course', 'student_id', 'course_id'], 
                                                                      logger=logger)
                                                calificacion.save()
                                            
                                            estadisticas['calificaciones_combinadas'] += 1
                                            
                                            if created:
                                                logger.info(f"✅ Calificación creada para {estudiante.username}")
                                            else:
                                                logger.info(f"✅ Calificación actualizada para {estudiante.username}")
                                    else:
                                        logger.warning(f"⚠️ Usuario no encontrado para calificación: student_id={student_id_real}")
                                    
                                    transaction.savepoint_commit(sid)
                                    
                                except Exception as e:
                                    transaction.savepoint_rollback(sid)
                                    logger.error(f"Error procesando calificación: {e}")
                                    continue
                    
                    elif (tabla.lower().find('nota') != -1 or 
                          tabla.lower().find('individual') != -1 or 
                          tabla == 'Docencia_individualnote' or 
                          tabla == 'principal_notaindividual'):
                        # PROCESAR NOTAS INDIVIDUALES (igual que función principal)
                        from principal.models import NotaIndividual, Curso
                        
                        with transaction.atomic():
                            for dato in datos_tabla:
                                try:
                                    sid = transaction.savepoint()
                                    datos = dato.datos_originales
                                    # USAR IDs REALES DE LOS DATOS
                                    student_id_real = datos.get('student_id', datos.get('user_id'))  # FK a auth_user.id
                                    course_id_real = datos.get('course_id')  # FK a course.id
                                    
                                    if student_id_real in mapeo_usuarios:
                                        estudiante = mapeo_usuarios[student_id_real]
                                        
                                        # Buscar curso usando mapeo correcto
                                        curso = None
                                        if course_id_real and course_id_real in mapeo_cursos:
                                            curso = mapeo_cursos[course_id_real]
                                        else:
                                            curso = Curso.objects.first()  # Fallback
                                        
                                        if curso:
                                            # Crear o actualizar nota individual
                                            nota_data = {
                                                'student': estudiante,
                                                'course': curso,
                                                'tipo_evaluacion': datos.get('tipo_evaluacion', datos.get('evaluation_type', 'examen')),
                                                'nota': datos.get('nota', datos.get('grade', 0)),
                                                'fecha': datos.get('fecha', datos.get('date', timezone.now().date())),
                                                'activo': datos.get('activo', datos.get('active', True))
                                            }
                                            
                                            # Buscar nota existente por criterios únicos
                                            nota_existente = NotaIndividual.objects.filter(
                                                student=estudiante,
                                                course=curso,
                                                tipo_evaluacion=nota_data['tipo_evaluacion'],
                                                fecha=nota_data['fecha']
                                            ).first()
                                            
                                            if nota_existente:
                                                # Actualizar nota existente
                                                copiar_campos_dinamicos(nota_existente, datos, 
                                                                      campos_excluir=['id', 'pk', 'student', 'course', 'student_id', 'course_id'], 
                                                                      logger=logger)
                                                nota_existente.save()
                                                logger.info(f"✅ Nota individual actualizada para {estudiante.username}")
                                            else:
                                                # Crear nueva nota
                                                nueva_nota = NotaIndividual(**nota_data)
                                                copiar_campos_dinamicos(nueva_nota, datos, 
                                                                      campos_excluir=['id', 'pk', 'student', 'course', 'tipo_evaluacion', 'nota', 'fecha', 'student_id', 'course_id'], 
                                                                      logger=logger)
                                                nueva_nota.save()
                                                logger.info(f"✅ Nota individual creada para {estudiante.username}")
                                            
                                            estadisticas['notas_combinadas'] += 1
                                    else:
                                        logger.warning(f"⚠️ Usuario no encontrado para nota individual: student_id={student_id_real}")
                                    
                                    transaction.savepoint_commit(sid)
                                    
                                except Exception as e:
                                    transaction.savepoint_rollback(sid)
                                    logger.error(f"Error procesando nota individual: {e}")
                                    continue
                    
                    else:
                        # OTRAS TABLAS - Procesamiento genérico básico (igual que función principal)
                        logger.info(f"Tabla {tabla} procesada de forma genérica")
                        estadisticas['otras_tablas'] += datos_tabla.count()
                
                # Actualizar progreso después de cada tabla
                actualizar_progreso_selectivo(f'Tabla {tabla} completada', paso_actual, pasos_totales, **estadisticas)
            
            # PASO FINAL: COMPLETAR
            actualizar_progreso_selectivo('Combinación selectiva completada exitosamente', pasos_totales, pasos_totales, **estadisticas)
            
            # Marcar como completado
            resultado_final = {
                'fecha_inicio': timezone.now().isoformat(),
                'fecha_fin': timezone.now().isoformat(),
                'tipo_combinacion': 'selectiva',
                'tablas_procesadas': tablas_seleccionadas,
                **estadisticas,
                'campos_agregados': 0  # Por ahora no agregamos campos dinámicamente
            }
            cache.set('ultima_combinacion_completada', resultado_final, timeout=300)
            cache.delete('combinacion_en_progreso')
            
            logger.info("=== COMBINACIÓN SELECTIVA COMPLETADA EXITOSAMENTE ===")
            logger.info(f"Tablas procesadas: {tablas_seleccionadas}")
            logger.info(f"Usuarios combinados: {estadisticas['usuarios_combinados']}")
            logger.info(f"Registros combinados: {estadisticas['registros_combinados']}")
            logger.info(f"Cursos académicos combinados: {estadisticas['cursos_academicos_combinados']}")
            logger.info(f"Cursos combinados: {estadisticas['cursos_combinados']}")
            logger.info(f"Matrículas combinadas: {estadisticas['matriculas_combinadas']}")
            logger.info(f"Asistencias combinadas: {estadisticas['asistencias_combinadas']}")
            logger.info(f"Calificaciones combinadas: {estadisticas['calificaciones_combinadas']}")
            logger.info(f"Notas individuales combinadas: {estadisticas['notas_combinadas']}")
            logger.info(f"Otras tablas procesadas: {estadisticas['otras_tablas']}")
                
        except InterruptedError as e:
            # Combinación selectiva interrumpida por el usuario
            logger.info(f"🛑 Combinación selectiva interrumpida: {str(e)}")
            
            cache.delete('combinacion_en_progreso')
            
            # Guardar estado de interrupción en cache
            interrupcion_info = {
                'estado': 'interrumpida',
                'mensaje': 'Combinación selectiva interrumpida por el usuario',
                'fecha_interrupcion': timezone.now().isoformat(),
                'tipo_combinacion': 'selectiva'
            }
            cache.set('combinacion_interrumpida_info', interrupcion_info, timeout=300)
            
        except Exception as e:
            # En caso de error, limpiar cache y registrar error
            logger.error(f"Error en combinación selectiva: {str(e)}")
            logger.error(f"Traceback: ", exc_info=True)
            
            cache.delete('combinacion_en_progreso')
            
            # Guardar error en cache para mostrar en frontend
            error_info = {
                'estado': 'error',
                'mensaje': str(e),
                'fecha_error': timezone.now().isoformat(),
                'tipo_combinacion': 'selectiva'
            }
            cache.set('combinacion_error', error_info, timeout=300)
            raise
    
    # Ejecutar en hilo separado
    import threading
    import time
    
    def wrapper():
        # Pequeña pausa para asegurar que la respuesta se envíe primero
        time.sleep(0.1)
        ejecutar_combinacion_seleccionada()
    
    thread = threading.Thread(target=wrapper)
    thread.daemon = True
    thread.start()
    
    # Respuesta inmediata
    return JsonResponse({
        'success': True, 
        'message': f'Combinación selectiva iniciada para {len(tablas_seleccionadas)} tablas. Puede seguir el progreso en tiempo real.',
        'tablas_seleccionadas': tablas_seleccionadas
    })